#!/usr/bin/env python3
"""
Versión de consola del Comparador de BOMs (sin interfaz gráfica)
Para usar cuando tkinter no está disponible
"""

import pandas as pd
import PyPDF2
import re
import openpyxl
from openpyxl.styles import PatternFill, Font
import sys
from pathlib import Path
import os
import glob

class BOMComparator:
    def __init__(self):
        self.bom_401 = {}
        self.bom_403 = {}
        
    def extract_text_from_pdf(self, pdf_path):
        """Extrae texto de un archivo PDF"""
        text = ""
        try:
            with open(pdf_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                for page in pdf_reader.pages:
                    text += page.extract_text() + "\n"
        except Exception as e:
            print(f"Error leyendo PDF {pdf_path}: {e}")
        return text
    
    def parse_part_line(self, line):
        """Extrae información de una línea de parte"""
        # Patrón para líneas que contienen información de partes
        # Busca patrones como: Part Number, Descripción, Embraer Code, Quantity, etc.
        
        # Patrón para números de parte (formato similar a EMB-0213-18-9, MS3476L22-41P, etc.)
        part_pattern = r'([A-Z0-9\-\/]+(?:\-[A-Z0-9]+)*)\s+'
        
        # Patrón más específico para las líneas de BOM
        bom_pattern = r'([A-Z0-9\-\/]+)\s+([^0-9]+?)\s+(\d+)\s+([LN])\s+([LN]?)\s*(\d+)\s+([A-Z]\d+)\s+(\d+)\s+(\d+)\s+(\d+)\s+(\d+)\s+(ST|UM)\s+([0-9,\.]+)\s*\(?\d*\.?\d*\)?\s*([0-9,\.]+)'
        
        match = re.search(bom_pattern, line)
        if match:
            return {
                'part_number': match.group(1).strip(),
                'description': match.group(2).strip(),
                'embraer_code': match.group(3),
                'status': match.group(4),
                'traceability': match.group(5),
                'position': match.group(6),
                'zone': match.group(7),
                'initial_sn': match.group(8),
                'final_sn': match.group(9),
                'quantity': int(match.group(10)),
                'unit': match.group(11),
                'rm_usage': match.group(12),
                'weight': match.group(13)
            }
        
        # Patrón alternativo más simple
        simple_pattern = r'([A-Z0-9\-\/]+)\s+(.+?)\s+(\d{6,7})\s+([LN])\s+'
        match = re.search(simple_pattern, line)
        if match:
            return {
                'part_number': match.group(1).strip(),
                'description': match.group(2).strip(),
                'embraer_code': match.group(3),
                'status': match.group(4),
                'quantity': 1,  # valor por defecto
                'unit': 'UM',
                'weight': '0'
            }
        
        return None
    
    def extract_bom_from_text(self, text, drawing_number):
        """Extrae información del BOM del texto del PDF"""
        bom_data = {}
        lines = text.split('\n')
        
        # Buscar líneas que contengan información de partes
        for line in lines:
            line = line.strip()
            if not line:
                continue
                
            # Saltar líneas de encabezado y metadatos
            if any(skip in line.upper() for skip in ['DRAWING', 'PART NUMBER', 'DESCRIPTION', 'EMBRAER', 'COPYRIGHT', 'PRINTED']):
                continue
            
            # Intentar extraer información de la parte
            part_info = self.parse_part_line(line)
            if part_info and part_info['part_number']:
                part_number = part_info['part_number']
                bom_data[part_number] = part_info
        
        return bom_data
    
    def load_boms(self, pdf_401_path, pdf_403_path):
        """Carga los BOMs de ambos PDFs"""
        print("Extrayendo texto del PDF 401...")
        text_401 = self.extract_text_from_pdf(pdf_401_path)
        self.bom_401 = self.extract_bom_from_text(text_401, "170-94864-401")
        
        print("Extrayendo texto del PDF 403...")
        text_403 = self.extract_text_from_pdf(pdf_403_path)
        self.bom_403 = self.extract_bom_from_text(text_403, "170-94864-403")
        
        print(f"BOM 401: {len(self.bom_401)} partes encontradas")
        print(f"BOM 403: {len(self.bom_403)} partes encontradas")
    
    def compare_boms(self):
        """Compara los BOMs y genera un reporte de diferencias"""
        all_parts = set(self.bom_401.keys()) | set(self.bom_403.keys())
        
        comparison_data = []
        
        for part in all_parts:
            in_401 = part in self.bom_401
            in_403 = part in self.bom_403
            
            if in_401 and in_403:
                # Parte existe en ambos - comparar detalles
                part_401 = self.bom_401[part]
                part_403 = self.bom_403[part]
                
                changes = []
                for key in ['description', 'quantity', 'embraer_code', 'weight']:
                    if key in part_401 and key in part_403:
                        if str(part_401[key]) != str(part_403[key]):
                            changes.append(f"{key}: {part_401[key]} → {part_403[key]}")
                
                status = "MODIFICADO" if changes else "SIN CAMBIOS"
                change_detail = "; ".join(changes) if changes else "No hay cambios"
                
                comparison_data.append({
                    'Part Number': part,
                    'Status': status,
                    'En 401': 'Sí',
                    'En 403': 'Sí',
                    'Description 401': part_401.get('description', ''),
                    'Description 403': part_403.get('description', ''),
                    'Quantity 401': part_401.get('quantity', ''),
                    'Quantity 403': part_403.get('quantity', ''),
                    'Embraer Code 401': part_401.get('embraer_code', ''),
                    'Embraer Code 403': part_403.get('embraer_code', ''),
                    'Weight 401': part_401.get('weight', ''),
                    'Weight 403': part_403.get('weight', ''),
                    'Changes': change_detail
                })
                
            elif in_401 and not in_403:
                # Parte removida en 403
                part_401 = self.bom_401[part]
                comparison_data.append({
                    'Part Number': part,
                    'Status': 'REMOVIDO',
                    'En 401': 'Sí',
                    'En 403': 'No',
                    'Description 401': part_401.get('description', ''),
                    'Description 403': '',
                    'Quantity 401': part_401.get('quantity', ''),
                    'Quantity 403': '',
                    'Embraer Code 401': part_401.get('embraer_code', ''),
                    'Embraer Code 403': '',
                    'Weight 401': part_401.get('weight', ''),
                    'Weight 403': '',
                    'Changes': 'Parte removida en 403'
                })
                
            elif not in_401 and in_403:
                # Parte nueva en 403
                part_403 = self.bom_403[part]
                comparison_data.append({
                    'Part Number': part,
                    'Status': 'NUEVO',
                    'En 401': 'No',
                    'En 403': 'Sí',
                    'Description 401': '',
                    'Description 403': part_403.get('description', ''),
                    'Quantity 401': '',
                    'Quantity 403': part_403.get('quantity', ''),
                    'Embraer Code 401': '',
                    'Embraer Code 403': part_403.get('embraer_code', ''),
                    'Weight 401': '',
                    'Weight 403': part_403.get('weight', ''),
                    'Changes': 'Parte nueva en 403'
                })
        
        return comparison_data
    
    def export_to_excel(self, comparison_data, output_path):
        """Exporta la comparación a un archivo Excel con formato"""
        df = pd.DataFrame(comparison_data)
        
        # Crear el archivo Excel
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='BOM Comparison', index=False)
            
            # Obtener el workbook y worksheet para aplicar formato
            workbook = writer.book
            worksheet = writer.sheets['BOM Comparison']
            
            # Definir colores para diferentes estados
            colors = {
                'NUEVO': PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid'),  # Verde claro
                'REMOVIDO': PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid'),  # Rosa claro
                'MODIFICADO': PatternFill(start_color='FFE4B5', end_color='FFE4B5', fill_type='solid'),  # Naranja claro
                'SIN CAMBIOS': PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')  # Gris claro
            }
            
            # Aplicar formato basado en el estado
            for row in range(2, len(df) + 2):  # Empezar desde la fila 2 (después del encabezado)
                status = worksheet[f'B{row}'].value  # Columna Status
                if status in colors:
                    for col in range(1, len(df.columns) + 1):
                        worksheet.cell(row=row, column=col).fill = colors[status]
            
            # Formato del encabezado
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
            
            # Ajustar ancho de columnas
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        print(f"Comparación exportada a: {output_path}")
    
    def generate_summary(self, comparison_data):
        """Genera un resumen de los cambios"""
        summary = {
            'NUEVO': 0,
            'REMOVIDO': 0,
            'MODIFICADO': 0,
            'SIN CAMBIOS': 0
        }
        
        for item in comparison_data:
            status = item['Status']
            summary[status] += 1
        
        print("\n=== RESUMEN DE CAMBIOS ===")
        print(f"Partes nuevas (403): {summary['NUEVO']}")
        print(f"Partes removidas (401): {summary['REMOVIDO']}")
        print(f"Partes modificadas: {summary['MODIFICADO']}")
        print(f"Partes sin cambios: {summary['SIN CAMBIOS']}")
        print(f"Total de partes: {sum(summary.values())}")
        
        return summary

def find_pdf_files():
    """Busca archivos PDF en el directorio actual"""
    pdf_files = glob.glob("*.pdf")
    return pdf_files

def select_from_list(files, prompt):
    """Permite al usuario seleccionar de una lista de archivos"""
    print(f"\n{prompt}")
    for i, file in enumerate(files, 1):
        print(f"{i}. {file}")
    print(f"{len(files) + 1}. Escribir ruta manualmente")
    
    while True:
        try:
            choice = int(input(f"\nSelecciona una opción (1-{len(files) + 1}): "))
            if 1 <= choice <= len(files):
                return files[choice - 1]
            elif choice == len(files) + 1:
                return input("Escribe la ruta del archivo: ").strip()
            else:
                print("Opción inválida.")
        except ValueError:
            print("Por favor ingresa un número válido.")

def main():
    """Función principal - versión de consola"""
    print("=== COMPARADOR DE BOMs EMBRAER (Versión Consola) ===")
    
    # Buscar PDFs en el directorio actual
    pdf_files = find_pdf_files()
    
    if pdf_files:
        print(f"\n📁 Se encontraron {len(pdf_files)} archivos PDF en el directorio actual:")
        for pdf in pdf_files:
            print(f"   • {pdf}")
        
        use_found = input("\n¿Deseas usar alguno de estos archivos? (s/n): ").lower()
        
        if use_found in ['s', 'si', 'sí', 'y', 'yes']:
            # Seleccionar de la lista
            pdf_401_path = select_from_list(pdf_files, "Selecciona el archivo del BOM 401 (versión anterior):")
            pdf_403_path = select_from_list(pdf_files, "Selecciona el archivo del BOM 403 (versión nueva):")
        else:
            # Escribir rutas manualmente
            pdf_401_path = input("Ingresa la ruta del PDF 401 (versión anterior): ").strip()
            pdf_403_path = input("Ingresa la ruta del PDF 403 (versión nueva): ").strip()
    else:
        print("\nNo se encontraron archivos PDF en el directorio actual.")
        pdf_401_path = input("Ingresa la ruta del PDF 401 (versión anterior): ").strip()
        pdf_403_path = input("Ingresa la ruta del PDF 403 (versión nueva): ").strip()
    
    # Verificar que los archivos existen
    if not Path(pdf_401_path).exists():
        print(f"❌ Error: No se encuentra el archivo {pdf_401_path}")
        return
    
    if not Path(pdf_403_path).exists():
        print(f"❌ Error: No se encuentra el archivo {pdf_403_path}")
        return
    
    # Mostrar archivos seleccionados
    print(f"\n📁 Archivo BOM 401: {Path(pdf_401_path).name}")
    print(f"📁 Archivo BOM 403: {Path(pdf_403_path).name}")
    
    try:
        # Crear instancia del comparador
        comparator = BOMComparator()
        
        # Cargar BOMs
        print("\n⏳ Procesando archivos...")
        comparator.load_boms(pdf_401_path, pdf_403_path)
        
        # Verificar que se encontraron partes
        if len(comparator.bom_401) == 0 and len(comparator.bom_403) == 0:
            print("❌ No se pudieron extraer partes de los PDFs. Verifica que los archivos tengan el formato correcto.")
            return
        
        # Realizar comparación
        print("🔍 Comparando BOMs...")
        comparison_data = comparator.compare_boms()
        
        # Generar resumen
        summary = comparator.generate_summary(comparison_data)
        
        # Exportar a Excel
        default_name = f"BOM_Comparison_{Path(pdf_401_path).stem}_vs_{Path(pdf_403_path).stem}.xlsx"
        
        custom_name = input(f"\nNombre del archivo de salida (presiona Enter para usar '{default_name}'): ").strip()
        output_path = custom_name if custom_name else default_name
        
        print(f"💾 Guardando comparación en Excel...")
        comparator.export_to_excel(comparison_data, output_path)
        
        print(f"\n✅ ¡Comparación completada exitosamente!")
        print(f"📄 Archivo generado: {output_path}")
        
        # Preguntar si abrir el archivo
        open_file = input("\n¿Deseas intentar abrir el archivo Excel? (s/n): ").lower()
        if open_file in ['s', 'si', 'sí', 'y', 'yes']:
            try:
                if sys.platform.startswith('win'):
                    os.startfile(output_path)
                elif sys.platform.startswith('darwin'):  # macOS
                    os.system(f'open "{output_path}"')
                else:  # Linux
                    os.system(f'xdg-open "{output_path}"')
                print("🔗 Abriendo archivo...")
            except Exception as e:
                print(f"❌ No se pudo abrir el archivo automáticamente: {e}")
                print(f"📍 Puedes encontrar el archivo en: {Path(output_path).absolute()}")
        
    except Exception as e:
        print(f"❌ Error durante la comparación: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()