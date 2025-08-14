import flet as ft
import pandas as pd
import PyPDF2
import re
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import os
from pathlib import Path
from datetime import datetime
import tkinter as tk
from tkinter import filedialog

# OCR Libraries
OCR_AVAILABLE = False
try:
    import pytesseract
    import pdf2image
    from PIL import Image, ImageEnhance
    
    # Configurar ruta de Tesseract para Windows
    pytesseract.pytesseract.tesseract_cmd = r'C:\Users\JVazquez\tesseract\tesseract.exe'
    
    OCR_AVAILABLE = True
    print("✅ OCR disponible - puede procesar PDFs imagen")
    print(f"📁 Tesseract configurado en: {pytesseract.pytesseract.tesseract_cmd}")
except ImportError as e:
    OCR_AVAILABLE = False
    print("⚠️ OCR no disponible - solo PDFs con texto")
    print(f"Instala: pip install pytesseract pdf2image pillow")
except Exception as e:
    OCR_AVAILABLE = False
    print(f"⚠️ Error configurando Tesseract: {e}")
    print("💡 Verifica que Tesseract esté instalado en: C:\\Users\\JVazquez\\tesseract\\tesseract.exe")

class BOMManager:
    """Gestor de datos de BOMs con soporte OCR"""
    
    def __init__(self):
        self.bom_401_df = pd.DataFrame()
        self.bom_403_df = pd.DataFrame()
        self.comparison_df = pd.DataFrame()
        
    def extract_text_from_pdf(self, pdf_path):
        """Extrae texto de PDF con fallback a OCR MEJORADO"""
        text = ""
        
        try:
            with open(pdf_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                for page in pdf_reader.pages:
                    page_text = page.extract_text()
                    if page_text.strip():
                        text += page_text + "\n"
                
                text_length = len(text.strip())
                print(f"✅ PDF con texto extraído: {text_length} caracteres")
                
                if text_length > 1000:
                    print(f"📝 PDF con texto nativo detectado")
                    return text
                elif text_length > 100 and text_length <= 1000:
                    print(f"⚠️ PDF con poco texto ({text_length} chars) - posible imagen")
                    if OCR_AVAILABLE:
                        print("🔄 Intentando OCR automáticamente...")
                        ocr_text = self.extract_text_with_ocr(pdf_path)
                        if len(ocr_text) > text_length * 2:
                            print(f"✅ OCR exitoso: {len(ocr_text)} caracteres")
                            return ocr_text
                    return text
                else:
                    print(f"🖼️ PDF imagen detectado ({text_length} chars) - forzando OCR")
                    
        except Exception as e:
            print(f"⚠️ Error extrayendo texto normal: {e}")
        
        if OCR_AVAILABLE:
            print("🔍 Procesando PDF imagen con OCR...")
            return self.extract_text_with_ocr(pdf_path)
        else:
            print("❌ PDF imagen detectado pero OCR no disponible")
            print("💡 Para procesar PDFs imagen instala: pip install pytesseract pdf2image pillow")
            return text if text else ""
    
    def extract_text_with_ocr(self, pdf_path):
        """Extrae texto usando OCR para PDFs imagen OPTIMIZADO PARA EMBRAER"""
        if not OCR_AVAILABLE:
            return ""
            
        try:
            print("📄 Convirtiendo PDF imagen a imágenes...")
            images = pdf2image.convert_from_path(
                pdf_path, 
                dpi=600,  # MUCHO MAYOR resolución para tablas
                first_page=1,
                last_page=15  # Más páginas
            )
            
            full_text = ""
            for i, image in enumerate(images):
                print(f"🔍 Procesando página {i+1} con OCR especializado para Embraer...")
                
                # PREPROCESSING SIMPLIFICADO (sin OpenCV para evitar dependencias)
                # 1. Convertir a escala de grises
                if image.mode != 'L':
                    image = image.convert('L')
                
                # 2. Aumentar contraste agresivamente
                enhancer = ImageEnhance.Contrast(image)
                processed_image = enhancer.enhance(3.0)  # Más contraste
                
                # 3. Aumentar brillo
                brightness_enhancer = ImageEnhance.Brightness(processed_image)
                processed_image = brightness_enhancer.enhance(1.2)
                
                # 4. Aumentar nitidez
                sharpness_enhancer = ImageEnhance.Sharpness(processed_image)
                processed_image = sharpness_enhancer.enhance(2.0)
                
                # MÚLTIPLES INTENTOS OCR con configuraciones específicas para EMBRAER
                ocr_configs = [
                    '--oem 3 --psm 6 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789-/.,()[]_ preserve_interword_spaces=1',
                    '--oem 3 --psm 4 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789-/.,()[]_',
                    '--oem 1 --psm 6 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789-/.,()[]_',
                    '--oem 3 --psm 8 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789-/.,()[]_',
                    '--oem 3 --psm 12 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789-/.,()[]_'
                ]
                
                best_text = ""
                best_config = ""
                
                for config in ocr_configs:
                    try:
                        # Intentar con imagen procesada
                        page_text = pytesseract.image_to_string(processed_image, config=config)
                        
                        # También intentar con imagen original mejorada
                        enhanced_original = ImageEnhance.Contrast(image).enhance(2.5)
                        page_text_orig = pytesseract.image_to_string(enhanced_original, config=config)
                        
                        # Tomar el mejor resultado
                        if len(page_text_orig) > len(page_text):
                            page_text = page_text_orig
                        
                        if len(page_text) > len(best_text):
                            best_text = page_text
                            best_config = config
                            print(f"   📝 Mejor config hasta ahora: {len(page_text)} chars")
                        
                    except Exception as e:
                        print(f"   ⚠️ Error con config: {e}")
                        continue
                
                if best_text:
                    full_text += best_text + "\n"
                    print(f"   ✅ Página {i+1}: {len(best_text)} caracteres")
                    
                    # Debug: mostrar líneas que parecen part numbers
                    lines = best_text.split('\n')
                    embraer_lines = []
                    for line in lines:
                        line = line.strip()
                        # Buscar líneas que contengan patrones típicos de Embraer
                        if re.search(r'(EMB-|MS\d|D\d+|CDEL\d|WFFR|\d{3}-\d+|\d{7,8})', line, re.IGNORECASE):
                            embraer_lines.append(line)
                    
                    print(f"   🎯 Líneas con patrones Embraer detectadas: {len(embraer_lines)}")
                    for j, line in enumerate(embraer_lines[:10]):  # Mostrar primeras 10
                        print(f"      {j+1}: {line[:80]}")
                        
                else:
                    print(f"   ❌ No se pudo extraer texto de página {i+1}")
            
            print(f"✅ OCR completado: {len(full_text)} caracteres totales")
            
            # Guardar texto extraído con timestamp único
            debug_filename = f"ocr_debug_embraer_403_{datetime.now().strftime('%H%M%S')}.txt"
            with open(debug_filename, "w", encoding="utf-8") as f:
                f.write(full_text)
            print(f"💾 Texto OCR guardado en: {debug_filename}")
            
            return full_text
            
        except Exception as e:
            print(f"❌ Error en OCR Embraer: {e}")
            import traceback
            traceback.print_exc()
            return ""
    
    def parse_part_line(self, line):
        """Extrae información de una línea de parte ESPECIALIZADO PARA EMBRAER BOM 403"""
        line = line.strip()
        if not line or len(line) < 3:
            return None
            
        # Limpiar caracteres OCR problemáticos
        line = re.sub(r'[|¡!]', 'I', line)
        line = re.sub(r'\s+', ' ', line)
        line = re.sub(r'[^\w\s\-\.,():/]', ' ', line)  # Limpiar caracteres raros
        
        # Patrones ESPECÍFICOS para BOM 403 de Embraer (basado en las imágenes)
        patterns = [
            # Patrón 1: EMB-0213-XX-X WIRE, ELECTRICAL, XETFE INSULATED - COPPER, NICKEL COATED XXXXXX L
            r'(EMB-\d{4}-\d{2}-\d)\s+(WIRE[^L]+)\s+(\d{6,8})\s+L',
            
            # Patrón 2: MS3476L22-41P CONNECTOR, ELECTRICAL, PLUG - ALUMINIUM XXXXXX L
            r'(MS\d+[A-Z]\d+-\d+[A-Z]?)\s+(CONNECTOR[^L]+)\s+(\d{6,8})\s+L',
            
            # Patrón 3: D38999/26JB99SN CONNECTOR,ELECTRICAL-STRAIGHT PLUG,CIRCULAR, XXXXXX L
            r'(D\d+/\d+[A-Z]+\d+[A-Z]+)\s+(CONNECTOR[^L]+)\s+(\d{6,8})\s+L',
            
            # Patrón 4: M85049/52-1-22N CLAMP, STRAIN RELIEF, STRAIGHT- FOR ELECTRICAL XXXXXX L
            r'(M\d+/\d+-\d+-\d+[A-Z]?)\s+(CLAMP[^L]+|STRAIN[^L]+)\s+(\d{6,8})\s+L',
            
            # Patrón 5: CDEL052-1 CONTACT, SOCKET 20 GA. XXXXXX L
            r'(CDEL\d+-\d+)\s+(CONTACT[^L]+)\s+(\d{6,8})\s+L',
            
            # Patrón 6: 3357 0001 301 CABLE CLAMP XXXXXX L
            r'(\d{4}\s+\d{4}\s+\d{3})\s+(CABLE[^L]+)\s+(\d{6,8})\s+L',
            
            # Patrón 7: WFFRBK250 SLEEVING,EXPLANDABLE-PROTECTION,PLASTIC XXXXXX L
            r'(WFFR[A-Z0-9]+)\s+(SLEEVE?[^L]+|SLEEVING[^L]+)\s+(\d{6,8})\s+L',
            
            # Patrón 8: Part numbers genéricos seguidos de descripción y código Embraer
            r'([A-Z0-9\-\/\.]{5,25})\s+([A-Z][^L]{5,50})\s+(\d{6,8})\s+L',
            
            # Patrón 9: Líneas que empiecen con código Embraer primero
            r'(\d{7,8})\s+([A-Z0-9\-\/\.]{5,25})\s+([A-Z][^L]{5,50})\s+L',
            
            # Patrón 10: Simplificado - cualquier part number + descripción + L al final
            r'([A-Z0-9\-\/\.]{3,20})\s+([A-Z].{5,40})\s+L\s*$'
        ]
        
        for i, pattern in enumerate(patterns):
            match = re.search(pattern, line, re.IGNORECASE)
            if match:
                groups = match.groups()
                try:
                    # Para patrón 9, el código Embraer está primero
                    if i == 8 and len(groups) >= 3:
                        embraer_code = groups[0].strip()
                        part_number = groups[1].strip()
                        description = groups[2].strip()
                    else:
                        part_number = groups[0].strip()
                        description = groups[1].strip() if len(groups) > 1 else 'PART DESCRIPTION'
                        embraer_code = groups[2].strip() if len(groups) > 2 and groups[2].isdigit() else ''
                    
                    # Validaciones específicas para Embraer
                    if len(part_number) < 3:
                        continue
                    
                    # Saltar códigos de notas
                    if re.match(r'^(NL|NS|NI|NOTE|DRAWING|PART|POSITION|ZONE|PROGRAM|MODEL)[0-9]*', part_number.upper()):
                        continue
                        
                    # Saltar si es solo números (excepto patrones específicos como "3357 0001 301")
                    if part_number.replace(' ', '').isdigit() and ' ' not in part_number:
                        continue
                    
                    # Limpiar descripción
                    description = re.sub(r'[^\w\s\-\.,():/]', '', description)
                    description = description.strip()
                    
                    # Validar descripción
                    if len(description) < 3:
                        description = "ELECTRICAL COMPONENT"
                    
                    # Limpiar part number de espacios múltiples
                    part_number = re.sub(r'\s+', ' ', part_number).strip()
                    
                    result = {
                        'Part_Number': part_number,
                        'Description': description,
                        'Embraer_Code': embraer_code,
                        'Status': 'L',
                        'Quantity': 1,
                        'Unit': 'UM',
                        'Weight': '0',
                    }
                    
                    print(f"🎯 Patrón {i+1} Embraer detectó: {part_number} - {description[:30]} [{embraer_code}]")
                    return result
                    
                except Exception as e:
                    print(f"⚠️ Error en patrón Embraer {i+1}: {e}")
                    continue
        
        # Si no se encontró nada, mostrar la línea para debug
        if len(line) > 20 and any(keyword in line.upper() for keyword in ['EMB-', 'MS', 'D38999', 'CDEL', 'WFFR', 'WIRE', 'CONNECTOR']):
            print(f"🔍 Línea no procesada (debug): {line[:80]}")
        
        return None














    def extract_bom_from_text(self, text, bom_version):
        """Extrae información del BOM del texto (normal o OCR) MEJORADO"""
        lines = text.split('\n')
        bom_data = []
        
        skip_keywords = [
            'DRAWING', 'PART NUMBER', 'DESCRIPTION', 'EMBRAER', 'COPYRIGHT', 
            'PRINTED', 'LAST REVISION', 'RELEASED', 'REVISION', 'STATUS',
            'POSITION', 'ZONE', 'PROGRAM', 'MODEL', 'INITIAL', 'FINAL',
            'QUANTITY', 'UNIT', 'WEIGHT', 'TRACEABILITY', 'ESPEC', 'PAGE',
            'PREVIEW', 'LIST', 'DATA', 'DOCUMENT', 'TYPE', 'ORIGIN'
        ]
        
        print(f"🔍 Analizando {len(lines)} líneas para BOM {bom_version}")
        
        print("📝 Primeras 20 líneas del texto:")
        for i, line in enumerate(lines[:20]):
            if line.strip():
                print(f"   {i:2d}: {line[:80]}")
        
        for line_num, line in enumerate(lines):
            line = line.strip()
            if not line or len(line) < 3:
                continue
                
            if any(keyword in line.upper() for keyword in skip_keywords):
                continue
            
            if re.match(r'^[\d\s\-\=\|_\.]+$', line):
                continue
                
            if len(line) < 5:
                continue
            
            if re.match(r'^(EMB|MS|D\d|CDEL|WFFR|\d{3}-\d|-\d+|[A-Z0-9]{3,}[\-\/])', line):
                print(f"🔍 Candidato línea {line_num}: {line[:60]}")
                
            part_info = self.parse_part_line(line)
            if part_info and part_info['Part_Number'] and len(part_info['Part_Number']) > 2:
                part_info['BOM_Version'] = bom_version
                part_info['Line_Number'] = line_num
                bom_data.append(part_info)
                print(f"✅ {bom_version}: {part_info['Part_Number']} - {part_info['Description'][:30]}")
        
        df = pd.DataFrame(bom_data)
        print(f"📊 BOM {bom_version} procesado: {len(df)} partes extraídas")
        
        if len(df) == 0:
            print(f"❌ NO SE EXTRAJERON PARTES DEL BOM {bom_version}")
            print("📝 TODAS las líneas no vacías para debug:")
            non_empty_lines = [line for line in lines if line.strip()]
            for i, line in enumerate(non_empty_lines[:50]):
                print(f"   {i:2d}: {line}")
        
        return df
    
    def load_boms(self, pdf_401_path, pdf_403_path):
        """Carga los BOMs de ambos PDFs con soporte OCR"""
        print("🔄 Iniciando extracción de BOMs...")
        
        print("📄 Procesando BOM 401...")
        text_401 = self.extract_text_from_pdf(pdf_401_path)
        if text_401:
            self.bom_401_df = self.extract_bom_from_text(text_401, "401")
        else:
            print("❌ No se pudo extraer texto del BOM 401")
            return False
        
        print("📄 Procesando BOM 403...")
        text_403 = self.extract_text_from_pdf(pdf_403_path)
        if text_403:
            self.bom_403_df = self.extract_bom_from_text(text_403, "403")
        else:
            print("❌ No se pudo extraer texto del BOM 403")
            return False
        
        if not self.bom_401_df.empty:
            self.bom_401_df = self.bom_401_df.drop_duplicates(subset=['Part_Number'])
        if not self.bom_403_df.empty:
            self.bom_403_df = self.bom_403_df.drop_duplicates(subset=['Part_Number'])
        
        print(f"✅ Extracción completada:")
        print(f"   📋 BOM 401: {len(self.bom_401_df)} partes únicas")
        print(f"   📋 BOM 403: {len(self.bom_403_df)} partes únicas")
        
        return True
    
    def compare_boms_right_join(self):
        """Compara BOMs usando lógica RIGHT JOIN"""
        if self.bom_401_df.empty and self.bom_403_df.empty:
            return pd.DataFrame()
        
        print(f"🔍 Ejecutando RIGHT JOIN...")
        comparison_data = []
        
        parts_401 = set(self.bom_401_df['Part_Number'].values) if not self.bom_401_df.empty else set()
        parts_403 = set(self.bom_403_df['Part_Number'].values) if not self.bom_403_df.empty else set()
        
        if not self.bom_403_df.empty:
            for _, part_403 in self.bom_403_df.iterrows():
                part_number = part_403['Part_Number']
                
                part_401_data = self.bom_401_df[self.bom_401_df['Part_Number'] == part_number]
                
                if not part_401_data.empty:
                    part_401 = part_401_data.iloc[0]
                    
                    changes = []
                    if str(part_401['Description']) != str(part_403['Description']):
                        changes.append(f"Desc: {part_401['Description'][:15]} → {part_403['Description'][:15]}")
                    if str(part_401['Quantity']) != str(part_403['Quantity']):
                        changes.append(f"Qty: {part_401['Quantity']} → {part_403['Quantity']}")
                    if str(part_401['Embraer_Code']) != str(part_403['Embraer_Code']):
                        changes.append(f"Code: {part_401['Embraer_Code']} → {part_403['Embraer_Code']}")
                    
                    status = "MODIFICADO" if changes else "SIN CAMBIOS"
                    change_detail = "; ".join(changes) if changes else "No hay cambios"
                else:
                    status = "NUEVO"
                    change_detail = "Parte nueva en 403"
                    part_401 = pd.Series({'Description': '', 'Quantity': '', 'Embraer_Code': ''})
                
                comparison_data.append({
                    'Part_Number': part_number,
                    'Status': status,
                    'En_401': 'SÍ' if not part_401_data.empty else 'NO',
                    'En_403': 'SÍ',
                    'Description_401': str(part_401['Description']),
                    'Description_403': str(part_403['Description']),
                    'Quantity_401': str(part_401['Quantity']),
                    'Quantity_403': str(part_403['Quantity']),
                    'Embraer_Code_401': str(part_401['Embraer_Code']),
                    'Embraer_Code_403': str(part_403['Embraer_Code']),
                    'Changes': change_detail
                })
        
        if not self.bom_401_df.empty:
            for _, part_401 in self.bom_401_df.iterrows():
                part_number = part_401['Part_Number']
                
                if part_number not in parts_403:
                    comparison_data.append({
                        'Part_Number': part_number,
                        'Status': 'REMOVIDO',
                        'En_401': 'SÍ',
                        'En_403': 'NO',
                        'Description_401': str(part_401['Description']),
                        'Description_403': '',
                        'Quantity_401': str(part_401['Quantity']),
                        'Quantity_403': '',
                        'Embraer_Code_401': str(part_401['Embraer_Code']),
                        'Embraer_Code_403': '',
                        'Changes': 'Parte removida en 403'
                    })
        
        self.comparison_df = pd.DataFrame(comparison_data)
        
        print(f"✅ RIGHT JOIN completado: {len(self.comparison_df)} comparaciones")
        if not self.comparison_df.empty:
            status_counts = self.comparison_df['Status'].value_counts()
            for status, count in status_counts.items():
                print(f"   {status}: {count}")
        
        return self.comparison_df
    
    def get_kpis(self):
        """Calcula KPIs de la comparación"""
        if self.comparison_df.empty:
            return {
                'total_401': len(self.bom_401_df),
                'total_403': len(self.bom_403_df),
                'nuevos': 0,
                'removidos': 0,
                'modificados': 0,
                'sin_cambios': 0,
                'total_comparacion': 0
            }
        
        status_counts = self.comparison_df['Status'].value_counts()
        
        return {
            'total_401': len(self.bom_401_df),
            'total_403': len(self.bom_403_df),
            'nuevos': status_counts.get('NUEVO', 0),
            'removidos': status_counts.get('REMOVIDO', 0),
            'modificados': status_counts.get('MODIFICADO', 0),
            'sin_cambios': status_counts.get('SIN CAMBIOS', 0),
            'total_comparacion': len(self.comparison_df)
        }
    
    def export_to_excel(self, output_path):
        """Exporta con las 3 hojas requeridas"""
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            
            if not self.bom_401_df.empty:
                export_401 = self.bom_401_df.copy()
                export_401 = export_401.drop(['BOM_Version', 'Line_Number'], axis=1, errors='ignore')
                export_401.to_excel(writer, sheet_name='BOM_401_VIEJO', index=False)
                self.format_sheet(writer.book['BOM_401_VIEJO'], "#2196F3", "BOM 401 (VIEJO)")
            
            if not self.bom_403_df.empty:
                export_403 = self.bom_403_df.copy()
                export_403 = export_403.drop(['BOM_Version', 'Line_Number'], axis=1, errors='ignore')
                export_403.to_excel(writer, sheet_name='BOM_403_NUEVO', index=False)
                self.format_sheet(writer.book['BOM_403_NUEVO'], "#FF6B35", "BOM 403 (NUEVO)")
            
            if not self.comparison_df.empty:
                self.comparison_df.to_excel(writer, sheet_name='COMPARACION_RIGHT_JOIN', index=False)
                self.format_comparison_sheet(writer.book['COMPARACION_RIGHT_JOIN'])
            
            self.create_summary_sheet(writer)
    
    def format_sheet(self, worksheet, color, title):
        """Aplica formato a hojas de BOM"""
        header_fill = PatternFill(start_color=color.replace('#', ''), end_color=color.replace('#', ''), fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=12)
        
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 3, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def format_comparison_sheet(self, worksheet):
        """Aplica formato especial a la hoja de comparación"""
        colors = {
            'NUEVO': PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid'),
            'REMOVIDO': PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid'),
            'MODIFICADO': PatternFill(start_color='FFE0B2', end_color='FFE0B2', fill_type='solid'),
            'SIN CAMBIOS': PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
        }
        
        for row in range(2, worksheet.max_row + 1):
            status_cell = worksheet[f'B{row}']
            status = status_cell.value
            
            if status in colors:
                for col in range(1, worksheet.max_column + 1):
                    worksheet.cell(row=row, column=col).fill = colors[status]
        
        header_fill = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=12)
        
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 3, 60)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def create_summary_sheet(self, writer):
        """Crea hoja de resumen ejecutivo"""
        kpis = self.get_kpis()
        ocr_status = "✅ OCR Activado" if OCR_AVAILABLE else "❌ Solo PDFs texto"
        
        summary_data = [
            ['🔧 RESUMEN EJECUTIVO - RIGHT JOIN BOM COMPARISON', ''],
            ['', ''],
            ['📊 TOTALES POR BOM', ''],
            ['BOM 401 (Viejo) - Total Partes', kpis['total_401']],
            ['BOM 403 (Nuevo) - Total Partes', kpis['total_403']],
            ['', ''],
            ['⚖️ ANÁLISIS DE CAMBIOS (RIGHT JOIN)', ''],
            ['🆕 Partes NUEVAS (solo en 403)', kpis['nuevos']],
            ['🗑️ Partes REMOVIDAS (solo en 401)', kpis['removidos']],
            ['🔄 Partes MODIFICADAS', kpis['modificados']],
            ['✅ Partes SIN CAMBIOS', kpis['sin_cambios']],
            ['📋 Total en Comparación', kpis['total_comparacion']],
            ['', ''],
            ['🔍 CAPACIDADES DEL SISTEMA', ''],
            ['Procesamiento OCR', ocr_status],
            ['Método de Comparación', 'RIGHT JOIN (como SQL)'],
            ['Base de Comparación', 'BOM 403 (Nuevo) + Removidos de 401'],
            ['Timestamp', datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
        ]
        
        summary_df = pd.DataFrame(summary_data, columns=['Métrica', 'Valor'])
        summary_df.to_excel(writer, sheet_name='RESUMEN_EJECUTIVO', index=False)
        
        ws = writer.book['RESUMEN_EJECUTIVO']
        ws['A1'].fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
        ws['A1'].font = Font(color='FFFFFF', bold=True, size=14)
        ws['B1'].fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
        
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 25


class BOMDashboard:
    """Dashboard para Comparación RIGHT JOIN de BOMs"""
    
    def __init__(self, page: ft.Page):
        self.page = page
        self.bom_manager = BOMManager()
        self.pdf_401_path = ""
        self.pdf_403_path = ""
        self.setup_page()
        
    def setup_page(self):
        """Configuración de página estilo dark"""
        self.page.title = "🔧 BOM RIGHT JOIN Comparator - Embraer"
        self.page.theme_mode = ft.ThemeMode.DARK
        self.page.bgcolor = "#0a0a0a"
        self.page.window_width = 1600
        self.page.window_height = 1000
        self.page.window_maximized = True
        self.page.padding = 0
        
        # Componentes
        self.status_text = ft.Text("🔄 Listo para análisis RIGHT JOIN...", color="#00bcd4", size=16)
        self.kpi_container = ft.Container()
        self.comparison_table_container = ft.Container()
        self.bom_401_container = ft.Container()
        self.bom_403_container = ft.Container()
        
        self.build_ui()
    
    def build_ui(self):
        """Construye interfaz estilo SQL JOIN"""
        
        # Header con tema SQL
        header = ft.Container(
            content=ft.Row([
                ft.Icon(ft.Icons.JOIN_RIGHT, size=60, color="#00bcd4"),
                ft.Column([
                    ft.Text("BOM RIGHT JOIN", size=36, weight=ft.FontWeight.BOLD, color="#ffffff"),
                    ft.Text("SQL-Style Comparison Analysis", size=18, color="#00bcd4")
                ], spacing=0),
                ft.Container(expand=True),
                ft.Row([
                    self.create_load_button("📂 BOM 401\n(VIEJO)", "#2196F3", self.load_bom_401),
                    ft.Container(width=10),
                    self.create_load_button("📂 BOM 403\n(NUEVO)", "#FF6B35", self.load_bom_403),
                    ft.Container(width=10),
                    self.create_action_button("⚖️ RIGHT\nJOIN", "#4CAF50", self.compare_boms),
                    ft.Container(width=10),
                    self.create_action_button("💾 EXPORT\nEXCEL", "#9C27B0", self.export_excel),
                ])
            ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
            padding=30,
            bgcolor="#1a1a1a",
            border=ft.border.only(bottom=ft.border.BorderSide(2, "#00bcd4"))
        )
        
        # SQL-style status bar
        status_bar = ft.Container(
            content=ft.Row([
                ft.Icon(ft.Icons.STORAGE, color="#4caf50"),
                self.status_text,
                ft.Container(expand=True),
                ft.Text("🔗 RIGHT JOIN Mode", color="#00bcd4", weight=ft.FontWeight.BOLD),
                ft.Text("🔍 OCR Support" if OCR_AVAILABLE else "📝 Text Only", color="#ff9800", size=12),
                ft.Text(f"⏰ {datetime.now().strftime('%H:%M:%S')}", color="#757575", size=14)
            ]),
            padding=15,
            bgcolor="#1a1a1a"
        )
        
        # Tabs principales
        tabs = ft.Tabs(
            selected_index=0,
            animation_duration=300,
            indicator_color="#00bcd4",
            label_color="#00bcd4",
            unselected_label_color="#757575",
            tabs=[
                ft.Tab(
                    text="📊 RIGHT JOIN ANALYSIS",
                    content=ft.Container(
                        content=ft.Column([
                            self.kpi_container,
                            self.comparison_table_container
                        ], spacing=20, scroll=ft.ScrollMode.AUTO),
                        padding=20
                    )
                ),
                ft.Tab(
                    text="📋 BOM 401 (VIEJO)",
                    content=ft.Container(
                        content=self.bom_401_container,
                        padding=20
                    )
                ),
                ft.Tab(
                    text="📋 BOM 403 (NUEVO)",
                    content=ft.Container(
                        content=self.bom_403_container,
                        padding=20
                    )
                )
            ]
        )
        
        # Layout principal
        self.page.add(
            ft.Column([
                header,
                status_bar,
                ft.Container(content=tabs, expand=True)
            ], spacing=0)
        )
    
    def create_load_button(self, text, color, on_click):
        """Crea botón de carga estilizado"""
        return ft.Container(
            content=ft.TextButton(
                text=text,
                style=ft.ButtonStyle(
                    color="#ffffff",
                    bgcolor=color,
                    shape=ft.RoundedRectangleBorder(radius=10)
                ),
                on_click=on_click
            ),
            width=100,
            height=70
        )
    
    def create_action_button(self, text, color, on_click):
        """Crea botón de acción estilizado"""
        return ft.Container(
            content=ft.TextButton(
                text=text,
                style=ft.ButtonStyle(
                    color="#ffffff",
                    bgcolor=color,
                    shape=ft.RoundedRectangleBorder(radius=10)
                ),
                on_click=on_click
            ),
            width=100,
            height=70
        )
    
    def select_file(self, title):
        """Selector de archivos"""
        root = tk.Tk()
        root.withdraw()
        
        file_path = filedialog.askopenfilename(
            title=title,
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")]
        )
        
        root.destroy()
        return file_path
    
    def load_bom_401(self, e):
        """Carga BOM 401 (viejo)"""
        file_path = self.select_file("Selecciona BOM 401 (VIEJO)")
        if file_path:
            self.pdf_401_path = file_path
            self.status_text.value = f"✅ BOM 401 (VIEJO): {Path(file_path).name}"
            self.show_snackbar(f"BOM 401 cargado: {Path(file_path).name}", "#2196F3")
            self.page.update()
    
    def load_bom_403(self, e):
        """Carga BOM 403 (nuevo)"""
        file_path = self.select_file("Selecciona BOM 403 (NUEVO)")
        if file_path:
            self.pdf_403_path = file_path
            self.status_text.value = f"✅ BOM 403 (NUEVO): {Path(file_path).name}"
            self.show_snackbar(f"BOM 403 cargado: {Path(file_path).name}", "#FF6B35")
            self.page.update()
    
    def compare_boms(self, e):
        """Ejecuta RIGHT JOIN comparison"""
        if not self.pdf_401_path or not self.pdf_403_path:
            self.show_snackbar("⚠️ Carga ambos BOMs primero", "#ff9800")
            return
        
        self.status_text.value = "🔄 Ejecutando RIGHT JOIN..."
        self.page.update()
        
        try:
            # Cargar BOMs
            if self.bom_manager.load_boms(self.pdf_401_path, self.pdf_403_path):
                # RIGHT JOIN comparison
                self.bom_manager.compare_boms_right_join()
                
                # Actualizar dashboard
                self.update_dashboard()
                self.update_bom_viewers()
                
                self.status_text.value = "✅ RIGHT JOIN completado exitosamente"
                self.show_snackbar("✅ Análisis RIGHT JOIN completado", "#4caf50")
            else:
                self.status_text.value = "❌ Error procesando BOMs"
                self.show_snackbar("❌ Error procesando PDFs", "#f44336")
        
        except Exception as error:
            self.status_text.value = f"❌ Error: {str(error)}"
            self.show_snackbar(f"❌ Error: {str(error)}", "#f44336")
        
        self.page.update()
    
    def update_dashboard(self):
        """Actualiza dashboard con resultados RIGHT JOIN"""
        kpis = self.bom_manager.get_kpis()
        
        # KPI Cards estilo SQL
        kpi_cards = [
            self.create_sql_kpi_card("📋 BOM 401", f"{kpis['total_401']:,}", "#2196F3", "Partes en VIEJO"),
            self.create_sql_kpi_card("📋 BOM 403", f"{kpis['total_403']:,}", "#FF6B35", "Partes en NUEVO"),
            self.create_sql_kpi_card("🆕 NUEVOS", f"{kpis['nuevos']:,}", "#4CAF50", "Solo en 403"),
            self.create_sql_kpi_card("🗑️ REMOVIDOS", f"{kpis['removidos']:,}", "#F44336", "Solo en 401"),
            self.create_sql_kpi_card("🔄 MODIFICADOS", f"{kpis['modificados']:,}", "#FF9800", "Diferentes"),
            self.create_sql_kpi_card("✅ IGUALES", f"{kpis['sin_cambios']:,}", "#9E9E9E", "Sin cambios")
        ]
        
        self.kpi_container.content = ft.Container(
            content=ft.Column([
                ft.Text("📊 RIGHT JOIN ANALYSIS RESULTS", size=24, weight=ft.FontWeight.BOLD, color="#ffffff"),
                ft.Text("🔗 Mostrando todas las partes de BOM 403 + partes removidas de BOM 401", color="#00bcd4", size=14),
                ft.Container(height=20),
                ft.Row(kpi_cards, spacing=20, wrap=True)
            ]),
            margin=ft.margin.only(bottom=30)
        )
        
        # Actualizar tabla de comparación
        self.update_comparison_table()
    
    def create_sql_kpi_card(self, title, value, color, subtitle):
        """Crea KPI card estilo SQL"""
        return ft.Container(
            content=ft.Column([
                ft.Text(title, size=12, weight=ft.FontWeight.BOLD, color="#bdbdbd"),
                ft.Text(value, size=28, weight=ft.FontWeight.BOLD, color=color),
                ft.Text(subtitle, size=10, color="#757575")
            ], spacing=8, horizontal_alignment=ft.CrossAxisAlignment.CENTER),
            width=250,
            height=120,
            padding=20,
            bgcolor="#1a1a1a",
            border=ft.border.all(2, color),
            border_radius=15,
            shadow=ft.BoxShadow(
                spread_radius=1,
                blur_radius=15,
                color="#424242",
                offset=ft.Offset(0, 5)
            )
        )
    
    def update_comparison_table(self):
        """Actualiza tabla de comparación RIGHT JOIN"""
        if self.bom_manager.comparison_df.empty:
            self.comparison_table_container.content = ft.Container(
                content=ft.Text("🔗 Ejecuta RIGHT JOIN para ver la comparación", color="#ffffff", size=16),
                padding=40,
                bgcolor="#1a1a1a",
                border_radius=15
            )
            return
        
        # Campo de búsqueda SQL-style
        search_field = ft.TextField(
            label="🔍 WHERE Part_Number LIKE '%search%' OR Description LIKE '%search%'",
            width=600,
            color="#ffffff",
            bgcolor="#2a2a2a",
            border_color="#00bcd4",
            on_change=self.search_comparison_table
        )
        
        # Crear tabla RIGHT JOIN
        self.create_right_join_table()
        
        self.comparison_table_container.content = ft.Column([
            ft.Text("🔗 RIGHT JOIN COMPARISON RESULTS", size=20, weight=ft.FontWeight.BOLD, color="#ffffff"),
            ft.Text(f"📊 SELECT * FROM (BOM_403 RIGHT JOIN BOM_401) - Total: {len(self.bom_manager.comparison_df):,} registros", 
                   color="#00bcd4", size=12),
            ft.Container(height=10),
            search_field,
            ft.Container(height=10),
            self.scrollable_right_join_table
        ])
    
    def create_right_join_table(self):
        """Crea tabla estilo RIGHT JOIN con colores SQL"""
        df = self.bom_manager.comparison_df
        
        rows = []
        for _, row in df.iterrows():
            # Colores estilo SQL
            status_colors = {
                'NUEVO': "#4CAF50",      # Verde - INSERT
                'REMOVIDO': "#F44336",   # Rojo - DELETE  
                'MODIFICADO': "#FF9800", # Naranja - UPDATE
                'SIN CAMBIOS': "#9E9E9E" # Gris - SELECT
            }
            status_color = status_colors.get(row['Status'], "#757575")
            
            # Iconos SQL
            status_icons = {
                'NUEVO': "➕",
                'REMOVIDO': "➖", 
                'MODIFICADO': "🔄",
                'SIN CAMBIOS': "✅"
            }
            status_icon = status_icons.get(row['Status'], "❓")
            
            rows.append(
                ft.DataRow(
                    cells=[
                        ft.DataCell(ft.Text(str(row['Part_Number']), color="#00bcd4", size=11, weight=ft.FontWeight.BOLD)),
                        ft.DataCell(ft.Container(
                            content=ft.Row([
                                ft.Text(status_icon, size=12),
                                ft.Text(str(row['Status']), color="#ffffff", size=10, weight=ft.FontWeight.BOLD)
                            ], spacing=5),
                            bgcolor=status_color,
                            padding=ft.padding.symmetric(horizontal=8, vertical=4),
                            border_radius=5
                        )),
                        ft.DataCell(ft.Text("✅" if row['En_401'] == 'SÍ' else "❌", color="#4CAF50" if row['En_401'] == 'SÍ' else "#F44336", size=12)),
                        ft.DataCell(ft.Text("✅" if row['En_403'] == 'SÍ' else "❌", color="#4CAF50" if row['En_403'] == 'SÍ' else "#F44336", size=12)),
                        ft.DataCell(ft.Text(str(row['Description_401'])[:25] + "..." if len(str(row['Description_401'])) > 25 else str(row['Description_401']), 
                                          color="#2196F3", size=10)),
                        ft.DataCell(ft.Text(str(row['Description_403'])[:25] + "..." if len(str(row['Description_403'])) > 25 else str(row['Description_403']), 
                                          color="#FF6B35", size=10)),
                        ft.DataCell(ft.Text(str(row['Quantity_401']), color="#2196F3", size=11)),
                        ft.DataCell(ft.Text(str(row['Quantity_403']), color="#FF6B35", size=11)),
                        ft.DataCell(ft.Text(str(row['Embraer_Code_401']), color="#9C27B0", size=10)),
                        ft.DataCell(ft.Text(str(row['Embraer_Code_403']), color="#9C27B0", size=10)),
                        ft.DataCell(ft.Text(str(row['Changes'])[:30] + "..." if len(str(row['Changes'])) > 30 else str(row['Changes']), 
                                          color="#FFEB3B", size=9))
                    ]
                )
            )
        
        right_join_table = ft.DataTable(
            columns=[
                ft.DataColumn(ft.Text("Part Number", weight=ft.FontWeight.BOLD, color="#ffffff", size=11)),
                ft.DataColumn(ft.Text("SQL Status", weight=ft.FontWeight.BOLD, color="#ffffff", size=11)),
                ft.DataColumn(ft.Text("401", weight=ft.FontWeight.BOLD, color="#2196F3", size=11)),
                ft.DataColumn(ft.Text("403", weight=ft.FontWeight.BOLD, color="#FF6B35", size=11)),
                ft.DataColumn(ft.Text("Description 401", weight=ft.FontWeight.BOLD, color="#2196F3", size=10)),
                ft.DataColumn(ft.Text("Description 403", weight=ft.FontWeight.BOLD, color="#FF6B35", size=10)),
                ft.DataColumn(ft.Text("Qty 401", weight=ft.FontWeight.BOLD, color="#2196F3", size=10)),
                ft.DataColumn(ft.Text("Qty 403", weight=ft.FontWeight.BOLD, color="#FF6B35", size=10)),
                ft.DataColumn(ft.Text("Code 401", weight=ft.FontWeight.BOLD, color="#9C27B0", size=10)),
                ft.DataColumn(ft.Text("Code 403", weight=ft.FontWeight.BOLD, color="#9C27B0", size=10)),
                ft.DataColumn(ft.Text("Changes", weight=ft.FontWeight.BOLD, color="#FFEB3B", size=10))
            ],
            rows=rows,
            bgcolor="#1a1a1a",
            border_radius=10,
            heading_row_color="#2a2a2a"
        )
        
        self.scrollable_right_join_table = ft.Container(
            content=ft.Column(
                controls=[right_join_table],
                scroll=ft.ScrollMode.ALWAYS,
                auto_scroll=True
            ),
            height=500,
            bgcolor="#1a1a1a",
            border_radius=10,
            padding=10,
            border=ft.border.all(2, "#00bcd4")
        )
    
    def search_comparison_table(self, e):
        """Búsqueda estilo SQL WHERE"""
        search_term = e.control.value.lower().strip()
        
        if not search_term:
            filtered_df = self.bom_manager.comparison_df.copy()
        else:
            mask = (
                self.bom_manager.comparison_df['Part_Number'].astype(str).str.lower().str.contains(search_term, na=False) |
                self.bom_manager.comparison_df['Description_401'].astype(str).str.lower().str.contains(search_term, na=False) |
                self.bom_manager.comparison_df['Description_403'].astype(str).str.lower().str.contains(search_term, na=False) |
                self.bom_manager.comparison_df['Embraer_Code_401'].astype(str).str.lower().str.contains(search_term, na=False) |
                self.bom_manager.comparison_df['Embraer_Code_403'].astype(str).str.lower().str.contains(search_term, na=False) |
                self.bom_manager.comparison_df['Status'].astype(str).str.lower().str.contains(search_term, na=False)
            )
            filtered_df = self.bom_manager.comparison_df[mask]
        
        self.create_filtered_right_join_table(filtered_df)
        self.page.update()
    
    def create_filtered_right_join_table(self, df):
        """Crea tabla filtrada estilo RIGHT JOIN"""
        rows = []
        for _, row in df.iterrows():
            status_colors = {
                'NUEVO': "#4CAF50",
                'REMOVIDO': "#F44336",
                'MODIFICADO': "#FF9800",
                'SIN CAMBIOS': "#9E9E9E"
            }
            status_color = status_colors.get(row['Status'], "#757575")
            
            status_icons = {
                'NUEVO': "➕",
                'REMOVIDO': "➖",
                'MODIFICADO': "🔄", 
                'SIN CAMBIOS': "✅"
            }
            status_icon = status_icons.get(row['Status'], "❓")
            
            rows.append(
                ft.DataRow(
                    cells=[
                        ft.DataCell(ft.Text(str(row['Part_Number']), color="#00bcd4", size=11, weight=ft.FontWeight.BOLD)),
                        ft.DataCell(ft.Container(
                            content=ft.Row([
                                ft.Text(status_icon, size=12),
                                ft.Text(str(row['Status']), color="#ffffff", size=10, weight=ft.FontWeight.BOLD)
                            ], spacing=5),
                            bgcolor=status_color,
                            padding=ft.padding.symmetric(horizontal=8, vertical=4),
                            border_radius=5
                        )),
                        ft.DataCell(ft.Text("✅" if row['En_401'] == 'SÍ' else "❌", color="#4CAF50" if row['En_401'] == 'SÍ' else "#F44336", size=12)),
                        ft.DataCell(ft.Text("✅" if row['En_403'] == 'SÍ' else "❌", color="#4CAF50" if row['En_403'] == 'SÍ' else "#F44336", size=12)),
                        ft.DataCell(ft.Text(str(row['Description_401'])[:25] + "..." if len(str(row['Description_401'])) > 25 else str(row['Description_401']), 
                                          color="#2196F3", size=10)),
                        ft.DataCell(ft.Text(str(row['Description_403'])[:25] + "..." if len(str(row['Description_403'])) > 25 else str(row['Description_403']), 
                                          color="#FF6B35", size=10)),
                        ft.DataCell(ft.Text(str(row['Quantity_401']), color="#2196F3", size=11)),
                        ft.DataCell(ft.Text(str(row['Quantity_403']), color="#FF6B35", size=11)),
                        ft.DataCell(ft.Text(str(row['Embraer_Code_401']), color="#9C27B0", size=10)),
                        ft.DataCell(ft.Text(str(row['Embraer_Code_403']), color="#9C27B0", size=10)),
                        ft.DataCell(ft.Text(str(row['Changes'])[:30] + "..." if len(str(row['Changes'])) > 30 else str(row['Changes']), 
                                          color="#FFEB3B", size=9))
                    ]
                )
            )
        
        filtered_table = ft.DataTable(
            columns=[
                ft.DataColumn(ft.Text("Part Number", weight=ft.FontWeight.BOLD, color="#ffffff", size=11)),
                ft.DataColumn(ft.Text("SQL Status", weight=ft.FontWeight.BOLD, color="#ffffff", size=11)),
                ft.DataColumn(ft.Text("401", weight=ft.FontWeight.BOLD, color="#2196F3", size=11)),
                ft.DataColumn(ft.Text("403", weight=ft.FontWeight.BOLD, color="#FF6B35", size=11)),
                ft.DataColumn(ft.Text("Description 401", weight=ft.FontWeight.BOLD, color="#2196F3", size=10)),
                ft.DataColumn(ft.Text("Description 403", weight=ft.FontWeight.BOLD, color="#FF6B35", size=10)),
                ft.DataColumn(ft.Text("Qty 401", weight=ft.FontWeight.BOLD, color="#2196F3", size=10)),
                ft.DataColumn(ft.Text("Qty 403", weight=ft.FontWeight.BOLD, color="#FF6B35", size=10)),
                ft.DataColumn(ft.Text("Code 401", weight=ft.FontWeight.BOLD, color="#9C27B0", size=10)),
                ft.DataColumn(ft.Text("Code 403", weight=ft.FontWeight.BOLD, color="#9C27B0", size=10)),
                ft.DataColumn(ft.Text("Changes", weight=ft.FontWeight.BOLD, color="#FFEB3B", size=10))
            ],
            rows=rows,
            bgcolor="#1a1a1a",
            border_radius=10,
            heading_row_color="#2a2a2a"
        )
        
        self.scrollable_right_join_table.content = ft.Column(
            controls=[filtered_table],
            scroll=ft.ScrollMode.ALWAYS,
            auto_scroll=True
        )
    
    def update_bom_viewers(self):
        """Actualiza visores de BOMs individuales"""
        # Visor BOM 401
        if not self.bom_manager.bom_401_df.empty:
            self.bom_401_container.content = self.create_bom_table(
                self.bom_manager.bom_401_df, 
                "📋 BOM 401 (VIEJO)", 
                "#2196F3",
                "Base de datos original - lo que HABÍA"
            )
        else:
            self.bom_401_container.content = ft.Text("Carga BOM 401 para ver datos", color="#757575")
        
        # Visor BOM 403  
        if not self.bom_manager.bom_403_df.empty:
            self.bom_403_container.content = self.create_bom_table(
                self.bom_manager.bom_403_df,
                "📋 BOM 403 (NUEVO)",
                "#FF6B35", 
                "Base de datos actualizada - lo que HAY"
            )
        else:
            self.bom_403_container.content = ft.Text("Carga BOM 403 para ver datos", color="#757575")
    
    def create_bom_table(self, df, title, color, subtitle):
        """Crea tabla de BOM individual"""
        rows = []
        for _, row in df.iterrows():
            rows.append(
                ft.DataRow(
                    cells=[
                        ft.DataCell(ft.Text(str(row['Part_Number']), color=color, size=11, weight=ft.FontWeight.BOLD)),
                        ft.DataCell(ft.Text(str(row['Description'])[:40] + "..." if len(str(row['Description'])) > 40 else str(row['Description']), 
                                          color="#ffffff", size=10)),
                        ft.DataCell(ft.Text(str(row['Embraer_Code']), color="#9C27B0", size=11)),
                        ft.DataCell(ft.Text(str(row['Quantity']), color="#4CAF50", size=11)),
                        ft.DataCell(ft.Text(str(row['Unit']), color="#FFEB3B", size=10)),
                        ft.DataCell(ft.Text(str(row['Weight']), color="#FF9800", size=10))
                    ]
                )
            )
        
        bom_table = ft.DataTable(
            columns=[
                ft.DataColumn(ft.Text("Part Number", weight=ft.FontWeight.BOLD, color="#ffffff", size=11)),
                ft.DataColumn(ft.Text("Description", weight=ft.FontWeight.BOLD, color="#ffffff", size=11)),
                ft.DataColumn(ft.Text("Embraer Code", weight=ft.FontWeight.BOLD, color="#ffffff", size=11)),
                ft.DataColumn(ft.Text("Quantity", weight=ft.FontWeight.BOLD, color="#ffffff", size=11)),
                ft.DataColumn(ft.Text("Unit", weight=ft.FontWeight.BOLD, color="#ffffff", size=11)),
                ft.DataColumn(ft.Text("Weight", weight=ft.FontWeight.BOLD, color="#ffffff", size=11))
            ],
            rows=rows,
            bgcolor="#1a1a1a",
            border_radius=10,
            heading_row_color="#2a2a2a"
        )
        
        scrollable_table = ft.Container(
            content=ft.Column(
                controls=[bom_table],
                scroll=ft.ScrollMode.ALWAYS,
                auto_scroll=True
            ),
            height=500,
            bgcolor="#1a1a1a",
            border_radius=10,
            padding=10,
            border=ft.border.all(2, color)
        )
        
        return ft.Column([
            ft.Text(title, size=20, weight=ft.FontWeight.BOLD, color="#ffffff"),
            ft.Text(subtitle, color=color, size=14),
            ft.Text(f"📊 Total: {len(df):,} partes", color="#bdbdbd", size=12),
            ft.Container(height=10),
            scrollable_table
        ])
    
    def export_excel(self, e):
        """Exporta con las 3 hojas: BOM 401, BOM 403, RIGHT JOIN"""
        if self.bom_manager.comparison_df.empty:
            self.show_snackbar("⚠️ Ejecuta RIGHT JOIN primero", "#ff9800")
            return
        
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"BOM_RIGHT_JOIN_Analysis_{timestamp}.xlsx"
            
            self.status_text.value = "🔄 Generando Excel con 3 hojas..."
            self.page.update()
            
            self.bom_manager.export_to_excel(filename)
            
            self.status_text.value = f"✅ Excel generado: {filename}"
            self.show_snackbar(f"✅ Excel RIGHT JOIN: {filename}", "#4caf50")
            
        except Exception as error:
            self.status_text.value = f"❌ Error: {str(error)}"
            self.show_snackbar(f"❌ Error: {str(error)}", "#f44336")
        
        self.page.update()
    
    def show_snackbar(self, message, color):
        """Muestra notificación"""
        self.page.snack_bar = ft.SnackBar(
            content=ft.Text(message, color="#ffffff"),
            bgcolor=color,
            duration=3000
        )
        self.page.snack_bar.open = True
        self.page.update()

def main(page: ft.Page):
    """Función principal del dashboard"""
    dashboard = BOMDashboard(page)

if __name__ == "__main__":
    print("🚀 INICIANDO BOM RIGHT JOIN DASHBOARD CON OCR")
    
    if not OCR_AVAILABLE:
        print("⚠️  IMPORTANTE: Para procesar PDFs imagen instala:")
        print("   pip install pytesseract pdf2image pillow")
        print("   También instala Tesseract: https://github.com/tesseract-ocr/tesseract")
        print("")
    
    print("🔗 SQL-Style Comparison Analysis")
    print("📊 Soporte para PDFs texto e imagen")
    print("🌐 Dashboard disponible en: http://localhost:8085")
    ft.app(target=main, port=8085)