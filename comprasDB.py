"""
About:  Company: EZ Air Interior.  Author/Owner: Jose Vazquez.  email: jose.vazquez@ezairinterior.com.
Phone: 614-173-80-04. Programming Language: Python. Creation Date: 09/07/2023.   All Rights Reserved.  se libera
"""
import re
import os
import sys
import pyodbc
import socket
import sqlite3
import webbrowser
import time
import os
import math
import numpy as np
import xlwings as xw
from datetime import timedelta
from datetime import datetime
import datetime as dt
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle, Font
import win32com.client
import win32com.client as win32
from PyQt5.QtWidgets import QPushButton
from PyQt5.QtGui import QIcon, QPixmap, QFont
import psutil
import calendar
from purch import Ui_MainWindow             #aqui
from purch_prd import Ui_Form               #aqui
from po_app import Ui_MainWindow as po      #aqui
from PyQt5.QtWidgets import QDialog
from PyQt5 import QtCore, QtWidgets
from cryptography.fernet import Fernet
from PyQt5.QtWidgets import  QCalendarWidget
from PyQt5.QtCore import QDate
from PyQt5.uic import loadUi 
from PyQt5 import QtCore
from PyQt5.QtGui import QIcon
from PyQt5.QtCore import QTimer
from PyQt5.QtGui import QKeyEvent
from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QMessageBox
from PyQt5.QtWidgets import QInputDialog, QLineEdit, QFileDialog
from cryptography.fernet import Fernet
from PyQt5.QtWidgets import  QMainWindow, QComboBox
from PyQt5.QtWidgets import QLabel
from PyQt5.QtCore import QTime
from PyQt5.QtWidgets import  QVBoxLayout, QWidget
from PyQt5.QtSql import QSqlDatabase, QSqlQuery, QSqlTableModel
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import  QTableWidget, QTableWidgetItem
from PyQt5.QtWidgets import QApplication, QTableView
from PyQt5.QtWidgets import QComboBox, QApplication, QMainWindow
from PyQt5.QtGui import QStandardItem, QStandardItemModel
from PyQt5.QtWidgets import QApplication, QMainWindow,QMessageBox
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from win32com.client import Dispatch
import pandas as pd
import xlwings as xw
from PyQt5.QtWidgets import QApplication, QMessageBox
import pythoncom
from win32com.client import Dispatch
from openpyxl.styles import PatternFill, Font
from datetime import date, timedelta
"""__________________________________GLOBAL VARIABLES________________________________________"""
# cpu = os.environ['USERNAME']
cpu = f"{os.environ['USERNAME']}.ONE"
today = dt.datetime.today().date()
datestring = dt.datetime.strftime(dt.datetime.now(), ' %m_%d_%Y')
"""__________________________________GENERAL PATHS__________________________________________"""

folder_db = "J:\\Departments\\Material Control\\Purchasing\\Tools\\ComprasDB\\"
path_conn = "J:\\Departments\\Material Control\\Purchasing\\Tools\\ComprasDB\\DataBase\\compras_DB.db"
path_conn_output = "J:\\Departments\\Material Control\\Purchasing\\Tools\\ComprasDB\\DataBase\\compras_db_reports_outputs.db"
icon_ezair = "J:\\Departments\\Operations\\Apps\\python3\\Icon\\LogoEZAirHD.JPG"
app_logo = "J:\\Departments\\Operations\\Apps\\python3\\Icon\\icons8-lego-64.ico"
body_mail = "J:\\Departments\\Operations\\Apps\\python3\\Icon\\icons8-composing-mail-96.png"
contact_mail = "J:\\Departments\\Operations\\Apps\\python3\\Icon\\icons8-contact-80.png"
open_order = "J:\\Departments\\Operations\\Apps\\python3\\Icon\\icons8-order-64.png"
cross_table = "J:\\Departments\\Operations\\Apps\\python3\\Icon\\icons8-data-transfer-100.png"
supplier = "J:\\Departments\\Operations\\Apps\\python3\\Icon\\icons8-delivery-64.png"

class LoginWindow(QMainWindow): #CLASS LOGIN
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Login")
        self.setGeometry(200, 200, 300, 300)  # Adjusted height to accommodate larger placeholders
        self.setWindowTitle("BUYER TOOLS")
        icon_ezair = "J:\\Departments\\Operations\\Apps\\python3\\Icon\\LogoEZAirHD.JPG"
        icon = QIcon(icon_ezair)
        self.setWindowIcon(icon)
        layout = QVBoxLayout()

        image_label = QLabel(self)
        pixmap = QPixmap(app_logo)  # Replace with your image path
        pixmap = pixmap.scaledToWidth(200)  # Adjust the width as needed
        image_label.setPixmap(pixmap)
        image_label.setAlignment(QtCore.Qt.AlignCenter)
        layout.addWidget(image_label)

        self.username_input = QLineEdit()
        self.username_input.setFont(QFont("Arial", 12))  # Set font and size for username input
        self.username_input.setPlaceholderText("Username")
        self.username_input.setAlignment(QtCore.Qt.AlignCenter)  # Center align text
        layout.addWidget(self.username_input)

        self.password_input = QLineEdit()
        self.password_input.setFont(QFont("Arial", 12))  # Set font and size for password input
        self.password_input.setEchoMode(QLineEdit.Password)
        self.password_input.setPlaceholderText("Password")
        self.password_input.setAlignment(QtCore.Qt.AlignCenter)  # Center align text
        layout.addWidget(self.password_input)

        login_button = QPushButton("Login")
        login_button.setFont(QFont("Arial", 12))  # Set font and size for the button
        login_button.clicked.connect(self.authenticate)
        layout.addWidget(login_button)

        self.error_label = QLabel()
        layout.addWidget(self.error_label)

        login_button.setFocus()  # Set focus to the login button after setting up the username input

        widget = QWidget()
        widget.setLayout(layout)
        self.setCentralWidget(widget)

    def authenticate(self):
        username = self.username_input.text()
        password = self.password_input.text()

        # Perform authentication - Replace this with your authentication logic
        valid_username = "Admin"
        valid_password = "pass18712"

        # open_order_username = 'OpenOrder'
        # open_order_password = 'pass18713'
        open_order_username = '1'
        open_order_password = '1'

        prd_username = 'Prod'
        prd_password = 'pass18712'

        po_username  = '4'
        po_password = '4'
        # prd_username = '2'
        # prd_password = '2'
        
        if username == valid_username and password == valid_password:
            self.close()  # Close the login window if authentication is successful
            self.main_app = MainApplication()  # Open the main application window
            self.main_app.show()
        elif username == open_order_username and password == open_order_password:
                self.close()  # Close the login window if authentication is successful
                self.main_app = MainApplication2()  # Open the main application window
                self.main_app.show()
        elif username == prd_username and password == prd_password:
                self.close()  # Close the login window if authentication is successful
                self.main_app = MainApplication3()  # Open the main application window
                self.main_app.show()
        


        else:
            self.error_label.setText("Invalid credentials")

class MainApplication(QMainWindow, Ui_MainWindow): #FULL BUYER APPLICATION
    def __init__(self):
        super().__init__()
        self.startTimer()  #CLOCK
        self.setupUi(self)
        icon = QIcon(f'{icon_ezair}')
        self.setWindowIcon(icon)
        self.setWindowTitle("BUYER TOOLS") 
        self.selected_date = None
        self.calendarWidget_2.clicked.connect(self.date_selected)
        self.pushButton.clicked.connect(self.create_expedite_table)
        self.pushButton_2.clicked.connect(self.create_in_9_2_table)
        self.pushButton_3.clicked.connect(self.create_po_3_51_table)
        self.pushButton_4.clicked.connect(self.create_oh_neteable_table)
        self.pushButton_5.clicked.connect(self.create_vendor_table)
        self.pushButton_6.clicked.connect(self.create_holidays_table)
        self.pushButton_7.clicked.connect(self.create_openOrder_table)
        self.pushButton_8.clicked.connect(self.create_rework_location_table)
        self.pushButton_9.clicked.connect(self.create_action_notes_table)
        self.pushButton_10.clicked.connect(self.validation_submit_path)
        self.pushButton_11.clicked.connect(self.registrar_exception_item_show)
        self.pushButton_12.clicked.connect(self.delete_item_exception_show)
        self.pushButton_13.clicked.connect(self.cancel_registro_button)
        self.pushButton_14.clicked.connect(self.create_std_entity_table)
        self.pushButton_15.clicked.connect(self.running_gemba_shortages)
        self.pushButton_16.clicked.connect(self.show_path_buttons_modification)
        self.pushButton_18.clicked.connect(self.cancel_change_path_button)
        self.pushButton_17.clicked.connect(self.modify_comments_path)
        self.pushButton_19.clicked.connect(self.create_projct_entity_table)
        self.pushButton_20.clicked.connect(self.running_pr_actions)
        self.pushButton_21.clicked.connect(self.create_transit_table)
        self.pushButton_28.clicked.connect(self.mail)
        self.pushButton_22.clicked.connect(self.create_buyer_table_fcst)
        self.pushButton_27.clicked.connect(self.send_supplier_email)
        self.pushButton_29.clicked.connect(self.mail2)
        self.pushButton_23.clicked.connect(self.create_ldna_oh)
        self.pushButton_24.clicked.connect(self.create_kit_project_table)
        self.pushButton_25.clicked.connect(self.create_qa_oh)
        self.pushButton_26.clicked.connect(self.create_table_dates)
        self.pushButton_32.clicked.connect(self.validate_good_liberation)
        self.pushButton_33.clicked.connect(self.create_wo_inquiry_table)
        self.pushButton_34.clicked.connect(self.new_future_shortage_all_forumas)
        self.pushButton_37.clicked.connect(self.create_where_use_table)
        self.pushButton_39.clicked.connect(self.find_information_selected_table)
        self.pushButton_41.clicked.connect(self.create_safety_stock_table)
        self.pushButton_42.clicked.connect(self.create_templete_safety_stock)
        self.pushButton_43.clicked.connect(self.create_should_be_templete_safety_stock)
        self.pushButton_44.clicked.connect(self.run_saefty_stock)
        self.pushButton_50.clicked.connect(self.create_shelf_life_table_material_in521)
        self.pushButton_51.clicked.connect(self.pr_all_materials)
        self.pushButton_52.clicked.connect(self.shelf_life_material_flag)
        self.pushButton_56.clicked.connect(self.add_venor_europeo_buttons)
        self.pushButton_59.clicked.connect(self.cancel_venors_europeo_buttons)
        self.pushButton_58.clicked.connect(self.registrar_vendor_europeo)
        self.pushButton_57.clicked.connect(self.show_delete_buttons_europeo)
        self.pushButton_60.clicked.connect(self.delete_vendor_europeo)
        self.label_2.setHidden(True)
        self.label_4.setHidden(True)
        self.label_5.setHidden(True)
        self.label_6.setHidden(True)
        self.label_8.setHidden(True)
        self.label_9.setHidden(True)
        self.label_10.setHidden(True)
        self.label_11.setHidden(True)
        self.label_12.setHidden(True)
        self.label_14.setHidden(True)
        self.label_20.setHidden(True)
        self.label_21.setHidden(True)
        self.label_34.setHidden(True)
        self.label_22.setHidden(True)
        self.label_60.setHidden(True)
        self.lineEdit.setHidden(True)
        self.lineEdit_2.setHidden(True)
        self.lineEdit_3.setHidden(True)
        self.pushButton_10.setHidden(True)
        self.pushButton_13.setHidden(True)
        self.pushButton_17.setHidden(True)
        self.pushButton_18.setHidden(True)
        self.pushButton_58.setHidden(True)
        self.pushButton_59.setHidden(True)
        self.pushButton_60.setHidden(True)
        self.lineEdit_8.setHidden(True)
        self.lineEdit_9.setHidden(True)
        self.label_23.setHidden(True)
        self.label_24.setHidden(True)
        self.label_25.setHidden(True)
        self.label_32.setHidden(True)
        self.label_36.setHidden(True)
        self.label_37.setHidden(True)
        self.label_13.setHidden(True)
        self.label_35.setHidden(True)    
        self.label_42.setHidden(True)  
        self.label_43.setHidden(True)  
        self.label_44.setHidden(True)  
        self.label_45.setHidden(True)  
        self.label_38.setHidden(True)  
        self.label_39.setHidden(True)  
        self.label_40.setHidden(True)  
        self.label_41.setHidden(True)
        self.label_53.setHidden(True)
        self.label_62.setHidden(True)
        self.label_59.setHidden(True)
        self.lineEdit.setPlaceholderText("Item")
        self.lineEdit_2.setPlaceholderText("Replace")
        self.show_last_actualization_table()
        self.show_exception_table()
        self.show_path_comments()
        self.show_supplier_email_table()
        self.show_supplier_email_table2()
        self.show_safety_stock_should_be()
        self.show_vendor_europeos()
        self.label_safety_stock_should_be()

        self.show_safety_stock()
        self.label_safety_stock()
        self.lineEdit_6.setPlaceholderText("Column Name")
        self.lineEdit_7.setPlaceholderText("Value to Find")
        self.lineEdit_8.setPlaceholderText("Vendor Code")
        self.lineEdit_9.setPlaceholderText(" Delete Vendor Code")

        pixmap = QPixmap(app_logo)
        scaled_pixmap = pixmap.scaledToWidth(200)
        self.label_48.setPixmap(scaled_pixmap)
        self.label_48.setScaledContents(True)
        pixmap = QPixmap(cross_table)
        scaled_pixmap = pixmap.scaledToWidth(300)
        self.label_57.setPixmap(scaled_pixmap)
        self.label_57.setScaledContents(True)


        """NO AVAILABLE TAB"""
        self.tabWidget.setTabEnabled(6, False)
        self.tabWidget.setTabEnabled(7, False)
        self.toolBox.setItemEnabled(1, False)
        self.tabWidget.setTabEnabled(4, False)

    """________________________________________INFORMATION DATABASE_____________________________________________________________"""

    def add_venor_europeo_buttons(self):
        self.pushButton_56.setHidden(True)
        self.pushButton_57.setHidden(True)
        self.pushButton_58.setHidden(False)
        self.pushButton_60.setHidden(True)
        self.pushButton_59.setHidden(False)
        self.lineEdit_8.setHidden(False)
        self.lineEdit_9.setHidden(True)

    def cancel_venors_europeo_buttons(self):
        self.pushButton_58.setHidden(True)
        self.pushButton_59.setHidden(True)
        self.pushButton_60.setHidden(True)
        self.lineEdit_8.setHidden(True)
        self.lineEdit_9.setHidden(True)
        self.pushButton_56.setHidden(False)
        self.pushButton_57.setHidden(False)

    def show_delete_buttons_europeo(self):
        self.pushButton_58.setHidden(True)
        self.pushButton_59.setHidden(False)
        self.pushButton_60.setHidden(False)
        self.lineEdit_9.setHidden(False)
        self.pushButton_56.setHidden(True)
        self.pushButton_57.setHidden(True)

    def initial_state(self):
        self.pushButton_56.setHidden(False)
        self.pushButton_57.setHidden(False)
        self.pushButton_58.setHidden(True)
        self.pushButton_60.setHidden(True)
        self.pushButton_59.setHidden(True)
        self.lineEdit_8.setHidden(True)
        self.lineEdit_9.setHidden(True)

    def create_shelf_life_table_material_in521(self):
        try:
            self.label_59.setHidden(False)
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()
            cursor.execute("""DELETE FROM in521""")
            conn.commit()
            cursor.execute("""CREATE TABLE IF NOT EXISTS in521(id INTEGER PRIMARY KEY,
                        Wh TEXT, ItemNo TEXT, Description TEXT, UM TEXT, Prod TEXT,ExpDate TEXT, Rec TEXT, QtyOH TEXT,
                        Bin TEXT, Lot TEXT, OrderBy TEXT, LRev TEXT, IRev TEXT, PlanType TEXT, EntCode TEXT, Proj TEXT,
                        ManufDate TEXT, Shelf_Life TEXT)""")
            conn.commit()
            try:
                conn = sqlite3.connect(path_conn)
                cursor = conn.cursor()
                expedite = 'in521_shelf_life.txt'
                
                df = pd.read_fwf(f'{folder_db}{expedite}', widths=[2,32,21,3,5,11,10,10,8,20,9,5,5,10,9,5,10,15], header=3)
                df.drop([0], axis=0, inplace=True)
                df = df.loc[df['Wh']!= 'En']
                df = df.rename(columns={'Item-No':'ItemNo','Exp-Date':'ExpDate','Rec.#':'Rec','Qty-OH':'QtyOH','Order By':'OrderBy','SL %':'Shelf_Life'})
                df.to_sql('in521', conn, if_exists='append', index=False)
                conn.commit()

                now = datetime.now()
                today = now.strftime("%m/%d/%Y %I:%M:%S")
                self.label_59.setText(f'{today}')
                
                
                self.last_day_shelf_life_tabel()
                self.show_last_actualization_table()

            except sqlite3.Error as error:
                QMessageBox.critical(self, "Error en base Expedite", f'No se pudo establecer coneccion {error}')
                self.label_59.setHidden(True)
            finally:
                conn.close()
            
        except sqlite3.Error as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
            self.label_59.setHidden(True)
        finally:
            conn.close()

    def create_expedite_table(self):
        self.label_23.setHidden(False)
        try:
            self.label_2.setHidden(False)
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()
            cursor.execute("""DELETE FROM expedite""")
            conn.commit()
            cursor.execute("""CREATE TABLE IF NOT EXISTS expedite(id INTEGER PRIMARY KEY
                        , EntityGroup TEXT
                        , Project TEXT
                        , AC TEXT
                        , ItemNo TEXT
                        , Description TEXT
                        , PlanTp TEXT
                        , Ref TEXT
                        , Sub TEXT
                        , FillDoc TEXT
                        , DemandType TEXT
                        , Sort TEXT, ReqQty TEXT, DemandSource TEXT, Unit TEXT
                        , Vendor TEXT, ReqDate TEXT, ShipDate TEXT, OH TEXT, MLIKCode TEXT, LT TEXT
                        , STDCost TEXT
                        , LotSize TEXT
                        , UOM TEXT)""")#creamos una tabla llamada expedite
            conn.commit()
            try:
                conn = sqlite3.connect(path_conn)
                cursor = conn.cursor()
                expedite = 'Expedite.csv'
                import_file = pd.read_csv(f'{folder_db}{expedite}')
                import_file = import_file[['Entity Group','Project','A/C','Item-No', 'Description'
                                        , 'PlanTp','Ref','Sub','Sort','Fill-Doc','Demand - Type','Req-Qty',
                                        'Demand-Source','Unit','Vendor','Req-Date','Ship-Date','OH','MLIK Code','LT','STD-Cost','Lot-Size','UOM']]

                import_file = import_file.rename(columns = {'Entity Group':'EntityGroup','A/C':'AC','Item-No':'ItemNo'
                                                ,'Req-Qty':'ReqQty','Demand-Source':'DemandSource','Req-Date':'ReqDate','Ship-Date':'ShipDate'
                                                ,'MLIK Code':'MLIKCode','STD-Cost':'STDCost','Lot-Size':'LotSize','Fill-Doc':'FillDoc','Demand - Type':'DemandType'})

                import_file.to_sql('expedite', conn, if_exists='append', index=False)
                conn.commit()

                now = datetime.now()
                today = now.strftime("%m/%d/%Y %I:%M:%S")
                self.label_2.setText(f'{today}')
                self.last_day_expedite()
                self.show_last_actualization_table()

            except sqlite3.Error as error:
                QMessageBox.critical(self, "Error en base Expedite", f'No se pudo establecer coneccion {error}')
                self.label_2.setHidden(True)
            finally:
                conn.close()
            
        except sqlite3.Error as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
            self.label_2.setHidden(True)
        finally:
            conn.close()
            self.label_23.setHidden(True)

    def create_in_9_2_table(self):
        try:
            self.label_4.setHidden(True)
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()
            cursor.execute("""DELETE FROM in92""")
            conn.commit()
            cursor.execute("""CREATE TABLE IF NOT EXISTS in92(id INTEGER PRIMARY KEY
                        , ItemNo TEXT, Vendor TEXT, LT TEXT, STDCost TEXT, PLANTYPE TEXT, MLIKCode TEXT, ACTIVE TEXT, LotSize TEXT)""")
            conn.commit()
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
            self.label_4.setHidden(True)
        finally:
            conn.close()
            
        try:
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()   
            in92 = 'in92.txt'
            import_file = (f'{folder_db}{in92}')
            with open(import_file, 'r') as f:
                lines = f.readlines()
            headers = lines[0].strip().split('@')
            data = [line.strip().split('@') for line in lines[1:] if len(line.strip().split('@')) == len(headers)]
            df = pd.DataFrame(data, columns=headers)

            df = df[['ITEM NUMBER','Vendor CD','L/T','STD COST 1','PLANTYPE','MLI','ACTIVE','LOT SIZE']]
            df = df.rename(columns={'ITEM NUMBER':'ItemNo','Vendor CD':'Vendor'
                                    ,'L/T':'LT','STD COST 1':'STDCost','MLI':'MLIKCode','LOT SIZE':'LotSize'})

            df['ItemNo'] = df['ItemNo'].str.upper()
            df['Vendor'] = df['Vendor'].str.upper()
            df['PLANTYPE'] = df['PLANTYPE'].str.upper()
            df['MLIKCode'] = df['MLIKCode'].str.upper()
            df['ACTIVE'] = df['ACTIVE'].str.upper()
            
            for col in df.columns:
                # Check if the column contains strings (i.e., if it's an object dtype)
                if df[col].dtype == 'object':
                    df[col] = df[col].str.strip()

        
            df.to_sql('in92', conn, if_exists='append', index=False)
            conn.commit()
            self.label_4.setHidden(False)
            now = datetime.now()
            today = now.strftime("%m/%d/%Y %I:%M:%S")
            self.label_4.setText(f'{today}')
            
            self.last_day_in92()
            self.create_shelf_life_by_item()
            self.show_last_actualization_table()
            QMessageBox.information(self, "Cargado de in92", "Se cargo el in 9 2 y in 9 2 shelf life correctamente")

        except sqlite3.Error as error:
            QMessageBox.critical(self, "Error en base in92", f'No se pudo establecer coneccion {error}')
            self.label_4.setHidden(True)
            conn.close()
        finally:
            conn.close()
            self.label_23.setHidden(True)

    def create_shelf_life_by_item(self):
        try:
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()
            cursor.execute("""DELETE FROM in92_Shelflife""")
            conn.commit()
            cursor.execute("""CREATE TABLE IF NOT EXISTS in92_Shelflife(id INTEGER PRIMARY KEY
                        , ItemNo TEXT, Vendor TEXT, LT TEXT,STD_Cost TEXT
                        , MLIKCode TEXT, ACTIVE TEXT, LotSize TEXT,SHELFLIFE TEXT)""")
            conn.commit()
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
    
        finally:
            conn.close()
            
        try:
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()   
            in92 = 'in92.txt'
            import_file = (f'{folder_db}{in92}')
            with open(import_file, 'r') as f:
                lines = f.readlines()
            headers = lines[0].strip().split('@')
            data = [line.strip().split('@') for line in lines[1:] if len(line.strip().split('@')) == len(headers)]
            df = pd.DataFrame(data, columns=headers)
            df = df[['ITEM NUMBER','Vendor CD','L/T','MLI','ACTIVE','LOT SIZE','SHELF-LIFE','STD COST 5']]
            df = df.rename(columns={'ITEM NUMBER':'ItemNo','Vendor CD':'Vendor'
                                    ,'L/T':'LT','STD COST 1':'STDCost'
                                    ,'MLI':'MLIKCode','LOT SIZE':'LotSize'
                                    ,'SHELF-LIFE':'SHELFLIFE','STD COST 5':'STD_Cost'})
            
            df['ItemNo'] = df['ItemNo'].str.upper()
            df['Vendor'] = df['Vendor'].str.upper()
            df['MLIKCode'] = df['MLIKCode'].str.upper()
            df['ACTIVE'] = df['ACTIVE'].str.upper()
            df['SHELFLIFE'] = df['SHELFLIFE'].str.upper()
            
            for col in df.columns:
                # Check if the column contains strings (i.e., if it's an object dtype)
                if df[col].dtype == 'object':
                    df[col] = df[col].str.strip()
            df.to_sql('in92_Shelflife', conn, if_exists='append', index=False)
            conn.commit()
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
    
        finally:
            conn.close()

    def create_po_3_51_table(self):
        try:
            self.label_5.setHidden(True)
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()
            cursor.execute("""DELETE FROM po351""")
            conn.commit()
            
            cursor.execute("""CREATE TABLE IF NOT EXISTS po351(id INTEGER PRIMARY KEY
                        , ItemNo TEXT
                        , Vendor TEXT
                        , FobVendor TEXT)""")
            conn.commit()
            
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
            self.label_4.setHidden(True)
        finally:
            conn.close()

        try:

            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()
            po351 = 'po351.txt'
            import_file = (f'{folder_db}{po351}')
            with open(import_file, 'r') as f:
                lines = f.readlines()
            headers = lines[0].strip().split('@')
            data = [line.strip().split('@') for line in lines[1:] if len(line.strip().split('@')) == len(headers)]
            df = pd.DataFrame(data, columns=headers)

            df = df.rename(columns={'Item-No':'ItemNo','endor-code':'Vendor','Fob-vendor[1]':'FobVendor'})

            for col in df.columns:
                if df[col].dtype == 'object':
                    df[col] = df[col].str.strip()

            df = df[['ItemNo','Vendor','FobVendor']]
            df.to_sql('po351', conn, if_exists='append', index=False)
            conn.commit()
            self.label_5.setHidden(False)
            now = datetime.now()
            today = now.strftime("%m/%d/%Y %I:%M:%S")
            self.label_5.setText(f'{today}')
            self.last_day_Po351()
            self.show_last_actualization_table()
            QMessageBox.information(self, "Cargado de PO 3 51", "Se cargo el PO 3 51 correctamente")

        except sqlite3.Error as error:
            QMessageBox.critical(self, "Error en base po 3 51", f'No se pudo establecer coneccion {error}')
            self.label_5.setHidden(True)
            conn.close()
        finally:
            conn.close()

    def create_oh_neteable_table_gemba_shortages(self):
        try:
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()
            cursor.execute("""DELETE FROM oh_gemba_shortages""")
            conn.commit()
            cursor.execute("""CREATE TABLE IF NOT EXISTS oh_gemba_shortages(id INTEGER PRIMARY KEY
                        , ItemNo TEXT
                        , Description TEXT
                        , Lot TEXT
                        , OH TEXT
                        , Bin Text)""")
            conn.commit()
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
            self.label_6.setHidden(True)
            conn.close()
        finally:
            conn.close()

        try:
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()
            oh_wq = 'ohNeteable.xlsx'
            ohfile = pd.read_excel(f'{folder_db}{oh_wq}',skiprows=4,dtype=str)
            
            for col in ohfile.columns:
                if ohfile[col].dtype == 'object':
                    ohfile[col] = ohfile[col].str.strip()
            
            # oh = oh.applymap(lambda x: x.replace("'", ""))
            for i, col in enumerate(ohfile.columns):
                ohfile.iloc[:, i] = ohfile.iloc[:, i].str.replace("'", '')
            ohfile = ohfile[['Item-Number','Description','Lot','OH','Bin']]
            ohfile['Bin'] = ohfile['Bin'].str.upper() 
            ohfile = ohfile.loc[ohfile['Bin']!='QA']
            ohfile = ohfile.rename(columns={'Item-Number':'ItemNo'})
            ohfile.to_sql('oh_gemba_shortages', conn, if_exists='append', index=False)
            conn.commit()
            self.label_6.setHidden(False)
            now = datetime.now()
            today = now.strftime("%m/%d/%Y %I:%M:%S")
            self.label_6.setText(f'{today}')
            
            self.last_day_ohnet()
            self.show_last_actualization_table()
            self.create_oh_pr_table()

        except sqlite3.Error as error:
            QMessageBox.critical(self, "Error en base OH", f'No se pudo establecer coneccion {error}')
            self.label_6.setHidden(True)
            conn.close()
        finally:
            conn.close()

    def create_oh_neteable_table(self):
        try:
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()
            cursor.execute("""DELETE FROM oh""")
            conn.commit()
            cursor.execute("""CREATE TABLE IF NOT EXISTS oh(id INTEGER PRIMARY KEY
                        , ItemNo TEXT
                        , Description TEXT
                        , Lot TEXT
                        , OH TEXT)""")
            conn.commit()
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
            self.label_6.setHidden(True)
            conn.close()
        finally:
            conn.close()

        try:
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()
            oh = 'ohNeteable.xlsx'
            oh = pd.read_excel(f'{folder_db}{oh}',skiprows=4,dtype=str)
            
            for col in oh.columns:
                if oh[col].dtype == 'object':
                    oh[col] = oh[col].str.strip()
            
            # oh = oh.applymap(lambda x: x.replace("'", ""))
            for i, col in enumerate(oh.columns):
                oh.iloc[:, i] = oh.iloc[:, i].str.replace("'", '')
            oh = oh[['Item-Number','Description','Lot','OH']]

            oh = oh.rename(columns={'Item-Number':'ItemNo'})
            oh.to_sql('oh', conn, if_exists='append', index=False)
            conn.commit()
            self.label_6.setHidden(False)
            now = datetime.now()
            today = now.strftime("%m/%d/%Y %I:%M:%S")
            self.label_6.setText(f'{today}')
            self.last_day_ohnet()
            self.show_last_actualization_table()
            self.create_oh_neteable_table_gemba_shortages()
            

        except sqlite3.Error as error:
            QMessageBox.critical(self, "Error en base OH", f'No se pudo establecer coneccion {error}')
            self.label_6.setHidden(True)
            conn.close()
        finally:
            conn.close()

    def create_vendor_table(self):
        try:
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='Vendor'")
            table_exists = cursor.fetchone()
            if table_exists:
                cursor.execute("""DELETE FROM Vendor""")
                conn.commit()
                cursor.execute("""CREATE TABLE IF NOT EXISTS Vendor(id INTEGER PRIMARY KEY
                        , Vendor TEXT
                        , Name TEXT
                        , Shortname TEXT
                        , Tactical TEXT
                        , Operational TEXT
                        , Process TEXT
                        , Strategic TEXT )""")
                conn.commit()
        except Exception as error:
            QMessageBox.critical(self, "DataBase Error", f'Impossible to reach the DataBase the error is: {error}')
        finally:
            conn.close()
        try:
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()
            vendor = 'vendor.xlsx'
            vendor = pd.read_excel(f'{folder_db}{vendor}',dtype=str)
            actual_vendor_columns = vendor.columns.to_list()
            vendor_columns_should_be = ['Vendor Code','Name','Shortname','Tactical','Operational','Process','Strategic']
            try:
                if actual_vendor_columns == vendor_columns_should_be:
                    for col in vendor.columns:
                        if vendor[col].dtype == 'object':
                            vendor[col] = vendor[col].str.strip()
                    for i, col in enumerate(vendor.columns):
                        vendor.iloc[:, i] = vendor.iloc[:, i].str.replace("'", '')
                    vendor = vendor.rename(columns={'Vendor Code':'Vendor','Sort-name':'Shortname'})
                    vendor.to_sql('Vendor', conn, if_exists='append', index=False)
                    conn.commit()
                    now = datetime.now()
                    today = now.strftime("%m/%d/%Y %I:%M:%S")
                    self.label_8.setHidden(False)
                    self.label_8.setText(f'{today}')
                    conn.close()
                    self.last_day_vendor()
                    self.show_last_actualization_table()
                    QMessageBox.information(self, "Cargado de Vendor", "Se cargo el Vendor correctamente")
                else:
                    QMessageBox.critical(self, "Error en archivo", f'el en el cargado  {error}')
            except Exception as error:
                QMessageBox.critical(self, "Error en archivo", f'el en el cargado  {error}')
            finally:
                pass
        except sqlite3.Error as error:
            QMessageBox.critical(self, "Error en base Vendor", f'No se pudo establecer coneccion {error}')
            self.label_8.setHidden(True)
            conn.close()
        finally:
            conn.close()

    def create_holidays_table(self):
        try:
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()
            cursor.execute("""DELETE FROM Holiday""")
            conn.commit()
            cursor.execute("""CREATE TABLE IF NOT EXISTS Holiday(id INTEGER PRIMARY KEY
                        , Holiday TEXT
                        , Day DATE )""")
            
            conn.commit()
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
            conn.close()
        finally:
            conn.close()
        try:
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()
            holyday = 'holiday.xlsx'
            holydays = pd.read_excel(f'{folder_db}{holyday}',dtype=str)
            
            for col in holydays.columns:
                if holydays[col].dtype == 'object':
                    holydays[col] = holydays[col].str.strip()
            
            # holydays = holydays.applymap(lambda x: x.replace("'", ""))
            for i, col in enumerate(holydays.columns):
                holydays.iloc[:, i] = holydays.iloc[:, i].str.replace("'", '')
            
            holydays.to_sql('Holiday', conn, if_exists='append', index=False)
            conn.commit()
            now = datetime.now()
            today = now.strftime("%m/%d/%Y %I:%M:%S")
            self.label_9.setHidden(False)
            self.label_9.setText(f'{today}')
            conn.close()
            self.last_day_holydays()
            self.show_last_actualization_table()
            QMessageBox.information(self, "Cargado de Holiday", "Se cargaron los Holidays correctamente")

        except sqlite3.Error as error:
            QMessageBox.critical(self, "Error en base Holiday", f'No se pudo establecer coneccion {error}')
            self.label_9.setHidden(True)
            conn.close()
        finally:
            conn.close()

    def create_openOrder_table(self):
        try:
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()
            cursor.execute("""DELETE FROM openOrder""")
            conn.commit()
            cursor.execute("""CREATE TABLE IF NOT EXISTS openOrder(id INTEGER PRIMARY KEY
                        , EntityCode TEXT, Proj TEXT, PONo TEXT, Ln TEXT, Type TEXT, WONo TEXT
                        , Vendor TEXT, VendorName TEXT, ItemNo TEXT, ItemDescription TEXT, Rev TEXT
                        , UM TEXT, OrdQ TEXT, RcvdQ TEXT, OpnQ TEXT, Stdcost TEXT, Unit$ TEXT, OpnVal TEXT, NetOpnVal TEXT, PODt TEXT, ReqDt TEXT
                        , PromDt TEXT, RevPrDt TEXT, BuyerName TEXT, LT TEXT, LineMemo TEXT, NoteStatus TEXT)""")
            conn.commit()
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
        finally:
            conn.close()
        try:
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()   
            oo = 'openOrder.xlsx'
            openOrder = pd.read_excel(f'{folder_db}{oo}',skiprows=3,dtype=str)
            for col in openOrder.columns:
                if openOrder[col].dtype == 'object':
                    openOrder[col] = openOrder[col].str.strip()
            # openOrder = openOrder.applymap(lambda x: x.replace("'", ""))
            
            for i, col in enumerate(openOrder.columns):
                openOrder.iloc[:, i] = openOrder.iloc[:, i].str.replace("'", '')
            
            
            
            openOrder['Item-Number'] = openOrder['Item-Number'].str.upper()
            # openOrder['left'] = openOrder['Item-Number'].str[-4:]
            # openOrder = openOrder.loc[openOrder['left']!='MISC']
            openOrder = openOrder[['Entity-Code','Proj', 'PO-No', 'Ln','Type','WO-No','Vendor #',
                                'Vendor Name', 'Item-Number','Item-Description', 'Rev', 'U/M', 'Ord-Q', 'Rcvd-Q', 'Opn-Q'
                                ,'Std-cost', 'Unit-$', 'Opn-Val', 'Net-Opn-Val','PO-Dt','Req-Dt', 'Prom-Dt', 'Rev-Pr-Dt','Buyer Name'
                                , 'LT', 'Line-Memo', 'Last Note/Status']]
            openOrder = openOrder.rename(columns={'Item-Number':'ItemNo','Entity-Code':'EntityCode','PO-No':'PONo'
                                                ,'WO-No':'WONo','Vendor #':'Vendor','Vendor Name':'VendorName'
                                                ,'Item-Description':'ItemDescription','U/M':'UM','Ord-Q':'OrdQ','Rcvd-Q':'RcvdQ'
                                                ,'Opn-Q':'OpnQ','Std-cost':'Stdcost','Unit-$':'Unit$'
                                                ,'Opn-Val':'OpnVal','Net-Opn-Val':'NetOpnVal'
                                                ,'PO-Dt':'PODt','Req-Dt':'ReqDt','Prom-Dt':'PromDt'
                                                ,'Rev-Pr-Dt':'RevPrDt','Buyer Name':'BuyerName'
                                                ,'Line-Memo':'LineMemo','Last Note/Status':'NoteStatus'})
            openOrder.to_sql('openOrder', conn, if_exists='append', index=False)
            conn.commit()
            self.label_10.setHidden(False)
            now = datetime.now()
            today = now.strftime("%m/%d/%Y %I:%M:%S")
            self.label_10.setText(f'{today}')
            self.last_day_openOrder()
            self.show_last_actualization_table()
            QMessageBox.information(self, "Cargado de OO", "Se cargo el Open Order correctamente")
        except sqlite3.Error as error:
            QMessageBox.critical(self, "Error en base OO", f'No se pudo establecer coneccion {error}')
            self.label_10.setHidden(True)
            conn.close()
        finally:
            conn.close()

    def create_rework_location_table_pr_actions(self):
        try:
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()
            cursor.execute("""DELETE FROM ReworkLoc_all""")
            conn.commit()
            cursor.execute("""CREATE TABLE IF NOT EXISTS ReworkLoc_all(id INTEGER PRIMARY KEY
                        , ItemNo TEXT
                        , Description TEXT
                        , Lot TEXT
                        , Expire_Date TEXT
                        , OH TEXT
                        , Bin TEXT)""")
            conn.commit()
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
            conn.close()
        finally:
            conn.close()

        try:
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()
            oh_rwk = 'reworkLoc.xlsx'
            rwk = pd.read_excel(f'{folder_db}{oh_rwk}',skiprows=2,dtype=str)
            
            for col in rwk.columns:
                if rwk[col].dtype == 'object':
                    rwk[col] = rwk[col].str.strip()

            for i, col in enumerate(rwk.columns):
                rwk.iloc[:, i] = rwk.iloc[:, i].str.replace("'", '')
            rwk = rwk[['Item-Number','Description','Lot','Expire-Date','OH','Bin']]
            
            current_date = datetime.today().date() 
            future_date = (pd.Timestamp(current_date) + pd.DateOffset(years=2)).normalize()
            rwk['Expire-Date'].fillna(future_date, inplace=True)
            rwk = rwk.rename(columns={'Item-Number':'ItemNo','Expire-Date':'Expire_Date'})
            rwk.to_sql('ReworkLoc_all', conn, if_exists='append', index=False)
            conn.commit()
        except sqlite3.Error as error:
            QMessageBox.critical(self, "Error en base OH", f'No se pudo establecer coneccion {error}')
            self.label_11.setHidden(True)
            conn.close()
        finally:
            conn.close()

    def create_rework_location_table(self):
        try:
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()
            cursor.execute("""DELETE FROM ReworkLoc""")
            conn.commit()
            cursor.execute("""CREATE TABLE IF NOT EXISTS ReworkLoc(id INTEGER PRIMARY KEY
                        , ItemNo TEXT
                        , Description TEXT
                        , Lot TEXT
                        , OH TEXT
                        , Bin TEXT)""")
            conn.commit()
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
            conn.close()
        finally:
            conn.close()

        try:
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()
            oh_rwk = 'reworkLoc.xlsx'
            rwk = pd.read_excel(f'{folder_db}{oh_rwk}',skiprows=2,dtype=str)
            
            for col in rwk.columns:
                if rwk[col].dtype == 'object':
                    rwk[col] = rwk[col].str.strip()
            # rwk = rwk.applymap(lambda x: x.replace("'", ""))

            for i, col in enumerate(rwk.columns):
                rwk.iloc[:, i] = rwk.iloc[:, i].str.replace("'", '')

            rwk = rwk[['Item-Number','Description','Lot','OH','Bin']]
            rwk = rwk.rename(columns={'Item-Number':'ItemNo'})
            rwk.to_sql('ReworkLoc', conn, if_exists='append', index=False)
            conn.commit()
            self.label_11.setHidden(False)
            now = datetime.now()
            today = now.strftime("%m/%d/%Y %I:%M:%S")
            self.label_11.setText(f'{today}')
            
            self.create_rework_location_table_pr_actions()
            self.last_day_rework()
            self.show_last_actualization_table()
            QMessageBox.information(self, "Cargado de material en Rework", "Se cargo el material en Rework correctamente")

        except sqlite3.Error as error:
            QMessageBox.critical(self, "Error en base OH", f'No se pudo establecer coneccion {error}')
            self.label_11.setHidden(True)
            conn.close()
        finally:
            conn.close()

    def create_action_notes_table(self):
        try:
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()
            cursor.execute("""DELETE FROM ActionMessages""")
            conn.commit()
            cursor.execute("""CREATE TABLE IF NOT EXISTS ActionMessages(id INTEGER PRIMARY KEY
                        , ItemNo TEXT
                        , ItemDesc TEXT
                        , PlanTp TEXT
                        , PlnrCode TEXT
                        , StdCost TEXT
                        , PO TEXT
                        , Line TEXT
                        , ACD TEXT
                        , ActQty TEXT
                        , FmQty TEXT
                        , ToQty TEXT      
                        , ADT TEXT
                        , REQDT TEXT
                        , Vendor TEXT
                        , Type TEXT
                        , OH TEXT
                        , MLI TEXT
                        , POLineNote TEXT
                        , PlanningNotes TEXT )""")
            conn.commit()
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
            conn.close()
        finally:
            conn.close()
        try:
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()
            actions = 'actionMessages.xlsx'
            am = pd.read_excel(f'{folder_db}{actions}',skiprows=3,dtype=str)
            for col in am.columns:
                if am[col].dtype == 'object':
                    am[col] = am[col].str.strip()
            for i, col in enumerate(am.columns):
                am.iloc[:, i] = am.iloc[:, i].str.replace("'", '')

            am = am[['ITEM-NO','ITEM-DESC','PlanTp','Plnr Code','Std-Cost','REF','SUB','A-CD'
            ,'Act-Qty','Fm-Qty','To-Qty','A-DT','REQ-DT','Vendor','Type','OH','MLI','PO Line Note','Planning Notes']]
            
            am = am.rename(columns={'ITEM-NO':'ItemNo'
                                    ,'ITEM-DESC':'ItemDesc'
                                    ,'Plnr Code':'PlnrCode'
                                    ,'Std-Cost': 'StdCost'
                                    ,'REF':'PO'
                                    ,'SUB':'Line'
                                    ,'A-CD':'ACD'
                                    ,'Act-Qty':'ActQty'
                                    ,'Fm-Qty':'FmQty'
                                    ,'To-Qty':'ToQty'
                                    ,'A-DT':'ADT'
                                    ,'REQ-DT':'REQDT'
                                    ,'PO Line Note':'POLineNote'
                                    ,'Planning Notes':'PlanningNotes'})

            am.to_sql('ActionMessages', conn, if_exists='append', index=False)
            conn.commit()
            now = datetime.now()
            today = now.strftime("%m/%d/%Y %I:%M:%S")
            self.label_12.setHidden(False)
            self.label_12.setText(f'{today}')
            conn.close()
            self.last_day_action_messages()
            self.show_last_actualization_table()
            QMessageBox.information(self, "Cargado de Action Messages", "Se cargo el Action Messages correctamente")

        except sqlite3.Error as error:
            QMessageBox.critical(self, "Error en base Action Messages", f'No se pudo establecer coneccion {error}')
            self.label_12.setHidden(True)
            conn.close()
        finally:
            conn.close()

    def validation_submit_path(self):
        if self.lineEdit.isVisible() and self.lineEdit_2.isVisible():
            if self.lineEdit.text() != '' and self.lineEdit_2.text() != '':
                try:
                    conexion = sqlite3.connect(path_conn)
                    cursor = conexion.cursor()
                    query = "INSERT INTO ExceptionItems (ItemNo,Replace) VALUES (?,?)"
                    cursor.execute(query, (self.lineEdit.text(),self.lineEdit_2.text()))
                    conexion.commit()
                    cursor.close()
                    QMessageBox.information(self, "Excepcion cargada ", f"registrado correctamente")
                except Exception as error:
                    QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
                finally:
                    conexion.close()
                    self.show_exception_table()
                    self.lineEdit.clear()
                    self.lineEdit_2.clear()
                    self.cancel_registro_button()
            else:
                QMessageBox.critical(self, 'Registro fallido', 'Favor de llenar todos los campos')
        elif self.lineEdit.isVisible():
            if  self.lineEdit.text() != '':
                try:
                    conexion = sqlite3.connect(path_conn)
                    cursor = conexion.cursor()
                    item_to_delete = self.lineEdit.text()
                    query = "Delete FROM ExceptionItems WHERE ItemNo = ?"
                    cursor.execute(query, (item_to_delete,))
                    conexion.commit()
                    cursor.close()

                    QMessageBox.information(self, "Excepcion cargada ", f"registrado correctamente")
                except Exception as error:
                    QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
                finally:
                    conexion.close()
                    self.show_exception_table()
                    self.lineEdit.clear()
                    self.lineEdit_2.clear()
                    self.cancel_registro_button()
            else:
                QMessageBox.critical(self, 'Registro fallido', 'Favor de llenar todos los campos')

    def create_std_entity_table(self):
        try:
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()
            cursor.execute("""DELETE FROM entity""")
            conn.commit()
            cursor.execute("""CREATE TABLE IF NOT EXISTS entity(id INTEGER PRIMARY KEY
                        , EntityGroup TEXT
                        , Name TEXT
                        , EntityName TEXT
                        , Commodity TEXT)""")
            conn.commit()
            self.last_day_entity()
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
        finally:
            conn.close()

        try:
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()
            entity_file = 'entity.xlsx'
            entity = pd.read_excel(f'{folder_db}{entity_file}',dtype=str)
            
            for col in entity.columns:
                if entity[col].dtype == 'object':
                    entity[col] = entity[col].str.strip()
            
        
            for i, col in enumerate(entity.columns):
                entity.iloc[:, i] = entity.iloc[:, i].str.replace("'", '')
            
            entity = entity[['Entity','Name','EntityName','Commodity']]
            entity = entity.rename(columns={'Entity':'EntityGroup'})
            entity.to_sql('entity', conn, if_exists='append', index=False)
            
            conn.commit()
            self.label_14.setHidden(False)
            now = datetime.now()
            today = now.strftime("%m/%d/%Y %I:%M:%S")
            self.label_14.setText(f'{today}')
            conn.close()
            self.last_day_entity()
            self.show_last_actualization_table()
            QMessageBox.information(self, "Cargado de STD Entity", "Se cargo el STD Entity correctamente")

        except sqlite3.Error as error:
            QMessageBox.critical(self, "Error en base Entity", f'No se pudo establecer coneccion {error}')
            self.label_14.setHidden(True)
            
        finally:
            conn.close()

    def create_transit_table(self):
        try:
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()
            cursor.execute("""DELETE FROM transit""")
            conn.commit()
            cursor.execute("""CREATE TABLE IF NOT EXISTS transit(id INTEGER PRIMARY KEY
                        , Vendor TEXT
                        , VendorName TEXT
                        , TransitTime TEXT)""")
            conn.commit()
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
            conn.close()
        finally:
            conn.close()
            print('DONE')
        try:
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()
            transit_table = 'transit_table.xlsx'
            transit = pd.read_excel(f'{folder_db}{transit_table}',dtype=str)
            for col in transit.columns:
                if transit[col].dtype == 'object':
                    transit[col] = transit[col].str.strip()
            for i, col in enumerate(transit.columns):
                transit.iloc[:, i] = transit.iloc[:, i].str.replace("'", '')
            
            
            transit.to_sql('transit', conn, if_exists='append', index=False)
            conn.commit()
            now = datetime.now()
            today = now.strftime("%m/%d/%Y %I:%M:%S")
            self.label_21.setHidden(False)
            self.label_21.setText(f'{today}')
            conn.close()


            self.last_day_transit_table()
            self.show_last_actualization_table()
            QMessageBox.information(self, "Cargado de transit table", "Se cargo la tabla correctamente")

        except sqlite3.Error as error:
            QMessageBox.critical(self, "Error en base Action Messages", f'No se pudo establecer coneccion {error}')
            self.label_21.setHidden(True)
            conn.close()
        finally:
            conn.close()

    def modify_comments_path(self):
        if self.lineEdit_3.text() !='':
            try:
                connection = sqlite3.connect(path_conn)
                cursor = connection.cursor()
                query = "UPDATE paths SET value = ? WHERE path = ?"
                cursor.execute(query, ( self.lineEdit_3.text(),'BuyerInformation'))
                connection.commit()
                self.cancel_change_path_button()
                self.show_path_comments()
                QMessageBox.information(self, "Modificacion de path", "Se actualizo el Path correctamente")
            except Exception as error:
                QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
            finally:
                connection.close()
        else:
            QMessageBox.critical(self, "Error en base de datos", f'introduce un path valido ')

    def create_projct_entity_table(self):
        try:
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()
            cursor.execute("""DELETE FROM entity_project""")
            conn.commit()
            cursor.execute("""CREATE TABLE IF NOT EXISTS entity_project(id INTEGER PRIMARY KEY
                        , EntityGroup TEXT
                        , Project TEXT)""")
            conn.commit()
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
        finally:
            conn.close()
        try:
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()
            project_entity_file = 'project_entity.xlsx'
            entity_project = pd.read_excel(f'{folder_db}{project_entity_file}',dtype=str)
            
            for col in entity_project.columns:
                if entity_project[col].dtype == 'object':
                    entity_project[col] = entity_project[col].str.strip()

            for i, col in enumerate(entity_project.columns):
                entity_project.iloc[:, i] = entity_project.iloc[:, i].str.replace("'", '')
            
            entity_project = entity_project[['EntityGroup','Project']]
            entity_project.to_sql('entity_project', conn, if_exists='append', index=False)
            conn.commit()
            self.label_20.setHidden(False)
            now = datetime.now()
            today = now.strftime("%m/%d/%Y %I:%M:%S")
            self.label_20.setText(f'{today}')
            conn.close()
            self.last_day_entity_project()
            self.show_last_actualization_table()
            QMessageBox.information(self, "Cargado de STD Entity", "Se cargo el STD Entity correctamente")

        except sqlite3.Error as error:
            QMessageBox.critical(self, "Error en base Entity", f'No se pudo establecer coneccion {error}')
            self.label_14.setHidden(True)
            
        finally:
            conn.close()

    def create_buyer_table_fcst(self):
        try:
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()
            cursor.execute("""DELETE FROM fcst_supplier""")
            conn.commit()
            cursor.execute("""CREATE TABLE IF NOT EXISTS fcst_supplier(id INTEGER PRIMARY KEY
                    , Vendor TEXT
                    , Shortname TEXT
                    , Owner1 TEXT
                    , Owner2 TEXT
                    , Owner3 TEXT
                    , email TEXT
                    , email2 TEXT
                    , email3 TEXT)""")
            conn.commit()
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
        finally:
            conn.close()
            print('DONE')
        try:
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()
            supplier_mail = 'supplier_mail.xlsx'
            supplier_mail = pd.read_excel(f'{folder_db}{supplier_mail}',dtype=str)
            for col in supplier_mail.columns:
                if supplier_mail[col].dtype == 'object':
                    supplier_mail[col] = supplier_mail[col].str.strip()
            
            # holydays = holydays.applymap(lambda x: x.replace("'", ""))
            for i, col in enumerate(supplier_mail.columns):
                supplier_mail.iloc[:, i] = supplier_mail.iloc[:, i].str.replace("'", '')
            supplier_mail.to_sql('fcst_supplier', conn, if_exists='append', index=False)
            conn.commit()
            now = datetime.now()
            today = now.strftime("%m/%d/%Y %I:%M:%S")
            self.label_22.setHidden(False)
            self.label_22.setText(f'{today}')
            self.last_day_supplier_email()
            self.show_last_actualization_table()
            self.show_supplier_email_table()
            self.show_supplier_email_table2()
            QMessageBox.information(self, "Cargado de Supplier mails", "Se cargaron correctamente")
        except sqlite3.Error as error:
            QMessageBox.critical(self, "Error en base Holiday", f'No se pudo establecer coneccion {error}')
            self.label_9.setHidden(True)
            
        finally:
            conn.close()

    def create_ldna_oh(self):
        try:
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()
            cursor.execute("""DELETE FROM LeanDNA""")
            conn.commit()
            cursor.execute("""CREATE TABLE IF NOT EXISTS LeanDNA(
                                id INTEGER PRIMARY KEY,
                                Item_Code TEXT,
                                Garden_Grove_C37_Inventory TEXT,
                                Bolsa_C91_Inventory TEXT,
                                Tijuana_C44_Inventory TEXT,
                                Santa_Maria_C61_Inventory TEXT,
                                Montreal_CG1_Inventory TEXT
                            )""")
            conn.commit()
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
        finally:
            conn.close()
        try:
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()
            leanDNA_reporte = 'LeanDNAoh.xlsx'
            leanDNAoh = pd.read_excel(f'{folder_db}{leanDNA_reporte}',dtype=str,skiprows=1)
            leanDNAoh = leanDNAoh.fillna(0)
            leanDNAoh = leanDNAoh[['Item Code','EZ Air Interior.1'
                                ,'Garden Grove C37 Inventory.1'
                                ,'Bolsa C91 Inventory.1'
                                ,'Tijuana C44 Inventory.1'
                                ,'Santa Maria C61 Inventory.1'
                                ,'Montreal CG1 Inventory.1']]
            leanDNAoh = leanDNAoh.rename(columns={'Item Code':'Item_Code'
                                                ,'Garden Grove C37 Inventory.1':'Garden_Grove_C37_Inventory'
                                                ,'Bolsa C91 Inventory.1':'Bolsa_C91_Inventory'
                                                ,'Tijuana C44 Inventory.1':'Tijuana_C44_Inventory'
                                                ,'Santa Maria C61 Inventory.1':'Santa_Maria_C61_Inventory'
                                                ,'Montreal CG1 Inventory.1':'Montreal_CG1_Inventory'})
            leanDNAoh = leanDNAoh[['Item_Code','Garden_Grove_C37_Inventory','Bolsa_C91_Inventory','Tijuana_C44_Inventory','Santa_Maria_C61_Inventory','Montreal_CG1_Inventory']]
            leanDNAoh.to_sql('LeanDNA', conn, if_exists='append', index=False)
            conn.commit()
            conn.close()
        except Exception as error:
            print(f'No se pudo establecer coneccion {error}')
        finally:
            conn.close()
            QMessageBox.information(self, "Cargado de OH LeanDNA", "Se cargo el OH de LeanDNA correctamente")
            self.label_23.setHidden(False)
            now = datetime.now()
            today = now.strftime("%m/%d/%Y %I:%M:%S")
            self.label_23.setText(f'{today}')
            self.last_day_LeanDNA()
            self.show_last_actualization_table()

    def create_kit_project_table(self):
        try:
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()
            cursor.execute("""DELETE FROM kpt""")
            conn.commit()
            cursor.execute("""CREATE TABLE IF NOT EXISTS kpt(
                                id INTEGER PRIMARY KEY,
                                EntityGroup TEXT,
                                Unit TEXT,
                                WorkCenter TEXT,
                                Sort TEXT,
                                Active TEXT,
                                leadtime TEXT,
                                Mfgtime TEXT,
                                RouteSort TEXT,
                                Sequence TEXT
                            )""")
            conn.commit()
        except Exception as error:
            print(f'el error es {error}')
        finally:
            conn.close()

        try:
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()
            df = 'KPT.txt'
            df = pd.read_fwf(f'J:\\Departments\\Material Control\\Purchasing\\Tools\\ComprasDB\\KPT.txt',widths=[9,13,21,21,10,7,9,9,10,12], header=3)
            df.drop([0], axis=0, inplace=True)
            df = df.loc[df['In-entity']!= 'End-of-Re']
            df = df.rename(columns = {'Entity-code':'EntityGroup','Sort-code':'Sort','Mfg-time':'Mfgtime','Route-Sort':'RouteSort'})
            df = df[['EntityGroup', 'Unit', 'WorkCenter', 'Sort', 'Active','leadtime', 'Mfgtime', 'RouteSort', 'Sequence']]
            df.to_sql('kpt', conn, if_exists='append', index=False)
            conn.commit()
        except Exception as error:
            print(f'No se pudo establecer coneccion {error}')
        finally:
            conn.close()
            QMessageBox.information(self, "Cargado de KIT PROJECT TABLE", "Se cargo el KPT correctamente")
            self.label_24.setHidden(False)
            now = datetime.now()
            today = now.strftime("%m/%d/%Y %I:%M:%S")
            self.label_24.setText(f'{today}')
            self.last_day_KPT()
            self.show_last_actualization_table()  

    def create_qa_oh(self):
        try:
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()
            cursor.execute("""DELETE FROM qa_oh""")
            conn.commit()
            cursor.execute("""CREATE TABLE IF NOT EXISTS qa_oh(
                                id INTEGER PRIMARY KEY,
                                ItemNo TEXT,
                                Description TEXT,
                                OH TEXT,
                                Bin TEXT,
                                Lot TEXT)""")
            conn.commit()
        except Exception as error:
            print(f'el error es {error}')
        finally:
            conn.close()

        try:
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()
            report = 'qaLoc.xlsx'
            df = pd.read_excel(f'J:\\Departments\\Material Control\\Purchasing\\Tools\\ComprasDB\\{report}',skiprows=3,dtype=str)
            df = df[['Item-Number','Description','OH','Bin','Lot']]
            df = df.rename(columns={'Item-Number':'ItemNo'})
            
            for col in df.columns:
                if df[col].dtype == 'object':
                    df[col] = df[col].str.strip()
            
            # oh = oh.applymap(lambda x: x.replace("'", ""))
            for i, col in enumerate(df.columns):
                df.iloc[:, i] = df.iloc[:, i].str.replace("'", '')

            df.to_sql('qa_oh', conn, if_exists='append', index=False)
            conn.commit()
        except Exception as error:
            print(f'No se pudo establecer coneccion {error}')
        finally:
            conn.close()
            QMessageBox.information(self, "Cargado de oh_QA", "Se cargo la localidad de QA correctamente")
            self.label_25.setHidden(False)
            now = datetime.now()
            today = now.strftime("%m/%d/%Y %I:%M:%S")
            self.label_25.setText(f'{today}')
            self.last_day_qa_oh()
            self.show_last_actualization_table()  

    def create_table_dates(self):
        try:
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()
            cursor.execute("""DELETE FROM table_yearweek""")
            conn.commit()
            cursor.execute("""CREATE TABLE IF NOT EXISTS table_yearweek(
                                id INTEGER PRIMARY KEY,
                                year_week TEXT,
                                Date TEXT)""")
            conn.commit()
        except Exception as error:
            print(f'el error es {error}')
        finally:
            conn.close()
        
        try:
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()
            report = 'year_week_table_dates.xlsx'
            df = pd.read_excel(f'J:\\Departments\\Material Control\\Purchasing\\Tools\\ComprasDB\\{report}',dtype=str)
            df = df[['year_week','Date']]
            df.to_sql('table_yearweek', conn, if_exists='append', index=False)
            conn.commit()
        except Exception as error:
            print(f'No se pudo establecer coneccion {error}')
        finally:
            conn.close()
            QMessageBox.information(self, "Cargado de table_yearweek", "Se cargo la table_yearweek correctamente")
            
            
            self.label_32.setHidden(False)
            now = datetime.now()
            today = now.strftime("%m/%d/%Y %I:%M:%S")
            self.label_32.setText(f'{today}')
            self.last_day_year_week_table()
            self.show_last_actualization_table()  

    def create_wo_inquiry_table(self):
        try:
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()
            cursor.execute("""DELETE FROM WO""")
            conn.commit()
            cursor.execute("""CREATE TABLE IF NOT EXISTS WOInquiry(
                                id INTEGER PRIMARY KEY, Entity TEXT, ProjectNo TEXT, WONo TEXT, SO_FCST TEXT, Sub TEXT
                                , ParentWO TEXT, ItemNumber TEXT, Rev TEXT, Description TEXT, AC TEXT, DueDt TEXT
                                , CreateDt TEXT, WoType TEXT, Srt TEXT, PlanType TEXT, OpnQ TEXT, St TEXT
                                , QAAprvl TEXT, PrtNo TEXT, PrtUser TEXT, Prtdate TEXT)""")
            conn.commit()
        except Exception as error:
            print(f'el error es {error}')
        finally:
            conn.close()
        
        try:
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()
            report = 'WO_Inquiry.xlsx'
            df = pd.read_excel(f'J:\\Departments\\Material Control\\Purchasing\\Tools\\ComprasDB\\{report}',dtype = str,skiprows = 3)
            
            df = df[['Entity','Project No','WO No','SO/FCST','Sub','Parent WO','Item-Number','Rev','Description'
                    ,'A/C','Due-Dt','Create-Dt','Wo-Type','Srt','Plan Type','Opn-Q','St','QA Aprvl','Prt-No'
                    ,'Prt-User','Prt-date']]
            
            df = df.rename(columns={'Project No':'ProjectNo','WO No':'WONo','SO/FCST':'SO_FCST'
                                    ,'Parent WO':'ParentWO','Item-Number':'ItemNumber','A/C':'AC'
                                    ,'Due-Dt':'DueDt','Create-Dt':'CreateDt','Wo-Type':'WoType'
                                    ,'Plan Type':'PlanType','Opn-Q':'OpnQ','QA Aprvl':'QAAprvl'
                                    ,'Prt-No':'PrtNo','Prt-User':'PrtUser','Prt-date':'Prtdate'})
            
            for col in df.columns:
                if df[col].dtype == 'object':
                    df[col] = df[col].str.strip()
            
            # oh = oh.applymap(lambda x: x.replace("'", ""))
            for i, col in enumerate(df.columns):
                df.iloc[:, i] = df.iloc[:, i].str.replace("'", '')
            
            
            
            
            df.to_sql('WOInquiry', conn, if_exists='append', index=False)
            conn.commit()
        except Exception as error:
            print(f'No se pudo establecer coneccion {error}')
        finally:
            conn.close()
            QMessageBox.information(self, "Cargado de WO_Inquiry", "Se cargo la WO_Inquiry correctamente")
            self.label_34.setHidden(False)
            now = datetime.now()
            today = now.strftime("%m/%d/%Y %I:%M:%S")
            self.label_34.setText(f'{today}')
            self.last_day_wo_inquiry()
            self.show_last_actualization_table()  

    def create_where_use_table(self):
        try:
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='where_use_table_plantype'")
            table_exists = cursor.fetchone()
            if table_exists:
                cursor.execute("""DELETE FROM where_use_table_plantype""")
                conn.commit()
                cursor.execute("""CREATE TABLE IF NOT EXISTS where_use_table_plantype(PlanTp TEXT, Area TEXT)""")
                conn.commit()
        except Exception as error:
            QMessageBox.critical(self, "DataBase Error", f'Impossible to reach the DataBase the error is: {error}')
        finally:
            conn.close()
        try:
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()
            where_use_file = 'where_use_plantype_table.xlsx'
            where_use = pd.read_excel(f'{folder_db}{where_use_file}',dtype=str)
            where_use_columns = where_use.columns.to_list()
            where_use_should_be_columns = ['PlanTp','Area']    
            if where_use_columns == where_use_should_be_columns:
                pass
            else:
                QMessageBox.critical(self, "Error en cargado del Where use table ", f'El archivo guardado no contiene las columnas esperadas: {where_use_should_be_columns} ')
                return
            for col in where_use.columns:
                if where_use[col].dtype == 'object':
                    where_use[col] = where_use[col].str.strip()
            for i, col in enumerate(where_use.columns):
                where_use.iloc[:, i] = where_use.iloc[:, i].str.replace("'", '')
            where_use = where_use.fillna('')
            where_use.to_sql('where_use_table_plantype', conn, if_exists='append', index=False)
            conn.commit()
            
            self.label_53.setHidden(False)
            now = datetime.now()
            today = now.strftime("%m/%d/%Y %I:%M:%S")
            self.label_53.setText(f'{today}')
            self.last_day_where_use_plantype_table()
            self.show_last_actualization_table()  
            QMessageBox.information(self, "Cargado del Where use", "Se cargo el Where use platype correctamente")
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion el error es:  {error}')
        finally:
            conn.close()

    def create_safety_stock_table(self):
        try:
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='safety_stock'")
            table_exists = cursor.fetchone()
            if table_exists:
                cursor.execute("""DELETE FROM safety_stock""")
                conn.commit()
                cursor.execute("""CREATE TABLE IF NOT EXISTS safety_stock(ItemNo TEXT, Qty TEXT)""")
                conn.commit()
        except Exception as error:
            QMessageBox.critical(self, "DataBase Error", f'Impossible to reach the DataBase the error is: {error}')
        finally:
            conn.close()
        try:
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()
            saefty_stock_file = 'safety_stock.txt'
            try:
                safety = pd.read_fwf(f'{folder_db}{saefty_stock_file}', widths=[9,32,30,4,13,13,13,6,9,100], header=3)
                safety = safety.fillna('')
                safety.drop([0], axis=0, inplace=True)
                safety = safety[(safety['Ent-Group']!='End-of-Re')]
                safety = safety[['Item-No','Safety-Stock']]
                safety = safety.rename(columns = {'Safety-Stock':'Qty','Item-No':'ItemNo'})
            except Exception as error:
                QMessageBox.critical(self, "DataBase Error", f'Impossible to reach the DataBase the error is: {error}')
                return
            finally:
                pass

            for col in safety.columns:
                if safety[col].dtype == 'object':
                    safety[col] = safety[col].str.strip()
            for i, col in enumerate(safety.columns):
                safety.iloc[:, i] = safety.iloc[:, i].str.replace("'", '')
            safety = safety.fillna('')
            safety.to_sql('safety_stock', conn, if_exists='append', index=False)
            conn.commit()
            
            self.label_60.setHidden(False)
            now = datetime.now()
            today = now.strftime("%m/%d/%Y %I:%M:%S")
            self.label_60.setText(f'{today}')
            
            
            self.last_day_safety_stock()
            self.show_safety_stock()
            self.show_last_actualization_table()

            QMessageBox.information(self, "Cargado del Safty Stock", "Se cargo el Safty Stock correctamente")
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion el error es:  {error}')
        finally:
            conn.close()

    def create_should_be_templete_safety_stock(self):
        try:
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='safety_stock_should_be'")
            table_exists = cursor.fetchone()
            if table_exists:
                cursor.execute("""DELETE FROM safety_stock_should_be""")
                conn.commit()
                cursor.execute("""CREATE TABLE IF NOT EXISTS safety_stock_should_be(ItemNo TEXT, New_Qty TEXT)""")
                conn.commit()
        except Exception as error:
            QMessageBox.critical(self, "DataBase Error", f'Impossible to reach the DataBase the error is: {error}')
        finally:
            conn.close()
        try:
            file_safety_stock, _ = QFileDialog.getOpenFileName(self, "Selecciona el Templete Safety", "", "Excel files (*.xlsx)")
            if file_safety_stock != "":
                self.file_safety_stock = file_safety_stock
                try:
                    conn = sqlite3.connect(path_conn)
                    cursor = conn.cursor()
                    df = pd.read_excel(f'{file_safety_stock}',dtype = str)
                    for col in df.columns:
                        if df[col].dtype == 'object':
                            df[col] = df[col].str.strip()
                    for i, col in enumerate(df.columns):
                        df.iloc[:, i] = df.iloc[:, i].str.replace("'", '')
                    df = df.fillna('')
                    df.to_sql('safety_stock_should_be', conn, if_exists='append', index=False)
                    conn.commit()
                except Exception as error:
                    print(error)
                finally:
                    pass
                self.show_safety_stock_should_be()
                self.last_day_safety_stock_should_be()
                QMessageBox.information(self, "Cargado del  Template Safty Stock", "Se cargo el Templete Safty Stock correctamente")
            else:
                QMessageBox.critical(self, "Seleccion de archivo", f'No se selecciono ningun archivo:  {error}')
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion el error es:  {error}')
        finally:
            conn.close()

    def create_oh_pr_table(self):
        try:
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()
            cursor.execute("""DELETE FROM oh_pr_table""")
            conn.commit()
            cursor.execute("""CREATE TABLE IF NOT EXISTS oh_pr_table(id INTEGER PRIMARY KEY
                        , ItemNo TEXT
                        , Description TEXT
                        , UM TEXT
                        , Lot TEXT
                        , Manuf_Date TEXT
                        , Expire_Date TEXT
                        , OH TEXT
                        , Bin Text)""")
            conn.commit()
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
            conn.close()
        finally:
            conn.close()
        try:
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()
            oh_wq = 'ohNeteable.xlsx'
            ohfile = pd.read_excel(f'{folder_db}{oh_wq}',skiprows=4,dtype=str)
            
            for col in ohfile.columns:
                if ohfile[col].dtype == 'object':
                    ohfile[col] = ohfile[col].str.strip()
            
            # oh = oh.applymap(lambda x: x.replace("'", ""))
            for i, col in enumerate(ohfile.columns):
                ohfile.iloc[:, i] = ohfile.iloc[:, i].str.replace("'", '')
            ohfile = ohfile[['Item-Number','Description','UM','Lot','Manuf-Date','Expire-Date','OH','Bin']]
            ohfile['Bin'] = ohfile['Bin'].str.upper() 
            
            ohfile = ohfile.rename(columns={'Item-Number':'ItemNo'
                                            ,'Manuf-Date':'Manuf_Date','Expire-Date':'Expire_Date'})
            
            # Convert 'Expire-Date' to datetime
            ohfile['Expire_Date'] = pd.to_datetime(ohfile['Expire_Date'])

            # Replace year 1900 with 2000
            ohfile['Expire_Date'] = ohfile['Expire_Date'].apply(lambda x: x.replace(year=x.year + 100 if x.year < 1950 else x.year))
            ohfile.to_sql('oh_pr_table', conn, if_exists='append', index=False)
            conn.commit()
            QMessageBox.information(self, "Cargado de On Hand", "Se cargo el OH Neteable correctamente")
        except sqlite3.Error as error:
            QMessageBox.critical(self, "Error en base OH", f'No se pudo establecer coneccion {error}')
            conn.close()
        finally:
            conn.close()

    """________________________________________SET LAST UPDATE__________________________________________________________________"""

    def last_day_shelf_life_tabel(self):
        try:
            connection = sqlite3.connect(path_conn)
            cursor = connection.cursor()
            query = "UPDATE actualizations SET LastDate = ? WHERE Report = ?"
            report = 'Shelf life table in 5 21'
            now = datetime.now()
            today = now.strftime("%m/%d/%Y %I:%M:%S")
            cursor.execute(query, (today, report))
            connection.commit()
            connection.close()
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
        finally:
            connection.close()

    def last_day_where_use_plantype_table(self):
        try:
            connection = sqlite3.connect(path_conn)
            cursor = connection.cursor()
            query = "UPDATE actualizations SET LastDate = ? WHERE Report = ?"
            report = 'Where use Plantype'
            now = datetime.now()
            today = now.strftime("%m/%d/%Y %I:%M:%S")
            cursor.execute(query, (today, report))
            connection.commit()
            connection.close()
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
        finally:
            connection.close()

    def last_day_wo_inquiry(self):
        try:
            connection = sqlite3.connect(path_conn)
            cursor = connection.cursor()
            query = "UPDATE actualizations SET LastDate = ? WHERE Report = ?"
            report = 'WO Inquiry'
            now = datetime.now()
            today = now.strftime("%m/%d/%Y %I:%M:%S")
            cursor.execute(query, (today, report))
            connection.commit()
            connection.close()
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
        finally:
            connection.close()

    def last_day_year_week_table(self):
        try:
            connection = sqlite3.connect(path_conn)
            cursor = connection.cursor()
            query = "UPDATE actualizations SET LastDate = ? WHERE Report = ?"
            report = 'Year week Date'
            now = datetime.now()
            today = now.strftime("%m/%d/%Y %I:%M:%S")
            cursor.execute(query, (today, report))
            connection.commit()
            connection.close()
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
        finally:
            connection.close()

    def last_day_qa_oh(self):
        try:
            connection = sqlite3.connect(path_conn)
            cursor = connection.cursor()
            query = "UPDATE actualizations SET LastDate = ? WHERE Report = ?"
            report = 'QA OH'
            now = datetime.now()
            today = now.strftime("%m/%d/%Y %I:%M:%S")
            cursor.execute(query, (today, report))
            connection.commit()
            connection.close()
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
        finally:
            connection.close()

    def last_day_KPT(self):
        try:
            connection = sqlite3.connect(path_conn)
            cursor = connection.cursor()
            query = "UPDATE actualizations SET LastDate = ? WHERE Report = ?"
            report = 'Kit Project Table'
            now = datetime.now()
            today = now.strftime("%m/%d/%Y %I:%M:%S")
            cursor.execute(query, (today, report))
            connection.commit()
            connection.close()
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
        finally:
            connection.close()

    def last_day_LeanDNA(self):
        try:
            connection = sqlite3.connect(path_conn)
            cursor = connection.cursor()
            query = "UPDATE actualizations SET LastDate = ? WHERE Report = ?"
            report = 'LeanDNA_OH'
            now = datetime.now()
            today = now.strftime("%m/%d/%Y %I:%M:%S")
            cursor.execute(query, (today, report))
            connection.commit()
            connection.close()
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
        finally:
            connection.close()

    def last_day_expedite(self):
        try:
            connection = sqlite3.connect(path_conn)
            cursor = connection.cursor()
            query = "UPDATE actualizations SET LastDate = ? WHERE Report = ?"
            report = 'Expedite'
            now = datetime.now()
            today = now.strftime("%m/%d/%Y %I:%M:%S")
            cursor.execute(query, (today, report))
            connection.commit()
            connection.close()
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
        finally:
            connection.close()

    def last_day_in92(self):
        try:
            connection = sqlite3.connect(path_conn)
            cursor = connection.cursor()
            query = "UPDATE actualizations SET LastDate = ? WHERE Report = ?"
            report = 'In 9 2'
            now = datetime.now()
            today = now.strftime("%m/%d/%Y %I:%M:%S")
            cursor.execute(query, (today, report))
            connection.commit()
            connection.close()
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
        finally:
            connection.close()

    def last_day_ohnet(self):
        try:
            connection = sqlite3.connect(path_conn)
            cursor = connection.cursor()
            query = "UPDATE actualizations SET LastDate = ? WHERE Report = ?"
            report = 'OH net'
            now = datetime.now()
            today = now.strftime("%m/%d/%Y %I:%M:%S")
            cursor.execute(query, (today, report))
            connection.commit()
            connection.close()
        
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
        finally:
            connection.close()

    def last_day_Po351(self):
        try:
            connection = sqlite3.connect(path_conn)
            cursor = connection.cursor()
            query = "UPDATE actualizations SET LastDate = ? WHERE Report = ?"
            report = 'Po 3 51'
            now = datetime.now()
            today = now.strftime("%m/%d/%Y %I:%M:%S")
            cursor.execute(query, (today, report))
            connection.commit()
            connection.close()
        
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
        finally:
            connection.close()

    def last_day_vendor(self):
        try:
            connection = sqlite3.connect(path_conn)
            cursor = connection.cursor()
            query = "UPDATE actualizations SET LastDate = ? WHERE Report = ?"
            report = 'Vendor'
            now = datetime.now()
            today = now.strftime("%m/%d/%Y %I:%M:%S")
            cursor.execute(query, (today, report))
            connection.commit()
            connection.close()
        
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
        finally:
            connection.close()

    def last_day_holydays(self):
        try:
            connection = sqlite3.connect(path_conn)
            cursor = connection.cursor()
            query = "UPDATE actualizations SET LastDate = ? WHERE Report = ?"
            report = 'Holiday'
            now = datetime.now()
            today = now.strftime("%m/%d/%Y %I:%M:%S")
            cursor.execute(query, (today, report))
            connection.commit()
            connection.close()
        
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
        finally:
            connection.close()

    def last_day_openOrder(self):
        try:
            connection = sqlite3.connect(path_conn)
            cursor = connection.cursor()
            query = "UPDATE actualizations SET LastDate = ? WHERE Report = ?"
            report = 'Open Order'
            now = datetime.now()
            today = now.strftime("%m/%d/%Y %I:%M:%S")
            cursor.execute(query, (today, report))
            connection.commit()
            connection.close()
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
        finally:
            connection.close()

    def last_day_rework(self):
        try:
            connection = sqlite3.connect(path_conn)
            cursor = connection.cursor()
            query = "UPDATE actualizations SET LastDate = ? WHERE Report = ?"
            report = 'Rework Loc'
            now = datetime.now()
            today = now.strftime("%m/%d/%Y %I:%M:%S")
            cursor.execute(query, (today, report))
            connection.commit()
            connection.close()
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
        finally:
            connection.close()

    def last_day_action_messages(self):
        try:
            connection = sqlite3.connect(path_conn)
            cursor = connection.cursor()
            query = "UPDATE actualizations SET LastDate = ? WHERE Report = ?"
            report = 'Action Message'
            now = datetime.now()
            today = now.strftime("%m/%d/%Y %I:%M:%S")
            cursor.execute(query, (today, report))
            connection.commit()
            connection.close()
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
        finally:
            connection.close()

    def last_day_entity(self):
        try:
            connection = sqlite3.connect(path_conn)
            cursor = connection.cursor()
            query = "UPDATE actualizations SET LastDate = ? WHERE Report = ?"
            report = 'Entity STD'
            now = datetime.now()
            today = now.strftime("%m/%d/%Y %I:%M:%S")
            cursor.execute(query, (today, report))
            connection.commit()
            connection.close()
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
        finally:
            connection.close()

    def last_day_entity_project(self):
        try:
            connection = sqlite3.connect(path_conn)
            cursor = connection.cursor()
            query = "UPDATE actualizations SET LastDate = ? WHERE Report = ?"
            report = 'Entity Project'
            now = datetime.now()
            today = now.strftime("%m/%d/%Y %I:%M:%S")
            cursor.execute(query, (today, report))
            connection.commit()
            connection.close()
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
        finally:
            connection.close()

    def last_day_supplier_email(self):
        try:
            connection = sqlite3.connect(path_conn)
            cursor = connection.cursor()
            query = "UPDATE actualizations SET LastDate = ? WHERE Report = ?"
            report = 'Supplier Email'
            now = datetime.now()
            today = now.strftime("%m/%d/%Y %I:%M:%S")
            cursor.execute(query, (today, report))
            connection.commit()
            connection.close()
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
        finally:
            connection.close()

    def last_day_transit_table(self):
        try:
            connection = sqlite3.connect(path_conn)
            cursor = connection.cursor()
            query = "UPDATE actualizations SET LastDate = ? WHERE Report = ?"
            report = 'Transit Table'
            now = datetime.now()
            today = now.strftime("%m/%d/%Y %I:%M:%S")
            cursor.execute(query, (today, report))
            connection.commit()
            connection.close()
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
        finally:
            connection.close()

    def last_day_report_table_dispersion_output(self):
        try:
            connection = sqlite3.connect(path_conn_output)
            cursor = connection.cursor()
            query = "UPDATE actualizations SET LastDate = ? WHERE Report = ?"
            report = 'Dispersion Table'
            now = datetime.now()
            today = now.strftime("%m/%d/%Y %I:%M:%S")
            cursor.execute(query, (today, report))
            connection.commit()
            connection.close()
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
        finally:
            connection.close()

    def last_day_safety_stock(self):
        try:
            connection = sqlite3.connect(path_conn)
            cursor = connection.cursor()
            query = "UPDATE actualizations SET LastDate = ? WHERE Report = ?"
            report = 'Safety Stock'
            now = datetime.now()
            today = now.strftime("%m/%d/%Y %I:%M:%S")
            cursor.execute(query, (today, report))
            connection.commit()
            connection.close()
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
        finally:
            connection.close()

    def last_day_safety_stock_should_be(self):
        try:
            connection = sqlite3.connect(path_conn)
            cursor = connection.cursor()
            query = "UPDATE actualizations SET LastDate = ? WHERE Report = ?"
            report = 'Safety Stock Should be'
            now = datetime.now()
            today = now.strftime("%m/%d/%Y %I:%M:%S")
            cursor.execute(query, (today, report))
            connection.commit()
            connection.close()
            self.label_62.setHidden(False)
            now = datetime.now()
            today = now.strftime("%m/%d/%Y %I:%M:%S")
            self.label_62.setText(f'{today}')
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
        finally:
            connection.close()

    """________________________________________REFRESH TABLES/GET INFORMATION___________________________________________________"""

    def show_safety_stock_should_be(self):
        try:
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()
            query = "SELECT * FROM safety_stock_should_be"
            result = cursor.execute(query)
            
            self.tableWidget_5.clear()
            self.tableWidget_5.setRowCount(0)
            self.tableWidget_5.setColumnCount(len(cursor.description))
            column_names = [description[0] for description in cursor.description]
            self.tableWidget_5.setHorizontalHeaderLabels(column_names)
            for row_number, row_data in enumerate(result):
                self.tableWidget_5.insertRow(row_number)
                for column_number, data in enumerate(row_data):
                    self.tableWidget_5.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
            
            # Ajustar automáticamente el ancho de las columnas
            self.tableWidget_5.resizeColumnsToContents()
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer conexión {error}')
        finally:
            conn.close()

    def show_safety_stock(self):
        try:
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()
            query = "SELECT * FROM safety_stock"
            result = cursor.execute(query)
            
            self.tableWidget_4.clear()
            self.tableWidget_4.setRowCount(0)
            self.tableWidget_4.setColumnCount(len(cursor.description))
            column_names = [description[0] for description in cursor.description]
            self.tableWidget_4.setHorizontalHeaderLabels(column_names)
            for row_number, row_data in enumerate(result):
                self.tableWidget_4.insertRow(row_number)
                for column_number, data in enumerate(row_data):
                    self.tableWidget_4.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
            
            # Ajustar automáticamente el ancho de las columnas
            self.tableWidget_4.resizeColumnsToContents()
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer conexión {error}')
        finally:
            conn.close()

    def show_supplier_email_table2(self):
        try:
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()
            query = "SELECT * FROM fcst_supplier"
            result = cursor.execute(query)
            self.tableWidget_3.clear()
            self.tableWidget_3.setRowCount(0)
            self.tableWidget_3.setColumnCount(len(cursor.description))
            column_names = [description[0] for description in cursor.description]
            self.tableWidget_3.setHorizontalHeaderLabels(column_names)
            for row_number, row_data in enumerate(result):
                self.tableWidget_3.insertRow(row_number)
                for column_number, data in enumerate(row_data):
                    self.tableWidget_3.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion holidays {error}')
        finally:
            conn.close()

    def show_supplier_email_table(self):
        try:
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()
            query = "SELECT * FROM fcst_supplier"
            result = cursor.execute(query)
            self.tableWidget_3.clear()
            self.tableWidget_3.setRowCount(0)
            self.tableWidget_3.setColumnCount(len(cursor.description))
            column_names = [description[0] for description in cursor.description]
            self.tableWidget_3.setHorizontalHeaderLabels(column_names)
            for row_number, row_data in enumerate(result):
                self.tableWidget_3.insertRow(row_number)
                for column_number, data in enumerate(row_data):
                    self.tableWidget_3.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion holidays {error}')
        finally:
            conn.close()

    def show_last_actualization_table(self):
        try:
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()
            query = "SELECT * FROM actualizations"
            result = cursor.execute(query)
            
            self.tableWidget.clear()
            self.tableWidget.setRowCount(0)
            self.tableWidget.setColumnCount(len(cursor.description))
            column_names = [description[0] for description in cursor.description]
            self.tableWidget.setHorizontalHeaderLabels(column_names)
            for row_number, row_data in enumerate(result):
                self.tableWidget.insertRow(row_number)
                for column_number, data in enumerate(row_data):
                    self.tableWidget.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
            
            # Ajustar automáticamente el ancho de las columnas
            self.tableWidget.resizeColumnsToContents()
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer conexión {error}')
        finally:
            conn.close()

    def show_exception_table(self):
        try:
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()
            query = "SELECT * FROM ExceptionItems"
            result = cursor.execute(query)
            
            self.tableWidget_2.clear()
            self.tableWidget_2.setRowCount(0)
            self.tableWidget_2.setColumnCount(len(cursor.description))
            column_names = [description[0] for description in cursor.description]
            self.tableWidget_2.setHorizontalHeaderLabels(column_names)
            for row_number, row_data in enumerate(result):
                self.tableWidget_2.insertRow(row_number)
                for column_number, data in enumerate(row_data):
                    self.tableWidget_2.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion holidays {error}')
        finally:
            conn.close()

    def get_oh_information_without_qa(self):
        try:
            conn = sqlite3.connect(path_conn)
            query_oh = "SELECT * FROM oh_gemba_shortages"
            query_exception_items = "SELECT * FROM ExceptionItems"
            oh = pd.read_sql_query(query_oh, conn)
            exception_items = pd.read_sql_query(query_exception_items, conn)
            oh['OH'] = oh['OH'].astype(float)
            oh.loc[:, 'ItemNo'] = oh['ItemNo'].str.upper()
            exception_items.loc[:, 'ItemNo'] = exception_items['ItemNo'].str.upper()
            oh = pd.merge(oh,exception_items,how='left',on = 'ItemNo')
            oh = oh.fillna('')
            def replace_items(DataFrame):
                for index,row in oh.iterrows():
                    if row['Replace'] !='':
                        DataFrame.at[index,'ItemNo'] = DataFrame.at[index,'Replace']
                    else:
                        pass
            replace_items(oh)

            piv_oh = pd.pivot_table(oh,index = 'ItemNo',values = 'OH',aggfunc = 'sum')
            piv_oh = piv_oh.reset_index()
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion OH {error}')
        finally:
            conn.close()
        return piv_oh

    def get_oh_table_information_pr(self):
        try:
            conn = sqlite3.connect(path_conn)
            query_oh = "SELECT * FROM oh_pr_table "
            oh = pd.read_sql_query(query_oh, conn)
        except Exception as error:
                QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion Expedite{error}')
        finally:
            conn.close()
        return oh

    def get_in521_information_table(self):
        try:
            conn = sqlite3.connect(path_conn)
            query_in521 = "SELECT * FROM in521 "
            in521 = pd.read_sql_query(query_in521, conn)
        except Exception as error:
                QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion Expedite{error}')
        finally:
            conn.close()
        return in521

    def get_oh_information(self):
        try:
            conn = sqlite3.connect(path_conn)
            query_oh = "SELECT * FROM oh"
            query_exception_items = "SELECT * FROM ExceptionItems"
            oh = pd.read_sql_query(query_oh, conn)
            exception_items = pd.read_sql_query(query_exception_items, conn)
            oh['OH'] = oh['OH'].astype(float)
            oh.loc[:, 'ItemNo'] = oh['ItemNo'].str.upper()
            exception_items.loc[:, 'ItemNo'] = exception_items['ItemNo'].str.upper()
            oh = pd.merge(oh,exception_items,how='left',on = 'ItemNo')
            oh = oh.fillna('')
            def replace_items(DataFrame):
                for index,row in oh.iterrows():
                    if row['Replace'] !='':
                        DataFrame.at[index,'ItemNo'] = DataFrame.at[index,'Replace']
                    else:
                        pass
            replace_items(oh)
            piv_oh = pd.pivot_table(oh,index = 'ItemNo',values = 'OH',aggfunc = 'sum')
            piv_oh = piv_oh.reset_index()
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion OH {error}')
        finally:
            conn.close()
        return piv_oh
    
    def get_vendor_europeo(self):
        try:
            conn = sqlite3.connect(path_conn)
            query_vendor_europeo = "SELECT * FROM europeos"
            
            vendor_europeos = pd.read_sql_query(query_vendor_europeo, conn)
            conn.close()
            return vendor_europeos
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion OH {error}')
        finally:
            conn.close()

    def get_expedite_information_clean_liberations(self):
        try:
            conn = sqlite3.connect(path_conn)
            query_expedite = "SELECT * FROM expedite WHERE expedite.FillDoc LIKE '%WO%'"
            expedite = pd.read_sql_query(query_expedite, conn)
        except Exception as error:
                QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion Expedite{error}')
        finally:
            conn.close()
        return expedite

    def get_expedite_information_for_gemba_shortages(self):
        try:
            conn = sqlite3.connect(path_conn)
            query_expedite = "SELECT * FROM expedite"
            expedite = pd.read_sql_query(query_expedite, conn)
            expedite.loc[:, 'EntityGroup'] = expedite['EntityGroup'].str.upper()
            expedite.loc[:, 'PlanTp'] = expedite['PlanTp'].str.upper()
            expedite.loc[:, 'MLIKCode'] = expedite['MLIKCode'].str.upper()
            expedite.loc[:, 'AC'] = expedite['AC'].str.upper()
            expedite['llave'] = expedite['EntityGroup'] + expedite['AC']  
            expedite['left'] = expedite['DemandSource'].str[:4]
            expedite['left'] = expedite['left'].str.upper()
            expedite['LT'] = expedite['LT'].astype(float)
            expedite = expedite.loc[(expedite['PlanTp'] !='VMIHDW')]
            expedite = expedite.loc[(expedite['MLIKCode'] =='L')]
            expedite = expedite.loc[(expedite['Ref'] !='Safety Stock')]
            expedite = expedite.loc[(expedite['left'] !='FCST')]
            expedite = expedite.fillna('')
            expedite['MLIKCode'] = expedite['MLIKCode'].str.upper()
            expedite[['Vendor', 'ItemNo']] = expedite[['Vendor', 'ItemNo']].apply(lambda x: x.str.upper().str.strip())
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion Expedite{error}')
        finally:
            conn.close()
        return expedite

    def shelf_table(self):
        try:
            conn = sqlite3.connect(path_conn)
            query_shelf = "SELECT * FROM in92_Shelflife"
            item_in92 = pd.read_sql_query(query_shelf, conn)
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion Expedite{error}')
        finally:
            conn.close()
        return item_in92

    def get_shelf_life_information_table(self):
        try:
            conn = sqlite3.connect(path_conn)
            query_shelf = "SELECT * FROM in92_Shelflife"
            shelf_table = pd.read_sql_query(query_shelf, conn)
            shelf_table = shelf_table.loc[shelf_table['SHELFLIFE']=='YES']
            

            query_shelf_vendor = "SELECT * FROM in92_Shelflife"
            shelf_table_vendor = pd.read_sql_query(query_shelf_vendor, conn)
            shelf_table_vendor = shelf_table_vendor[['ItemNo','Vendor']]
            shelf_table_vendor = shelf_table_vendor.drop_duplicates()
            
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion Expedite{error}')
        finally:
            conn.close()
        return shelf_table,shelf_table_vendor

    def get_expedite_information(self):
            try:
                conn = sqlite3.connect(path_conn)
                query_expedite = "SELECT * FROM expedite"
                expedite = pd.read_sql_query(query_expedite, conn)
                expedite.loc[:, 'EntityGroup'] = expedite['EntityGroup'].str.upper()
                expedite.loc[:, 'PlanTp'] = expedite['PlanTp'].str.upper()
                expedite.loc[:, 'MLIKCode'] = expedite['MLIKCode'].str.upper()
                expedite.loc[:, 'AC'] = expedite['AC'].str.upper()
                expedite['llave'] = expedite['EntityGroup'] + expedite['AC']  
                expedite['left'] = expedite['DemandSource'].str[:4]
                expedite['left'] = expedite['left'].str.upper()
                expedite['LT'] = expedite['LT'].astype(float)
                expedite = expedite.loc[(expedite['PlanTp'] !='VMIHDW')]
                expedite = expedite.loc[(expedite['MLIKCode'] =='L')]
                #expedite = expedite.loc[(expedite['Ref'] !='Safety Stock')]
                #expedite = expedite.loc[(expedite['left'] !='FCST')]
                expedite = expedite.fillna('')
                expedite['MLIKCode'] = expedite['MLIKCode'].str.upper()
                expedite[['Vendor', 'ItemNo']] = expedite[['Vendor', 'ItemNo']].apply(lambda x: x.str.upper().str.strip())
            except Exception as error:
                QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion Expedite{error}')
            finally:
                conn.close()
            return expedite

    def get_transit_information(self):
            try:
                conn = sqlite3.connect(path_conn)
                query_transit = "SELECT * FROM transit"
                transit = pd.read_sql_query(query_transit, conn)
                transit = transit.fillna('')
                transit = transit[['Vendor','TransitTime']]
                transit[['Vendor','TransitTime']] = transit[['Vendor','TransitTime']].apply(lambda x: x.str.upper().str.strip())
                transit = transit.drop_duplicates()
            except Exception as error:
                QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion in92 {error}')
            finally:
                conn.close()
            return transit

    def get_entity_information(self):
        try:
            conn = sqlite3.connect(path_conn)
            query_entity = "SELECT * FROM entity"
            entity = pd.read_sql_query(query_entity, conn)
            entity = entity.fillna('')
            entity.loc[:, 'EntityGroup'] = entity['EntityGroup'].str.upper()
            entity = entity[['EntityGroup','EntityName','Commodity']]
            entity = entity[['EntityGroup','EntityName','Commodity']].apply(lambda x: x.str.upper().str.strip())
            entity = entity.drop_duplicates()
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion entity {error}')
        finally:
            conn.close()
        return entity

    def get_vendor_information(self):
        try:
            conn = sqlite3.connect(path_conn)
            query_vendor = "SELECT * FROM Vendor"
            buyer = pd.read_sql_query(query_vendor, conn)
            buyer = buyer.fillna('')
            buyer['Vendor']  = buyer['Vendor'].str.upper().str.strip()
            buyer['Shortname']  = buyer['Shortname'].str.upper().str.strip()
            buyer = buyer[['Vendor','Shortname','Tactical']]
            buyer = buyer[['Vendor','Shortname','Tactical']].apply(lambda x: x.str.upper().str.strip())
            buyer = buyer.drop_duplicates()
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion vendor {error}')
        finally:
            conn.close()
        return buyer
    
    def get_vendor_information2(self):
        try:
            conn = sqlite3.connect(path_conn)
            query_vendor = "SELECT * FROM Vendor"
            buyer = pd.read_sql_query(query_vendor, conn)
            buyer = buyer.fillna('')
            buyer['Vendor']  = buyer['Vendor'].str.upper().str.strip()
            buyer['Shortname']  = buyer['Shortname'].str.upper().str.strip()
            buyer = buyer[['Vendor','Shortname','Tactical','Operational']]
            buyer = buyer[['Vendor','Shortname','Tactical','Operational']].apply(lambda x: x.str.upper().str.strip())
            buyer = buyer.drop_duplicates()
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion vendor {error}')
        finally:
            conn.close()
        return buyer

    def get_entity_project_information(self):
        try:
            conn = sqlite3.connect(path_conn)
            query_entity_project = "SELECT * FROM entity_project"
            project = pd.read_sql_query(query_entity_project, conn)
            project = project.fillna('')
            project['EntityGroup']  = project['EntityGroup'].str.upper().str.strip()
            project['Project']  = project['Project'].str.upper().str.strip()
            project = project.drop_duplicates()
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion vendor {error}')
        finally:
            conn.close()
        return project

    def show_path_comments(self):
        try:
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()
            query = "SELECT value FROM paths WHERE path = ?"
            selected_path = ('BuyerInformation',)  # Put the value in a tuple
            cursor.execute(query, selected_path)
            result = cursor.fetchone()  # Assuming you expect one row as result
            
            if result is not None:
                column_value = result[0]  # Assuming the value you want is in the first column
                self.label_16.setText(column_value)
            else:
                self.label_16.setText("No matching record found")
            
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer conección {error}')
        finally:
            conn.close()

    def date_selected(self):
        selected_date = self.calendarWidget_2.selectedDate()
        if not selected_date.isNull():
            self.selected_date = selected_date.toString("yyyy-MM-dd")
        else:
            QMessageBox.critical(self, "Error", "Please select a date.")

    def get_open_order_information(self):
        try:
            conn = sqlite3.connect(path_conn)
            open_order_query = "SELECT * FROM openOrder"
            oo = pd.read_sql_query(open_order_query, conn)
            oo = oo.fillna('')
            oo['ItemNo']  = oo['ItemNo'].str.upper().str.strip()
            return oo
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion vendor {error}')
        finally:
            conn.close()

    def get_rwk_information(self):
        try:
            conn = sqlite3.connect(path_conn)
            rework_query = "SELECT * FROM ReworkLoc"
            rwk = pd.read_sql_query(rework_query, conn)
            rwk = rwk.fillna('')
            rwk['ItemNo']  = rwk['ItemNo'].str.upper().str.strip()
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion vendor {error}')
        finally:
            conn.close()
        return rwk

    def get_po351_information(self):
        try:
            conn = sqlite3.connect(path_conn)
            query_po351 = "SELECT * FROM po351"
            po351 = pd.read_sql_query(query_po351, conn)
            po351 = po351.fillna('')
            po351 = po351[['ItemNo','Vendor','FobVendor']]
            po351 = po351[['ItemNo','Vendor','FobVendor']].apply(lambda x: x.str.upper().str.strip())
            po351 = po351.drop_duplicates()
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion po351 {error}')
        finally:
            conn.close()
        return po351

    def get_in92_information(self):
        try:
            conn = sqlite3.connect(path_conn)
            query_in92 = "SELECT * FROM in92"
            in_9_2 = pd.read_sql_query(query_in92, conn)
            in_9_2 = in_9_2.fillna('')
            in_9_2 = in_9_2[['ItemNo','Vendor','MLIKCode','LT']]
            in_9_2[['ItemNo','Vendor','MLIKCode','LT']] = in_9_2[['ItemNo','Vendor','MLIKCode','LT']].apply(lambda x: x.str.upper().str.strip())
            in_9_2 = in_9_2.fillna('')
            in_9_2 = in_9_2.loc[(in_9_2['MLIKCode']=='L')]
            # in_9_2 = in_9_2.loc[(in_9_2['Vendor']!='')]
            in_9_2 = in_9_2.drop_duplicates()
            return in_9_2
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion in92 {error}')
        finally:
            conn.close()

    def get_in92_information_send_mail(self):
        try:
            conn = sqlite3.connect(path_conn)
            query_in92 = "SELECT * FROM in92"
            in_9_2 = pd.read_sql_query(query_in92, conn)
            in_9_2 = in_9_2.fillna('')
            in_9_2 = in_9_2[['ItemNo','Vendor','MLIKCode']]
            in_9_2[['ItemNo','Vendor','MLIKCode']] = in_9_2[['ItemNo','Vendor','MLIKCode']].apply(lambda x: x.str.upper().str.strip())
            in_9_2 = in_9_2.fillna('')
            in_9_2 = in_9_2.loc[(in_9_2['MLIKCode']=='L')]
            # in_9_2 = in_9_2.loc[(in_9_2['Vendor']!='')]
            in_9_2 = in_9_2.drop_duplicates()
            return in_9_2
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion in92 {error}')
        finally:
            conn.close()

    def get_holidays_information(self):
        try:
            conn = sqlite3.connect(path_conn)
            query_holiday = "SELECT * FROM Holiday"
            holidays = pd.read_sql_query(query_holiday, conn)
            holidays = holidays.fillna('')
            holidays = holidays.drop_duplicates()
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion holidays {error}')
        finally:
            conn.close()
        return holidays

    def get_action_information(self):
        try:
            conn = sqlite3.connect(path_conn)
            query_actions = "SELECT * FROM ActionMessages"
            action = pd.read_sql_query(query_actions, conn)
            action = action.fillna('')
            action.loc[action['PlanningNotes']!='']
            action = action[['ItemNo','PlanningNotes']]
        except Exception as e:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion al acction {e}')
        finally:
            conn.close()
        return action
    
    def get_action_information2(self):
        try:
            conn = sqlite3.connect(path_conn)
            query_actions = "SELECT * FROM ActionMessages"
            action = pd.read_sql_query(query_actions, conn)
            action = action.fillna('')
            # action.loc[action['PlanningNotes']!='']
            action = action[['ItemNo','ACD','PlanningNotes']]
        except Exception as e:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion al acction {e}')
        finally:
            conn.close()
        return action

    def get_supplier_email(self):
        try:
            conn = sqlite3.connect(path_conn)
            query_actions = "SELECT * FROM fcst_supplier"
            buyer = pd.read_sql_query(query_actions, conn)
            buyer = buyer[['Vendor', 'Owner1','Owner2','Owner3','email','email2','email3','Shortname']]
            buyer['Owner1'] = buyer['Owner1'].str.split('@').str[0]
            return buyer
        except Exception as e:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion al acction {e}')
        finally:
            conn.close()

    def get_kpt_information(self):
        try:
            conn = sqlite3.connect(path_conn)
            query_actions = "SELECT * FROM kpt"
            kpt = pd.read_sql_query(query_actions, conn)
        except Exception as e:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion al acction {e}')
        finally:
            conn.close()
        return kpt

    def get_LeanDNA_information(self):
        try:
            conn = sqlite3.connect(path_conn)
            query_actions = "SELECT * FROM LeanDNA"
            leanDNA = pd.read_sql_query(query_actions, conn)
        except Exception as e:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion al acction {e}')
        finally:
            conn.close()
        return leanDNA

    def get_oh_qa_information(self):
        try:
            conn = sqlite3.connect(path_conn)
            query_qa_oh = "SELECT * FROM qa_oh"
            qa_oh = pd.read_sql_query(query_qa_oh, conn)
        except Exception as e:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion al acction {e}')
        finally:
            conn.close()
        return qa_oh

    def get_year_week_information(self):
        try:
            conn = sqlite3.connect(path_conn)
            query_yearweek_table = "SELECT * FROM table_yearweek"
            virtual_df = pd.read_sql_query(query_yearweek_table, conn)
        except Exception as e:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion al acction {e}')
        finally:
            conn.close()
        return virtual_df

    def get_wo_inquiry_information(self):
        try:
            conn = sqlite3.connect(path_conn)
            query_wo = "SELECT * FROM WOInquiry"
            woinquiry = pd.read_sql_query(query_wo, conn)
        except Exception as e:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion al acction {e}')
        finally:
            conn.close()
        return woinquiry
    
    def get_exception_items(self):
        try:
            conn = sqlite3.connect(path_conn)
            query_exception_items = "SELECT * FROM ExceptionItems"
            exception_items = pd.read_sql_query(query_exception_items, conn)
        except Exception as e:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion al acction {e}')
        finally:
            conn.close()
        return exception_items

    def get_rework_all_information_table(self):
        try:
            conn = sqlite3.connect(path_conn)
            rework_query = "SELECT * FROM ReworkLoc_all"
            rwk = pd.read_sql_query(rework_query, conn)
            rwk = rwk.fillna('')
            rwk['ItemNo']  = rwk['ItemNo'].str.upper().str.strip()
            rwk['OH']  = rwk['OH'].astype(float)
            return rwk
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion vendor {error}')
        finally:
            conn.close()
        return rwk

    def get_clean_expedite_information_future_shortages(self):
        in92  = self.get_in92_information()
        buyer = self.get_vendor_information_future_shortages()
        try:
            conn = sqlite3.connect(path_conn)
            query_expedite = "SELECT * FROM expedite"
            expedite = pd.read_sql_query(query_expedite, conn)
            expedite.loc[:, 'EntityGroup'] = expedite['EntityGroup'].str.upper()
            expedite.loc[:, 'PlanTp'] = expedite['PlanTp'].str.upper()
            expedite.loc[:, 'MLIKCode'] = expedite['MLIKCode'].str.upper()
            expedite.loc[:, 'AC'] = expedite['AC'].str.upper()
            expedite['llave'] = expedite['EntityGroup'] + expedite['AC']  
            expedite['left'] = expedite['DemandSource'].str[:4]
            expedite['left'] = expedite['left'].str.upper()
            expedite['LT'] = expedite['LT'].astype(float)
            expedite = expedite.loc[(expedite['PlanTp'] !='VMIHDW')]
            expedite = expedite.loc[(expedite['MLIKCode'] =='L')]
            expedite = expedite.loc[(expedite['DemandType'] !='SAFE')]
            
            expedite = expedite.fillna('')
            expedite['MLIKCode'] = expedite['MLIKCode'].str.upper()
            expedite[['Vendor', 'ItemNo']] = expedite[['Vendor', 'ItemNo']].apply(lambda x: x.str.upper().str.strip())

            expedite['ReqDate'] = pd.to_datetime(expedite['ReqDate'])
            expedite['week'] = expedite['ReqDate'].dt.isocalendar().week
            expedite['year'] = expedite['ReqDate'].dt.isocalendar().year
            expedite_fs = expedite.copy()
            expedite = expedite[['ItemNo', 'Description', 'UOM','ReqQty','STDCost','Vendor','PlanTp','LT','MLIKCode','ReqDate','year','week']]

            expedite_fs = expedite_fs[['ItemNo','Sort']]
            expedite_fs['Sort'] = expedite_fs['Sort'].str.upper()
            expedite_fs = expedite_fs.loc[(expedite_fs['Sort']=='FS')|(expedite_fs['Sort']=='FS1')|(expedite_fs['Sort']=='FS2')|(expedite_fs['Sort']=='KB')|(expedite_fs['Sort']=='KB4')|(expedite_fs['Sort']=='KB2')|(expedite_fs['Sort']=='KB1')]
            expedite_fs = expedite_fs.drop_duplicates(subset='ItemNo', keep='first')
            
            in92 = in92[['ItemNo','Vendor','MLIKCode','LT']]
            expedite = pd.merge(expedite,in92,how = 'left',on = 'ItemNo')
            
            expedite = expedite.fillna('')
            def select_correct_vendor(dataframe):
                dataframe['Correct_vendor'] = dataframe.apply(lambda row: row['Vendor_y'] if row['Vendor_y'] != '' else row['Vendor_x'], axis=1)
            select_correct_vendor(expedite)
            
            

            expedite['Correct_vendor'] = expedite['Correct_vendor'].replace('','No vendor')
            expedite['week'] = expedite['week'].astype(str)
            
            expedite = expedite[['ItemNo', 'Description', 'UOM', 'ReqQty', 'STDCost','PlanTp', 'MLIKCode_y', 'ReqDate', 'year', 'week', 'Vendor_y', 'LT_y','Vendor_x','Correct_vendor']]
            expedite = expedite.rename(columns = {'LT_y':'LT'})

            def validate_len(DataFrame):
                for index,row in DataFrame.iterrows():
                    value = len(row['week'])
                    if len(row['week']) == 1:
                        value2  = row['week']
                        DataFrame.at[index,'week'] = f"'0{value2}"
                    else:
                        pass
            validate_len(expedite)

            today = datetime.today()
            expedite['ReqDate'] = pd.to_datetime(expedite['ReqDate'])
            current_date = pd.to_datetime('today').normalize()
            last_monday = current_date - pd.DateOffset(days=current_date.dayofweek)
            expedite['week'] = expedite.apply(lambda row: '01_Past due' if row['ReqDate'] < last_monday else row['week'], axis=1)


            def remove_year_in_past_due(DataFrame):
                for index,row in DataFrame.iterrows():
                    if row['week'] == '01_Past due':
                        DataFrame.at[index,'year'] = np.nan
                    else:
                        pass
            remove_year_in_past_due(expedite)

            expedite['year_week'] = expedite['year'].astype(str) + expedite['week'].astype(str)
            expedite['year_week'] = expedite['year_week'].str.replace("'",'')
            expedite['year_week'] = expedite['year_week'].str.replace("<NA>",'')
            expedite['ItemNo'] = expedite['ItemNo'].str.upper()

            column_to_drop = 'Vendor_x'
            expedite.drop(column_to_drop, axis=1, inplace=True)
            column_to_drop = 'Vendor_y'
            expedite.drop(column_to_drop, axis=1, inplace=True)
            column_to_drop = 'MLIKCode_y'
            expedite.drop(column_to_drop, axis=1, inplace=True)
            column_to_drop = 'year'
            expedite.drop(column_to_drop, axis=1, inplace=True)   
            column_to_drop = 'week'
            expedite.drop(column_to_drop, axis=1, inplace=True)   
            expedite = expedite.rename(columns={'Correct_vendor':'Vendor'})
            expedite['ReqQty'] = expedite['ReqQty'].astype(float)
            
            expedite_1 = pd.merge(expedite,buyer,how = 'left',on = 'Vendor')
            expedite_1 = expedite_1[['ItemNo','Description','UOM','STDCost','Vendor','Shortname','PlanTp','LT','Owner1','Owner3']]
            expedite_1 = expedite_1.drop_duplicates()
            
            expedite_pivot = pd.pivot_table(expedite,index='ItemNo',columns = 'year_week',values = 'ReqQty',aggfunc='sum')

            return expedite_1,expedite_pivot,expedite_fs
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion Expedite{error}')
        finally:
            conn.close()
        return expedite_1,expedite_pivot,expedite_fs

    def get_safety_stock(self):
        try:
            conn = sqlite3.connect(path_conn)
            query_expedite = "SELECT * FROM expedite"
            expedite = pd.read_sql_query(query_expedite, conn)
            expedite.loc[:, 'EntityGroup'] = expedite['EntityGroup'].str.upper()
            expedite.loc[:, 'PlanTp'] = expedite['PlanTp'].str.upper()
            expedite.loc[:, 'MLIKCode'] = expedite['MLIKCode'].str.upper()
            expedite.loc[:, 'AC'] = expedite['AC'].str.upper()
            expedite['llave'] = expedite['EntityGroup'] + expedite['AC']  
            expedite['left'] = expedite['DemandSource'].str[:4]
            expedite['left'] = expedite['left'].str.upper()
            expedite['LT'] = expedite['LT'].astype(float)
            expedite['ReqQty'] = expedite['ReqQty'].astype(float)
            expedite = expedite.loc[(expedite['PlanTp'] !='VMIHDW')]
            expedite = expedite.loc[(expedite['MLIKCode'] =='L')]
            expedite = expedite.loc[(expedite['Ref'] =='Safety Stock')]
            expedite = expedite.fillna('')
            expedite['MLIKCode'] = expedite['MLIKCode'].str.upper()
            expedite[['Vendor', 'ItemNo']] = expedite[['Vendor', 'ItemNo']].apply(lambda x: x.str.upper().str.strip())
            piv_safe = pd.pivot_table(expedite,index = 'ItemNo',values = 'ReqQty',aggfunc='sum')
            piv_safe = piv_safe.rename(columns = {'ReqQty':'SS'})
            conn.close()
            return piv_safe
        
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion Expedite{error}')
        finally:
            pass

    def get_rework_information_table(self):
        try:
            conn = sqlite3.connect(path_conn)
            rework_query = "SELECT * FROM ReworkLoc"
            rwk = pd.read_sql_query(rework_query, conn)
            rwk = rwk.fillna('')
            rwk['ItemNo']  = rwk['ItemNo'].str.upper().str.strip()
            rwk['OH']  = rwk['OH'].astype(float)
            piv_rwk_location = pd.pivot_table(rwk,index = 'ItemNo',values = 'OH',aggfunc = 'sum')
            piv_rwk_location = piv_rwk_location.rename(columns = {'OH':'RWK OH'})
            piv_rwk_location = piv_rwk_location.reset_index()
            return piv_rwk_location
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion vendor {error}')
        finally:
            conn.close()
        return piv_rwk_location

    def get_action_information_open_order(self):
        try:
            conn = sqlite3.connect(path_conn)
            query_actions = "SELECT * FROM ActionMessages"
            action = pd.read_sql_query(query_actions, conn)
            action = action.fillna('')
            
        except Exception as e:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion al acction {e}')
        finally:
            conn.close()
        return action

    def get_vendor_information_future_shortages(self):
        try:
            conn = sqlite3.connect(path_conn)
            query_vendor = "SELECT * FROM Vendor"
            buyer = pd.read_sql_query(query_vendor, conn)
            buyer = buyer.fillna('')
            buyer['Vendor']  = buyer['Vendor'].str.upper().str.strip()
            buyer['Shortname']  = buyer['Shortname'].str.upper().str.strip()
            buyer = buyer[['Vendor','Shortname','Tactical','Strategic']]
            buyer = buyer[['Vendor','Shortname','Tactical','Strategic']].apply(lambda x: x.str.upper().str.strip())
            buyer = buyer[['Vendor','Shortname','Tactical','Strategic']]
            buyer = buyer.rename(columns={'Tactical':'Owner1','Strategic':'Owner3'})

            buyer = buyer.drop_duplicates()
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion vendor {error}')
        finally:
            conn.close()
        return buyer

    def get_clean_open_order(self):
        buyer = self.get_vendor_information_future_shortages()
        try:
            conn = sqlite3.connect(path_conn)
            open_order_query = "SELECT * FROM openOrder"
            oo = pd.read_sql_query(open_order_query, conn)
            oo = oo.fillna('')
            oo['ItemNo']  = oo['ItemNo'].str.upper().str.strip()
            
            oo = oo[['PONo', 'Ln','Vendor','VendorName','ItemNo','ItemDescription','Rev','UM','OrdQ'
                        ,'RcvdQ','OpnQ','Unit$','NetOpnVal','PODt','ReqDt','PromDt', 'RevPrDt','LT','NoteStatus']]

            oo = oo.applymap(lambda x: x.strip().upper() if isinstance(x, str) else x)

            oo['PODt'] = pd.to_datetime(oo['PODt'])
            oo['ReqDt'] = pd.to_datetime(oo['ReqDt'])
            oo['PromDt'] = pd.to_datetime(oo['PromDt'])
            oo['RevPrDt'] = pd.to_datetime(oo['RevPrDt'])

            today = datetime.today()
            monday = today - timedelta(days=today.weekday())
            monday_plus_4 = monday + timedelta(days=4)
            bday = pd.offsets.BDay()
            monday_plus_4_plus_11 = monday_plus_4 + 11 * bday

            def po_status(DataFrame):
                DataFrame['PO STATUS'] = ''
                for index, row in DataFrame.iterrows():
                    if row['RevPrDt'] < monday:
                        DataFrame.at[index, 'PO STATUS'] = 'Past due'
                    elif row['RevPrDt'] > monday and row['RevPrDt'] < monday_plus_4:
                        DataFrame.at[index, 'PO STATUS'] = 'Should be delivered'
                    elif row['RevPrDt'] >= monday_plus_4 and row['RevPrDt'] <= monday_plus_4_plus_11:
                        DataFrame.at[index, 'PO STATUS'] = 'Should be in Transit'
                    else:
                        DataFrame.at[index, 'PO STATUS'] = 'On time'

            po_status(oo)

            def po_recovery_week(DataFrame):
                DataFrame['PO Recovery WK'] = ''
                for index,row in DataFrame.iterrows():
                    if row['RevPrDt']< monday:
                        DataFrame.at[index,'PO Recovery WK'] = '01_Past due'
                    else:
                        if row['RevPrDt']>=monday and row['RevPrDt']<monday_plus_4:
                            DataFrame.at[index,'PO Recovery WK'] = '02_Should be delivered'
                        else:
                            if row['RevPrDt']>=monday_plus_4 and row['RevPrDt']<=monday_plus_4_plus_11:
                                DataFrame.at[index,'PO Recovery WK'] = '03_Should be in Transit'
                            else:
                                week = pd.Series(row['RevPrDt']).dt.isocalendar().week
                                year = pd.Series(row['RevPrDt']).dt.isocalendar().year
                                formatted_week = str(week.values[0]).zfill(2)
                                DataFrame.at[index, 'PO Recovery WK'] = f'{year.values[0]}{formatted_week}'
            
            po_recovery_week(oo)

            oo['key'] = oo['PONo'] +'-'+ oo['Ln']
            oo['PODt'] = oo['PODt'].dt.strftime("%m/%d/%Y")
            oo['ReqDt'] = oo['ReqDt'].dt.strftime("%m/%d/%Y")
            oo['PromDt'] = oo['PromDt'].dt.strftime("%m/%d/%Y")
            oo['RevPrDt'] = oo['RevPrDt'].dt.strftime("%m/%d/%Y")
            oo['Supplier Comments'] = ''
            oo['Confirmed Recovery'] = ''
            oo['Tracking'] = ''
            oo['Qty In Transit'] = ''
            oo['Supplier Comments'] = ''
            oo = oo.fillna('')
            oo['Confirm Status'] = oo['NoteStatus'].apply(lambda x: 'CONFIRMED' if pd.notna(x) and x != '' else 'NOT CONFIRMED')
            oo['OpnQ'] = oo['OpnQ'].astype(float)
            oo_pivot = pd.pivot_table(oo,index='ItemNo',columns = 'PO Recovery WK',values = 'OpnQ',aggfunc='sum')
            oo_pivot = oo_pivot.reset_index()                        
            
            oo = pd.merge(oo,buyer,how = 'left',on = 'Vendor')
            oo['Owner1'] = oo['Owner1'].fillna('No buyer') 
            oo = oo.rename(columns={'Owner1':'BUYER'})

            oo['AM'] = ''
            oo['EZA Req Date'] = ''
            oo['CONFIRMED RECOVERY'] = ''
            oo['Total Open QTY'] = ''
            oo['ETA'] = ''
            oo['Recibos'] = ''
            oo['Control'] = ''
            oo['Impo'] = ''

            oo = oo[['BUYER','key','PONo', 'Ln','Vendor', 'VendorName', 'ItemNo', 'ItemDescription',
                        'Rev', 'UM', 'OrdQ', 'RcvdQ', 'OpnQ', 'Unit$', 'NetOpnVal', 'PODt',
                        'ReqDt', 'PromDt', 'RevPrDt','AM','EZA Req Date', 'LT', 'NoteStatus','CONFIRMED RECOVERY','PO STATUS',
                        'PO Recovery WK','Total Open QTY','Qty In Transit','Tracking','ETA','Recibos','Control','Impo','Supplier Comments',
                        'Owner3','Shortname']]

            return oo,oo_pivot
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion vendor {error}')
        finally:
            conn.close()
        return oo,oo_pivot

    def get_where_use_information_table(self):
        try:
            conn = sqlite3.connect(path_conn)
            rework_query = "SELECT * FROM where_use_table_plantype"
            where_use = pd.read_sql_query(rework_query, conn)
            where_use = where_use.fillna('')
            where_use['PlanTp']  = where_use['PlanTp'].str.upper().str.strip()
            where_use['Area']  = where_use['Area'].str.upper().str.strip()
            conn.close()
            return where_use
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion vendor {error}')
        finally:
            conn.close()

    def find_information_selected_table(self):
        
        if self.radioButton_12.isChecked(): #IN 9 2
            if self.lineEdit_6.text() != '' and self.lineEdit_7.text() != '':
                query = f"""SELECT * FROM in92 WHERE {self.lineEdit_6.text()} = '{self.lineEdit_7.text()}'"""
            else:
                query = "SELECT * FROM in92"
            try:
                conn = sqlite3.connect(path_conn)
                cursor = conn.cursor()
                result = cursor.execute(query)
                self.tableWidget_7.clear()
                self.tableWidget_7.setRowCount(0)
                self.tableWidget_7.setColumnCount(len(cursor.description))
                column_names = [description[0] for description in cursor.description]
                self.tableWidget_7.setHorizontalHeaderLabels(column_names)
                for row_number, row_data in enumerate(result):
                    self.tableWidget_7.insertRow(row_number)
                    for column_number, data in enumerate(row_data):
                        self.tableWidget_7.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
                    # Ajustar automáticamente el ancho de las columnas
                self.tableWidget.resizeColumnsToContents()
            except Exception as error:
                QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
            finally:
                conn.close()
        
        elif self.radioButton_13.isChecked(): #OH NET
            if self.lineEdit_6.text() != '' and self.lineEdit_7.text() != '':
                query = f"""SELECT * FROM oh WHERE {self.lineEdit_6.text()} = '{self.lineEdit_7.text()}'"""
            else:
                query = "SELECT * FROM oh"
            try:
                conn = sqlite3.connect(path_conn)
                cursor = conn.cursor()
                result = cursor.execute(query)
                self.tableWidget_7.clear()
                self.tableWidget_7.setRowCount(0)
                self.tableWidget_7.setColumnCount(len(cursor.description))
                column_names = [description[0] for description in cursor.description]
                self.tableWidget_7.setHorizontalHeaderLabels(column_names)
                for row_number, row_data in enumerate(result):
                    self.tableWidget_7.insertRow(row_number)
                    for column_number, data in enumerate(row_data):
                        self.tableWidget_7.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
                # Ajustar automáticamente el ancho de las columnas
                self.tableWidget.resizeColumnsToContents()
            except Exception as error:
                QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
            finally:
                conn.close()

        elif self.radioButton_14.isChecked(): #QA_OH
            if self.lineEdit_6.text() != '' and self.lineEdit_7.text() != '':
                query = f"""SELECT * FROM qa_oh WHERE {self.lineEdit_6.text()} = '{self.lineEdit_7.text()}'"""
            else:
                query = "SELECT * FROM qa_oh"
            try:
                conn = sqlite3.connect(path_conn)
                cursor = conn.cursor()
                result = cursor.execute(query)
                self.tableWidget_7.clear()
                self.tableWidget_7.setRowCount(0)
                self.tableWidget_7.setColumnCount(len(cursor.description))
                column_names = [description[0] for description in cursor.description]
                self.tableWidget_7.setHorizontalHeaderLabels(column_names)
                for row_number, row_data in enumerate(result):
                    self.tableWidget_7.insertRow(row_number)
                    for column_number, data in enumerate(row_data):
                        self.tableWidget_7.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
                # Ajustar automáticamente el ancho de las columnas
                self.tableWidget.resizeColumnsToContents()
            except Exception as error:
                QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
            finally:
                conn.close()
        
        elif self.radioButton_15.isChecked(): #LEANDNA
            if self.lineEdit_6.text() != '' and self.lineEdit_6.text() != '':
                query = f"""SELECT * FROM LeanDNA WHERE {self.lineEdit_6.text()} = '{self.lineEdit_6.text()}'"""
            else:
                query = "SELECT * FROM LeanDNA"
            try:
                conn = sqlite3.connect(path_conn)
                cursor = conn.cursor()
                result = cursor.execute(query)
                self.tableWidget_7.clear()
                self.tableWidget_7.setRowCount(0)
                self.tableWidget_7.setColumnCount(len(cursor.description))
                column_names = [description[0] for description in cursor.description]
                self.tableWidget_7.setHorizontalHeaderLabels(column_names)
                for row_number, row_data in enumerate(result):
                    self.tableWidget_7.insertRow(row_number)
                    for column_number, data in enumerate(row_data):
                        self.tableWidget_7.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
                # Ajustar automáticamente el ancho de las columnas
                self.tableWidget.resizeColumnsToContents()
            except Exception as error:
                QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
            finally:
                conn.close()

        elif self.radioButton_16.isChecked(): #PO351
            if self.lineEdit_6.text() != '' and self.lineEdit_7.text() != '':
                query = f"""SELECT * FROM po351 WHERE {self.lineEdit_6.text()} = '{self.lineEdit_7.text()}'"""
            else:
                query = "SELECT * FROM po351"
            try:
                conn = sqlite3.connect(path_conn)
                cursor = conn.cursor()
                result = cursor.execute(query)
                self.tableWidget_7.clear()
                self.tableWidget_7.setRowCount(0)
                self.tableWidget_7.setColumnCount(len(cursor.description))
                column_names = [description[0] for description in cursor.description]
                self.tableWidget_7.setHorizontalHeaderLabels(column_names)
                for row_number, row_data in enumerate(result):
                    self.tableWidget_7.insertRow(row_number)
                    for column_number, data in enumerate(row_data):
                        self.tableWidget_7.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
                # Ajustar automáticamente el ancho de las columnas
                self.tableWidget.resizeColumnsToContents()
            except Exception as error:
                QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
            finally:
                conn.close()

        elif self.radioButton_20.isChecked(): #VENDOR
            if self.lineEdit_6.text() != '' and self.lineEdit_7.text() != '':
                query = f"""SELECT * FROM Vendor WHERE {self.lineEdit_6.text()} = '{self.lineEdit_7.text()}'"""
            else:
                query = "SELECT * FROM Vendor"
            try:
                conn = sqlite3.connect(path_conn)
                cursor = conn.cursor()
                result = cursor.execute(query)
                self.tableWidget_7.clear()
                self.tableWidget_7.setRowCount(0)
                self.tableWidget_7.setColumnCount(len(cursor.description))
                column_names = [description[0] for description in cursor.description]
                self.tableWidget_7.setHorizontalHeaderLabels(column_names)
                for row_number, row_data in enumerate(result):
                    self.tableWidget_7.insertRow(row_number)
                    for column_number, data in enumerate(row_data):
                        self.tableWidget_7.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
                # Ajustar automáticamente el ancho de las columnas
                self.tableWidget.resizeColumnsToContents()
            except Exception as error:
                QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
            finally:
                conn.close()
        
        elif self.radioButton_21.isChecked():#HOLIDAY
            if self.lineEdit_6.text() != '' and self.lineEdit_7.text() != '':
                query = f"""SELECT * FROM Holiday WHERE {self.lineEdit_6.text()} = '{self.lineEdit_7.text()}'"""
            else:
                query = "SELECT * FROM Holiday"
            try:
                conn = sqlite3.connect(path_conn)
                cursor = conn.cursor()
                result = cursor.execute(query)
                self.tableWidget_7.clear()
                self.tableWidget_7.setRowCount(0)
                self.tableWidget_7.setColumnCount(len(cursor.description))
                column_names = [description[0] for description in cursor.description]
                self.tableWidget_7.setHorizontalHeaderLabels(column_names)
                for row_number, row_data in enumerate(result):
                    self.tableWidget_7.insertRow(row_number)
                    for column_number, data in enumerate(row_data):
                        self.tableWidget_7.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
                # Ajustar automáticamente el ancho de las columnas
                self.tableWidget.resizeColumnsToContents()
            except Exception as error:
                QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
            finally:
                conn.close()

        elif self.radioButton_22.isChecked():#OPEN ORDER
            if self.lineEdit_6.text() != '' and self.lineEdit_7.text() != '':
                query = f"""SELECT * FROM openOrder WHERE {self.lineEdit_6.text()} = '{self.lineEdit_7.text()}'"""
            else:
                query = "SELECT * FROM openOrder"
            try:
                conn = sqlite3.connect(path_conn)
                cursor = conn.cursor()
                result = cursor.execute(query)
                self.tableWidget_7.clear()
                self.tableWidget_7.setRowCount(0)
                self.tableWidget_7.setColumnCount(len(cursor.description))
                column_names = [description[0] for description in cursor.description]
                self.tableWidget_7.setHorizontalHeaderLabels(column_names)
                for row_number, row_data in enumerate(result):
                    self.tableWidget_7.insertRow(row_number)
                    for column_number, data in enumerate(row_data):
                        self.tableWidget_7.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
                # Ajustar automáticamente el ancho de las columnas
                self.tableWidget.resizeColumnsToContents()
            except Exception as error:
                QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
            finally:
                conn.close()
    
        elif self.radioButton_23.isChecked():#LOCALIDAD REWORK
            if self.lineEdit_6.text() != '' and self.lineEdit_7.text() != '':
                query = f"""SELECT * FROM ReworkLoc WHERE {self.lineEdit_6.text()} = '{self.lineEdit_7.text()}'"""
            else:
                query = "SELECT * FROM ReworkLoc"
            try:
                conn = sqlite3.connect(path_conn)
                cursor = conn.cursor()
                result = cursor.execute(query)
                self.tableWidget_7.clear()
                self.tableWidget_7.setRowCount(0)
                self.tableWidget_7.setColumnCount(len(cursor.description))
                column_names = [description[0] for description in cursor.description]
                self.tableWidget_7.setHorizontalHeaderLabels(column_names)
                for row_number, row_data in enumerate(result):
                    self.tableWidget_7.insertRow(row_number)
                    for column_number, data in enumerate(row_data):
                        self.tableWidget_7.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
                # Ajustar automáticamente el ancho de las columnas
                self.tableWidget.resizeColumnsToContents()
            except Exception as error:
                QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
            finally:
                conn.close()
    
        elif self.radioButton_24.isChecked():# ACTION MESSAGES
            if self.lineEdit_6.text() != '' and self.lineEdit_7.text() != '':
                query = f"""SELECT * FROM ActionMessages WHERE {self.lineEdit_6.text()} = '{self.lineEdit_7.text()}'"""
            else:
                query = "SELECT * FROM ActionMessages"
            try:
                conn = sqlite3.connect(path_conn)
                cursor = conn.cursor()
                result = cursor.execute(query)
                self.tableWidget_7.clear()
                self.tableWidget_7.setRowCount(0)
                self.tableWidget_7.setColumnCount(len(cursor.description))
                column_names = [description[0] for description in cursor.description]
                self.tableWidget_7.setHorizontalHeaderLabels(column_names)
                for row_number, row_data in enumerate(result):
                    self.tableWidget_7.insertRow(row_number)
                    for column_number, data in enumerate(row_data):
                        self.tableWidget_7.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
                # Ajustar automáticamente el ancho de las columnas
                self.tableWidget.resizeColumnsToContents()
            except Exception as error:
                QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
            finally:
                conn.close()
        
        elif self.radioButton_25.isChecked():#ENTITY
            if self.lineEdit_6.text() != '' and self.lineEdit_7.text() != '':
                query = f"""SELECT * FROM entity WHERE {self.lineEdit_6.text()} = '{self.lineEdit_7.text()}'"""
            else:
                query = "SELECT * FROM entity"
            try:
                conn = sqlite3.connect(path_conn)
                cursor = conn.cursor()
                result = cursor.execute(query)
                self.tableWidget_7.clear()
                self.tableWidget_7.setRowCount(0)
                self.tableWidget_7.setColumnCount(len(cursor.description))
                column_names = [description[0] for description in cursor.description]
                self.tableWidget_7.setHorizontalHeaderLabels(column_names)
                for row_number, row_data in enumerate(result):
                    self.tableWidget_7.insertRow(row_number)
                    for column_number, data in enumerate(row_data):
                        self.tableWidget_7.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
                # Ajustar automáticamente el ancho de las columnas
                self.tableWidget.resizeColumnsToContents()
            except Exception as error:
                QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
            finally:
                conn.close()
        
        elif self.radioButton_26.isChecked(): #ENTITY PROJECT
            if self.lineEdit_6.text() != '' and self.lineEdit_7.text() != '':
                query = f"""SELECT * FROM entity_project WHERE {self.lineEdit_6.text()} = '{self.lineEdit_7.text()}'"""
            else:
                query = "SELECT * FROM entity_project"
            try:
                conn = sqlite3.connect(path_conn)
                cursor = conn.cursor()
                result = cursor.execute(query)
                self.tableWidget_7.clear()
                self.tableWidget_7.setRowCount(0)
                self.tableWidget_7.setColumnCount(len(cursor.description))
                column_names = [description[0] for description in cursor.description]
                self.tableWidget_7.setHorizontalHeaderLabels(column_names)
                for row_number, row_data in enumerate(result):
                    self.tableWidget_7.insertRow(row_number)
                    for column_number, data in enumerate(row_data):
                        self.tableWidget_7.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
                # Ajustar automáticamente el ancho de las columnas
                self.tableWidget.resizeColumnsToContents()
            except Exception as error:
                QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
            finally:
                conn.close()
        
        elif self.radioButton_27.isChecked():#TRANSIT TABLE
            if self.lineEdit_6.text() != '' and self.lineEdit_7.text() != '':
                query = f"""SELECT * FROM transit WHERE {self.lineEdit_6.text()} = '{self.lineEdit_7.text()}'"""
            else:
                query = "SELECT * FROM transit"
            try:
                conn = sqlite3.connect(path_conn)
                cursor = conn.cursor()
                result = cursor.execute(query)
                self.tableWidget_7.clear()
                self.tableWidget_7.setRowCount(0)
                self.tableWidget_7.setColumnCount(len(cursor.description))
                column_names = [description[0] for description in cursor.description]
                self.tableWidget_7.setHorizontalHeaderLabels(column_names)
                for row_number, row_data in enumerate(result):
                    self.tableWidget_7.insertRow(row_number)
                    for column_number, data in enumerate(row_data):
                        self.tableWidget_7.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
                # Ajustar automáticamente el ancho de las columnas
                self.tableWidget.resizeColumnsToContents()
            except Exception as error:
                QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
            finally:
                conn.close()
        
        elif self.radioButton_28.isChecked():#SUPPLIER TABLE
            if self.lineEdit_6.text() != '' and self.lineEdit_7.text() != '':
                query = f"""SELECT * FROM fcst_supplier WHERE {self.lineEdit_6.text()} = '{self.lineEdit_7.text()}'"""
            else:
                query = "SELECT * FROM fcst_supplier"
            try:
                conn = sqlite3.connect(path_conn)
                cursor = conn.cursor()
                result = cursor.execute(query)
                self.tableWidget_7.clear()
                self.tableWidget_7.setRowCount(0)
                self.tableWidget_7.setColumnCount(len(cursor.description))
                column_names = [description[0] for description in cursor.description]
                self.tableWidget_7.setHorizontalHeaderLabels(column_names)
                for row_number, row_data in enumerate(result):
                    self.tableWidget_7.insertRow(row_number)
                    for column_number, data in enumerate(row_data):
                        self.tableWidget_7.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
                # Ajustar automáticamente el ancho de las columnas
                self.tableWidget.resizeColumnsToContents()
            except Exception as error:
                QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
            finally:
                conn.close()
        
        elif self.radioButton_29.isChecked():#KIT PROJECT TABLE 
            if self.lineEdit_6.text() != '' and self.lineEdit_7.text() != '':
                query = f"""SELECT * FROM kpt WHERE {self.lineEdit_6.text()} = '{self.lineEdit_7.text()}'"""
            else:
                query = "SELECT * FROM kpt"
            try:
                conn = sqlite3.connect(path_conn)
                cursor = conn.cursor()
                result = cursor.execute(query)
                self.tableWidget_7.clear()
                self.tableWidget_7.setRowCount(0)
                self.tableWidget_7.setColumnCount(len(cursor.description))
                column_names = [description[0] for description in cursor.description]
                self.tableWidget_7.setHorizontalHeaderLabels(column_names)
                for row_number, row_data in enumerate(result):
                    self.tableWidget_7.insertRow(row_number)
                    for column_number, data in enumerate(row_data):
                        self.tableWidget_7.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
                # Ajustar automáticamente el ancho de las columnas
                self.tableWidget.resizeColumnsToContents()
            except Exception as error:
                QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
            finally:
                conn.close()
        
        elif self.radioButton_30.isChecked():#YEAR WEEK TABLE 
            if self.lineEdit_6.text() != '' and self.lineEdit_7.text() != '':
                query = f"""SELECT * FROM table_yearweek  WHERE {self.lineEdit_6.text()} = '{self.lineEdit_7.text()}'"""
            else:
                query = "SELECT * FROM table_yearweek"
            try:
                conn = sqlite3.connect(path_conn)
                cursor = conn.cursor()
                result = cursor.execute(query)
                self.tableWidget_7.clear()
                self.tableWidget_7.setRowCount(0)
                self.tableWidget_7.setColumnCount(len(cursor.description))
                column_names = [description[0] for description in cursor.description]
                self.tableWidget_7.setHorizontalHeaderLabels(column_names)
                for row_number, row_data in enumerate(result):
                    self.tableWidget_7.insertRow(row_number)
                    for column_number, data in enumerate(row_data):
                        self.tableWidget_7.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
                # Ajustar automáticamente el ancho de las columnas
                self.tableWidget.resizeColumnsToContents()
            except Exception as error:
                QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
            finally:
                conn.close()

        elif self.radioButton_31.isChecked():#WHERE USE TABLE
            if self.lineEdit_6.text() != '' and self.lineEdit_7.text() != '':
                query = f"""SELECT * FROM where_use_table_plantype  WHERE {self.lineEdit_6.text()} = '{self.lineEdit_7.text()}'"""
            else:
                query = "SELECT * FROM where_use_table_plantype"
            try:
                conn = sqlite3.connect(path_conn)
                cursor = conn.cursor()
                result = cursor.execute(query)
                self.tableWidget_7.clear()
                self.tableWidget_7.setRowCount(0)
                self.tableWidget_7.setColumnCount(len(cursor.description))
                column_names = [description[0] for description in cursor.description]
                self.tableWidget_7.setHorizontalHeaderLabels(column_names)
                for row_number, row_data in enumerate(result):
                    self.tableWidget_7.insertRow(row_number)
                    for column_number, data in enumerate(row_data):
                        self.tableWidget_7.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
                # Ajustar automáticamente el ancho de las columnas
                self.tableWidget.resizeColumnsToContents()
            except Exception as error:
                QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
            finally:
                conn.close()

        else:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion selecciona un archivo')

    """________________________________________SECUNDARY_ACTIONS________________________________________________________________"""

    def create_historic_folder(self):
        try:
            carpeta_historico = os.path.join(f'J:\\Departments\\Material Control\\Purchasing\\Tools\\ComprasDB\\Fcst Supplier Historico','Historico '+ datestring)
            # carpeta_historico = os.path.join(f'C:\\Users\\{cpu}\\Desktop', 'Historico '+ datestring)
            if not os.path.exists(carpeta_historico):
                # Crear la carpeta con la fecha actual como nombre
                os.makedirs(carpeta_historico)
        except Exception as error:
                QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion al acction {error}')
        finally:
            pass

    def mail(self):
        global mail_value
        mail_value = self.plainTextEdit.toPlainText()
        print(mail_value)
        QMessageBox.information(self, "Éxito", "Fue cargado el cuerpo del correo que sera enviado a los proveedores correctamente")

    def mail2(self):
        global mail_value2
        mail_value2 = self.plainTextEdit_2.toPlainText()
        print(mail_value2)
        QMessageBox.information(self, "Éxito", "Fue cargado el cuerpo del correo que sera enviado a los proveedores correctamente")

    def clean_action_messages(self):
        action_to_clean = self.get_action_information()
        
        action_to_clean['ACD'] = action_to_clean['ACD'].str.upper()
        
        def clean_AD_logic(DataFrame):
            for index,row in DataFrame.iterrows():
                if DataFrame.at[index,'ACD'] == 'AD':
                    DataFrame.at[index,'ACD'] = ''
                    DataFrame.at[index,'REQDT'] = ''
                else:
                    pass
        clean_AD_logic(action_to_clean)
        
        def clean_CN_logic(DataFrame):
            for index,row in DataFrame.iterrows():
                if DataFrame.at[index,'ACD'] == 'CN':
                    DataFrame.at[index,'REQDT'] = ''
                else:
                    pass
        clean_CN_logic(action_to_clean)
        
        wb = xw.Book()
        sheet1 = wb.sheets.add()
        sheet1.name = "validar"
        sheet1.range('A1').value = action_to_clean

    def label_safety_stock(self):
        try:
            conn = sqlite3.connect(path_conn)
            rework_query = "SELECT LastDate FROM actualizations WHERE Report = 'Safety Stock'"
            safe_value = pd.read_sql_query(rework_query, conn)
            # For label_safety_stock
            safe_value = pd.read_sql_query(rework_query, conn)
            if not safe_value.empty:
                last_date = safe_value['LastDate'].iloc[0]  # Extracting the first date from the DataFrame
                self.label_60.setHidden(False)
                self.label_60.setText(str(last_date))  # Setting the extracted date as text

        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion vendor {error}')
        finally:
            conn.close()

    def label_safety_stock_should_be(self):
        try:
            conn = sqlite3.connect(path_conn)
            rework_query = "SELECT LastDate FROM actualizations WHERE Report = 'Safety Stock Should be'"
            safe_value_should_be = pd.read_sql_query(rework_query, conn)
            if not safe_value_should_be.empty:
                last_date_should_be = safe_value_should_be['LastDate'].iloc[0]
                self.label_62.setHidden(False)
                self.label_62.setText(str(last_date_should_be))
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion vendor {error}')
        finally:
            conn.close()

    def create_templete_safety_stock(self):
        
        wb = xw.Book()    
        sheet1 = wb.sheets.add()
        sheet1.name = "Safety_Stock"
        sheet1.range('A1').value = 'ItemNo'
        sheet1.range('B1').value = 'New_Qty'
        QMessageBox.information(self, "Templete Safety Stock", "Se creo el templete del Safty Stock correctamente, llenalo para cargarlo ")

    """________________________________________SHOW/HIDE BUTTONS________________________________________________________________"""


    def show_vendor_europeos(self):
        try:
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()
            query = "SELECT * FROM europeos"
            result = cursor.execute(query)
            self.tableWidget_8.clear()
            self.tableWidget_8.setRowCount(0)
            self.tableWidget_8.setColumnCount(len(cursor.description))
            column_names = [description[0] for description in cursor.description]
            self.tableWidget_8.setHorizontalHeaderLabels(column_names)
            for row_number, row_data in enumerate(result):
                self.tableWidget_8.insertRow(row_number)
                for column_number, data in enumerate(row_data):
                    self.tableWidget_8.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
            
            # Ajustar automáticamente el ancho de las columnas
            self.tableWidget_8.resizeColumnsToContents()
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer conexión {error}')
        finally:
            conn.close()

    def registrar_vendor_europeo(self):
        try:
            if self.lineEdit_8.text() != '' :
                conexion = sqlite3.connect(path_conn)
                cursor = conexion.cursor()
                query = "INSERT INTO europeos (Vendor) VALUES (?)"
                cursor.execute(query, (self.lineEdit_8.text().upper(),))
                conexion.commit()
                cursor.close()
                conexion.close()
                QMessageBox.information(self, "Vendor Cargado ", f"registrado correctamente")
                self.lineEdit_8.clear()
                self.show_vendor_europeos()
                self.initial_state()
            else:
                QMessageBox.critical(self, 'Registro fallido', 'Favor de llenar todos los campos')
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer conexión {error}')
        finally:
            conexion.close()

    def delete_vendor_europeo(self):
        try:
            vendor_to_delete = self.lineEdit_9.text().upper()
            conexion = sqlite3.connect(path_conn)
            cursor = conexion.cursor()
            delete_query = "DELETE FROM europeos WHERE Vendor = ? "
            cursor.execute(delete_query, (vendor_to_delete,))
            conexion.commit()
            conexion.close()
            QMessageBox.information(self, "Eliminacion de Vendor", f"Vendor eliminado correctamente")
            self.lineEdit_9.clear()
            self.show_vendor_europeos()
            self.initial_state()
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer conexión {error}')
        finally:
            conexion.close()

    def registrar_exception_item_show(self):
        self.lineEdit.setHidden(False)
        self.lineEdit_2.setHidden(False)
        self.pushButton_10.setHidden(False)
        self.pushButton_13.setHidden(False)
        self.pushButton_11.setHidden(True)
        self.pushButton_12.setHidden(True)

    def delete_item_exception_show(self):
        self.lineEdit.setHidden(False)
        self.pushButton_10.setHidden(True)
        self.pushButton_11.setHidden(True)
        self.pushButton_12.setHidden(True)
        self.pushButton_10.setHidden(False)
        self.pushButton_13.setHidden(False)

    def cancel_registro_button(self):
        self.pushButton_11.setHidden(False)
        self.pushButton_12.setHidden(False)
        self.pushButton_10.setHidden(True)
        self.pushButton_13.setHidden(True)
        self.lineEdit.setHidden(True)
        self.lineEdit_2.setHidden(True)
        self.lineEdit.clear()
        self.lineEdit_2.clear()

    def show_path_buttons_modification(self):
        self.lineEdit_3.setHidden(False)
        self.pushButton_17.setHidden(False)
        self.pushButton_18.setHidden(False)
        self.pushButton_16.setHidden(True)
        self.lineEdit_3.setPlaceholderText("Replace path")

    def cancel_change_path_button(self):
        self.lineEdit_3.setHidden(True)
        self.lineEdit_3.clear()
        self.pushButton_17.setHidden(True)
        self.pushButton_18.setHidden(True)
        self.pushButton_16.setHidden(False)

    """________________________________________CROSS INFORMATION REPORTS________________________________________________________"""

    def running_gemba_shortages(self):
        if self.selected_date != None and self.checkBox_7.isChecked() and self.checkBox_4.isChecked() and self.checkBox_5.isChecked() and self.checkBox_6.isChecked() and self.checkBox_9.isChecked() and  self.checkBox_8.isChecked():
            
            self.label_42.setHidden(False)
            self.label_44.setHidden(False)
            now = datetime.now()
            today = now.strftime("%m/%d/%Y %I:%M:%S")
            self.label_44.setText(f'{today}')
            QApplication.processEvents()
            where_use = self.get_where_use_information_table()
            piv_oh = self.get_oh_information_without_qa() #aqui se modifico para que solo tome material fuera de QA
            expedite = self.get_expedite_information_for_gemba_shortages()
            in_9_2 = self.get_in92_information()
            entity = self.get_entity_information()
            buyer = self.get_vendor_information()
            project = self.get_entity_project_information()
            leanDNA = self.get_LeanDNA_information()
            qa_oh = self.get_oh_qa_information()
            qa_oh['OH'] = qa_oh['OH'].astype(float) 


            leanDNA = leanDNA[['Item_Code','Garden_Grove_C37_Inventory','Bolsa_C91_Inventory','Tijuana_C44_Inventory','Santa_Maria_C61_Inventory','Montreal_CG1_Inventory']]
            new_base = pd.merge(expedite,in_9_2,how = 'left',on = 'ItemNo').fillna('')
            new_base['NEW_VENDOR'] = new_base.apply(lambda row: row['Vendor_y'] if row['Vendor_y'] != '' else row['Vendor_x'], axis=1)
            new_base = new_base.rename(columns={'NEW_VENDOR':'Vendor','MLIKCode_x':'MLIKCode'})
            expedite = new_base
            expedite['ItemNo'] = expedite['ItemNo'].astype(str)
            expedite['ItemNo'] = expedite['ItemNo'].str.upper()
            expedite = expedite.rename(columns = {'LT_y':'LT'})

            expedite = expedite[['EntityGroup', 'Project', 'AC', 'ItemNo', 'Description', 'PlanTp','Ref', 'FillDoc'
                                , 'Sort', 'ReqQty', 'DemandSource', 'Unit', 'ReqDate', 'ShipDate', 'MLIKCode', 'LT', 'STDCost'
                                , 'LotSize','llave', 'left', 'Vendor']]

            qa_oh_piv = pd.pivot_table(qa_oh,index = 'ItemNo',values = 'OH',aggfunc='sum')
            qa_oh_piv = qa_oh_piv.reset_index()
            qa_oh_piv = qa_oh_piv.rename(columns={'OH':'QA loc'})

            exp_oh = pd.merge(expedite,piv_oh,how = 'left', on = 'ItemNo')
            exp_oh['OH'] = exp_oh['OH'].fillna(0)
            filtered_df = exp_oh.loc[pd.to_datetime(exp_oh['ReqDate']) <= self.selected_date]
            filtered_df.loc[:, 'EntityGroup'] = filtered_df['EntityGroup'].str.upper()
            filtered_df_entity = pd.merge(filtered_df,entity,how = 'left',on = 'EntityGroup')
            filtered_df_entity['ItemNo'] = filtered_df_entity['ItemNo'].str.upper() 
            filtered_df_entity['EntityName'] = filtered_df_entity['EntityName'].fillna('Agregar Entity') 
            df = filtered_df_entity
            df['ItemNo'] = df['ItemNo'].str.upper() 
            df = df.sort_values(['ItemNo', 'ReqDate'], ascending=[True, True])
            df['OH'] = df['OH'].astype(float)
            df['ReqQty'] = df['ReqQty'].astype(float)
            prev_part_num = None
            prev_balance = 0
            for index, row in df.iterrows():
                curr_part_num = row['ItemNo']
                req_qty = row['ReqQty']
                avail_material = row['OH']
                if prev_part_num == curr_part_num:
                    balance = prev_balance - req_qty
                    df.at[index, 'Balance'] = balance
                    prev_balance = balance
                else:
                    balance = avail_material - req_qty
                    df.at[index, 'Balance'] = balance
                    prev_part_num = curr_part_num
                    prev_balance = balance

            df['Balance'] = df['Balance'].astype(float)
            df['cob'] = df['Balance'].apply(lambda x: 'shortage' if x < 0 else 'coverage by oh')

            df_buyer = pd.merge(df,buyer,how = 'left',on = 'Vendor')
            dupCodes = df_buyer
            dupCodes = dupCodes[['ItemNo','Vendor']]
            dupCodes = dupCodes.drop_duplicates()

            pivdupCodes = pd.pivot_table(dupCodes,index = 'ItemNo',values = 'Vendor',aggfunc = 'count')
            pivdupCodes = pivdupCodes.reset_index()
            pivdupCodes = pivdupCodes.loc[pivdupCodes['Vendor'] != 1]
            df_buyer['Vendor'] = df_buyer['Vendor'].replace('', 'item without VendorCode')
            df_buyer = df_buyer.fillna('')

            def consolidate_name(DataFrame):
                df_buyer['Vendor_name'] = ''
                for index,row in df_buyer.iterrows():
                    if row["Vendor"] == 'item without VendorCode':
                        DataFrame.at[index, "Vendor_name"] = 'item without VendorCode'
                    else:
                        if row["Vendor"] != 'item without VendorCode':
                            DataFrame.at[index, "Vendor_name"]  = row['Shortname']
                        else:
                            if row['Shortname'] == '':
                                DataFrame.at[index, "Vendor_name"] = 'add vendor to bc_File'
            consolidate_name(df_buyer)

            df_buyer = df_buyer.fillna('')
            df_buyer['Vendor_name'] = df_buyer['Vendor_name'].replace('', 'add vendor to bc_File')
            df_buyer['Tactical'] = df_buyer['Tactical'].replace('', 'without owner')
            """Concatenation file"""
            
            try:
                conn = sqlite3.connect(path_conn)
                cursor = conn.cursor()
                query = "SELECT value FROM paths WHERE path = ?"
                selected_path = ('BuyerInformation',)  # Put the value in a tuple
                cursor.execute(query, selected_path)
                result = cursor.fetchone()  # Assuming you expect one row as result
                
                if result is not None:
                    path = result[0]  # Assuming the value you want is in the first column
                else:
                    QMessageBox.critical(self, "Error en Path de comentarios", f'No se pudo encontrar el path de los comentarios{error}')
                
            except Exception as error:
                QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer conección {error}')
            finally:
                conn.close()
            dfs = []
            # path = path.replace("\\", "\\\\")
            print(path)
            
            
            # try:
            #     for file in os.listdir(path):
            #         if file.endswith('.xlsx') or file.endswith('.xls'):
            #             df = pd.read_excel(os.path.join(path, file))
            #             dfs.append(df)
            #     df_concat = pd.concat(dfs)
            #     df_concat1 = df_concat
            # except Exception as error:
            #     QMessageBox.critical(self, "Error", f'No se pudo tomar los cometarios el error es: {error}')
            # finally:
            #     pass

        
            path = 'J:\\Departments\\Material Control\\Purchasing\\Tools\\ComprasDB\\Gemba Shortages Historico\\GEMBA COMMENTS'
            file_name = 'EZA Shortages Last WK.xlsx'
            df_concat = pd.read_excel(f'{path}\\{file_name}')
            pepo = df_concat.copy()
            df_concat['ItemNo'] = df_concat['ItemNo'].astype(str)
            df_concat['ItemNo'] = df_concat['ItemNo'].str.upper()
            df_concat['ItemNo'] = df_concat['ItemNo'].str.replace('4500611','04500611')
            
            df_concat['ItemNo'] = df_concat['ItemNo'].str.upper().str.strip()
            df_concat = df_concat[['ItemNo','Shortage date','Root Cause','Comments','ETA']]
            df_concat = df_concat.fillna('')    
            df_buyer_concat = pd.merge(df_buyer,df_concat,how = 'left', on = 'ItemNo')

            gemba = df_buyer_concat

            gemba = gemba.loc[gemba['cob'] != 'coverage by oh']
            gemba['ItemNo'] = gemba['ItemNo'].str.upper() 
            gemba['impacts'] = gemba.groupby('ItemNo')['EntityName'].transform(lambda x: ', '.join(x))
            gemba = gemba[['Tactical','ItemNo','Description','LT','Vendor_name','impacts','cob']].drop_duplicates()
            gemba['impacts'] = gemba['impacts'].str.split(', ').apply(set).str.join(', ') # type: ignore

            df_buyer_concat_project = df_buyer_concat
            df_buyer_concat_project = df_buyer_concat_project.loc[(df_buyer_concat_project['cob'] != 'coverage by oh')]
            df_buyer_concat_project['ItemNo'] = df_buyer_concat_project['ItemNo'].str.upper()

            base2 = pd.merge(df_buyer_concat_project,project,how = 'left',on = 'EntityGroup')
            base2 = base2.fillna('')
            base2['Project_x'] = base2['Project_x'].replace('','190')
            base2['Project_x'] = base2['Project_x'].str.upper()
            def fill_project_y(DataFrame):
                e2 = 'OEM E2'
                spares1 = 'SPR'
                e1 = 'OEM E1'
                for index,row in DataFrame.iterrows():
                    if row['Project_y'] == '':
                        if row['Project_x'] == '190':
                            DataFrame.at[index,'Project_y'] = e1
                        else:
                            if row['Project_x'] == 'SCD':
                                DataFrame.at[index,'Project_y'] = spares1
                            else:
                                if row['Project_x'] == 'SEB':
                                    DataFrame.at[index,'Project_y'] = spares1
                                else:
                                    if row['Project_x'] == 'PE2':
                                        DataFrame.at[index,'Project_y'] = e2
                                    else:
                                        pass

            fill_project_y(base2)
            base2['Project_y'] = base2['Project_y'].astype(str)
            base2['Project_y'] = base2['Project_y'].str.upper()
            
            base2_piv = pd.pivot_table(base2,values = 'Balance',aggfunc = 'min',index = 'ItemNo',columns = 'Project_y')
            base2_piv = base2_piv.reset_index()
            base2_piv = base2_piv.fillna(0)
            
            
            # Assuming base2_piv is your DataFrame
            base2_piv['total_qty'] = base2_piv[['CD-NOTE', 'OEM E1', 'OEM E2', 'SPR']].min(axis=1)

            # base2_piv['total_qty'] = base2_piv['CD-NOTE'] + base2_piv['OEM E1'] + base2_piv['OEM E2'] + base2_piv['SPR']

            gemba['ItemNo'] = gemba['ItemNo'].str.upper().str.upper()
            gemba_proj = pd.merge(gemba,base2_piv,how = 'left',on = 'ItemNo')

            datereqProject = df_buyer_concat
            datereqProject = datereqProject.loc[(datereqProject['cob'] != 'coverage by oh')]
            datereqProject = datereqProject[['EntityGroup','ReqDate','ShipDate','ItemNo','ReqQty']]
            datereqProject1 = pd.merge(datereqProject,project,how = 'left',left_on = 'EntityGroup',right_on = 'EntityGroup')
            datereqProject1 = datereqProject1[['EntityGroup','ReqDate','ShipDate','ItemNo','Project']]
            datereqProject1['ItemNo'] = datereqProject1['ItemNo'].str.upper()
            
            min_date_rows = datereqProject1.groupby(['Project', 'ItemNo'])['ReqDate'].agg(min).reset_index()
            min_date_rows = min_date_rows[['Project','ItemNo','ReqDate']]
            result = pd.merge(min_date_rows, datereqProject1, on=['Project', 'ItemNo', 'ReqDate'], how='inner')
            result['ItemNo'] = result['ItemNo'].str.upper()
            result = result.drop_duplicates()

            resultpiv = pd.pivot_table(result,columns = 'Project',values = 'ReqDate',aggfunc = 'min',index = 'ItemNo')
            resultpiv = resultpiv.reset_index()
            resultpiv = resultpiv.fillna('')

            resultpiv['ItemNo'] = resultpiv['ItemNo'].astype(str)
            resultpiv['ItemNo'] = resultpiv['ItemNo'].str.upper()
            gemba_proj_datesproj= pd.merge(gemba_proj,resultpiv,how = 'left',on = 'ItemNo')

            date_columns = ['CD-NOTE_y', 'OEM E1_y', 'OEM E2_y', 'SPR_y']
            gemba_proj_datesproj[date_columns] = gemba_proj_datesproj[date_columns].apply(pd.to_datetime, errors='coerce')
            gemba_proj_datesproj['min_date'] = gemba_proj_datesproj[date_columns].min(axis=1)

            gemba_proj_datesproj['CD-NOTE_y'] = gemba_proj_datesproj['CD-NOTE_y'].apply(lambda x: x.strftime('%m/%d/%Y') if pd.notnull(x) else '')
            gemba_proj_datesproj['OEM E1_y'] = gemba_proj_datesproj['OEM E1_y'].apply(lambda x: x.strftime('%m/%d/%Y') if pd.notnull(x) else '')
            gemba_proj_datesproj['OEM E2_y'] = gemba_proj_datesproj['OEM E2_y'].apply(lambda x: x.strftime('%m/%d/%Y') if pd.notnull(x) else '')
            gemba_proj_datesproj['SPR_y'] = gemba_proj_datesproj['SPR_y'].apply(lambda x: x.strftime('%m/%d/%Y') if pd.notnull(x) else '')
            gemba_proj_datesproj['min_date'] = gemba_proj_datesproj['min_date'].apply(lambda x: x.strftime('%m/%d/%Y') if pd.notnull(x) else '')
            datestring = datetime.now().strftime('%m/%d/%Y')# type: ignore

            final = pd.merge(gemba_proj_datesproj,df_concat,how='left',on = 'ItemNo')
            final = final.fillna('')
            
            print(final.columns)
            
            final['Shortage date'] = final['Shortage date'].replace('',f'New! {datestring}')

            def format_eta(val):
                if pd.isnull(val) or val == "":
                    return val
                else:
                    try:
                        return pd.to_datetime(val).strftime("%m/%d/%Y")
                    except ValueError:
                        return val
            final["ETA"] = final["ETA"].apply(lambda x: format_eta(x))
            
            final1 = pd.merge(final,leanDNA,how = 'left',left_on = 'ItemNo',right_on = 'Item_Code')
            final1 = final1.rename(columns={'Garden_Grove_C37_Inventory':'Garden_Grove_OH','Bolsa_C91_Inventory':'Bolsa_OH','Tijuana_C44_Inventory':'Tijuana_OH'
                                            ,'Santa_Maria_C61_Inventory':'Santa_Maria_OH','Montreal_CG1_Inventory':'Montreal_OH'})
            final2 = pd.merge(final1,qa_oh_piv,how = 'left',on = 'ItemNo')
            # try:
            #     dfs = []
            #     path = path.replace("\\", "\\\\")
            #     for file in os.listdir(path):
            #         if file.endswith('.xlsx') or file.endswith('.xls'):
            #             df = pd.read_excel(os.path.join(path, file))
            #             dfs.append(df)
                
            #     comments_df = pd.concat(dfs)
            #     comments_df['ItemNo'] = comments_df['ItemNo'].str.upper().str.strip()
            #     comments_df = comments_df[['ItemNo','FLAG','Priority 1 Qty']]
            # except Exception as error:
            #     QMessageBox.critical(self, "Error", f'No se pudo tomar los cometarios el error es: {error}')
            # finally:
            #     pass
            pepo['ItemNo'] = pepo['ItemNo'].str.upper().str.strip()
            pepo = pepo[['ItemNo','FLAG','Priority 1 Qty']]


            final3 = pd.merge(final2,pepo,how = 'left',on = 'ItemNo')
            
            # final3['Point of use'] = ''
            final3['TOTAL IMPO'] = ''
            final3['Impo Lunes'] = ''
            final3['Impo Martes'] = ''
            final3['Impo Miercoles'] = ''
            final3['Impo Jueves'] = ''
            final3['Impo Viernes'] = ''
            final3['Impo Sabado'] = ''
            
            final3 = final3.rename(columns = {'Status':'Shortage date','impacts':'Commodity'})
            
            final3 = final3[['ItemNo','Description','Vendor_name','Shortage date','Tactical','Commodity','ETA'
                            ,'Priority 1 Qty','total_qty','Root Cause','Comments','FLAG','QA loc'
                            ,'TOTAL IMPO','Impo Lunes','Impo Martes','Impo Miercoles','Impo Jueves','Impo Viernes','Impo Sabado'
                            ,'Garden_Grove_OH','Bolsa_OH','Tijuana_OH','Santa_Maria_OH','Montreal_OH']]

            all_materials_plantype = df_buyer_concat.copy()
            

            all_materials_plantype = all_materials_plantype[['ItemNo','PlanTp']]
            all_materials_plantype['ItemNo'] = all_materials_plantype['ItemNo'].str.upper().str.strip()
            all_materials_plantype['PlanTp'] = all_materials_plantype['PlanTp'].str.upper().str.strip()
            all_materials_plantype = all_materials_plantype.drop_duplicates(subset='ItemNo', keep='first')
            
            base1 = pd.merge(all_materials_plantype,where_use,how ='left', on= 'PlanTp')
            base2 = pd.merge(final3,base1,how ='left', on= 'ItemNo')
            base2 = base2.rename(columns = {'Area':'Point of use'})
            base2 = base2[['ItemNo', 'Description', 'Vendor_name', 'Shortage date', 'Tactical','Commodity','Point of use', 'ETA', 'Priority 1 Qty', 'total_qty', 'Root Cause',
                        'Comments', 'FLAG', 'QA loc', 'TOTAL IMPO', 'Impo Lunes', 'Impo Martes','Impo Miercoles', 'Impo Jueves'
                        , 'Impo Viernes', 'Impo Sabado','Garden_Grove_OH', 'Bolsa_OH', 'Tijuana_OH', 'Santa_Maria_OH','Montreal_OH']]
            

            base2['total_qty'] = abs(base2['total_qty'])


            wb = xw.Book()
            sheet1 = wb.sheets.add()
            sheet1.name = "GembaShortages"
            sheet1.range('A1').value = base2
            column_letter = 'A'
            column_range = sheet1.range(f'{column_letter}:{column_letter}')
            column_range.delete()
            sheet2 = wb.sheets.add()
            sheet2.name = "allmaterials"
            sheet2.range('A1').value = df_buyer_concat
            sheet3 = wb.sheets.add()
            sheet3.name = "checkPN"
            sheet3.range('A1').value = pivdupCodes            
            white_text_color = (240, 242, 245)
            blue_value = (3, 46, 84)
            yellow_value = (245, 241, 10)
            red_value_text = (240, 15, 7)
            green_value = (7, 171, 51)
            orange_value = (230, 130, 0)
            black_value = (20, 20, 20)
            red_value = (207, 23, 2)
            

            for cell in sheet1.range('A1:G1'):
                cell.color = blue_value
                cell.api.Font.Color = xw.utils.rgb_to_int(white_text_color)
            
            for cell in sheet1.range('J1'):
                cell.color = blue_value
                cell.api.Font.Color = xw.utils.rgb_to_int(white_text_color)
            
            for cell in sheet1.range('N1'):
                cell.color = blue_value
                cell.api.Font.Color = xw.utils.rgb_to_int(white_text_color)
            
            for cell in sheet1.range('I1'):
                cell.color = yellow_value
                cell.api.Font.Color = xw.utils.rgb_to_int(red_value_text)
        
            for cell in sheet1.range('H1'):
                cell.color = green_value
                cell.api.Font.Color = xw.utils.rgb_to_int(white_text_color)

            for cell in sheet1.range('K1:L1'):
                cell.color = green_value
                cell.api.Font.Color = xw.utils.rgb_to_int(white_text_color)

            for cell in sheet1.range('P1:U1'):
                cell.color = yellow_value
                cell.api.Font.Color = xw.utils.rgb_to_int(red_value_text)
            
            for cell in sheet1.range('V1:Z1'):
                cell.color = orange_value
                cell.api.Font.Color = xw.utils.rgb_to_int(black_value)
            
            for cell in sheet1.range('M1'):
                cell.color = red_value
                cell.api.Font.Color = xw.utils.rgb_to_int(white_text_color)   
                
            for cell in sheet1.range('O1'):
                cell.color = orange_value
                cell.api.Font.Color = xw.utils.rgb_to_int(red_value_text)  

            datestringsave = datetime.now().strftime('%m_%d')# type: ignore
            writer = pd.ExcelWriter(f'J:\\Departments\\Material Control\\Purchasing\\Tools\\ComprasDB\\Gemba Shortages Historico\\EZA Shortages {datestringsave}.xlsx')            
            base2.to_excel(writer, sheet_name='GembaShortages', index=False)
            df_buyer_concat.to_excel(writer, sheet_name='allmaterials', index=False)
            pivdupCodes.to_excel(writer, sheet_name='checkPN', index=False)
            writer.close()

            self.label_43.setHidden(False)
            self.label_45.setHidden(False)
            now = datetime.now()
            today = now.strftime("%m/%d/%Y %I:%M:%S")
            self.label_45.setText(f'{today}')

            QMessageBox.information(self, "Reporte guardado", "Valida el archivo guardado en la carpeta Historico")
        else:
            QMessageBox.critical(self, "Error", "Selecciona el rango de covertura del reporte asi como la confirmacion de la actualizacion de la Base")

    def running_pr_actions(self):
        if self.checkBox_10.isChecked() and self.checkBox_13.isChecked() and self.checkBox_11.isChecked() and self.checkBox_12.isChecked() and self.checkBox_14.isChecked() and self.checkBox_15.isChecked() and self.checkBox_16.isChecked() and self.checkBox_17.isChecked() and self.checkBox_18.isChecked():
            
            self.label_38.setHidden(False)
            self.label_40.setHidden(False)
            now = datetime.now()
            today = now.strftime("%m/%d/%Y %I:%M:%S")
            self.label_40.setText(f'{today}')
            QApplication.processEvents()
            
            
            expedite = self.get_expedite_information()
            in_9_2 = self.get_in92_information()
            oh_piv = self.get_oh_information()
            oo  = self.get_open_order_information()
            rwk = self.get_rwk_information()
            transit = self.get_transit_information()
            po351 = self.get_po351_information()
            holidays = self.get_holidays_information()
            action = self.get_action_information()
            virtual_df = self.get_year_week_information()

            oo = oo[['ItemNo','OpnQ']]
            oo['OpnQ'] = oo['OpnQ'].astype(float)
            oo = pd.pivot_table(oo,index = 'ItemNo',values = 'OpnQ',aggfunc = 'sum')
            oo = oo.reset_index()
            oo = oo.rename(columns={'OpnQ':'POQty'})

            rwk = rwk[['ItemNo','OH']]
            rwk['OH'] = rwk['OH'].astype(float)
            rwk = pd.pivot_table(rwk,index = 'ItemNo',values = 'OH',aggfunc = 'sum')
            rwk = rwk.reset_index()
            rwk = rwk.rename(columns={'OH':'RwkQty'})

            po351['key'] = po351['ItemNo'] + po351['Vendor']
            po351 = po351[['key','FobVendor']]

            expedite_demand = expedite[['ItemNo','ReqQty','ReqDate']]
            item_stdcost = expedite[['ItemNo','STDCost']].fillna('')
            item_stdcost = item_stdcost.loc[item_stdcost['STDCost']!='']
            item_stdcost = item_stdcost.drop_duplicates()
            
            in_9_2[['ItemNo','Vendor']]
            
            action = action[['ItemNo','PlanningNotes']]
            action = action.loc[action['PlanningNotes']!='']

            """logic of vendor"""
            expedite_unique_item = expedite[['ItemNo','Vendor','MLIKCode','LotSize','LT']]
            expedite_unique_item = expedite_unique_item.drop_duplicates() #4700 registros
            
            exp_in92 = pd.merge(expedite_unique_item,in_9_2,how = 'left',on = 'ItemNo')
            exp_in92['NEW_VENDOR'] = exp_in92.apply(lambda row: row['Vendor_y'] if row['Vendor_y'] != '' else row['Vendor_x'], axis=1)
            exp_in92['NEW_VENDOR'] = exp_in92['NEW_VENDOR'].fillna('no vendor')
            exp_in92 = exp_in92[['ItemNo','LotSize','LT_x','MLIKCode_x','NEW_VENDOR']]
            exp_in92 = exp_in92.rename(columns={'MLIKCode_x':'MLIKCode','NEW_VENDOR':'Vendor','LT_x':'LT'})
            exp_in92 = exp_in92.drop_duplicates()
            exp_in92 = pd.merge(exp_in92,item_stdcost,how = 'left',on = 'ItemNo')
            exp_in92 = exp_in92.drop_duplicates()
            exp_in92['key'] = exp_in92['ItemNo']  + exp_in92['Vendor'] 
            
            try:
                exp_in92 = exp_in92[['key','ItemNo','LotSize','LT','Vendor','STDCost']]
                exp_in92['key'] = exp_in92['key'].str.upper().str.strip()
                exp_in92['ItemNo'] = exp_in92['ItemNo'].str.upper().str.strip()
                exp_in92['LotSize'] = exp_in92['LotSize'].str.upper().str.strip()
                exp_in92['Vendor'] = exp_in92['Vendor'].str.upper().str.strip()
                exp_in92['STDCost'] = exp_in92['STDCost'].str.strip()
            except Exception as error:
                QMessageBox.critical(self, "Error", f"error en el tipo de datos {error}")
            finally:
                pass

            expedite_demand = expedite[['ItemNo','ReqQty','ReqDate']]
            expedite_demand['ReqDate'] = pd.to_datetime(expedite_demand['ReqDate'])
            expedite_demand['FormattedReqDate'] = expedite_demand['ReqDate'].dt.strftime('%m/%d/%Y')
            expedite_demand['WeekNum'] = expedite_demand['ReqDate'].dt.isocalendar().week
            expedite_demand['year'] = expedite_demand['ReqDate'].dt.year

            expedite_demand.loc[:, 'ReqQty'] = expedite_demand['ReqQty'].astype(float)
            expedite_demand.loc[:, 'WeekNum'] = expedite_demand['WeekNum'].astype(str)


            def get_pastdue_information(DataFrame):
                today = datetime.today() 
                first_day_of_current_week = today - timedelta(days=today.weekday())
                last_day_of_previous_week = first_day_of_current_week - timedelta(days=1)
                
                DataFrame['WeekNum'] = DataFrame.apply(lambda row: 'Past due' if row['ReqDate'] < first_day_of_current_week else row['WeekNum'], axis=1)
                DataFrame['WeekNum'] = DataFrame['WeekNum'].astype(str).str.zfill(2)
                DataFrame['year'] = DataFrame.apply(lambda row: 0 if row['ReqDate'] < first_day_of_current_week else row['year'], axis=1)
            get_pastdue_information(expedite_demand)    
        
            exp_req_date = expedite_demand
            expedite_demand_grouped = expedite_demand.groupby(['ItemNo', 'year', 'WeekNum'], as_index=False)['ReqQty'].sum()
            base = pd.merge(exp_in92,expedite_demand_grouped,how = 'left',on  = 'ItemNo')
            base = pd.merge(base,oh_piv,how = 'left',on  = 'ItemNo')
            base = pd.merge(base,oo,how = 'left',on = 'ItemNo')
            base = pd.merge(base,rwk,how = 'left',on  = 'ItemNo')
            
            base = base.fillna('')
            base['RwkQty'] = base['RwkQty'].replace('',0)
            base['POQty'] = base['POQty'].replace('',0)
            base['OH'] = base['OH'].replace('',0)

            base['first_key'] = 'y'
            duplicates_mask = base['ItemNo'].duplicated(keep='first') #aqui modifique 
            base.loc[duplicates_mask, 'first_key'] = 'n'

            prev_balance = 0
            for index, row in base.iterrows():
                req_qty = row['ReqQty']
                if row['first_key'] == 'y':
                    avail_material = row['OH'] + row['POQty'] + row['RwkQty']
                    balance = avail_material - req_qty
                    base.at[index, 'Balance'] = balance
                    prev_balance = balance
                else:
                    if row['first_key'] == 'n':
                        balance = prev_balance - req_qty
                        base.at[index, 'Balance'] = balance
                        prev_balance = balance
            
            req_qty_shortage = base.copy()
            """--------------------------------------------------------------aqui va el cambio"""
            
            virtual_df = virtual_df.rename(columns={'year_week':'Year_Week','Date':'1stImpact2'})
            
            # date_range = pd.date_range(start='2010-01-01', end='2050-01-01', freq='W-Mon')
            # year_week = date_range.strftime('%Y%U')
            # formatted_dates = date_range.strftime('%m/%d/%Y')
            # virtual_df = pd.DataFrame({'Year_Week': year_week, '1stImpact2': formatted_dates})
            """--------------------------------------------------------------aqui va el cambio"""
            base2 = base.copy()
            base2 = base2.loc[base2['Balance']<0]
            base2['first_key2'] = 'y'
            duplicates_mask = base2['ItemNo'].duplicated(keep='first') #aqui modifique
            base2.loc[duplicates_mask, 'first_key2'] = 'n'
            base2 = base2.loc[base2['first_key2'] == 'y']

            base2['year'] = base2['year'].astype(str)
            base2['WeekNum'] = base2['WeekNum'].astype(str)
            base2['1stImpact'] = base2['year'] + base2['WeekNum']
            base2 = base2[['ItemNo','1stImpact']]

            base = pd.merge(base,base2,how = 'left',on='ItemNo')
            base = pd.merge(base,virtual_df,how = 'left',left_on ='1stImpact',right_on = 'Year_Week')
            
            
            base['Balance'] = base['Balance'].astype(float)
            base = base.loc[base['Balance']<0.0]
            
            today = datetime.today()
            
            
            monday = today - timedelta(days=today.weekday())
            
            base5 = base.copy()
            def check_pastdue(DataFrame):
                for index,row in DataFrame.iterrows():
                    if row['1stImpact'] == '0Past due':
                        DataFrame.at[index,'1stImpact2'] = monday
                    else:
                        pass
            check_pastdue(base5)        
            
            base5 = base5[['ItemNo','1stImpact','1stImpact2']]
            base5 = base5.drop_duplicates()
            df = base.copy()
            df = df[['ItemNo','LT','year','WeekNum','Balance']]
            df['LT'] = df['LT'].astype(int)
            df['Table_week'] = round(df['LT']/int(5))
            df['first_key2'] = 'y'
            duplicates_mask = df['ItemNo'].duplicated(keep='first')
            df.loc[duplicates_mask, 'first_key2'] = 'n'
            df = df.loc[df['Balance']<0]

            def magic(df):
                df['Count_intex'] = ''
                prev_item = None
                count = 0
                for index, row in df.iterrows():
                    current_item = row['ItemNo']
                    if current_item == prev_item:
                        count += 1
                    else:
                        count = 1 
                    df.at[index, 'Count_intex'] = count
                    prev_item = current_item

            magic(df)

            def magic_2(df):
                df['select_week_shortage'] = ''
                for index,row in df.iterrows():
                    if row['Table_week'] == row['Count_intex']:
                        df.at[index,'select_week_shortage'] = 'select_week'
                    else:
                        pass
            magic_2(df)

            df = df.sort_values(by=['ItemNo', 'WeekNum'])
            last_occurrences = df.groupby('ItemNo').tail(1).index
            df.loc[last_occurrences, 'last_row_item'] = 'select_last'
            df = df.fillna('')
            def magic_3(df):
                df['Final_select'] = ''
                for index,row in df.iterrows():
                    if row['select_week_shortage'] != '':
                        df.at[index,'Final_select'] = '*'
                    else:
                        if row['last_row_item'] != '':
                            df.at[index,'Final_select'] = '*'
                        else:
                            pass
            magic_3(df)

            df = df.loc[df['Final_select']=='*']           
            def magic_4(df):
                df['Count_index'] = ''
                prev_item = None
                count = 0
                for index, row in df.iterrows():
                    current_item = row['ItemNo']
                    if current_item == prev_item:
                        count += 1
                    else:
                        count = 1 
                    df.at[index, 'Count_index'] = count
                    prev_item = current_item

            magic_4(df)

            df = df.loc[df['Count_index']==1] 
            df['year'] = df['year'].astype(str)
            df['WeekNum'] = df['WeekNum'].astype(str)
            df['year_week'] = df['year'] + df['WeekNum']
            base3 = pd.pivot_table(base,index=['key','Vendor','ItemNo','LotSize','LT','STDCost'
                                            ,'OH','POQty','RwkQty'] #,'1stImpact','1stImpact2'
                                ,columns=['year','WeekNum'],values=['Balance'],aggfunc='sum')
            
            base3.reset_index(inplace=True) 
            base3 = base3.fillna(0)
            
            req_qty_shortage = req_qty_shortage[['ItemNo','year','WeekNum','Balance']]
            req_qty_shortage['Balance'] = req_qty_shortage['Balance'].astype(float)
            req_qty_shortage = req_qty_shortage.loc[req_qty_shortage['Balance']<0 ]
            range_formula_qty_shortage = req_qty_shortage.shape[0]+1

            req_qty_shortage['first_key2'] = 'y'
            duplicates_mask = req_qty_shortage['ItemNo'].duplicated(keep='first') #aqui modifique
            req_qty_shortage.loc[duplicates_mask, 'first_key2'] = 'n'
            req_qty_shortage = req_qty_shortage.loc[req_qty_shortage['first_key2'] == 'y']


            range_formula  = base3.shape[0] + 3
            print(f'the range can be: {range_formula}')

            
            wb = xw.Book()

            sheet1 = wb.sheets.add()
            sheet1.name = "1st_impacts_table"
            sheet1.range('A1').value = base5

            sheet2 = wb.sheets.add()
            sheet2.name = "Base"
            sheet2.range('A1').value = base3

            sheet3 = wb.sheets.add()
            sheet3.name = "check"
            sheet3.range('A1').value = base

            sheet4 = wb.sheets.add()
            sheet4.name = "FobVendor"
            sheet4.range('A1').value = po351

            sheet5 = wb.sheets.add()
            sheet5.name = "Transit"
            sheet5.range('A1').value = transit
            
            sheet6 = wb.sheets.add()
            sheet6.name = "Qty Shortage"
            sheet6.range('A1').value = df
            # sheet6.api.Columns(1).Insert()
            # sheet6.range(f'A1').value = 'key'
            # sheet6.range(f'A2:A{range_formula_qty_shortage}').formula = """=C2&D2&E2"""

            sheet7 = wb.sheets.add()
            sheet7.name = "Holidays"
            sheet7.range('A1').value = holidays
            
            sheet8 = wb.sheets.add()
            sheet8.name = "Action Notes"
            sheet8.range('A1').value = action
            
            # sheet9 = wb.sheets.add()
            # sheet9.name = "base_base"
            # sheet9.range('A1').value = pepo

            column_to_delete = "A" 
            column_range = sheet2.range(f"{column_to_delete}:{column_to_delete}")
            column_range.delete()
            
            sheet2.api.Columns(7).Insert()
            sheet2.range('G3').value = 'UnitCost'
            sheet2.range(f'G4:G{range_formula}').formula ="""=IFERROR(VLOOKUP(A4,FobVendor!B:C,2,0),"noVendorFob")"""
            
            sheet2.api.Columns(11).Insert()
            sheet2.range('K3').value = '1st_Impact_YW'
            sheet2.range(f'K4:K{range_formula}').formula ="""=IFERROR(IF(VLOOKUP(C4,'1st_impacts_table'!B:D,3,0)="","",VLOOKUP(C4,'1st_impacts_table'!B:D,3,0)),"")"""
            

            sheet2.api.Columns(12).Insert()
            sheet2.range('L3').value = 'Transit_time'
            sheet2.range(f'L4:L{range_formula}').formula = """=IFERROR(VLOOKUP(B4,'Transit'!B:C,2,0),"")"""

            sheet2.api.Columns(13).Insert()
            sheet2.range('M3').value = 'PR'
            sheet2.range(f'M4:M{range_formula}').formula ="""=IF(K4="","",IF(L4="",WORKDAY(K4,-SUM(E4,5)),WORKDAY(K4,-SUM(E4,L4))))"""

            sheet2.api.Columns(14).Insert()
            sheet2.range('N3').value = 'PR_sin_pd'
            sheet2.range(f'N4:N{range_formula}').formula = """=IF(M23="","",IF(M23<$M$2,$M$2,M23))"""

            sheet2.api.Columns(15).Insert()
            sheet2.range('O3').value = 'YW PR'
            sheet2.range(f'O4:O{range_formula}').formula = """=IF(N4="","",YEAR(N4)&WEEKNUM((N4)))"""

            sheet2.api.Columns(16).Insert()
            sheet2.range('P3').value = 'PR PD + LT'
            sheet2.range(f'P4:P{range_formula}').formula = """=IF(N4="","",IF(L4="",WORKDAY(N4,SUM(E4,5)),WORKDAY(N4,SUM(E4,L4))))"""
            
            sheet2.api.Columns(17).Insert()
            sheet2.range('Q3').value = 'WK Impactada'
            sheet2.range(f'Q4:Q{range_formula}').formula = """=IF(P4="",O4,YEAR(P4)&WEEKNUM(P4))"""   
            
            sheet2.api.Columns(18).Insert()
            sheet2.range('R3').value = 'Qty Shortage'
            sheet2.range(f'R4:R{range_formula}').formula = """=IFERROR(VLOOKUP(C4,'Qty Shortage'!B:F,5,0),"")"""

            sheet2.api.Columns(19).Insert()
            sheet2.range('S3').value = 'PR Qty'
            sheet2.range(f'S4:S{range_formula}').formula = """=IF(R4="","",IF(ABS(R4)>D4,ROUNDUP(ABS(R4),0),D4))"""

            sheet2.api.Columns(20).Insert()
            sheet2.range('T3').value = 'Delivery Date'
            sheet2.range(f'T4:T{range_formula}').formula = """=IF(P4="","",WORKDAY(P4,E4,Holidays!C2:C100))"""

            sheet2.api.Columns(21).Insert()
            now = datetime.now()
            current_year = now.year
            sheet2.range('U3').value = f'Validar {current_year}?'
            sheet2.range(f'U4:U{range_formula}').formula = """"""

            sheet2.api.Columns(22).Insert()
            sheet2.range('V3').value = 'Planning Notes'
            sheet2.range(f'V4:V{range_formula}').formula = """=IFERROR(VLOOKUP(C4,'Action Notes'!B:C,2,0),"")"""

            sheet2.api.Columns(23).Insert()
            sheet2.range('W3').formula = "=Y3"
            sheet2.range('W2').formula = "=Y2"
            sheet2.range('W1').formula = "=Y1"
            
            sheet2.range(f'W4:W{range_formula}').formula = """=X4+Y4"""
            
            today = datetime.today()
            first_day_of_current_week = today - timedelta(days=today.weekday())
            last_day_of_previous_week = first_day_of_current_week - timedelta(days=1)
            sheet2.range(f'M2').formula = first_day_of_current_week

            start_row_w = 4  # Replace with the actual start row
            end_row_w = range_formula  # Replace with the actual end row
            range_address_w = f'W{start_row_w}:W{end_row_w}'
            range_to_paste_special_w = sheet2.range(range_address_w)

            range_to_paste_special_w.api.Copy()
            range_to_paste_special_w.api.PasteSpecial(Paste=-4163)  # xlPasteValues constant

            cell_to_paste_special = sheet2.range('W1')
            cell_to_paste_special.api.Copy()
            cell_to_paste_special.api.PasteSpecial(Paste=-4163) 
            cell_to_paste_special = sheet2.range('W2')
            cell_to_paste_special.api.Copy()
            cell_to_paste_special.api.PasteSpecial(Paste=-4163)
            cell_to_paste_special = sheet2.range('W3')
            cell_to_paste_special.api.Copy()
            cell_to_paste_special.api.PasteSpecial(Paste=-4163)
            
            column_to_delete = 24  # Column number (W)
            sheet2.api.Columns(column_to_delete).Delete()
            column_to_delete = 24  # Column number (W)
            sheet2.api.Columns(column_to_delete).Delete()


            #FORMAT DATES
            column_letter = 'K'
            short_date_format = 'm/d/yyyy'  # Example short date format
            range_to_format = f'{column_letter}:{column_letter}'
            sheet2.range(range_to_format).number_format = short_date_format

            column_letter = 'M'
            short_date_format = 'm/d/yyyy'  # Example short date format
            range_to_format = f'{column_letter}:{column_letter}'
            sheet2.range(range_to_format).number_format = short_date_format

            column_letter = 'N'
            short_date_format = 'm/d/yyyy'  # Example short date format
            range_to_format = f'{column_letter}:{column_letter}'
            sheet2.range(range_to_format).number_format = short_date_format

            column_letter = 'P'
            short_date_format = 'm/d/yyyy'  # Example short date format
            range_to_format = f'{column_letter}:{column_letter}'
            sheet2.range(range_to_format).number_format = short_date_format

            column_letter = 'T'
            short_date_format = 'm/d/yyyy'  # Example short date format
            range_to_format = f'{column_letter}:{column_letter}'
            sheet2.range(range_to_format).number_format = short_date_format

            column_letter = 'C'
            short_date_format = 'm/d/yyyy'  # Example short date format
            range_to_format = f'{column_letter}:{column_letter}'
            sheet7.range(range_to_format).number_format = short_date_format

            #CLEAN TITLES
            sheet2.range(f'A1:F1').copy()
            sheet2.range(f'A3').paste()
            sheet2.range(f'A1:F1').clear()
            sheet2.range(f'H1:J1').copy()
            sheet2.range(f'H3').paste()
            sheet2.range(f'H1:J1').clear()

            self.label_39.setHidden(False)
            self.label_41.setHidden(False)
            now = datetime.now()
            today = now.strftime("%m/%d/%Y %I:%M:%S")
            self.label_41.setText(f'{today}')

            QMessageBox.information(self, "Generacion de reporte", "Se Genero el PR correctamente")
        else:
            QMessageBox.critical(self, "Error", "Selecciona  la confirmacion de la actualizacion de la Base")

    def send_supplier_email(self):
        self.label_23.setHidden(False)
        if self.checkBox_3.isChecked() and self.checkBox_2.isChecked() and self.checkBox.isChecked() and self.plainTextEdit.toPlainText() != '':
            expedite = self.get_expedite_information()
            in_9_2 = self.get_in92_information_send_mail()
            buyer = self.get_supplier_email()
            self.create_historic_folder()
            
            Historico = 'Historico'
            output_dir = f'J:\\Departments\\Material Control\\Purchasing\\Tools\\ComprasDB\\Fcst Supplier Historico\\{Historico} {datestring}'
            expedite = expedite.applymap(lambda x: x.strip().upper() if isinstance(x, str) else x)
            buyer = buyer.applymap(lambda x: x.strip().upper() if isinstance(x, str) else x)
            
            expedite = expedite[['EntityGroup', 'Project', 'AC', 'ItemNo', 'Description', 'PlanTp','Ref', 'FillDoc'
                                , 'Sort', 'ReqQty', 'DemandSource', 'Unit', 'Vendor','ReqDate', 'ShipDate', 'OH'
                                , 'MLIKCode', 'LT', 'STDCost', 'LotSize','UOM']]
            
            expedite = expedite.fillna('')
            expedite = expedite.loc[(expedite['FillDoc'] == '')]
            expedite['ItemNo'] = expedite['ItemNo'].str.upper()
            expedite['Description'] = expedite['Description'].str.upper()
            expedite['MLIKCode'] = expedite['MLIKCode'].str.upper()
            expedite['Vendor'] = expedite['Vendor'].str.upper()
            expedite['ReqQty'] = expedite['ReqQty'].astype(float)
            expedite = expedite.loc[(expedite['MLIKCode'] == 'L')]
            expedite['ReqDate'] = pd.to_datetime(expedite['ReqDate'])
            expedite['new_reqDate'] = expedite['ReqDate'].apply(lambda x: today if x.date() < today else x)

            exp_in9_2 = pd.merge(expedite,in_9_2,how = 'left',on = 'ItemNo')
            exp_in9_2 = exp_in9_2.fillna('')
            def select_vendor(DataFrame):
                DataFrame['new_vendor'] = ''
                for index,row in DataFrame.iterrows():
                    if DataFrame.at[index,'Vendor_y']!='':
                        DataFrame.at[index,'new_vendor'] = DataFrame.at[index,'Vendor_y']
                    else:
                        DataFrame.at[index,'new_vendor'] = DataFrame.at[index,'Vendor_x']
            select_vendor(exp_in9_2)
            
            exp_in9_2 = exp_in9_2.rename(columns={'new_vendor':'Vendor','MLIKCode_x':'MLIKCode'})
            expedite = exp_in9_2.copy()

            expedite = expedite[['EntityGroup', 'Project', 'AC', 'ItemNo', 'Description', 'PlanTp','Ref', 'FillDoc'
                                , 'Sort', 'ReqQty', 'DemandSource', 'Unit', 'Vendor','ReqDate', 'ShipDate', 'OH'
                                , 'MLIKCode', 'LT', 'STDCost', 'LotSize','UOM','new_reqDate']]

            expedite['new_reqDate'] = pd.to_datetime(expedite['ReqDate'])
            buyer = buyer[['Vendor','Shortname','Owner1','Owner2','Owner3','email','email2','email3']]
            buyer['Vendor'] = buyer['Vendor'].astype(str)
            buyer['Shortname'] = buyer['Shortname'].astype(str)
            buyer['Vendor'] = buyer['Vendor'].str.upper()
            buyer['Shortname'] = buyer['Shortname'].str.upper()
            
            exp_buy = pd.merge(expedite,buyer,how = 'left',on = 'Vendor')
            exp_buy['Shortname'] = exp_buy['Shortname'].fillna('')
            exp_buy['Shortname'] = exp_buy['Shortname'].replace('','Check Vendor code')
            exp_buy['month#'] = exp_buy['new_reqDate'].dt.month
            exp_buy['year'] = exp_buy['new_reqDate'].dt.year
            exp_buy['month'] = exp_buy['new_reqDate'].dt.strftime('%B')
            exp_buy['ItemNo'] = exp_buy['ItemNo'].str.upper()
            exp_buy['UOM'] = exp_buy['UOM'].str.upper()
            exp_buy['Description'] = exp_buy['Description'].str.upper()

            base = exp_buy.groupby(['ItemNo','Description','Vendor','UOM','month#','month','year'],as_index=False)['ReqQty'].sum()
            base = base.sort_values(by=['ItemNo','year'], ascending=[False,True]) # type: ignore
            vendor_codes = base['Vendor'].unique()

            if not os.path.exists(output_dir):
                os.makedirs(output_dir)
            for vendor_code in vendor_codes:
                vendor_df = base[base['Vendor'] == vendor_code]
                #summary
                items_unicos = vendor_df['ItemNo'].drop_duplicates()
                items_unicos = items_unicos.count()
                output_filename = os.path.join(output_dir, f'FORECAST EZAIR {vendor_code}'+datestring+'.xlsx') # type: ignore
                vendor_df.to_excel(output_filename, index=False)

                buyer = buyer.fillna('')
                for index, row in buyer.iterrows():
                    if row['Vendor'] == vendor_code:
                        shortName = row['Shortname']
                        destinatarios = [row['email'], row['email2'], row['email3']]  # Lista de destinatarios
                        copiados = [row['Owner1'], row['Owner2']]  # Lista de personas copiadas

                        # Remove empty strings from destinatarios and copiados
                        destinatarios = [email for email in destinatarios if email]
                        copiados = [email for email in copiados if email]

                        # Ensure there is at least one recipient
                        if not destinatarios:
                            print(f"No valid recipients for vendor {vendor_code}")
                            continue

                        # Try to send the email
                        try:
                            # Inicializar el cliente de Outlook
                            pythoncom.CoInitialize()
                            outlook = Dispatch('Outlook.Application')
                            mail = outlook.CreateItem(0)
                            # List all accounts and find the desired account
                            namespace = outlook.GetNamespace("MAPI")
                            accounts = namespace.Accounts
                            desired_account = 'forecast.ezair@ezairinterior.com'  # replace with your desired email address
                            for account in accounts:
                                if account.SmtpAddress == desired_account:
                                    mail._oleobj_.Invoke(*(64209, 0, 8, 0, account))
                                    break

                            # Establecer destinatarios y personas copiadas
                            mail.To = ";".join(destinatarios)  # Unir los destinatarios con ";"
                            mail.CC = ";".join(copiados)  # Unir las personas copiadas con ";"
                            mail.Subject = f'FCST EZAir {shortName}'
                            mail.Body = f"{mail_value}"
                            attachment = mail.Attachments.Add(output_filename)
                            mail.Send()
                            del outlook
                        except Exception as e:
                            print(f"Failed to send email to {vendor_code} due to: {e}")
                            continue
        else:
            QMessageBox.critical(self, "confirmar actualizacion", f'Es necesario confirmar la actualizacion ')
            self.label_23.setHidden(True)
            print("done")

    def validate_good_liberation(self):
        expedite = self.get_expedite_information_clean_liberations()
        woinquiry = self.get_wo_inquiry_information()
        
        woinquiry['key_wo']  = woinquiry['SO_FCST'] +'-'+ woinquiry['Sub'] 
        woinquiry = woinquiry[['key_wo','WONo','Srt','SO_FCST','Sub']]
        woinquiry['Srt'] = woinquiry['Srt'].astype(str)
        woinquiry['Srt'] = woinquiry['Srt'].str.upper()
        woinquiry['left'] = woinquiry['Srt'].apply(lambda x: x[0] if x else None)
        woinquiry['left'] = woinquiry['Srt'].apply(lambda x: x[0:1] if x else None)
        woinquiry['Srt'] = woinquiry['Srt'].fillna('Default Value')  # Replace NaN with a default value
        woinquiry = woinquiry.loc[woinquiry['left']=='G']

        expedite = expedite.loc[expedite['DemandType'] != 'WOM']
        data = expedite.copy()
        df = pd.DataFrame(data)

        # Dividir la columna 'Part Number' y expandirla en nuevas filas
        df['FillDoc'] = df['FillDoc'].str.split(r'[,;]')
        df = df.explode('FillDoc')

        # Restablecer el índice después de la expansión
        df = df.reset_index(drop=True)
        df = df.drop_duplicates()
        df = df.loc[(df['FillDoc']!= 'WO') & (df['FillDoc']!= '0') & (df['FillDoc']!= 'IN') & (df['FillDoc']!= 'PO')]
        df = df[['DemandType','EntityGroup','DemandSource','ItemNo','AC','Ref','Sub','FillDoc']]
        
        def validation_demand_source(DataFrame):
            DataFrame['validation'] = ''
            for index,row in DataFrame.iterrows():
                if DataFrame.at[index,'DemandSource'] == DataFrame.at[index,'ItemNo']:
                    DataFrame.at[index,'validation'] = 'y'
                else:
                    DataFrame.at[index,'validation'] = 'n'
        validation_demand_source(df)

        df = df.loc[df['validation'] == 'y']
        df['key_exp'] = df['Ref'] +'-'+ df['Sub'] 
        df = df[['key_exp','DemandType','EntityGroup','DemandSource','AC','Ref','Sub','FillDoc']]
        df = df.drop_duplicates()

        df = pd.merge(df,woinquiry,how = 'left',left_on = 'FillDoc',right_on = 'WONo')
        df = df.fillna('')
        df = df.loc[df['key_wo'] != '']
        df = df.drop_duplicates()

        def validation_demand_source(DataFrame):
            DataFrame['validation'] = ''
            for index,row in DataFrame.iterrows():
                if DataFrame.at[index,'key_wo'] == DataFrame.at[index,'key_exp']:
                    DataFrame.at[index,'validation'] = 'MAKE TO ORDER ASIGNED WELL'
                else:
                    DataFrame.at[index,'validation'] = 'MAKE TO ORDER ASIGNED WRONG'
        validation_demand_source(df)


        wb = xw.Book()
        sheet1 = wb.sheets.add()
        sheet1.name = "validation"
        sheet1.range('A1').value = df

        sheet2 = wb.sheets.add()
        sheet2.name = "WO"
        sheet2.range('A1').value = woinquiry

    def new_future_shortage_split_demand(self):
        if self.checkBox_22.isChecked() and self.checkBox_23.isChecked() and self.checkBox_24.isChecked() and self.checkBox_25.isChecked() and self.checkBox_26.isChecked() and self.checkBox_27.isChecked():
    
            self.label_36.setHidden(False)    
            self.label_13.setHidden(False)
            now = datetime.now()
            today = now.strftime("%m/%d/%Y %I:%M:%S")
            self.label_13.setText(f'{today}')
            QApplication.processEvents()
            
            oo ,piv_oo = self.get_clean_open_order()
            expedite_1,expedite_pivot,expedite_fs  = self.get_clean_expedite_information_future_shortages() 
            buyer = self.get_vendor_information_future_shortages()

            piv_safe = self.get_safety_stock()
            piv_rwk_location = self.get_rework_information_table()
            oh = self.get_oh_information()
            lean_dna = self.get_LeanDNA_information()
            range_formula = expedite_1.shape[0]+1

    def new_future_shortage_all_forumas(self):
        if self.checkBox_22.isChecked() and self.checkBox_23.isChecked() and self.checkBox_24.isChecked() and self.checkBox_25.isChecked() and self.checkBox_26.isChecked() and self.checkBox_27.isChecked():
        
            self.label_36.setHidden(False)    
            self.label_13.setHidden(False)
            now = datetime.now()
            today = now.strftime("%m/%d/%Y %I:%M:%S")
            self.label_13.setText(f'{today}')
            QApplication.processEvents()
            
            oo ,piv_oo = self.get_clean_open_order()
            expedite_1,expedite_pivot,expedite_fs  = self.get_clean_expedite_information_future_shortages() 
            buyer = self.get_vendor_information_future_shortages()

            piv_safe = self.get_safety_stock()
            piv_rwk_location = self.get_rework_information_table()
            oh = self.get_oh_information()
            lean_dna = self.get_LeanDNA_information()
            range_formula = expedite_1.shape[0]+1
            
            base = pd.merge(expedite_1,piv_safe,how = 'left',on = 'ItemNo')
            base1 = pd.merge(base,piv_rwk_location,how = 'left',on = 'ItemNo')
            base1 = base1.fillna('')    
            base1 = base1.rename(columns = {'Owner1':'BUYER','Owner3':'BS'})

            base1['SS'] = base1['SS'].fillna(0)
            base1['SS'] = base1['SS'].replace('', 0)
            base1['RWK OH'] = base1['RWK OH'].fillna(0)
            base1['RWK OH'] = base1['RWK OH'].replace('', 0)
            base1 = base1.fillna('')

            def fill_no_buyer(DataFrame):
                for index,row in DataFrame.iterrows():
                    if row['BUYER'] =='':
                        DataFrame.at[index,'BUYER'] = 'No Buyer'
                    else:
                        pass
            fill_no_buyer(base1)
            
            base2 = pd.merge(base1,oh,how = 'left',on = 'ItemNo')
            base2 = pd.merge(base2,expedite_fs,how = 'left',on = 'ItemNo')
            base2 = base2.rename(columns = {'Sort':'FS'})
            selected_columns_positions_exp_piv = [0,1,2,3,4,5,6,7,8,9,10] 
            selected_columns = expedite_pivot.columns[selected_columns_positions_exp_piv]
            selected_data_expedite_pivot = expedite_pivot[selected_columns]

            need_another_columns = selected_data_expedite_pivot.columns.to_list()
            columns_to_remove = ['ItemNo','01_Past due']
            need_another_columns = [col for col in need_another_columns if col not in columns_to_remove]
            need_another_columns = [col.replace('_y', '_z') for col in need_another_columns]
            new_df = pd.DataFrame(columns = need_another_columns,index = None )
            selected_data_expedite_pivot = selected_data_expedite_pivot.fillna('0')
            select_piv_oo  = [0,1,2,3,4,5,6,7,8,9,10,11]
            selected_columns_oo = piv_oo.columns[select_piv_oo]
            selected_data_oo = piv_oo[selected_columns_oo]
            selected_data_oo = selected_data_oo.fillna('0')
            
            base3 = pd.merge(base2,selected_data_expedite_pivot,how = 'left',on = 'ItemNo')
            
            """-------------------------------------------------------------------aqui debe de ser agregadas las columnas que no se usan """
            base3['Received'] = '0'
            base3['Incoming'] = '0' 
            
            base4 = pd.merge(base3,selected_data_oo,how = 'left',on = 'ItemNo')   
            base4 = base4.fillna('0')
            base4['FS'] = base4['FS'].str.replace('0','NO')
            base4['LT'] = pd.to_numeric(base4['LT'], errors='coerce')

            def lt_logic(DataFrame):
                for index, row in DataFrame.iterrows():
                    if pd.isnull(row['LT']):  # Check if value is NaN
                        DataFrame.at[index, 'LT'] = 0  # Replace NaN with 0
                    else:
                        DataFrame.at[index, 'LT'] = math.ceil(row['LT'] / 5)

            lt_logic(base4)           
            
            file_path = r"J:\\Departments\Material Control\Purchasing\Tools\ComprasDB\PythonCode\future_shortages.xlsm"
            
            def is_file_open(file_path):
                for proc in psutil.process_iter():
                    try:
                        files = proc.open_files()
                        for f in files:
                            if file_path.lower() in f.path.lower():
                                return True
                    except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
                        pass
                return False
            counter = 1
            while is_file_open(file_path):
                file_path = file_path.replace('.xlsm', f'_{counter}.xlsm')
                counter += 1
            wb = xw.Book(file_path)
            
            sheet1 = wb.sheets.add()
            sheet1.name = "OOR"
            sheet2 = wb.sheets.add()
            sheet2.name = "PIVOT OOR"
            sheet3 = wb.sheets.add()
            sheet3.name = "Vendor"
            sheet3.range('A1').value = buyer


            sheet4 = None
            for sheet in wb.sheets:
                if sheet.name == 'FUTURE SHORTAGES REPORT':
                    sheet4 = sheet
                    break

            if sheet4 is not None:

                sheet4.range('A3').value = base4
                chart = base4.copy()
                sheet4.range('AO3').value = new_df
            else:
                print("Sheet 'FUTURE SHORTAGES REPORT' not found.")

            column_to_delete = "AO" 
            column_range = sheet4.range(f"{column_to_delete}:{column_to_delete}")
            column_range.delete()

            sheet4.range('AO2').value = 2
            sheet4.range('AP2').value = 3
            sheet4.range('AQ2').value = 4
            sheet4.range('AR2').value = 5
            sheet4.range('AS2').value = 6
            sheet4.range('AT2').value = 7
            sheet4.range('AU2').value = 8
            sheet4.range('AV2').value = 9
            sheet4.range('AW2').value = 10
            sheet4.range('AX2').value = 11

            sheet4.api.Columns(32).Insert()
            sheet4.api.Columns(33).Insert()
            sheet4.api.Columns(34).Insert()
            
            cell_Q3 = sheet4.range('Q3').value
            cell_R3 = sheet4.range('R3').value
            cell_S3 = sheet4.range('S3').value

            sheet4.range('AF3').value = cell_Q3
            sheet4.range('AG3').value = cell_R3
            sheet4.range('AH3').value = cell_S3

            white_text_color = (240, 242, 245)
            black_value = (20, 20, 20)
            purple_value = (89, 17, 245)
            gary_value = (92, 92, 91)
            blue_value = (31, 100, 161)
            orange_value = (245, 114, 66)

            for cell in sheet4.range('A3:O3'):
                cell.color = black_value
                cell.api.Font.Color = xw.utils.rgb_to_int(white_text_color)
            
            for cell in sheet4.range('P3:Z3'):
                cell.color = purple_value
                cell.api.Font.Color = xw.utils.rgb_to_int(white_text_color)

            sheet4.range('P1').value = "NEED WK"
            for cell in sheet4.range('P1'):
                cell.color = purple_value
                cell.api.Font.Color = xw.utils.rgb_to_int(white_text_color)
            try:
                sheet4.range('P1:Z1').api.Merge()
            except Exception as e:
                print(f"An error occurred: {e}")

            for cell in sheet4.range('AA3:AQ3'):
                cell.color = gary_value
                cell.api.Font.Color = xw.utils.rgb_to_int(white_text_color)
            
            sheet4.range('AA1').value = "OOR"
            for cell in sheet4.range('AA1'):
                cell.color = gary_value
                cell.api.Font.Color = xw.utils.rgb_to_int(white_text_color)
            sheet4.range('AA1:AQ1').api.Merge()

            for cell in sheet4.range('AR3:BA3'):
                cell.color = black_value
                cell.api.Font.Color = xw.utils.rgb_to_int(white_text_color)
            
            for cell in sheet4.range('BB3:BL3'):
                cell.color = blue_value
                cell.api.Font.Color = xw.utils.rgb_to_int(white_text_color)

            for cell in sheet4.range('BJ3:BN3'):
                cell.color = orange_value
                cell.api.Font.Color = xw.utils.rgb_to_int(white_text_color)

            sheet4.range('BJ1').value = "LeanDNA"
            for cell in sheet4.range('BJ1'):
                cell.color = orange_value
                cell.api.Font.Color = xw.utils.rgb_to_int(white_text_color)
            sheet4.range('BJ1:BN1').api.Merge()

            sheet4.range('BG3').value = "Action"
            sheet4.range('BH3').value = "RC"
            sheet4.range('BI3').value = "Comments Buyer"

            sheet4.range('BJ3').value = "Garden Grove C37 Inventory"
            sheet4.range('BK3').value = "Bolsa C91 Inventory"
            sheet4.range('BL3').value = "Tijuana C44 Inventory"
            sheet4.range('BM3').value = "Santa Maria C61 Inventory"
            sheet4.range('BN3').value = "Montreal CG1 Inventory"

            sheet5 = wb.sheets.add()
            sheet5.name = "piv_exp"
            sheet5.range('A1').value = expedite_pivot

            sheet6 = wb.sheets.add()
            sheet6.name = "safe"
            sheet6.range('A1').value = piv_safe

            sheet7 = wb.sheets.add()
            sheet7.name = "fs"
            sheet7.range('A1').value = expedite_fs

            sheet8 = wb.sheets.add()
            sheet8.name = "leanDna"
            sheet8.range('A1').value = lean_dna

            sheet9 = wb.sheets.add()
            sheet9.name = "RECIBOS"
            sheet9.range('A1').value = 'conca'
            sheet9.range('B1').value = 'Project-no'
            sheet9.range('C1').value = 'Entity-code'
            sheet9.range('D1').value = 'GL-code'
            sheet9.range('E1').value = 'Item-no'
            sheet9.range('F1').value = 'Description'
            sheet9.range('G1').value = 'Plantype'
            sheet9.range('H1').value = 'Qty'
            sheet9.range('I1').value = 'Receipt-Date'
            sheet9.range('J1').value = 'Receipt-no'
            sheet9.range('K1').value = 'Receipt-Line-no'
            sheet9.range('L1').value = 'PO-No'
            sheet9.range('M1').value = 'Line-no'
            sheet9.range('N1').value = 'Po-Date'
            sheet9.range('O1').value = 'Promise-Date'
            sheet9.range('P1').value = 'Revised-Promised-Date'
            sheet9.range('Q1').value = 'Vendor-Code'
            sheet9.range('R1').value = 'Vendor-Name'
            sheet9.range('S1').value = 'cnet'
            sheet9.range('T1').value = 'cfob'
            sheet9.range('U1').value = 'Fob-Vendor'
            sheet9.range('V1').value = 'cqtyr'
            sheet9.range('W1').value = 'cval'
            sheet9.range('X1').value = 'AC-No'
            sheet9.range('Y1').value = 'Waybill No'
            sheet9.range('Z1').value = 'Reference'
            sheet9.range('AA1').value = 'User-Id'
            sheet9.range('AB1').value = 'Buyer'
            sheet9.range('AC1').value = 'Aging'
            sheet9.range('AD1').value = 'STATUS'
            sheet9.range('AE1').value = 'Cur Std'
            sheet9.range('AF1').value = 'Total'

            sheet10 = wb.sheets.add()
            sheet10.name = "IMPOS"
            sheet10.range('A1').value = 'Part Number'
            sheet10.range('B1').value = 'Description'
            sheet10.range('C1').value = 'qty'
            sheet10.range('D1').value = 'UOM'
            sheet10.range('E1').value = 'Unit weight'
            sheet10.range('F1').value = 'Unit Value(USD)'
            sheet10.range('G1').value = 'Unit Value(USD)'
            sheet10.range('H1').value = 'Packaging Type'
            sheet10.range('I1').value = 'Packaging Quantity'
            sheet10.range('J1').value = 'Gross weight'
            sheet10.range('K1').value = 'Packaging Weight'
            sheet10.range('L1').value = 'Net Weight'
            sheet10.range('M1').value = 'Country of Origin'
            sheet10.range('N1').value = 'US IMP HTS Code'
            sheet10.range('O1').value = 'US EXP HTS Code'
            sheet10.range('P1').value = 'Control'
            sheet10.range('Q1').value = 'PO'
            sheet10.range('R1').value = 'Line'

            sheet10.range('S1').value = 'POLine'
            sheet10.range('T1').value = 'License No.'
            sheet10.range('U1').value = 'License Exception'

            sheet11 = wb.sheets.add()
            sheet11.name = "RWK OH"
            sheet11.range('A1').value = piv_rwk_location

            sheet4.activate()
            range_to_search = 'P3:AQ3'  
            search_string = "_x"
            replace_string = "" 
            sheet4.range(range_to_search).api.Replace(
                What=search_string,
                Replacement=replace_string,
                LookAt=xw.constants.LookAt.xlPart, 
                SearchOrder=xw.constants.SearchOrder.xlByRows, 
                MatchCase=False, 
                SearchFormat=False,
                ReplaceFormat=False)
            
            range_to_search = 'P3:AQ3'
            search_string = "_y"
            replace_string = "" 
            sheet4.range(range_to_search).api.Replace(
                What=search_string,
                Replacement=replace_string,
                LookAt=xw.constants.LookAt.xlPart, 
                SearchOrder=xw.constants.SearchOrder.xlByRows, 
                MatchCase=False, 
                SearchFormat=False,
                ReplaceFormat=False)

            chart_buyer = chart[['BUYER']].drop_duplicates()
                    
            sheet12 = wb.sheets.add()
            sheet12.name = "Summary"
            sheet12.range('B4').value = chart_buyer
            
            sheet4.activate()
            range_to_copy = 'AR3:BA3'
            val = sheet4.range(range_to_copy).value 
            sheet12.activate()
            sheet12.range('D4').value = val

            cell_value_d4 = sheet12.range('D4').value
            before_decimal_d4 = str(cell_value_d4).split('.')[0][-2:]
            final_cell_value_d4 = f'Wk{before_decimal_d4}'
            time.sleep(.5)
            print(final_cell_value_d4)
            sheet12.range('D4').value = final_cell_value_d4
            cell_value_e4 = sheet12.range('E4').value
            before_decimal_e4 = str(cell_value_e4).split('.')[0][-2:]
            final_cell_value_e4 = f'Wk{before_decimal_e4}'
            time.sleep(.5)
            print(final_cell_value_e4)
            sheet12.range('e4').value = final_cell_value_e4
            cell_value_f4 = sheet12.range('F4').value
            before_decimal_f4 = str(cell_value_f4).split('.')[0][-2:]
            final_cell_value_f4 = f'Wk{before_decimal_f4}'
            time.sleep(.5)
            print(final_cell_value_f4)
            sheet12.range('F4').value = final_cell_value_f4
            cell_value_g4 = sheet12.range('G4').value
            before_decimal_g4 = str(cell_value_g4).split('.')[0][-2:]
            final_cell_value_g4 = f'Wk{before_decimal_g4}'
            time.sleep(.5)
            print(final_cell_value_g4)
            sheet12.range('G4').value = final_cell_value_g4
            cell_value_h4 = sheet12.range('H4').value
            before_decimal_h4 = str(cell_value_h4).split('.')[0][-2:]
            final_cell_value_h4 = f'Wk{before_decimal_h4}'
            time.sleep(.5)
            print(final_cell_value_h4)
            sheet12.range('H4').value = final_cell_value_h4
            cell_value_i4 = sheet12.range('I4').value
            before_decimal_i4 = str(cell_value_i4).split('.')[0][-2:]
            final_cell_value_i4 = f'Wk{before_decimal_i4}'
            time.sleep(.5)
            print(final_cell_value_i4)
            sheet12.range('I4').value = final_cell_value_i4
            cell_value_j4 = sheet12.range('J4').value
            before_decimal_j4 = str(cell_value_j4).split('.')[0][-2:]
            final_cell_value_j4 = f'Wk{before_decimal_j4}'
            time.sleep(.5)
            print(final_cell_value_j4)
            sheet12.range('J4').value = final_cell_value_j4
            cell_value_k4 = sheet12.range('K4').value
            before_decimal_k4 = str(cell_value_k4).split('.')[0][-2:]
            final_cell_value_k4 = f'Wk{before_decimal_k4}'
            time.sleep(.5)
            print(final_cell_value_k4)
            sheet12.range('K4').value = final_cell_value_k4
            cell_value_l4 = sheet12.range('L4').value
            before_decimal_l4 = str(cell_value_l4).split('.')[0][-2:]
            final_cell_value_l4 = f'Wk{before_decimal_l4}'
            time.sleep(.5)
            print(final_cell_value_l4)
            sheet12.range('L4').value = final_cell_value_l4
            cell_value_m4 = sheet12.range('M4').value
            before_decimal_m4 = str(cell_value_m4).split('.')[0][-2:]
            final_cell_value_m4 = f'Wk{before_decimal_m4}'
            time.sleep(.5)
            print(final_cell_value_m4)
            sheet12.range('M4').value = final_cell_value_m4

            data_range = 'C4:M12' 
            chart = sheet12.charts.add()
            chart.chart_type = 'line'
            chart.set_source_data(sheet12.range(data_range))
            chart.top = sheet12.range('C17').top
            chart.left = sheet12.range('C17').left

            range_to_copy = 'D4:M4'
            val = sheet12.range(range_to_copy).value 
            
            sheet12.range('Q4').value = val
            sheet12.range('P4').value = 'Type'
            sheet12.range('P5').value = 'Shortage'
            sheet12.range('P6').value = 'Coverage'


            sheet12.activate()
            data_range = 'P4:Z6' 

            chart = sheet12.charts.add()
            chart.chart_type = 'bar_clustered'

            chart.set_source_data(sheet12.range(data_range))
            chart.top = sheet12.range('P8').top
            chart.left = sheet12.range('P8').left
            sheet_to_hide = wb.sheets['leanDna']
            sheet_to_hide.api.Visible = False 
            sheet_to_hide = wb.sheets['fs']
            sheet_to_hide.api.Visible = False 
            sheet_to_hide = wb.sheets['safe']
            sheet_to_hide.api.Visible = False 
            sheet_to_hide = wb.sheets['piv_exp']
            sheet_to_hide.api.Visible = False 
            sheet_to_hide = wb.sheets['Vendor']
            sheet_to_hide.api.Visible = False 
            sheet_to_hide = wb.sheets['RWK OH']
            sheet_to_hide.api.Visible = False 

            column_index_to_insert = 62
            sheet4.api.Columns(column_index_to_insert).Insert()
            column_index_to_insert = 63
            sheet4.api.Columns(column_index_to_insert).Insert()
            column_index_to_insert = 64
            sheet4.api.Columns(column_index_to_insert).Insert()
            column_index_to_insert = 65
            sheet4.api.Columns(column_index_to_insert).Insert()
            column_index_to_insert = 66
            sheet4.api.Columns(column_index_to_insert).Insert()
            column_index_to_insert = 67
            sheet4.api.Columns(column_index_to_insert).Insert()
            column_index_to_insert = 68
            sheet4.api.Columns(column_index_to_insert).Insert()
            column_index_to_insert = 69
            sheet4.api.Columns(column_index_to_insert).Insert()
            column_index_to_insert = 70
            sheet4.api.Columns(column_index_to_insert).Insert()
            column_index_to_insert = 71
            sheet4.api.Columns(column_index_to_insert).Insert()
            sheet4.range('BJ3').value = val
            red_value = (140, 11, 35)
            sheet4.range('BJ1').value = 'RWK COVERAGE'
            for cell in sheet4.range('BJ1:BS1'):
                cell.color = red_value
                cell.api.Font.Color = xw.utils.rgb_to_int(white_text_color)

            for cell in sheet4.range('BJ3:BS3'):
                cell.color = red_value
                cell.api.Font.Color = xw.utils.rgb_to_int(white_text_color)

            today = datetime.today()
            monday = today - timedelta(days=today.weekday())
            monday_plus_4 = monday + timedelta(days=4)
            bday = pd.offsets.BDay()
            monday_plus_4_plus_11 = monday_plus_4 + 11 * bday
            buyer = self.get_vendor_information_future_shortages()
            buyer = buyer[['Vendor','Owner1']]
            buyer = buyer.drop_duplicates()
            try:
                conn = sqlite3.connect(path_conn)
                open_order_query = "SELECT * FROM openOrder"
                oo = pd.read_sql_query(open_order_query, conn)
                oo = oo.fillna('')
            except Exception as error:
                QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion al acction {e}')
            finally:
                pass
            try:
                conn = sqlite3.connect(path_conn)
                query_actions = "SELECT * FROM ActionMessages"
                action = pd.read_sql_query(query_actions, conn)
                action = action.fillna('')
                action['PO'] = action['PO'].astype(str)
                action['Line'] = action['Line'].astype(str)
                action['key'] = action['PO'] +"-"+ action['Line']
                action = action[['key','ACD','REQDT']]
                action['ACD'] = action['ACD'].str.upper()
                action = action.loc[action['ACD']!='AD']
                action = action.loc[action['ACD']!='PR']
                action = action.drop_duplicates()
            except Exception as e:
                QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion al acction {e}')
            finally:
                conn.close()
            oo['ItemNo']  = oo['ItemNo'].str.upper().str.strip()
            oo = oo[['PONo', 'Ln','Vendor','VendorName','ItemNo','ItemDescription','Rev','UM','OrdQ','RcvdQ','OpnQ','Unit$','NetOpnVal'
                                        ,'PODt','ReqDt','PromDt', 'RevPrDt','LT','NoteStatus']]
            oo = oo.applymap(lambda x: x.strip().upper() if isinstance(x, str) else x)
            
            oo['PONo'] = oo['PONo'].astype(str)
            oo['Ln'] = oo['Ln'].astype(str)
            oo['key'] = oo['PONo'] +"-"+ oo['Ln']
            oo['PODt'] = pd.to_datetime(oo['PODt'])
            oo['ReqDt'] = pd.to_datetime(oo['ReqDt'])
            oo['PromDt'] = pd.to_datetime(oo['PromDt'])
            oo['RevPrDt'] = pd.to_datetime(oo['RevPrDt'])
            base = pd.merge(oo,action,how = 'left',on = 'key')
            
            
            """1-	En el action eliminar AD, una vez eliminado hacer el cruce con PO&LN
                2-	CN – Eliminar las fechas de la columna EZA Req. Date, para PN Misc eliminar el action CN
                3-	PD – EZ Req. Date siguiente Mes y Rev-Pr-Dt que sea antes que EZ Req. Date poner RO
                PD – EZ Req. Date siguiente Mes y Rev-Pr-Dt que sea después que EZ Req. Date poner RI
                PD – EZ Req. Date después de la semana actual y en el pasado en Rev-Pr-Dt poner RO
                PD –Rev-Pr-Dt después de esta semana y en EZ Req. Date que este en el pasado poner RI
                PD- si no cumple con ninguna de las siguientes, eliminar Action PD y fecha EZ Req. Date
                4-	RI- EZ Req. Date fecha past due, poner al Lunes de la WK Actual
                5-	RO – Si la fecha EZ Req. Date esta para hoy, eliminar el Action y la fecha EZ Req. Date
                6-	#N/A – Eliminar 
                7-	Para los PN Misc que NO tienen Buyer, poner “Servicios” en Buyer
            """
            base1 = pd.merge(base,buyer,how = 'left',on = 'Vendor')
            base1 = base1.rename(columns = {'Owner1':'Buyer'})
            base1['ItemNo'] = base1['ItemNo'].str.upper()
            base1 = base1.fillna('')
            def clean_misc_cn_fst_step(DataFrame):
                for index, row in DataFrame.iterrows():
                    if row['ItemNo'][-4:] == 'MISC' and row['ACD'] == 'CN':
                        DataFrame.at[index, 'ACD'] = '' 
                        DataFrame.at[index, 'REQDT'] = '' 
                    else:
                        if row['ACD'] == 'CN':
                            DataFrame.at[index, 'REQDT'] = '' 
                        else:
                            pass
            clean_misc_cn_fst_step(base1)
            base1 = base1.fillna('')
            def Function_process(DataFrame, last_day_of_month, next_monday_date, monday_date, friday_date):
                for index, row in DataFrame.iterrows():
                    if row['ACD'] == 'PD':
                        req_date = row['REQDT']
                        rev_pr_date = row['RevPrDt']
                        if req_date > last_day_of_month and rev_pr_date <= req_date:
                            DataFrame.at[index, 'ACD'] = 'RO' 
                        
                        elif req_date > last_day_of_month and rev_pr_date >= req_date:
                            DataFrame.at[index, 'ACD'] = 'RI' 
                        
                        elif req_date > pd.Timestamp(next_monday_date) and req_date <= pd.Timestamp(monday_date):
                            DataFrame.at[index, 'ACD'] = 'RO'
                        
                        elif rev_pr_date > pd.Timestamp(next_monday_date) and req_date <= pd.Timestamp(monday_date):
                            DataFrame.at[index, 'ACD'] = 'RI'
                        # else:
                        #     DataFrame.at[index, 'ACD'] = ''
                        #     DataFrame.at[index, 'REQDT'] = ''
            base1 = base1.fillna('')
            def Function_process2(DataFrame, monday_date):
                for index, row in DataFrame.iterrows():
                    if row['ACD'] == 'RI':
                        req_date = row['REQDT']
                        if req_date <= pd.Timestamp(monday_date):
                            DataFrame.at[index, 'REQDT'] = monday_date
            base1 = base1.fillna('')

            def Function_process3(DataFrame, monday_date, friday_date):
                for index, row in DataFrame.iterrows():
                    if row['ACD'] == 'RI':
                        rev_pr_date = row['RevPrDt']
                        if rev_pr_date >= pd.Timestamp(monday_date) and rev_pr_date <= pd.Timestamp(friday_date):
                            DataFrame.at[index, 'ACD'] = ''
                            DataFrame.at[index, 'REQDT'] = ''
                        else:
                            pass
            base1 = base1.fillna('')
            
            def Function_process4(DataFrame, monday_date):
                for index, row in DataFrame.iterrows():
                    if row['ACD'] == 'RO':
                        req_date = row['REQDT']
                        if req_date == pd.Timestamp(monday_date):
                            DataFrame.at[index, 'ACD'] = ''
                            DataFrame.at[index, 'REQDT'] = ''
                        else:
                            pass
            base1 = base1.fillna('')
            def Function_process5(DataFrame, monday_date):
                for index, row in DataFrame.iterrows():
                    if row['ACD'] == 'RI':
                        req_date = row['REQDT']
                        if req_date == pd.Timestamp(monday_date):
                            DataFrame.at[index, 'ACD'] = ''
                            DataFrame.at[index, 'REQDT'] = ''
                        else:
                            pass
                            
            base1 = base1.fillna('')
            def Function_process6(DataFrame, monday_date):
                for index, row in DataFrame.iterrows():
                    if row['ACD'] == 'RI':
                        rev_pr_date = row['RevPrDt']
                        if rev_pr_date == pd.Timestamp(monday_date):
                            DataFrame.at[index, 'ACD'] = ''
                            DataFrame.at[index, 'REQDT'] = ''
                        else:
                            pass
            base1 = base1.fillna('')
            def Function_process7(DataFrame, monday_date):
                for index, row in DataFrame.iterrows():
                    if row['ACD'] == 'RI':
                        rev_pr_date = pd.Timestamp(row['RevPrDt'])
                        req_date = pd.Timestamp(row['REQDT'])
                        if rev_pr_date < req_date:
                            DataFrame.at[index, 'ACD'] = ''
                            DataFrame.at[index, 'REQDT'] = ''
                        else:
                            pass
            base1 = base1.fillna('')
            def service_misc(DataFrame):
                for index, row in DataFrame.iterrows():
                    if row['ItemNo'][-4:] == 'MISC' and row['Buyer'] =='':
                        DataFrame.at[index, 'Buyer'] = 'SERVICIOS' 
                    else:
                        pass
            service_misc(base1)
            base1 = base1.fillna('')

            # Convert columns to datetime format
            base1['REQDT'] = pd.to_datetime(base1['REQDT'])
            base1['RevPrDt'] = pd.to_datetime(base1['RevPrDt'])

            today = datetime.today()
            if today.month == 12:
                last_day_of_month = datetime(today.year + 1, 1, 1) - timedelta(days=1)
            else:
                last_day_of_month = datetime(today.year, today.month + 1, 1) - timedelta(days=1)
            next_monday = today + timedelta(days=(7 - today.weekday()))
            next_monday_date = next_monday.date()
            monday_of_current_week = today - timedelta(days=today.weekday())
            monday_date = monday_of_current_week.date()

            monday_of_current_week1 = today - timedelta(days=today.weekday())
            friday_of_current_week = monday_of_current_week1 + timedelta(days=4)
            friday_date = friday_of_current_week.date()

            Function_process(base1, last_day_of_month, next_monday_date, monday_date, friday_date)
            Function_process2(base1, monday_date)
            Function_process3(base1, monday_date, friday_date)
            Function_process4(base1, monday_date)
            Function_process5(base1, monday_date)      
            Function_process6(base1, monday_date)  
            Function_process7(base1, monday_date)                  

            base1['CONFIRMED RECOVERY'] = ''
            base1['PO STATUS'] = ''
            base1['PO RECOVERY WK'] = ''
            base1['TOTAL OPEN QTY'] = ''
            base1['Qty In Transit'] = ''
            base1['Tracking'] = ''
            base1['ETA'] = ''
            base1['Recibos'] = ''
            base1['Control'] = ''
            base1['Impo'] = ''
            base1['comments'] = ''
            
            """ESTO SE MODIFICO"""
            base1['Comments Manager'] = ''
            base1['Critical Qty'] = ''
            base1['Shrt Wk'] = ''
            base1['Action'] = ''
            base1['GAP'] = ''
            base1['PO Status2'] = ''
            base1['Aging PO Status2'] = ''
            base1['Aging'] = ''
            base1['Estatus Confirmacion'] = ''
            
            base1 = base1[['Buyer','key','PONo', 'Ln', 'Vendor', 'VendorName', 'ItemNo', 'ItemDescription','Rev', 'UM', 'OrdQ',
                        'RcvdQ', 'OpnQ', 'Unit$', 'NetOpnVal', 'PODt','ReqDt', 'PromDt', 'RevPrDt', 'ACD', 'REQDT','LT', 'NoteStatus',
                        'CONFIRMED RECOVERY','PO STATUS','PO RECOVERY WK','TOTAL OPEN QTY','Qty In Transit','Tracking','ETA','Recibos'
                        ,'Control','Impo','comments','Comments Manager','Critical Qty',
                        'Shrt Wk','Action','GAP','PO Status2','Aging PO Status2','Aging','Estatus Confirmacion']]

            base1_qty_rows = base1.shape[0]+1
            sheet1.range('A4').value = base1
            column_letter = 'A'
            column_range = sheet1.range(f'{column_letter}:{column_letter}')
            column_range.delete()
            
            sheet1.range('R1').value = 'Fecha Lunes'
            sheet1.range('S1').value = monday.strftime("%m/%d/%Y")
            sheet1.range('Y1').formula = "=S1+4"                        #monday_plus_4.strftime("%m/%d/%Y")
            sheet1.range('AA1').formula = "=WORKDAY(Y1,11)"             #monday_plus_4_plus_11.strftime("%m/%d/%Y")

            #NOTE: AQUI ESTA LA PRIMER MODIFICACION 
            range_formula_oo = base1.shape[0]+1
            sheet1.range(f'A{range_formula_oo+4}').value = '.'
            sheet1.range(f'A{range_formula_oo+5}').value = '.'
            sheet1.range(f'A{range_formula_oo+6}').value = '.'
            
            today = datetime.now()
            start_date = pd.Timestamp(today.year, today.month, 1)
            end_date = pd.Timestamp(today.year, today.month, calendar.monthrange(today.year, today.month)[1])
            dates = pd.date_range(start=start_date, end=end_date, freq='D')
            mondays = dates[dates.dayofweek == 0]

            # Get the first, second, and third Monday and format them
            first_monday = mondays[0].strftime('%m/%d/%Y')
            second_monday = mondays[1].strftime('%m/%d/%Y')
            third_monday = mondays[2].strftime('%m/%d/%Y')
                        
            sheet1.range(f'X{range_formula_oo+4}').value = first_monday
            sheet1.range(f'X{range_formula_oo+5}').value = second_monday
            sheet1.range(f'X{range_formula_oo+6}').value = third_monday
            
            column_index = 30
            sheet4.api.Columns(column_index).Insert()
            sheet4.range('AD3').value = 'In transit'            
            sheet4.range(f'AD4:AD{range_formula+2}').formula = "=SUMIF(OOR!G:G,'FUTURE SHORTAGES REPORT'!B:B,OOR!AB:AB)"
            
            
            column_letter = 'AR'
            column_range = sheet4.range(f'{column_letter}:{column_letter}')
            column_range.delete()
            sheet_to_cut = wb.sheets['FUTURE SHORTAGES REPORT']
            sheet_to_cut.range('O1:O10000').api.Cut()
            sheet_to_cut.range('L1').api.Insert(xw.constants.InsertShiftDirection.xlShiftToRight)
            wb.app.calculate()
            
            sheet4.range(f'AR4:AR{range_formula+2}').formula = '=IF($AR$2>I4,0,(O4+AA4)-(Q4+P4+M4))'
            sheet4.range(f'AS4:AS{range_formula+2}').formula = '=IF($AS$2>I4,0,(AR4+AH4+AG4+AD4+AB4)-R4)'
            sheet4.range(f'AT4:AT{range_formula+2}').formula = '=IF(AT$2>I4,0,AS4-S4+AI4)'
            sheet4.range(f'AU4:AU{range_formula+2}').formula = '=IF(AU$2>I4,0,AT4-T4+AJ4)'
            sheet4.range(f'AV4:AV{range_formula+2}').formula = '=IF(AV$2>I4,0,AU4-U4+AK4)'
            sheet4.range(f'AW4:AW{range_formula+2}').formula = '=IF(AW$2>I4,0,AV4-V4+AL4)'
            sheet4.range(f'AX4:AX{range_formula+2}').formula = '=IF(AX$2>I4,0,AW4-W4+AM4)' 
            sheet4.range(f'AY4:AY{range_formula+2}').formula = '=IF(AY$2>I4,0,AX4-X4+AN4)'
            sheet4.range(f'AZ4:AZ{range_formula+2}').formula = '=IF(AZ$2>I4,0,AY4-Y4+AO4)'
            sheet4.range(f'BA4:BA{range_formula+2}').formula = '=IF(BA$2>I4,0,AZ4-Z4+AP4)'
            
            
            source_column = "BA"
            target_column = "BB"
            sheet4[f"{target_column}:{target_column}"].api.EntireColumn.Copy()
            sheet4[f"{source_column}:{source_column}"].api.EntireColumn.PasteSpecial(
                Paste=-4163,  # xlPasteFormats
                Operation=-4142,  # xlNone
                SkipBlanks=False,
                Transpose=False,
            )

            xw.apps.active.api.CutCopyMode = False

            sheet4.range('BB3').value = "FLAG"                  
            
            sheet4.range('BC3').value = "SHT WK"
            sheet4.range(f'BC4:BC{range_formula+2}').formula = '=IF(AR4<0,$AR$3,IF(AS4<0,$AS$3,IF(AT4<0,$AT$3,IF(AU4<0,$AU$3,IF(AV4<0,$AV$3,IF(AW4<0,$AW$3,IF(AX4<0,$AX$3,IF(AY4<0,$AY$3,IF(AZ4<0,$AZ$3,IF(BA4<0,$BA$3,"zz"))))))))))'
            sheet4.range('BD3').value = "GAP"
            sheet4.range(f'BD4:BD{range_formula+2}').formula ='=COUNTIF(AR4:BA4,"<0")'
            sheet4.range('BE3').value = "Total shortage Qty"
            sheet4.range(f'BE4:BE{range_formula+2}').formula = '=IF(MIN(AR4:BA4)>0,0,MIN(AR4:BA4))'
            sheet4.range('BF3').value = "RWK Coverage wks"
            

            sheet4.range(f'AC4:AC{range_formula+2}').formula ="=+IFERROR(VLOOKUP($B4,'PIVOT OOR'!$A:$X,2,0),0)"
            sheet4.range(f'AD4:AD{range_formula+2}').formula = "=SUMIF(OOR!G:G,'FUTURE SHORTAGES REPORT'!B:B,OOR!AB:AB)"
            sheet4.range(f'AE4:AE{range_formula+2}').formula = "=+IFERROR(VLOOKUP($B4,'PIVOT OOR'!$A:$X,3,0),0)"
            sheet4.range(f'AF4:AF{range_formula+2}').formula = "=+IFERROR(VLOOKUP($B4,'PIVOT OOR'!$A:$X,4,0),0)"
            
            
            sheet4.range(f'AG4:AG{range_formula+2}').formula = "=+IFERROR(VLOOKUP($B4,'PIVOT OOR'!$A:$X,5,0),0)"
            sheet4.range(f'AH4:AH{range_formula+2}').formula = "=+IFERROR(VLOOKUP($B4,'PIVOT OOR'!$A:$X,6,0),0)"
            sheet4.range(f'AI4:AI{range_formula+2}').formula = "=+IFERROR(VLOOKUP($B4,'PIVOT OOR'!$A:$X,7,0),0)"
            sheet4.range(f'AJ4:AJ{range_formula+2}').formula = "=+IFERROR(VLOOKUP($B4,'PIVOT OOR'!$A:$X,8,0),0)"
            sheet4.range(f'AK4:AK{range_formula+2}').formula = "=+IFERROR(VLOOKUP($B4,'PIVOT OOR'!$A:$X,9,0),0)"
            sheet4.range(f'AL4:AL{range_formula+2}').formula = "=+IFERROR(VLOOKUP($B4,'PIVOT OOR'!$A:$X,10,0),0)"
            sheet4.range(f'AM4:AM{range_formula+2}').formula = "=+IFERROR(VLOOKUP($B4,'PIVOT OOR'!$A:$X,11,0),0)"
            sheet4.range(f'AN4:AN{range_formula+2}').formula = "=+IFERROR(VLOOKUP($B4,'PIVOT OOR'!$A:$X,12,0),0)"
            sheet4.range(f'AO4:AO{range_formula+2}').formula = "=+IFERROR(VLOOKUP($B4,'PIVOT OOR'!$A:$X,13,0),0)"
            sheet4.range(f'AP4:AP{range_formula+2}').formula = "=+IFERROR(VLOOKUP($B4,'PIVOT OOR'!$A:$X,14,0),0)"
            sheet4.range(f'AQ4:AQ{range_formula+2}').formula = "=+IFERROR(VLOOKUP($B4,'PIVOT OOR'!$A:$X,15,0),0)"
            
            sheet4.range(f'BJ4:BJ{range_formula+2}').formula = '=IFERROR(VLOOKUP($B4,leanDna!$C:$H,2,0),0)'
            sheet4.range(f'BK4:BN{range_formula+2}').formula = '=IFERROR(VLOOKUP($B4,leanDna!$C:$H,3,0),0)'
            sheet4.range(f'BL4:BL{range_formula+2}').formula = '=IFERROR(VLOOKUP($B4,leanDna!$C:$H,4,0),0)'
            sheet4.range(f'BM4:BM{range_formula+2}').formula = '=IFERROR(VLOOKUP($B4,leanDna!$C:$H,5,0),0)'
            sheet4.range(f'BN4:BN{range_formula+2}').formula = '=IFERROR(VLOOKUP($B4,leanDna!$C:$H,6,0),0)'

            sheet4.range(f'BJ4:BJ{range_formula+2}').formula ="""=IF(M4>0,(M4+N4)-(O4+P4),0)"""
            sheet4.range(f'BK4:BK{range_formula+2}').formula ="""=IF($N4=0,0,(BJ4-Q4))"""
            sheet4.range(f'BL4:BL{range_formula+2}').formula ="""=IF($N4=0,0,(BK4-R4))"""
            sheet4.range(f'BM4:BM{range_formula+2}').formula ="""=IF($N4=0,0,(BL4-S4))"""
            sheet4.range(f'BN4:BN{range_formula+2}').formula ="""=IF($N4=0,0,(BM4-T4))"""
            sheet4.range(f'BO4:BO{range_formula+2}').formula ="""=IF($N4=0,0,(BN4-U4))"""
            sheet4.range(f'BP4:BP{range_formula+2}').formula ="""=IF($N4=0,0,(BO4-V4))"""
            sheet4.range(f'BQ4:BQ{range_formula+2}').formula ="""=IF($N4=0,0,(BP4-W4))"""
            sheet4.range(f'BR4:BR{range_formula+2}').formula ="""=IF($N4=0,0,(BQ4-X4))"""
            sheet4.range(f'BS4:BS{range_formula+2}').formula ="""=IF($N4=0,0,(BR4-Y4))"""
            

            sheet4.range(f'AA4:AA{range_formula+2}').formula ="""=IF(M5-SUM(AB5,AE5+AG5)<0,0,M5-SUM(AB5,AE5+AG5))"""
            sheet4.range(f'AB4:AB{range_formula+2}').formula ="""=IF(AA4>=SUMIF(OOR!G:G,'FUTURE SHORTAGES REPORT'!B:B,OOR!AG:AG),0,SUMIF(OOR!G:G,'FUTURE SHORTAGES REPORT'!B:B,OOR!AG:AG))"""
            sheet4.range(f'BT4:BT{range_formula+2}').formula ="""=IFERROR(VLOOKUP($B4,leanDna!$C:$H,2,0),0)"""
            sheet4.range(f'BU4:BU{range_formula+2}').formula ="""=IFERROR(VLOOKUP($B4,leanDna!$C:$H,3,0),0)"""
            
            sheet4.range(f'BV4:BV{range_formula+2}').formula ="""=IFERROR(VLOOKUP($B4,leanDna!$C:$H,4,0),0)"""
            sheet4.range(f'BW4:BW{range_formula+2}').formula ="""=IFERROR(VLOOKUP($B4,leanDna!$C:$H,5,0),0)"""
            sheet4.range(f'BX4:BX{range_formula+2}').formula ="""=IFERROR(VLOOKUP($B4,leanDna!$C:$H,6,0),0)"""

            column_to_delete = "BA" 
            column_range = sheet4.range(f"{column_to_delete}:{column_to_delete}")
            column_range.delete()
            sheet4.range(f'BB4:BB{range_formula+2}').formula ="""=IF(AR4<0,$AR$3,IF(AS4<0,$AS$3,IF(AT4<0,$AT$3,IF(AU4<0,$AU$3,IF(AV4<0,$AV$3,IF(AW4<0,$AW$3,IF(AX4<0,$AX$3,IF(AY4<0,$AY$3,IF(AZ4<0,$AZ$3,IF(AZ4<0,$AZ$3,"zz"))))))))))"""
            range_row_buyers = chart_buyer.shape[0]+1
            sheet12.range(f'D5:D{range_row_buyers+3}').formula ="""=COUNTIFS('FUTURE SHORTAGES REPORT'!$J:$J,$C5,'FUTURE SHORTAGES REPORT'!AR:AR,"<0")"""
            sheet12.range(f'E5:E{range_row_buyers+3}').formula = """=COUNTIFS('FUTURE SHORTAGES REPORT'!$J:$J,$C5,'FUTURE SHORTAGES REPORT'!AS:AS,"<0")"""
            sheet12.range(f'F5:F{range_row_buyers+3}').formula = """=COUNTIFS('FUTURE SHORTAGES REPORT'!$J:$J,$C5,'FUTURE SHORTAGES REPORT'!AT:AT,"<0")"""
            sheet12.range(f'G5:G{range_row_buyers+3}').formula = """=COUNTIFS('FUTURE SHORTAGES REPORT'!$J:$J,$C5,'FUTURE SHORTAGES REPORT'!AU:AU,"<0")"""
            sheet12.range(f'H5:H{range_row_buyers+3}').formula = """=COUNTIFS('FUTURE SHORTAGES REPORT'!$J:$J,$C5,'FUTURE SHORTAGES REPORT'!AV:AV,"<0")"""
            sheet12.range(f'I5:I{range_row_buyers+3}').formula = """=COUNTIFS('FUTURE SHORTAGES REPORT'!$J:$J,$C5,'FUTURE SHORTAGES REPORT'!AW:AW,"<0")"""
            sheet12.range(f'J5:J{range_row_buyers+3}').formula = """=COUNTIFS('FUTURE SHORTAGES REPORT'!$J:$J,$C5,'FUTURE SHORTAGES REPORT'!AX:AX,"<0")"""
            sheet12.range(f'K5:K{range_row_buyers+3}').formula = """=COUNTIFS('FUTURE SHORTAGES REPORT'!$J:$J,$C5,'FUTURE SHORTAGES REPORT'!AY:AY,"<0")"""
            sheet12.range(f'L5:L{range_row_buyers+3}').formula = """=COUNTIFS('FUTURE SHORTAGES REPORT'!$J:$J,$C5,'FUTURE SHORTAGES REPORT'!AZ:AZ,"<0")"""
            sheet12.range(f'M5:M{range_row_buyers+3}').formula = """=COUNTIFS('FUTURE SHORTAGES REPORT'!$J:$J,$C5,'FUTURE SHORTAGES REPORT'!BA:BA,"<0")"""
            
            sheet12.range('Q5').formula = """=COUNTIF('FUTURE SHORTAGES REPORT'!AR4:AR10000,"<0")"""
            sheet12.range('R5').formula = """=COUNTIF('FUTURE SHORTAGES REPORT'!AS4:AS10000,"<0")"""
            sheet12.range('S5').formula = """=COUNTIF('FUTURE SHORTAGES REPORT'!AT4:AT10000,"<0")"""
            sheet12.range('T5').formula = """=COUNTIF('FUTURE SHORTAGES REPORT'!AU4:AU10000,"<0")"""
            sheet12.range('U5').formula = """=COUNTIF('FUTURE SHORTAGES REPORT'!AV4:AV10000,"<0")"""
            sheet12.range('V5').formula = """=COUNTIF('FUTURE SHORTAGES REPORT'!AW4:AW10000,"<0")"""
            sheet12.range('W5').formula = """=COUNTIF('FUTURE SHORTAGES REPORT'!AX4:AX10000,"<0")"""
            sheet12.range('X5').formula = """=COUNTIF('FUTURE SHORTAGES REPORT'!AY4:AY10000,"<0")"""
            sheet12.range('Y5').formula = """=COUNTIF('FUTURE SHORTAGES REPORT'!AZ4:AZ10000,"<0")"""
            sheet12.range('Z5').formula = """=COUNTIF('FUTURE SHORTAGES REPORT'!BA4:BA10000,"<0")"""
            sheet12.range('Q6').formula = """=COUNTIF('FUTURE SHORTAGES REPORT'!AR4:AR10000,">=0")"""
            sheet12.range('R6').formula = """=COUNTIF('FUTURE SHORTAGES REPORT'!AS4:AS10000,">=0")"""
            sheet12.range('S6').formula = """=COUNTIF('FUTURE SHORTAGES REPORT'!AT4:AT10000,">=0")"""
            sheet12.range('T6').formula = """=COUNTIF('FUTURE SHORTAGES REPORT'!AU4:AU10000,">=0")"""
            sheet12.range('U6').formula = """=COUNTIF('FUTURE SHORTAGES REPORT'!AV4:AV10000,">=0")"""
            sheet12.range('V6').formula = """=COUNTIF('FUTURE SHORTAGES REPORT'!AW4:AW10000,">=0")"""
            sheet12.range('W6').formula = """=COUNTIF('FUTURE SHORTAGES REPORT'!AX4:AX10000,">=0")"""
            sheet12.range('X6').formula = """=COUNTIF('FUTURE SHORTAGES REPORT'!AY4:AY10000,">=0")"""
            sheet12.range('Y6').formula = """=COUNTIF('FUTURE SHORTAGES REPORT'!AZ4:AZ10000,">=0")"""
            sheet12.range('Z6').formula = """=COUNTIF('FUTURE SHORTAGES REPORT'!BA4:BA10000,">=0")"""
        
            try:
                conn = sqlite3.connect(path_conn)
                query = "SELECT * FROM where_use_table_plantype"
                where_use = pd.read_sql_query(query, conn)
                sheet13 = wb.sheets.add()
                sheet13.name = "where_use_table"
                sheet13.range('A1').value = where_use
                sheet4.range('A3').value = "Area"
                sheet4.range(f'A4:A{range_formula+2}').formula ="""=IFERROR(VLOOKUP(H4,where_use_table!B:C,2,0),"")"""

                conn.close()
            except Exception as error:
                QMessageBox.critical(self, "Error", f'ocurrio un error{error}')
            finally:
                conn.close()


            sheet1.range(f'Y5:Y{base1_qty_rows+3}').formula ="""=IF(R5<$S$1,"01_Past due",IF(AND(R5>$S$1,R5<$Y$1),"02_Should be delivered",IF(AND(R5>=$Y$1,R5<=$AA$1),"03_Should be in Transit","On time")))"""
            sheet1.range(f'Z5:Z{base1_qty_rows+6}').formula ="""=IF(X5="",IF(S5<$S$1,"01_Past due",IF(AND(S5>=$S$1,S5<$Y$1),"02_Should be delivered",IF(AND(S5>=$Y$1,S5<=$AA$1),"03_Should be in Transit",YEAR(S5)&TEXT(WEEKNUM(S5),"00")))),YEAR(X5)&TEXT(WEEKNUM(X5),"00"))"""
            sheet1.range(f'AA5:AA{base1_qty_rows+3}').formula ="""=IF(M5-SUM(AB5,AE5+AG5)<0,0,M5-SUM(AB5,AE5+AG5))"""
            sheet1.range(f'AE5:AE{base1_qty_rows+3}').formula ="=SUMIF(RECIBOS!A:A,OOR!B:B,RECIBOS!H:H)"
            sheet1.range(f'AG5:AG{base1_qty_rows+3}').formula ="=SUMIF(IMPOS!S:S,OOR!B:B,IMPOS!C:C)"
            


            """ESTO SE MODIFICO"""
            sheet1.range(f'AO5:AO{base1_qty_rows+3}').formula ="=+NETWORKDAYS(S5,$S$1)"
            sheet1.range(f'AP5:AP{base1_qty_rows+3}').formula ="=+NETWORKDAYS(P5,$AP$1)"
            sheet1.range(f'AQ5:AQ{base1_qty_rows+3}').formula ='=IF(W5="","Warning","Ok")'
            
            """Friday of previous week"""
            current_date = datetime.now()
            days_difference = (current_date.weekday() - 4) % 7
            if days_difference == 0:
                days_difference = 7
            previous_friday = current_date - timedelta(days=days_difference)
            formatted_previous_friday = previous_friday.strftime('%m/%d/%Y')
            sheet1.range('AP1').value = formatted_previous_friday
            sheet4.range(f'BA4:BA{range_formula+2}').formula = '=IF(SUM(AC4:AO4)=0,"No PO",IF(BD4=0,"",IF(H4="VMIHDW","VMI",IF(AC4>0,"Past due",IF(AE4>0,"OTD",IF(AF4>0,"Due","Expedite"))))))'
            sheet4.range(f'AA4:AA{range_formula+2}').formula = "=SUMIF(RECIBOS!E:E,'FUTURE SHORTAGES REPORT'!B:B,RECIBOS!H:H)"
            
            white_text_color = (240, 242, 245)
            black_value = (20, 20, 20)
            blue_value = (3, 73, 252)
            yellow_value = (245, 241, 10)
            green_value = (7, 171, 51)

            for cell in sheet1.range('A4:W4'):
                cell.color = black_value
                cell.api.Font.Color = xw.utils.rgb_to_int(white_text_color)
            
            for cell in sheet1.range('Y4:AA4'):
                cell.color = black_value
                cell.api.Font.Color = xw.utils.rgb_to_int(white_text_color)
            
            for cell in sheet1.range('AE4'):
                cell.color = black_value
                cell.api.Font.Color = xw.utils.rgb_to_int(white_text_color)
            
            for cell in sheet1.range('AI4:AQ4'):
                cell.color = black_value
                cell.api.Font.Color = xw.utils.rgb_to_int(white_text_color)

            for cell in sheet1.range('X4'):
                cell.color = blue_value
                cell.api.Font.Color = xw.utils.rgb_to_int(white_text_color)
            
            for cell in sheet1.range('AB4:AD4'):
                cell.color = blue_value
                cell.api.Font.Color = xw.utils.rgb_to_int(white_text_color)
            
            for cell in sheet1.range('AF4:AH4'):
                cell.color = blue_value
                cell.api.Font.Color = xw.utils.rgb_to_int(white_text_color)
            
            for cell in sheet9.range('A1:H1'):
                cell.color = green_value
                cell.api.Font.Color = xw.utils.rgb_to_int(white_text_color)
                
            for cell in sheet9.range('J1:AA1'):
                cell.color = green_value
                cell.api.Font.Color = xw.utils.rgb_to_int(white_text_color)
                    
            for cell in sheet9.range('AB1:AF1'):
                cell.color = blue_value
                cell.api.Font.Color = xw.utils.rgb_to_int(white_text_color)
            
            for cell in sheet9.range('I1'):
                cell.color = yellow_value
                cell.api.Font.Color = xw.utils.rgb_to_int(black_value)
            
            sheet9.range(f'A2:A1000').formula ="=L2&M2"
            
            for cell in sheet10.range('A1:O1'):
                cell.color = green_value
                cell.api.Font.Color = xw.utils.rgb_to_int(white_text_color)
            
            for cell in sheet10.range('S1'):
                cell.color = green_value
                cell.api.Font.Color = xw.utils.rgb_to_int(white_text_color)
            
            for cell in sheet10.range('P1:R1'):
                cell.color = yellow_value
                cell.api.Font.Color = xw.utils.rgb_to_int(black_value)
            
            for cell in sheet10.range('T1:U1'):
                cell.color = yellow_value
                cell.api.Font.Color = xw.utils.rgb_to_int(black_value)

            sheet1.range(f'AO5:AO{base1_qty_rows+3}').formula ="=+NETWORKDAYS(S5,$S$1)"
            sheet1.range(f'B5:B{base1_qty_rows+3}').formula = "=C5&D5"

            
            self.label_37.setHidden(False)
            self.label_35.setHidden(False)
            now = datetime.now()
            today = now.strftime("%m/%d/%Y %I:%M:%S")
            self.label_35.setText(f'{today}')
            today = now.strftime("%m_%d_%Y")
            save_path_future_shortages = "J:\\Departments\\Material Control\\Purchasing\\Tools\\ComprasDB\\Future Shortages Historico"

            wb.save(save_path_future_shortages + f"\\FUTURE SHORTAGES_V2_test {today}.xlsm")  # Replace 'file_name.xlsx' with your desired file name
            wb.close()
            
            QMessageBox.information(self, "Future Shortages ", "Valida el archivo se abrio un excel")

        else:
            QMessageBox.critical(self, "confirmar actualizacion", f'Es necesario confirmar la actualizacion ')

    def run_saefty_stock(self):
        try:
            conn = sqlite3.connect(path_conn)
            query_actual_safety = "SELECT * FROM safety_stock"
            actual_safety = pd.read_sql_query(query_actual_safety, conn)
            query_should_be_safety = "SELECT * FROM safety_stock_should_be"
            safety_should_be = pd.read_sql_query(query_should_be_safety, conn)
            conn.close()
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion Expedite{error}')
        finally:
            conn.close()
            
        base = pd.merge(safety_should_be,actual_safety,how = 'left',on = 'ItemNo')
        base = base.fillna('')
        def status_safety_stock(DataFrame):
            base['Status'] = ''
            for index,row in DataFrame.iterrows():
                if row['Qty'] != '':
                    DataFrame.at[index,'Status'] = 'F'
                else:
                    DataFrame.at[index,'Status'] = 'A'
        status_safety_stock(base)
        

        wb = xw.Book()
        sheet1 = wb.sheets.add()
        sheet1.name = "OOR"
        sheet1.range('A1').value = base
        sheet1.range('A1').value = 'Index'
        sheet1.range('K1').value = 'Start'
        base['ItemNo'] = base['ItemNo'].astype(str)
        base['Status'] = base['Status'].astype(str) 
        
        def place_arrange(DataFrame):
            i = 1
            j = 1
            for index,row in DataFrame.iterrows():  
                if row['Status'] == 'A':
                    sheet1.range(f'K{j+1}').value = row['Status'] + row['ItemNo']
                    sheet1.range(f'K{j+2}').value = row['New_Qty']
                    sheet1.range(f'K{j+3}').value = ''
                    sheet1.range(f'K{j+4}').value = ''
                    sheet1.range(f'K{j+5}').value = ''
                    sheet1.range(f'K{j+6}').value = ''
                    j = j + 6
                else:
                    if row['Status'] == 'F':
                        sheet1.range(f'K{j+1}').value = row['Status'] + row['ItemNo']
                        sheet1.range(f'K{j+2}').value = ''
                        sheet1.range(f'K{j+3}').value = row['New_Qty']
                        sheet1.range(f'K{j+4}').value = ''
                        sheet1.range(f'K{j+5}').value = ''
                        sheet1.range(f'K{j+6}').value = ''
                        sheet1.range(f'K{j+7}').value = ''
                        sheet1.range(f'K{j+8}').value = ''
                        sheet1.range(f'K{j+9}').value = ''
                        j = j + 7
            i = i + 1
        place_arrange(base)

    def shelf_life_material_flag(self):
        expedite = self.get_expedite_information_for_gemba_shortages()
        shelf_table,shelf_table_vendor = self.get_shelf_life_information_table()
        openOrder = self.get_open_order_information()
        in521 = self.get_in521_information_table()
        
        
        
        
        print('all information is load')
        df = pd.merge(expedite,shelf_table,how = 'left',on = 'ItemNo')
        df = df.loc[df['SHELFLIFE'] == 'YES']



        df['ReqDate'] = pd.to_datetime(df['ReqDate'])
        in521['ExpDate'] = pd.to_datetime(in521['ExpDate'])
        today = datetime.today()
        last_day_current_month = today.replace(day=1, month=today.month+1) - timedelta(days=1)
        df_current_month = df[df['ReqDate'] <= last_day_current_month]

        next_month = today.replace(day=28) + timedelta(days=4)  # to safely add 4 days for all months
        next_month = next_month.replace(day=1, month=next_month.month+1)
        last_day_next_month = next_month - timedelta(days=1)

        df_next_month = df[(df['ReqDate'] > last_day_current_month) & (df['ReqDate'] <= last_day_next_month)]
        last_day_following_month = last_day_next_month.replace(day=1, month=last_day_next_month.month+2) - timedelta(days=1)
        df_following_month = df[(df['ReqDate'] > last_day_next_month) & (df['ReqDate'] <= last_day_following_month)]

        
        df_current_month['ReqQty'] = df_current_month['ReqQty'].astype(float)
        df_next_month['ReqQty'] = df_next_month['ReqQty'].astype(float)
        df_following_month['ReqQty'] = df_following_month['ReqQty'].astype(float)
        
        
        current_demand = pd.pivot_table(df_current_month, index=['ItemNo'], values='ReqQty', aggfunc='sum')#, fill_value=0
        next_month_demand = pd.pivot_table(df_next_month, index=['ItemNo'], values='ReqQty', aggfunc='sum')
        following_month_demand  = pd.pivot_table(df_following_month, index=['ItemNo'], values='ReqQty', aggfunc='sum')

        base = pd.merge(in521,current_demand,how = 'left', on = 'ItemNo')
        base = base.rename(columns = {"ReqQty":"Pdue + Current" })
        base1 = pd.merge(base,next_month_demand,how = 'left', on = 'ItemNo')
        base1 = base1.rename(columns = {"ReqQty":"Next_month" })
        base2 = pd.merge(base1,following_month_demand,how = 'left', on = 'ItemNo')
        base2 = base2.rename(columns = {"ReqQty":"Third_month"})

        base2['Pdue + Current'] = base2['Pdue + Current'].fillna(0)
        base2['Next_month'] = base2['Next_month'].fillna(0)
        base2['Third_month'] = base2['Third_month'].fillna(0)

        in521['Shelf_Life'] = in521['Shelf_Life'].str.replace('%', '').astype(float)
        in521['Shelf_Life'] = pd.to_numeric(in521['Shelf_Life'], errors='coerce')

        in521_obsolet = in521.loc[in521['Shelf_Life']== 0]
        in521_exp_this_month = in521[(in521['Shelf_Life']!=0) & (in521['ExpDate'] <= last_day_current_month)]
        in521_exp_next_month_and_more = in521[(in521['ExpDate'] >= last_day_current_month)]


        in521_obsolet['QtyOH'] = in521_obsolet['QtyOH'].astype(float)
        piv_obs = pd.pivot_table(in521_obsolet, index=['ItemNo'], values='QtyOH', aggfunc='sum')
        in521_exp_this_month['QtyOH'] = in521_exp_this_month['QtyOH'].astype(float)
        piv_this_month = pd.pivot_table(in521_exp_this_month, index=['ItemNo'], values='QtyOH', aggfunc='sum')
        in521_exp_next_month_and_more['QtyOH'] = in521_exp_next_month_and_more['QtyOH'].astype(float)
        piv_next_month = pd.pivot_table(in521_exp_next_month_and_more, index=['ItemNo'], values='QtyOH', aggfunc='sum')

        # Convert 'ExpDate' to datetime format
        base2['ExpDate'] = pd.to_datetime(base2['ExpDate'])

        result_unique = base2.groupby(['ItemNo', 'Description']).agg(ExpDate_1st=('ExpDate', 'min'), ExpDate_last=('ExpDate', 'max')).reset_index()
        result_unique_obs = pd.merge(result_unique,piv_obs,how = 'left',on = 'ItemNo')
        result_unique_obs = result_unique_obs.rename(columns = {'QtyOH':'Qty_obs'})
        result_unique_obs['Qty_obs'] = result_unique_obs['Qty_obs'].fillna(0)

        result_unique_this_month = pd.merge(result_unique_obs,piv_this_month,how = 'left',on = 'ItemNo')
        result_unique_this_month = result_unique_this_month.rename(columns = {'QtyOH':'Qty_this_month'})
        result_unique_this_month['Qty_this_month'] = result_unique_this_month['Qty_this_month'].fillna(0)

        result_unique_third_month = pd.merge(result_unique_this_month,piv_next_month,how = 'left',on = 'ItemNo')
        result_unique_third_month = result_unique_third_month.rename(columns = {'QtyOH':'Qty_third_month_+'})
        result_unique_third_month['Qty_third_month_+'] = result_unique_third_month['Qty_third_month_+'].fillna(0)

        base2['Shelf_Life'] = base2['Shelf_Life'].str.replace('%', '').astype(float)
        base2['Shelf_Life'] = pd.to_numeric(base2['Shelf_Life'], errors='coerce')


        base2['ExpDate'] = pd.to_datetime(base2['ExpDate'])
        def columns_lots (DataFrame):
            DataFrame['columns_lots'] = ''
            for index, row in DataFrame.iterrows():
                if row['Shelf_Life'] == 0:
                    DataFrame.at[index,'columns_lots']= 'Obsoleto'
                else:
                    if row['ExpDate'] >= last_day_current_month:
                        DataFrame.at[index,'columns_lots']= 'Third_month_lots'
                    else:
                        if row['ExpDate'] <= last_day_current_month:
                            DataFrame.at[index,'columns_lots']= 'This_month_lots'
        columns_lots(base2)

        def concatenate_lots(lots):
            return ', '.join(lots)

        base_demand = base2[['ItemNo','Pdue + Current','Next_month','Third_month']]
        base_demand = base_demand.drop_duplicates()

        piv_base_lots = pd.pivot_table(base2, index=['ItemNo'], values='Lot', columns='columns_lots', aggfunc=concatenate_lots, fill_value='')
        piv_base_lots = piv_base_lots.fillna('')

        piv_base_lots['Obsoleto'] = piv_base_lots['Obsoleto'].replace('','.')
        piv_base_lots['Third_month_lots'] = piv_base_lots['Third_month_lots'].replace('','.')
        piv_base_lots['This_month_lots'] = piv_base_lots['This_month_lots'].replace('','.')
        piv_base_lots = piv_base_lots.reset_index()

        piv_base_lots['.'] = '.'
        piv_base_lots = piv_base_lots[['ItemNo','Obsoleto','This_month_lots','Third_month_lots','.']]

        base_demand_uniqe = pd.merge(result_unique_third_month,base_demand,how = 'left',on = 'ItemNo')
        base_demand_uniqe_lots = pd.merge(base_demand_uniqe,piv_base_lots,how = 'left',on = 'ItemNo')
        base_demand_uniqe_lots['Qty_this_month'] = base_demand_uniqe_lots['Qty_this_month'].astype(float)
        base_demand_uniqe_lots['Qty_third_month_+'] = base_demand_uniqe_lots['Qty_third_month_+'].astype(float)
        base_demand_uniqe_lots['Pdue + Current'] = base_demand_uniqe_lots['Pdue + Current'].astype(float)
        base_demand_uniqe_lots['Next_month'] = base_demand_uniqe_lots['Next_month'].astype(float)
        base_demand_uniqe_lots['Third_month'] = base_demand_uniqe_lots['Third_month'].astype(float)

        def flag(DataFrame):
            DataFrame['Flag'] = ''
            DataFrame['Comment'] = ''
            for index,row in DataFrame.iterrows():
                total_available = row['Qty_this_month'] + row['Qty_third_month_+'] 
                total_demand = row['Pdue + Current'] + row['Next_month'] + row['Third_month']
                two_months_demand = row['Pdue + Current'] + row['Next_month']
                one_month = row['Pdue + Current']
                
                if row['Pdue + Current'] == 0 and row['Next_month'] == 0 and row['Third_month'] == 0:
                    DataFrame.at[index,'Flag'] ='No demand'
                else:
                    if total_available>=total_demand:
                        DataFrame.at[index,'Flag'] =''
                    else:
                        if total_available>= two_months_demand:
                            DataFrame.at[index,'Flag'] ='Material Two Months'
                        else:
                            if total_available>= one_month:
                                DataFrame.at[index,'Flag'] ='Material One Month'
                            else:
                                value =  round(one_month - total_available,2)
                                DataFrame.at[index,'Flag'] = 'Check Supply'
                                DataFrame.at[index,'Comment'] = f'Need Qty for this month: {value}'

        flag(base_demand_uniqe_lots)


        def flag_2 (DataFrame):
            for index,row in DataFrame.iterrows():
                total_available = row['Qty_this_month'] + row['Qty_third_month_+'] 
                one_month = row['Pdue + Current']
                value =  round(one_month - total_available,2)
                
                if row['Flag'] == 'Check Supply' and row['Qty_obs'] != 0:
                    if value <= row['Qty_obs']:
                        DataFrame.at[index,'Flag'] = 'Check Supply-Recertificar OBS'
                    else:
                        pass
        flag_2(base_demand_uniqe_lots)


        openOrder['PONo'] = openOrder['PONo'].astype(str)
        openOrder['Ln'] = openOrder['Ln'].astype(str)
        openOrder['OpnQ'] = round(openOrder['OpnQ'])
        openOrder['OpnQ'] = openOrder['OpnQ'].astype(str)
        openOrder['Concat'] = openOrder['PONo'] + openOrder['Ln']+"_"+'qty:'+openOrder['OpnQ']
        openOrder = openOrder[['Concat','ItemNo','PONo','Ln','OpnQ','RevPrDt']]
        openOrder['RevPrDt'] = pd.to_datetime(openOrder['RevPrDt'])

        openOrder_unique_dates = openOrder.groupby(['ItemNo']).agg(first_arrival=('RevPrDt', 'min'), last_arrival=('RevPrDt', 'max')).reset_index()


        next_month = today.replace(day=28) + timedelta(days=4)  # to safely add 4 days for all months
        next_month = next_month.replace(day=1, month=next_month.month+1)
        last_day_next_month = next_month - timedelta(days=1)


        def calsification_po_rpd(DataFrame):
            DataFrame['Clasification_rpd'] = ''
            for index,row in DataFrame.iterrows():
                if row['RevPrDt']<= last_day_current_month:
                    DataFrame.at[index,'Clasification_rpd'] = 'Pd + Current_month'
                else:
                    if row['RevPrDt']>last_day_current_month and row['RevPrDt']<=last_day_next_month:
                        DataFrame.at[index,'Clasification_rpd'] = 'Next Month'
                    else:
                        if row['RevPrDt']>last_day_next_month:
                            DataFrame.at[index,'Clasification_rpd'] = 'Third_month_+'
        calsification_po_rpd(openOrder)


        def concatenate_lots(lots):
            return ', '.join(lots)



        piv_Clasification_rpd = pd.pivot_table(openOrder, index=['ItemNo'], values='Concat', columns='Clasification_rpd', aggfunc=concatenate_lots, fill_value='')
        piv_Clasification_rpd = piv_Clasification_rpd.fillna('')

        piv_Clasification_rpd['Next Month'] = piv_Clasification_rpd['Next Month'].replace('','.')
        piv_Clasification_rpd['Pd + Current_month'] = piv_Clasification_rpd['Pd + Current_month'].replace('','.')
        piv_Clasification_rpd['Third_month_+'] = piv_Clasification_rpd['Third_month_+'].replace('','.')
        piv_Clasification_rpd['.'] = '.'
        piv_Clasification_rpd = piv_Clasification_rpd.reset_index()
        piv_Clasification_rpd = piv_Clasification_rpd[['ItemNo','Pd + Current_month','Next Month','Third_month_+','.']]
        piv_Clasification_rpd_dates = pd.merge(openOrder_unique_dates,piv_Clasification_rpd,how = 'left',on='ItemNo')

        final_base = pd.merge(base_demand_uniqe_lots,piv_Clasification_rpd_dates,how='left',on='ItemNo')
        final_base1 = pd.merge(final_base,shelf_table_vendor,how = 'left',on = 'ItemNo')
        wb = xw.Book()
        sheet1 = wb.sheets['Sheet1']
        sheet1.range('A1').value = final_base1

    def pr_only_material_shelf_life(self):

        current_date = datetime.today().date()
        current_datetime = datetime.today()
        iso_calendar_year, iso_week_num, _ = current_datetime.isocalendar()
        iso_week_str = str(iso_week_num).zfill(2)
        iso_week_string = f"{iso_calendar_year}-W{iso_week_str}"

        df = self.get_expedite_information()
        shelf_table = self.shelf_table()
        openOrder = self.get_open_order_information()
        in521 = self.get_in521_information_table()


        df['ReqDate'] = pd.to_datetime(df['ReqDate'])
        days_to_subtract = current_date.weekday()
        monday_of_current_week = current_date - timedelta(days=days_to_subtract)

        std_cost_table = shelf_table.copy()
        std_cost_table = std_cost_table[['ItemNo','STD_Cost']]

        def eliminate_past_due(df, monday_of_current_week):
            for index, row in df.iterrows():
                req_date = pd.to_datetime(row['ReqDate'])  # Convert 'ReqDate' to Timestamp
                if req_date < pd.Timestamp(monday_of_current_week):
                    df.at[index, 'ReqDate'] = pd.Timestamp(monday_of_current_week)
            return df

        df = eliminate_past_due(df, monday_of_current_week)
        df['Week_Str'] = df['ReqDate'].dt.strftime('%Y-%V')
        df['ReqQty'] = df['ReqQty'].astype(float)
        df2 = df.copy()
        df3 = pd.merge(df2,shelf_table,how = 'left',on = 'ItemNo')

        df3 = df3[['ItemNo', 'Description','ReqQty','Week_Str','Vendor_y','LotSize_y', 'LT_y']]
        df3 = df3.rename(columns={'Vendor_y':'Vendor','LotSize_y':'LotSize','LT_y':'LT'})
        df3 = df3.groupby(['ItemNo','Description', 'Week_Str','Vendor','LotSize','LT'], as_index=False)['ReqQty'].sum()
        df4 = df3.copy()
        
        df4 = df4[['ItemNo','Description','Vendor','LotSize','LT']]
        df4 = df4.drop_duplicates()

        pivot_table_demand = pd.pivot_table(df, values='ReqQty', index='ItemNo', columns='Week_Str', aggfunc='sum')
        pivot_table_demand = pivot_table_demand.reset_index()

        pivot_table_demand_shelf_yes = pd.merge(pivot_table_demand,shelf_table,how = 'left',on = 'ItemNo')
        pivot_table_demand_shelf_yes['SHELFLIFE'] = pivot_table_demand_shelf_yes['SHELFLIFE'].str.upper()
        pivot_table_demand_shelf_yes = pivot_table_demand_shelf_yes.loc[pivot_table_demand_shelf_yes['SHELFLIFE'] =='YES']

        columns_to_drop = ['id', 'MLIKCode', 'ACTIVE','SHELFLIFE']
        pivot_table_demand_shelf_yes = pivot_table_demand_shelf_yes.drop(columns_to_drop, axis=1)
        openOrder['RevPrDt'] = pd.to_datetime(openOrder['RevPrDt'])
        openOrder['PONo'] = openOrder['PONo'].astype(str)
        openOrder['Ln'] = openOrder['Ln'].astype(str)
        openOrder['OpnQ'] = openOrder['OpnQ'].astype(float)

        def eliminate_past_due(openOrder, monday_of_current_week):
            for index, row in openOrder.iterrows():
                req_date = pd.to_datetime(row['RevPrDt'])  # Convert 'ReqDate' to Timestamp
                if req_date < pd.Timestamp(monday_of_current_week):
                    openOrder.at[index, 'RevPrDt'] = pd.Timestamp(monday_of_current_week)
            return openOrder

        openOrder = eliminate_past_due(openOrder, monday_of_current_week)
        openOrder['Week_Str'] = openOrder['RevPrDt'].dt.strftime('%Y-%V')
        openOrder1 = openOrder.copy()
        openOrder1 = openOrder1[['ItemNo','Week_Str','OpnQ']]
        
        openOrder1 = pd.pivot_table(openOrder1,index = ['ItemNo','Week_Str'],values = 'OpnQ',aggfunc='sum')
        openOrder1 = openOrder1.reset_index()
    
        pivot_table_openOrder = pd.pivot_table(openOrder, values='OpnQ', index='ItemNo', columns='Week_Str', aggfunc='sum')
        pivot_table_openOrder = pivot_table_openOrder.reset_index()
        
        current_week = current_date.strftime('%Y-%V')
        coverage_data = {'ItemNo': [], 'Week': [], 'QtyOH': []}
        in521['QtyOH'] = in521['QtyOH'].astype(float)

        current_date -= timedelta(days=current_date.weekday())
        future_date = current_date + timedelta(days=3*365)
        coverage_data = {'ItemNo': [], 'Week': [], 'QtyOH': []}

        for index, row in in521.iterrows():
            exp_date = datetime.strptime(row['ExpDate'], '%m/%d/%Y').date()
            if current_date <= exp_date:
                weeks_until_exp = (exp_date - current_date).days // 7
                for i in range(weeks_until_exp + 1):
                    week_start = (current_date + timedelta(weeks=i)).strftime('%Y-%V')
                    coverage_data['ItemNo'].append(row['ItemNo'])
                    coverage_data['Week'].append(week_start)
                    coverage_data['QtyOH'].append(row['QtyOH'])

        coverage_df = pd.DataFrame(coverage_data)
        inv_coverage = coverage_df.groupby(['ItemNo', 'Week'])['QtyOH'].sum()
        inv_coverage = inv_coverage.reset_index()
        current_date -= timedelta(days=current_date.weekday())
        future_date = current_date + timedelta(days=1*365)
        week_numbers = []

        while current_date <= future_date:
            iso_week_date = current_date.isocalendar()
            week_number = f"{iso_week_date[0]}-W{iso_week_date[1]:02d}"
            week_numbers.append(week_number)
            current_date += timedelta(weeks=1)

        week_df = pd.DataFrame(week_numbers, columns=['Week_Str'])       
        base = pd.merge(df4,week_df, how='cross')
        base['key'] = base['ItemNo'] +"_"+ base['Week_Str']
        df3['key'] = df3['ItemNo'] +"_"+ df3['Week_Str']
        openOrder1['key'] = openOrder1['ItemNo'] +"_"+ openOrder1['Week_Str']
        inv_coverage['key'] = inv_coverage['ItemNo'] +"_"+ inv_coverage['Week']

        df3 = df3[['key','ReqQty']]
        df3 = df3.rename(columns = {'ReqQty':'Demand'})

        openOrder1 = openOrder1[['key','OpnQ']]
        openOrder1 = openOrder1.rename(columns = {'OpnQ':'PO_Qty'})

        inv_coverage = inv_coverage[['key','QtyOH']]
        inv_coverage = inv_coverage.rename(columns = {'QtyOH':'OH_Qty'})
        
        base = base.drop_duplicates()
        openOrder1 = openOrder1.drop_duplicates()
        inv_coverage = inv_coverage.drop_duplicates()
        
        
        final_base = pd.merge(base,df3,how = 'left',on = 'key')
        final_base = pd.merge(final_base,openOrder1,how = 'left',on = 'key')
        final_base = pd.merge(final_base,inv_coverage,how = 'left',on = 'key')
        
        final_base['Demand'] = final_base['Demand'].fillna('0')
        final_base['PO_Qty'] = final_base['PO_Qty'].fillna('0')
        final_base['OH_Qty'] = final_base['OH_Qty'].fillna('0')
        

        final_base['Logic_item'] = 'y'
        duplicates_mask = final_base['ItemNo'].duplicated(keep='first') #aqui modifique 
        final_base.loc[duplicates_mask, 'Logic_item'] = 'n'
        
        final_base['Logic_oh'] = 'y'
        duplicates_mask = final_base['OH_Qty'].duplicated(keep='first') #aqui modifique 
        final_base.loc[duplicates_mask, 'Logic_oh'] = 'n'
        final_base = final_base[['Logic_item', 'Logic_oh','key','Week_Str','ItemNo', 'Description'
                                , 'Vendor', 'LotSize', 'LT','OH_Qty', 'Demand', 'PO_Qty']]
        
        def adjust_current_week(DataFrame):
            for index,row in DataFrame.iterrows():
                if row['Week_Str'] == iso_week_string:
                    DataFrame.at[index,'Logic_oh'] = 'y'
                else:
                    pass
        adjust_current_week(final_base)

        def calculate_the_real_balance(df):
            df['Balance'] = ''
            current_balance = ''
            for index, row in df.iterrows():
                req_qty = float(row['Demand'])
                po_qty = float(row['PO_Qty'])
                oh_available = float(row['OH_Qty'])
                if row['Logic_item'] == 'y' and row['Logic_oh'] == 'y':
                    balance_qty = oh_available + po_qty - req_qty
                    df.at[index, 'Balance'] = balance_qty
                    current_balance = balance_qty
                else:
                    if row['Logic_item'] == 'n' and row['Logic_oh'] == 'n':
                        balance_qty = current_balance + po_qty - req_qty
                        current_balance = balance_qty
                        df.at[index, 'Balance'] = current_balance
                    else:
                        if row['Logic_item'] == 'n' and row['Logic_oh'] == 'y':
                            if current_balance < 0:
                                balance_qty = current_balance + po_qty - req_qty
                                current_balance = balance_qty
                                df.at[index, 'Balance'] = current_balance
                            else:
                                balance_qty = oh_available  + po_qty - req_qty
                                df.at[index, 'Balance'] = balance_qty
                                current_balance = balance_qty  # Update current balance
                        else:
                            pass
            return df

        calculate_the_real_balance(final_base)
        
        def mark_first_negative_balance(df):
            first_impact = []
            prev_item = None
            for index, row in df.iterrows():
                if prev_item != row['ItemNo']:
                    first_less_than_0 = False
                    prev_item = row['ItemNo']
                if not first_less_than_0 and row['Balance'] < 0:
                    first_impact.append('x')
                    first_less_than_0 = True
                else:
                    first_impact.append('')

            df['1st_Impact'] = first_impact
            return df


        final_base = mark_first_negative_balance(final_base)

        first_impact_table = final_base.copy()
        first_impact_table = first_impact_table[['key','Week_Str','Balance','LT','1st_Impact']]
        first_impact_table = first_impact_table.loc[first_impact_table['1st_Impact']=='x']
        first_impact_table['LT'] = first_impact_table['LT'].astype(float)
        
        def working_days_to_weeks(working_days):
            # Assuming a standard workweek of 5 days
            weeks = working_days / 5
            return weeks
        first_impact_table['#_weeks'] = first_impact_table['LT'].apply(working_days_to_weeks)

        # Function to convert 'Year-Week' to a date
        def week_str_to_date(week_str):
            year, week = map(int, week_str.split('-W'))
            first_day_of_year = datetime(year, 1, 1)
            days_to_first_week = (7 - first_day_of_year.weekday()) % 7
            first_week_start = first_day_of_year + timedelta(days=days_to_first_week)
            return first_week_start + timedelta(weeks=week-1)

        # Function to convert a date back to 'Year-Week'
        def date_to_week_str(date):
            iso_week_date = date.isocalendar()
            return f"{iso_week_date[0]}-W{iso_week_date[1]:02d}"

        # Convert 'Week_Str' to datetime and subtract '#_weeks'
        first_impact_table['Week_Start_Date'] = first_impact_table['Week_Str'].apply(week_str_to_date)
        first_impact_table['YW PR'] = first_impact_table.apply(lambda row: row['Week_Start_Date'] - timedelta(weeks=row['#_weeks']), axis=1)

        # Convert adjusted date back to 'Year-Week'
        first_impact_table['YW PR'] = first_impact_table['YW PR'].apply(date_to_week_str)

        # Drop intermediate columns if not needed
        first_impact_table.drop(columns=['Week_Start_Date', 'YW PR'], inplace=True)
        line_of_balance = pd.pivot_table(final_base,index=['ItemNo','Description','Vendor','LotSize','LT'],columns=['Week_Str'],values = 'Balance',aggfunc='sum')

        final_base['Demand'] = final_base['Demand'].astype(float)
        final_base['PO_Qty'] = final_base['PO_Qty'].astype(float)
        final_base['OH_Qty'] = final_base['OH_Qty'].astype(float)


        main_base = final_base.copy()
        main_base = main_base[['ItemNo','LotSize','LT','Vendor']]
        main_base = main_base.drop_duplicates()


        po_pivot = pd.pivot_table(final_base,index=['ItemNo'],columns=['Week_Str'],values = 'PO_Qty',aggfunc='sum')
        po_pivot = po_pivot.reset_index()

        demand_pivot = pd.pivot_table(final_base,index=['ItemNo'],columns=['Week_Str'],values = 'Demand',aggfunc='sum')
        demand_pivot = demand_pivot.reset_index()

        oh = pd.pivot_table(final_base,index=['ItemNo'],columns=['Week_Str'],values = 'OH_Qty',aggfunc='sum')
        oh = oh.reset_index()
        base_pivot = pd.merge(main_base,std_cost_table,how = 'left',on = 'ItemNo')
        base_pivot = pd.merge(main_base,demand_pivot,how = 'left',on = 'ItemNo')
        base_pivot = pd.merge(base_pivot,po_pivot,how = 'left',on = 'ItemNo')
        base_pivot = pd.merge(base_pivot,oh,how = 'left',on = 'ItemNo')
        

        wb = xw.Book()
        
        sheet8 = wb.sheets.add()
        sheet8.name = "Logic_base"
        sheet8.range('A1').value = final_base
        
        sheet9 = wb.sheets.add()
        sheet9.name = "Line of Balance"
        sheet9.range('A1').value = line_of_balance
        
        sheet10 = wb.sheets.add()
        sheet10.name = "YW 1st Impact"
        sheet10.range('A1').value = first_impact_table
        
        sheet6 = wb.sheets.add()
        sheet6.name = "orig_demand"
        sheet6.range('A1').value = base

        sheet7 = wb.sheets.add()
        sheet7.name = "demand_x_week"
        sheet7.range('A1').value = df3

        sheet4 = wb.sheets.add()
        sheet4.name = "OpenOrder"
        sheet4.range('A1').value = openOrder1

        sheet3 = wb.sheets.add()
        sheet3.name = "IN521_Coverage"
        sheet3.range('A1').value = inv_coverage

        sheet1 = wb.sheets.add()
        sheet1.name = "extended_line_of_balance"
        sheet1.range('A1').value = base_pivot

        # sheet2 = wb.sheets.add()
        # sheet2.name = "OpneOrder_pivot"
        # sheet2.range('A1').value = pivot_table_openOrder

        sheet5 = wb.sheets.add()
        sheet5.name = "table_week"
        sheet5.range('A1').value = week_df
    
    def pr_all_materials(self):
        vendor_europeos = self.get_vendor_europeo()
        current_date = datetime.today().date()
        current_datetime = datetime.today()
        iso_calendar_year, iso_week_num, _ = current_datetime.isocalendar()
        iso_week_str = str(iso_week_num).zfill(2)
        iso_week_string = f"{iso_calendar_year}-{iso_week_str}"

        expedite = self.get_expedite_information()
        in_9_2 = self.shelf_table()
        buyer = self.get_vendor_information2()

        openOrder = self.get_open_order_information()
        oh = self.get_oh_table_information_pr()
        transit_table = self.get_transit_information()
        rework_table = self.get_rework_all_information_table()
        exception_items = self.get_exception_items()
        po351 = self.get_po351_information()
        action = self.get_action_information2()
        
        actions_by_item = action.copy()
        actions_by_item['ACD'] = actions_by_item['ACD'].str.upper()
        actions_by_item = actions_by_item.groupby('ItemNo')['ACD'].apply(lambda x: ','.join(x.unique())).reset_index()
        actions_by_item = actions_by_item.drop_duplicates()
        
        action_notes = action.copy()
        action_notes = action_notes[['ItemNo','PlanningNotes']]
        action_notes = action_notes.fillna('')
        action_notes = action_notes.loc[action_notes['PlanningNotes'] !='']
        action_notes = action_notes.drop_duplicates()
        
        """---------------------parse the itemNo-------------------------------------------------"""
        po351['key'] = po351['ItemNo'] +"_"+ po351['Vendor']
        po351 = po351[['key','ItemNo','Vendor','FobVendor']]
        po351 = po351.rename(columns = {'FobVendor':'Unit Cost'})
        po351 = po351[['key','Unit Cost']]
        
        exception_items['ItemNo'] = exception_items['ItemNo'].str.strip().str.upper()
        expedite['ItemNo'] = expedite['ItemNo'].str.strip().str.upper()
        expedite = pd.merge(expedite,exception_items,how = 'left',on = 'ItemNo')
        expedite = expedite.fillna('')
        expedite_project = expedite.copy()
        expedite_project['Project'] = expedite_project['Project'].str.upper()
        expedite_project = expedite_project.loc[expedite_project['Project']!='']
        expedite_project = expedite_project.groupby('ItemNo')['Project'].apply(lambda x: ','.join(x.unique())).reset_index()
        openOrder['ItemNo'] = openOrder['ItemNo'].str.strip().str.upper()
        openOrder = pd.merge(openOrder,exception_items,how = 'left',on = 'ItemNo')
        openOrder = openOrder.fillna('')
        
        in_9_2['ItemNo'] = in_9_2['ItemNo'].str.strip().str.upper()
        in_9_2 = pd.merge(in_9_2,exception_items,how = 'left',on = 'ItemNo')
        in_9_2 = in_9_2.fillna('')
        in_9_2['MLIKCode'] = in_9_2['MLIKCode'].str.strip().str.upper()
        in_9_2 = in_9_2.loc[in_9_2['MLIKCode'] =='L']
        
        oh['ItemNo'] = oh['ItemNo'].str.strip().str.upper()
        oh = pd.merge(oh,exception_items,how = 'left',on = 'ItemNo')
        oh = oh.fillna('')
        rework_table = pd.merge(rework_table,exception_items,how = 'left',on = 'ItemNo')
        rework_table = rework_table.fillna('')

        def exception_item(df):
            for index, row in df.iterrows():
                if row['Replace'] !='':
                    df.at[index,'ItemNo'] = row['Replace']
                else:
                    pass
        exception_item(rework_table)

        def exception_item(df):
            for index, row in df.iterrows():
                if row['Replace'] !='':
                    df.at[index,'ItemNo'] = row['Replace']
                else:
                    pass
        exception_item(expedite)

        def exception_item(df):
            for index, row in df.iterrows():
                if row['Replace'] !='':
                    df.at[index,'ItemNo'] = row['Replace']
                else:
                    pass
        exception_item(openOrder)

        def exception_item(df):
            for index, row in df.iterrows():
                if row['Replace'] !='':
                    df.at[index,'ItemNo'] = row['Replace']
                else:
                    pass
        exception_item(in_9_2)

        def exception_item(df):
            for index, row in df.iterrows():
                if row['Replace'] !='':
                    df.at[index,'ItemNo'] = row['Replace']
                else:
                    pass
        exception_item(oh)
        
        rwork_total = rework_table.copy()
        rwork_total['OH'] = rwork_total['OH'].astype(float)
        rwork_total = pd.pivot_table(rwork_total,index = 'ItemNo',values='OH',aggfunc='sum')
        rwork_total = rwork_total.reset_index()
        
        """---------------------Clean result expedite-------------------------------------------------"""
        expedite['ReqDate'] = pd.to_datetime(expedite['ReqDate'])
        expedite['ReqQty'] = expedite['ReqQty'].astype(float)
        days_to_subtract = current_date.weekday()
        monday_of_current_week = current_date - timedelta(days=days_to_subtract)

        current_week = current_date.strftime('%Y-%V')
        
        def eliminate_past_due(expedite, monday_of_current_week):
            for index, row in expedite.iterrows():
                req_date = pd.to_datetime(row['ReqDate'])  # Convert 'ReqDate' to Timestamp
                if req_date < pd.Timestamp(monday_of_current_week):
                    expedite.at[index, 'ReqDate'] = pd.Timestamp(monday_of_current_week)
            return expedite
        
        expedite = eliminate_past_due(expedite, monday_of_current_week)
        
        expedite['ReqDate'] = pd.to_datetime(expedite['ReqDate'])
        expedite['Week_Str'] = expedite['ReqDate'].dt.strftime('%Y-%V')
        expedite['ReqQty'] = expedite['ReqQty'].astype(float)
        expedite = expedite.groupby(['ItemNo','Week_Str'], as_index=False)['ReqQty'].sum()
        
        items_in_demand = expedite.copy()
        items_in_demand = items_in_demand[['ItemNo']]
        items_in_demand = items_in_demand.drop_duplicates()

        items_in_demand = pd.merge(items_in_demand,in_9_2,how = 'left',on = 'ItemNo')
        items_in_demand = items_in_demand[['ItemNo','LotSize','LT','STD_Cost','Vendor']]
        buyer = buyer[['Vendor','Shortname','Tactical','Operational']]
        buyer  = buyer.rename(columns = {'Shortname':'Vendor_Name'})
        
        items_in_demand = pd.merge(items_in_demand,buyer,how = 'left',on = 'Vendor')
        items_in_demand['Operational'] = items_in_demand['Operational'].str.replace('@ezairinterior.com', '', regex=False)
        items_in_demand = items_in_demand.fillna('')
        
        items_in_demand['Vendor'] = items_in_demand['Vendor'].replace('','No vendor')
        items_in_demand['Tactical'] = items_in_demand['Tactical'].replace('','No vendor')
        items_in_demand['Operational'] = items_in_demand['Operational'].replace('','No vendor')
        items_in_demand['Vendor_Name'] = items_in_demand['Vendor_Name'].replace('','No vendor')
        
        pivot_table_demand = pd.pivot_table(expedite, values='ReqQty', index='ItemNo', columns='Week_Str', aggfunc='sum')
        pivot_table_demand = pivot_table_demand.reset_index()
        
        """-------------------------------OPEN ORDER-------------------------------------------------"""
        openOrder['RevPrDt'] = pd.to_datetime(openOrder['RevPrDt'])
        openOrder['PONo'] = openOrder['PONo'].astype(str)
        openOrder['Ln'] = openOrder['Ln'].astype(str)
        openOrder['OpnQ'] = openOrder['OpnQ'].astype(float)
        
        def eliminate_past_due(openOrder, monday_of_current_week):
            for index, row in openOrder.iterrows():
                req_date = pd.to_datetime(row['RevPrDt'])  # Convert 'ReqDate' to Timestamp
                if req_date < pd.Timestamp(monday_of_current_week):
                    openOrder.at[index, 'RevPrDt'] = pd.Timestamp(monday_of_current_week)
            return openOrder
        
        openOrder = eliminate_past_due(openOrder, monday_of_current_week)
        openOrder['Week_Str'] = openOrder['RevPrDt'].dt.strftime('%Y-%V')
        openOrder1 = openOrder.copy()
        openOrder1 = openOrder1[['ItemNo','Week_Str','OpnQ']]
        
        openOrder1 = pd.pivot_table(openOrder1,index = ['ItemNo','Week_Str'],values = 'OpnQ',aggfunc='sum')
        openOrder1 = openOrder1.reset_index()
        
        piv_openOrder1_total = openOrder1.copy()
        piv_openOrder1_total['OpnQ'] = piv_openOrder1_total['OpnQ'].astype(float)
        piv_openOrder1_total = piv_openOrder1_total[['ItemNo','OpnQ']]
        
        
        piv_openOrder1_total = pd.pivot_table(piv_openOrder1_total,index = 'ItemNo',values = 'OpnQ',aggfunc='sum')
        piv_openOrder1_total = piv_openOrder1_total.reset_index()
        piv_openOrder1_total = piv_openOrder1_total.rename(columns = {'OpnQ':'PO_Qty'})
        
        piv_oh_total = oh.copy()
        piv_oh_total = piv_oh_total[['ItemNo','OH']]
        piv_oh_total['OH'] = piv_oh_total['OH'].astype(float)
        piv_oh_total = pd.pivot_table(piv_oh_total,index = 'ItemNo',values = 'OH',aggfunc='sum')
        piv_oh_total = piv_oh_total.reset_index()

        pivot_table_openOrder = pd.pivot_table(openOrder, values='OpnQ', index='ItemNo', columns='Week_Str', aggfunc='sum')
        pivot_table_openOrder = pivot_table_openOrder.reset_index()
        
    
        """---------------------------------INVENTORY COVERAGE ---------------------------------------"""

        def get_week_string(date):
            """Converts a datetime object to a string representing the week."""
            return date.strftime("%Y-%W")

        # Your existing code
        oh['Expire_Date'] = pd.to_datetime(oh['Expire_Date'])
        current_date = datetime.today().date()
        future_date = (pd.Timestamp(current_date) + pd.DateOffset(years=3)).normalize()
        future_date = pd.to_datetime(future_date)


        oh['Expire_Date'].fillna(future_date, inplace=True)
        oh['Expire_Date'] = pd.to_datetime(oh['Expire_Date'])
        oh.loc[oh['Expire_Date'] > future_date, 'Expire_Date'] = future_date
        oh['OH'] = oh['OH'].astype(float)

        current_date -= timedelta(days=current_date.weekday())
        coverage_data = {'ItemNo': [], 'Week': [], 'QtyOH': []}

        for index, row in oh.iterrows():
            exp_date = row['Expire_Date'].date()
            if current_date <= exp_date:
                weeks_until_exp = (exp_date - current_date).days // 7
                for i in range(weeks_until_exp + 1):
                    week_start = get_week_string(current_date + timedelta(weeks=i))
                    coverage_data['ItemNo'].append(row['ItemNo'])
                    coverage_data['Week'].append(week_start)
                    coverage_data['QtyOH'].append(row['OH'])
                    

        coverage_df = pd.DataFrame(coverage_data)

        coverage_df['Week_Number'] = coverage_df['Week'].apply(lambda x: x.split('-')[1])
        coverage_df['Week_Number'] = coverage_df['Week_Number'].astype(str)
        coverage_df['QtyOH'] = coverage_df['QtyOH'].astype(float)
        
        def mark_week53(df):
            df['drop'] = ''
            for index,row in df.iterrows():
                if row['Week_Number'] == '53'and row['QtyOH'] == '0':
                    df.at[index,'drop'] = 'x'
                else:
                    pass
        mark_week53(coverage_df)
        coverage_df = coverage_df.loc[coverage_df['drop']!='x']


        coverage_df['ItemNo'] = coverage_df['ItemNo'].astype(str)
        coverage_df['Week'] = coverage_df['Week'].astype(str)
        coverage_df['key'] = coverage_df['ItemNo']  +"_"+ coverage_df['Week']
        coverage_df['QtyOH'] = coverage_df['QtyOH'].astype(float)

        pivot_inv_coverage = pd.pivot_table(coverage_df,index =['key'],values = 'QtyOH',aggfunc = 'sum' )
        pivot_inv_coverage = pivot_inv_coverage.reset_index()
        pivot_inv_coverage = pivot_inv_coverage.rename(columns = {'QtyOH':'OH_Qty'})
        oh_per_week  = pd.pivot_table(coverage_df,index =['ItemNo'],columns = 'Week',values = 'QtyOH',aggfunc = 'sum' )
        oh_per_week = oh_per_week.reset_index()
        
        """------------------------------------------------------------REWORK LOCATION"""
        rework_table['OH'] = rework_table['OH'].astype(float)
        rework_table['Expire_Date'] = pd.to_datetime(rework_table['Expire_Date'])
        current_date -= timedelta(days=current_date.weekday())
        future_date = current_date + timedelta(days=3*365)
        coverage_data = {'ItemNo': [], 'Week': [], 'QtyOH': []}

        for index, row in rework_table.iterrows():
            exp_date = row['Expire_Date'].date()  # Convert to date object
            if current_date <= exp_date:
                weeks_until_exp = (exp_date - current_date).days // 7
                for i in range(weeks_until_exp + 1):
                    week_start = (current_date + timedelta(weeks=i)).strftime('%Y-%V')
                    coverage_data['ItemNo'].append(row['ItemNo'])
                    coverage_data['Week'].append(week_start)
                    coverage_data['QtyOH'].append(row['OH'])

        coverage_rework_table = pd.DataFrame(coverage_data)
        coverage_rework_table['ItemNo'] = coverage_rework_table['ItemNo'].astype(str)
        coverage_rework_table['Week'] = coverage_rework_table['Week'].astype(str)
        coverage_rework_table['key'] = coverage_rework_table['ItemNo']  +"_"+ coverage_rework_table['Week']
        coverage_rework_table['QtyOH'] = coverage_rework_table['QtyOH'].astype(float)

        rework_x_week = coverage_rework_table.copy()
        rework_x_week = pd.pivot_table(rework_x_week,index =['ItemNo'],columns = 'Week',values = 'QtyOH',aggfunc = 'sum' )
        coverage_rework_table_pivot = pd.pivot_table(coverage_rework_table,index =['key'],values = 'QtyOH',aggfunc = 'sum' )
        coverage_rework_table_pivot = coverage_rework_table_pivot.reset_index()
        coverage_rework_table_pivot = coverage_rework_table_pivot.rename(columns = {'QtyOH':'RwkOH'})
        """--------------------------------------------------------------------------------------------"""
        
        current_date = datetime.now()
        def get_week_string(date):
            year = date.year
            week = date.isocalendar()[1]
            if week == 53 and datetime(year, 1, 1).weekday() < 3:
                week = 1
                year += 1
            elif week == 1 and date.month == 12:
                year += 1
            week_str = f"{year}-{str(week).zfill(2)}"
            return week_str

        week_strings = []
        weeks_to_generate = 52 * 3  # 52 weeks/year * 3 years
        for i in range(weeks_to_generate):
            week_date = current_date + timedelta(weeks=i)
            week_string = get_week_string(week_date)
            week_strings.append(week_string)

        week_df = pd.DataFrame(week_strings, columns=['Week_Str'])
        
        week_df['Week_Number'] = week_df['Week_Str'].apply(lambda x: x.split('-')[1])
        week_df['Week_Number'] = week_df['Week_Number'].astype(str)
    
        
        def mark_week53(df):
            df['drop'] = ''
            for index,row in df.iterrows():
                if row['Week_Number'] == '53':
                    df.at[index,'drop'] = 'x'
                else:
                    pass
        mark_week53(week_df)
        week_df = week_df.loc[week_df['drop']!='x']
        week_df = week_df[['Week_Str']]
        base = pd.merge(items_in_demand,week_df, how='cross')
        
        
        """--------------------------------------CREATE KEYS---------------------------------------"""
        base['key'] = base['ItemNo'] +"_"+ base['Week_Str']#BASE
        expedite['key'] = expedite['ItemNo'] +"_"+ expedite['Week_Str'] #DEMAND
        expedite = expedite[['key','ReqQty']]
        expedite = expedite.rename(columns = {'ReqQty':'Demand'})
    
        openOrder1['key'] = openOrder1['ItemNo'] +"_"+ openOrder1['Week_Str'] #OPENORDER
        openOrder1 = openOrder1[['key','OpnQ']]
        openOrder1 = openOrder1.rename(columns = {'OpnQ':'PO_Qty'})

        base = pd.merge(base,pivot_inv_coverage,how = 'left',on ='key')
        base = pd.merge(base,expedite,how = 'left',on ='key')
        base = pd.merge(base,piv_openOrder1_total,how = 'left',on ='ItemNo')
        base = pd.merge(base,rwork_total,how = 'left',on ='ItemNo')
        base = base.rename(columns={'OH':'RwkOH'})
        
        base['Demand'] = base['Demand'].fillna('0')
        base['PO_Qty'] = base['PO_Qty'].fillna('0')
        base['OH_Qty'] = base['OH_Qty'].fillna('0')
        base['RwkOH'] = base['RwkOH'].fillna('0')
        base['Demand'] = base['Demand'].astype(float)
        base['PO_Qty'] = base['PO_Qty'].astype(float)
        base['OH_Qty'] = base['OH_Qty'].astype(float)
        base['RwkOH'] = base['RwkOH'].astype(float)
        
        # Initial setup for 'Logic_item'
        base['Logic_item'] = 'y'
        duplicates_mask_item = base['ItemNo'].duplicated(keep='first')
        base.loc[duplicates_mask_item, 'Logic_item'] = 'n'

        # Logic for 'Logic_oh'
        base['Logic_oh'] = 'y'
        # Create a mask for changes in 'OH_Qty'
        changes_mask_oh = base['OH_Qty'] != base['OH_Qty'].shift()
        base.loc[~changes_mask_oh, 'Logic_oh'] = 'n'

        # Select specific columns
        base = base[['Logic_item', 'Logic_oh', 'key', 'Week_Str', 'ItemNo', 'LotSize',
                    'LT', 'STD_Cost', 'Vendor', 'Tactical', 'Operational',
                    'Vendor_Name', 'OH_Qty', 'RwkOH', 'Demand', 'PO_Qty']]

        def adjust_current_week(DataFrame):
            for index,row in DataFrame.iterrows():
                if row['Week_Str'] == iso_week_string:
                    DataFrame.at[index,'Logic_oh'] = 'y'
                else:
                    pass
        adjust_current_week(base)

        def calculate_qty_to_rest(df):
            oh_qty = None
            qty_to_rest = []
            for index, row in df.iterrows():
                if row['Logic_item'] == 'n' and row['Logic_oh'] == 'y':
                    if oh_qty is not None:
                        qty_to_rest.append(oh_qty - row['OH_Qty'])
                    else:
                        qty_to_rest.append(None)
                else:
                    qty_to_rest.append(None)
                oh_qty = row['OH_Qty']
            df['qty_to_rest'] = qty_to_rest
            return df

        base = calculate_qty_to_rest(base)
        base = base.fillna('')

        # Function to calculate balance_only_oh
        def balance_only_oh(df):
            df['balance_only_oh'] = 0.0  # Initialize with 0.0
            now_current_balance = 0.0  # Initialize outside the loop
            for index, row in df.iterrows():
                oh_shelf = float(row['OH_Qty'])
                demand = float(row['Demand'])
                # qty_to_rest = float(row['qty_to_rest']) if pd.notna(row['qty_to_rest']) else 0.0  # Handle NaN values
                qty_to_rest = float(row['qty_to_rest']) if pd.notna(row['qty_to_rest']) and row['qty_to_rest'] != '' else 0.0
                if row['Logic_item'] == 'y' and row['Logic_oh'] == 'y':
                    now_current_balance = oh_shelf - demand
                    df.at[index, 'balance_only_oh'] = now_current_balance
                elif row['Logic_item'] == 'n' and row['Logic_oh'] == 'n':
                    now_current_balance -= demand
                    df.at[index, 'balance_only_oh'] = now_current_balance
                elif row['Logic_item'] == 'n' and row['Logic_oh'] == 'y':
                    if now_current_balance >= qty_to_rest:
                        now_current_balance -= (qty_to_rest + demand)
                        df.at[index, 'balance_only_oh'] = now_current_balance
                    else:
                        if 0.0 < now_current_balance <= qty_to_rest:
                            now_current_balance = 0.0 - demand
                            df.at[index, 'balance_only_oh'] = now_current_balance
                        else:
                            now_current_balance -= demand
                            df.at[index, 'balance_only_oh'] = now_current_balance
            return df

        # Calculate balance_only_oh
        base = balance_only_oh(base)

        # Function to calculate real_balance_oh_shelf_life
        def coverage_oh_real_with_shelf_life(df):
            df['real_balance_oh_shelf_life'] = ''  # Initialize with 0.0 if NaN
            now_current_balance = 0.0  # Initialize outside the loop
            past_oh = 0.0
            past_balance = 0
            for index, row in df.iterrows():
                oh_shelf = float(row['OH_Qty'])
                demand = float(row['Demand'])
                po_qty = float(row['PO_Qty'])
                rwk_qty = float(row['RwkOH'])
                balance_oh = float(row['balance_only_oh'])
                # qty_to_rest = float(row['qty_to_rest']) if pd.notna(row['qty_to_rest']) else 0.0  # Handle NaN values
                qty_to_rest = float(row['qty_to_rest']) if pd.notna(row['qty_to_rest']) and row['qty_to_rest'] != '' else 0.0
                if row['Logic_item'] == 'y' and row['Logic_oh'] == 'y':
                    
                    now_current_balance = 0.0
                    past_oh = 0.0
                    past_balance = 0
                    prev_oh = float(row['balance_only_oh'])
                    now_current_balance = oh_shelf + po_qty + rwk_qty - demand 
                    df.at[index, 'real_balance_oh_shelf_life'] = now_current_balance
                
                elif row['Logic_item'] == 'n' and row['Logic_oh'] == 'n':
                    now_current_balance -= demand
                    df.at[index, 'real_balance_oh_shelf_life'] = now_current_balance
                    past_oh = balance_oh
                    past_balance = now_current_balance
                    prev_oh = float(row['balance_only_oh'])
                    
                elif row['Logic_item'] == 'n' and row['Logic_oh'] == 'y':
                
                    if now_current_balance < 0.0:
                        now_current_balance = now_current_balance - demand
                        df.at[index, 'real_balance_oh_shelf_life'] = now_current_balance
                        
                    elif prev_oh == past_balance and balance_oh >= qty_to_rest:
                            prev_oh = float(row['balance_only_oh'])
                            now_current_balance = now_current_balance - qty_to_rest - demand
                            df.at[index, 'real_balance_oh_shelf_life'] = now_current_balance
                            
                    elif qty_to_rest > past_oh and past_oh == past_balance:
                        now_current_balance = now_current_balance - demand - prev_oh
                        df.at[index, 'real_balance_oh_shelf_life'] = now_current_balance
                        
                    elif balance_oh < 0.0 and now_current_balance < 0.0:
                        prev_oh = float(row['balance_only_oh'])
                        now_current_balance = now_current_balance - demand
                        df.at[index, 'real_balance_oh_shelf_life'] = now_current_balance
                        
                    elif now_current_balance > 0.0 and balance_oh < 0.0 and prev_oh > 0.0:
                        # prev_oh = float(row['balance_only_oh'])
                        now_current_balance = now_current_balance - demand - prev_oh
                        df.at[index, 'real_balance_oh_shelf_life'] = now_current_balance
                        
                    elif past_oh < 0.0 :
                        prev_oh = float(row['balance_only_oh'])
                        now_current_balance = now_current_balance - demand
                        df.at[index, 'real_balance_oh_shelf_life'] = now_current_balance
                        
                    elif past_oh > qty_to_rest and past_oh > 0.0:
                        prev_oh = float(row['balance_only_oh'])
                        now_current_balance = now_current_balance - demand - qty_to_rest
                        df.at[index, 'real_balance_oh_shelf_life'] = now_current_balance
                        
                    elif past_oh <= qty_to_rest:
                        prev_oh = float(row['balance_only_oh'])
                        now_current_balance = now_current_balance - demand - past_oh
                        df.at[index, 'real_balance_oh_shelf_life'] = now_current_balance
                        
                    else:
                        pass
            return df    
        base = coverage_oh_real_with_shelf_life(base)
        base = base.rename(columns = {'qty_to_rest':'qty_Expired','real_balance_oh_shelf_life':'Balance_OH_sl%'})

        def mark_first_negative_balance(df):
            first_impact = []
            prev_item = None
            for index, row in df.iterrows():
                if prev_item != row['ItemNo']:
                    first_less_than_0 = False
                    prev_item = row['ItemNo']
                if not first_less_than_0 and row['Balance_OH_sl%'] < 0:
                    first_impact.append('x')
                    first_less_than_0 = True
                else:
                    first_impact.append('')
            df['1st_Impact'] = first_impact
            return df
        
        base['Balance_OH_sl%'] = base['Balance_OH_sl%'].astype(float)
        base = mark_first_negative_balance(base)
        
        transit_table = transit_table[['Vendor','TransitTime']]
        transit_table['TransitTime'] = transit_table['TransitTime'].astype(float)
        
        base['Demand'] = base['Demand'].astype(float)
        base['PO_Qty'] = base['PO_Qty'].astype(float)
        base['OH_Qty'] = base['OH_Qty'].astype(float)
        base['RwkOH'] = base['RwkOH'].astype(float)
        base['Balance_OH_sl%'] = base['Balance_OH_sl%'].astype(float)
        
        piv_balance = base.copy()
        piv_balance = pd.pivot_table(piv_balance,index = 'ItemNo',columns = 'Week_Str',values='Balance_OH_sl%',aggfunc='sum')
        piv_balance = piv_balance.reset_index()
        
        main_base = base.copy()
        main_base = main_base[['ItemNo','LotSize','LT','STD_Cost','Vendor','Tactical','Operational','Vendor_Name','RwkOH']]
        main_base = main_base.drop_duplicates()
        main_base['LT'].replace('', np.nan, inplace=True)
        main_base['LT']  = main_base['LT'].astype(float)
        main_base = pd.merge(main_base,transit_table,how = 'left',on = 'Vendor')
        
        main_base['TransitTime'] = main_base['TransitTime'].fillna('0')
        main_base['TransitTime'] = main_base['TransitTime'].astype(float)
        main_base['LT_t'] = main_base['LT']  + main_base['TransitTime'] 
        main_base['TransitTime'] = main_base['TransitTime'].astype(str)
        
        def check_transit_time(DataFrame):
            DataFrame['Logic_Transit'] = ''
            for index,row in DataFrame.iterrows():
                if row['TransitTime'] !='0.0':
                    DataFrame.at[index,'Logic_Transit'] = 'Apply'
                else:
                    DataFrame.at[index,'Logic_Transit'] = 'n/a'
        check_transit_time(main_base)
        
        main_base = pd.merge(main_base,piv_oh_total,how = 'left',on = 'ItemNo')
        main_base = pd.merge(main_base,piv_openOrder1_total,how = 'left',on = 'ItemNo')

        main_base['OH'] = main_base['OH'].fillna('0')
        main_base["PO_Qty"] = main_base["PO_Qty"].fillna('0')
        
        impact_table = base.copy()
        past_due_imact = base.copy()
        impact_table['auto_increment'] = range(1, len(impact_table) + 1)
        impact_table['auto_increment'] = impact_table['auto_increment'].astype(int)
        impact_table['Balance_OH_sl%'] = impact_table['Balance_OH_sl%'].astype(float)
        impact_table = impact_table.loc[impact_table['Balance_OH_sl%']<0.0]

        def mas_7(DataFrame):
            DataFrame['mas 7'] = ''
            for index,row in DataFrame.iterrows():
                if row['1st_Impact'] =='x':
                    DataFrame.at[index,'mas 7'] = row['auto_increment'] + 7
                else:
                    pass
        mas_7(impact_table)

        impact_table_clarendon = impact_table.loc[impact_table['Vendor']=='EZ0663']
        validate_shape = impact_table_clarendon
        validate_shape = validate_shape.loc[validate_shape['1st_Impact']=='x']


        impact_table = impact_table.loc[impact_table['Vendor']!='EZ0663']
        impact_table = impact_table.loc[impact_table['1st_Impact']=='x']
        impact_table = impact_table[['auto_increment','ItemNo','Week_Str','Balance_OH_sl%']]

        impact_table_clarendon['ItemNo'] = impact_table_clarendon['ItemNo'].astype(str)
        impact_table_clarendon['mas 7'] = impact_table_clarendon['mas 7'] .astype(str)
        impact_table_clarendon['key2'] = impact_table_clarendon['ItemNo'] +"_"+ impact_table_clarendon['mas 7'] 

        df1 = impact_table_clarendon.copy()
        df2 = impact_table_clarendon.copy()

        df1 = df1.fillna('')
        df1 = df1.loc[df1['mas 7']!='']
        df1['key'] = df1['ItemNo'] +"_"+ df1['mas 7'] 
        df1 = df1[['key','ItemNo','Week_Str','Balance_OH_sl%']]

        df2 = df2.fillna('')
        df2['auto_increment'] = df2['auto_increment'].astype(str)
        df2['key'] = df2['ItemNo'] +"_"+ df2['auto_increment'] 
        df2 = df2[['key','Balance_OH_sl%']]

        df3 = pd.merge(df1,df2,how ='left' ,on = 'key')

        impact_table['Balance_OH_sl%_y'] = ''
        impact_table = impact_table.rename(columns = {'Balance_OH_sl%':'Balance_OH_sl%_x'})
        impact_table = impact_table[['ItemNo','Week_Str','Balance_OH_sl%_x','Balance_OH_sl%_y']]

        df3 = df3[['ItemNo','Week_Str','Balance_OH_sl%_x','Balance_OH_sl%_y']]
        df4 = pd.concat([df3,impact_table])


        base_pivot = pd.pivot_table(base,index = 'ItemNo',columns = 'Week_Str',values='Balance_OH_sl%',aggfunc='sum')
        
        main_base = main_base[['ItemNo','LotSize','LT','STD_Cost','Vendor','Tactical'
                            ,'Operational','Vendor_Name','Logic_Transit','OH','PO_Qty','RwkOH']]
        main_base = pd.merge(main_base,df4,how = 'left',on = 'ItemNo')
        

        """# Get the current date and calculate the start of the current week"""
        current_date = datetime.today().date()
        current_week_start = current_date - timedelta(days=current_date.weekday())
        data = []
        for i in range(156):
            week_start_date = current_week_start + timedelta(weeks=i)
            year, week_num, _ = week_start_date.isocalendar()
            week_string = f"{year}-{week_num:02d}"
            data.append([week_string, week_start_date])
        monday_table = pd.DataFrame(data, columns=['Week_Str', '1st Impact'])
        monday_table['1st Impact'] = pd.to_datetime(monday_table['1st Impact'])
        monday_table['1st Impact'] = monday_table['1st Impact'].dt.date

        main_base  = pd.merge(main_base,monday_table,how = 'left',on = 'Week_Str')
        main_base = main_base.rename(columns = {'Balance_OH_sl%_y':'Qty Sh Clarendon'
                                                ,'Balance_OH_sl%_x':'Qty Shortage','Week_Str':'YW 1st Impact'})

        def pr_column(data_frame):
            data_frame['PR'] = ''
            data_frame['YW PR'] = ''
            data_frame['PR sin Past Due'] = ''
            current_date = pd.Timestamp.now().normalize()
            current_week_start = current_date - pd.Timedelta(days=current_date.dayofweek)
            
            for index, row in data_frame.iterrows():
                lead_time = row['LT']
                if pd.isna(lead_time):
                    lead_time = 0  # Handle NaN by setting lead_time to 0 or another appropriate default value
                else:
                    lead_time = int(lead_time)
                
                if not pd.isnull(row['1st Impact']) and not pd.isna(row['1st Impact']):
                    first_impact_date = pd.Timestamp(row['1st Impact'])
                    if first_impact_date is not pd.NaT:
                        first_impact_date = first_impact_date.normalize()
                        if not pd.isnull(row['YW 1st Impact']):
                            if row['Logic_Transit'] == 'Apply':
                                pr_value = first_impact_date - pd.offsets.BDay(lead_time + 20)
                            else:
                                pr_value = first_impact_date - pd.offsets.BDay(lead_time + 15)
                            data_frame.at[index, 'PR'] = pr_value
                            pr_value_as_timestamp = pd.Timestamp(pr_value)  # Ensure PR is a Timestamp
                            if pr_value_as_timestamp < current_week_start:
                                data_frame.at[index, 'PR sin Past Due'] = current_week_start
                            else:
                                pass
                            data_frame.at[index, 'YW PR'] = pr_value.strftime('%Y-%W')

        pr_column(main_base)
        

        main_base = main_base.fillna('')

        def pr_pd_leadtime(DataFrame):
            DataFrame['PR PD + LT'] = ''
            for index, row in DataFrame.iterrows():
                impact_date = pd.Timestamp(row['PR sin Past Due'])
                lead_time = float(row['LT'])
                if row['PR sin Past Due'] !='':
                    if row['Logic_Transit'] == 'Apply':
                        DataFrame.at[index, 'PR PD + LT'] = impact_date + pd.offsets.BDay(lead_time + 20)
                    else:
                        DataFrame.at[index, 'PR PD + LT'] = impact_date + pd.offsets.BDay(lead_time + 15)
        pr_pd_leadtime(main_base)
        main_base = main_base.fillna('')

        def wk_impactada(DataFrame):
            DataFrame['WK Impactada'] = ''
            for index, row in DataFrame.iterrows():
                if row['PR sin Past Due'] == '':
                    if row['YW PR'] == '':
                        pass
                    else:
                        DataFrame.at[index,'WK Impactada'] = row['YW 1st Impact']
                else:
                    pr_sin_past_due = pd.Timestamp(row['PR PD + LT'])
                    DataFrame.at[index, 'WK Impactada'] = pr_sin_past_due.strftime('%Y-%W')
        wk_impactada(main_base)

        main_base['PR PD + LT'] = main_base['PR PD + LT'].astype(str).replace('NaT', '')

        past_due_imact = past_due_imact[['key','Balance_OH_sl%']]
        main_base['key'] = main_base['ItemNo'] +"_"+ main_base['WK Impactada']
        main_base = pd.merge(main_base,past_due_imact,how = 'left',on = 'key')
        def clean_qty_shortage(DataFrame):
            DataFrame['Qty Shortage2'] = ''
            for index,row in DataFrame.iterrows():
                if row['Vendor'] != 'EZ0663' and row['PR PD + LT'] =='':
                    DataFrame.at[index,'Qty Shortage2'] = row['Qty Shortage']
                else:
                    if row['Vendor'] == 'EZ0663' and row['PR PD + LT'] =='':
                        DataFrame.at[index,'Qty Shortage2'] = row['Qty Sh Clarendon']
                    else:
                        if row['PR PD + LT'] !='':
                            DataFrame.at[index,'Qty Shortage2'] = row['Balance_OH_sl%']
                        else:
                            pass
        clean_qty_shortage(main_base)
        

        main_base['Europeos'] = main_base['Vendor'].isin(vendor_europeos['Vendor'])
        main_base['Europeos'] = main_base['Europeos'].map({True: 'Yes', False: 'No'})
        main_base = main_base.fillna('')
        
        columns_to_drop = ['key', 'Balance_OH_sl%']
        main_base = main_base.drop(columns=columns_to_drop)
        main_base = main_base.fillna('')
        
        main_base['YW 1st Impact'] = main_base['YW 1st Impact'].replace('','OH')
        main_base['WK Impactada'] = main_base['WK Impactada'].replace('','OH')
        main_base['Qty Shortage2'] = main_base['Qty Shortage2'].replace('',0.0)
        
        def moq_qty_selected(DataFrame):
            DataFrame['PR_Qty'] = ''
            for index, row in DataFrame.iterrows():
                if row['YW 1st Impact'] == 'OH':
                    pass
                else:
                    lot_size = abs(float(row['LotSize']))
                    qty_shortage = abs(float(row['Qty Shortage2']))
                    if lot_size > math.ceil(qty_shortage):
                            DataFrame.at[index, 'PR_Qty'] = lot_size
                    else:
                        DataFrame.at[index, 'PR_Qty'] = math.ceil(qty_shortage)

        moq_qty_selected(main_base)
        
        main_base = pd.merge(main_base,expedite_project,how = 'left',on = 'ItemNo')
        main_base['key'] = main_base['ItemNo'] +"_"+ main_base['Vendor'] 
        main_base = pd.merge(main_base,po351,how = 'left',on = 'key')
        main_base = pd.merge(main_base,action_notes,how = 'left',on = 'ItemNo')
        main_base = pd.merge(main_base,actions_by_item,how = 'left',on = 'ItemNo')
        
        main_base = main_base.rename(columns = {'Vendor':'Vendor Code','Qty Shortage':'qs'
                                                ,'Qty Shortage2':'Qty Shortage'
                                                ,'PlanningNotes':'Planning Notes','ACD':'Action Message'})

        main_base = main_base[['ItemNo','LotSize','LT','STD_Cost','Unit Cost','Project','Vendor Code'
                            ,'Vendor_Name','Tactical','Operational','Logic_Transit','OH','PO_Qty','RwkOH'
                            ,'YW 1st Impact','1st Impact','PR','PR sin Past Due','YW PR','PR PD + LT'
                            ,'WK Impactada','Qty Shortage','PR_Qty','Europeos','Planning Notes','Action Message']]
        

        base_pivot = pd.pivot_table(base,index = 'ItemNo',columns = 'Week_Str',values='Balance_OH_sl%',aggfunc='sum')
        
        def fill_pr_sin_pastdue(df):
            for index,row in df.iterrows():
                if row['PR sin Past Due'] !='':
                    df.at[index,'PR sin Past Due'] = row['PR sin Past Due']
                else:
                    df.at[index,'PR sin Past Due'] = row['PR']
        fill_pr_sin_pastdue(main_base)
        
        main_base = main_base.fillna('')
        def delivery_date(df):
            df['Delivery Date'] = ''
            for index,row in df.iterrows():
                lead_time = float(row['LT'])
                if row['PR sin Past Due'] == '':
                    pass
                else:
                    impact_date = pd.Timestamp(row['PR sin Past Due'])
                    df.at[index, 'Delivery Date'] = impact_date + pd.offsets.BDay(lead_time)
        delivery_date(main_base)


        main_base = main_base.fillna('')
        def new_year_week(df):
            for index,row in df.iterrows():
                if row['PR sin Past Due'] !='':
                    first_impact_date = pd.Timestamp(row['PR sin Past Due'])
                    df.at[index, 'YW PR'] = first_impact_date.strftime('%Y-%W')
                else:
                    pass
        new_year_week(main_base) 
        main_base = main_base.fillna('')
        
        # Replace week 52 and 53 with 51 using regular expressions
        main_base['YW PR'] = main_base['YW PR'].str.replace(r'-(52|53)$', '-51', regex=True)
        main_base['YW 1st Impact'] = main_base['YW 1st Impact'].str.replace(r'-(52|53)$', '-51', regex=True)
        main_base['WK Impactada'] = main_base['WK Impactada'].str.replace(r'-(52|53)$', '-51', regex=True)

        def lt_burned(df):
            current_date = datetime.today().date()
            days_to_subtract = current_date.weekday()
            monday_of_current_week = current_date - timedelta(days=days_to_subtract)
            # Initialize 'LT Quemado' column
            df['LT Quemado'] = ''
            for index, row in df.iterrows():
                try:
                    pr_date = pd.to_datetime(row['PR'], errors='coerce').date()  # Convert PR to datetime, coerce invalid dates to NaT
                    if pd.isna(pr_date):
                        continue  # Skip rows with invalid dates
                    if pr_date < monday_of_current_week:
                        df.at[index, 'LT Quemado'] = 'LT Quemado'
                    else:
                        df.at[index, 'LT Quemado'] = 'No'
                except Exception as e:
                    continue  # Skip rows if any unexpected error occurs
        lt_burned(main_base)

        main_base = pd.merge(main_base,base_pivot,how = 'left',on = 'ItemNo')
        main_base = main_base.fillna('')
        
        
        datestringsave = datetime.now().strftime('%m_%d')# type: ignore
        writer = pd.ExcelWriter(f'J:\\Departments\\Material Control\\Purchasing\\Tools\\ComprasDB\\PR Actions with SL% Historico\\PR Actions SL% {datestringsave}.xlsx')            
        main_base.to_excel(writer, sheet_name='PR Actions', index=False)
        base.to_excel(writer, sheet_name='Shelf Life Logic', index=False)
        writer.close()

        def get_or_create_sheet(wb, sheet_name):
            """Returns the sheet if it exists, otherwise creates it."""
            if sheet_name in [sheet.name for sheet in wb.sheets]:
                return wb.sheets[sheet_name]
            else:
                return wb.sheets.add(name=sheet_name)



        wb = xw.Book("J:\\Departments\\Material Control\\Purchasing\\Tools\\ComprasDB\\PythonCode\\pr_file.xlsx")    
        

        sheet_pr_actions = get_or_create_sheet(wb, 'PR Actions')
        sheet_pr_actions.range('A1').value = [main_base.columns.tolist()] + main_base.to_numpy().tolist()
        sheet_shelf_life_logic = get_or_create_sheet(wb, 'Shelf Life Logic')
        sheet_shelf_life_logic.range('A1').value = [base.columns.tolist()] + base.to_numpy().tolist()
        path_save_pr_format = f"J:\\Departments\\Material Control\\Purchasing\\Tools\\ComprasDB\\PR Actions with SL% Historico\\PR With Fromat\\PR Actions SL% {datestringsave}.xlsx"
        wb.save(path_save_pr_format)   
        QMessageBox.information(self, "Creacion de PR Actions", "Se creo el historico PR Actions correctamente")

    """-------------------------------clock-------------------------------"""
    def setupUI(self): #CLOCK
        self.label_7 = QLabel(self)
        self.label_7.setGeometry(10, 10, 180, 80)
        font = QFont("Consolas", 32, QFont.Bold)
        self.label_7.setFont(font)

    def updateTime(self): #CLOCK
        current_time = QTime.currentTime().toString("hh:mm:ss")
        self.label_7.setText(current_time)

    def startTimer(self): #CLOCK
        timer = QTimer(self, interval=1000, timeout=self.updateTime)
        timer.start()

class MainApplication2(QMainWindow, Ui_MainWindow): # SEND OPEN ORDER REPORT
    
    def __init__(self):
        super().__init__()
        self.startTimer()  #CLOCK
        self.setupUi(self)
        icon = QIcon(f'{icon_ezair}')
        self.setWindowIcon(icon)
        self.setWindowTitle("BUYER TOOLS")
        self.folder_path = None
        self.pushButton_31.clicked.connect(self.create_a_historic_folder)
        self.pushButton_29.clicked.connect(self.mail2)
        self.pushButton_38.clicked.connect(self.find_information_selected_table)
        self.pushButton_40.clicked.connect(self.load_open_order_table)
        self.pushButton_35.clicked.connect(self.load_contact_supplier)
        self.pushButton_30.clicked.connect(self.run_open_order_send_supplier_email)
        self.pushButton_46.clicked.connect(self.create_folder_downloads)
        self.pushButton_45.clicked.connect(self.create_template_upload_PO)
        self.pushButton_47.clicked.connect(self.validate_run_uploadpo)
        self.pushButton_48.clicked.connect(self.load_upload_po)
        
        pixmap = QPixmap(app_logo)
        scaled_pixmap = pixmap.scaledToWidth(200)
        self.label_48.setPixmap(scaled_pixmap)
        self.label_48.setScaledContents(True)
        pixmap = QPixmap(body_mail)
        scaled_pixmap = pixmap.scaledToWidth(200)
        self.label_49.setPixmap(scaled_pixmap)
        self.label_49.setScaledContents(True)
        pixmap = QPixmap(contact_mail)
        scaled_pixmap = pixmap.scaledToWidth(200)
        self.label_50.setPixmap(scaled_pixmap)
        self.label_50.setScaledContents(True)
        pixmap = QPixmap(supplier)
        
        pixmap = QPixmap(open_order)
        scaled_pixmap = pixmap.scaledToWidth(200)
        self.label_31.setPixmap(scaled_pixmap)
        self.label_31.setScaledContents(True)
        
        """NO AVAILABLE TAB"""
        self.tabWidget.setTabEnabled(0, False)
        self.tabWidget.setTabEnabled(1, False)
        self.tabWidget.setTabEnabled(2, False)
        self.tabWidget.setTabEnabled(3, False)
        self.tabWidget.setTabEnabled(4, False)
        self.tabWidget.setTabEnabled(5, False)
        self.tabWidget.setTabEnabled(6, False)
        self.tabWidget.setTabEnabled(8, False)
        self.tabWidget.setTabEnabled(9, False)
        self.label_54.setHidden(True)
        self.label_55.setHidden(True)
        """place holders"""
        self.lineEdit_4.setPlaceholderText("Column Name")
        self.lineEdit_5.setPlaceholderText("Value to Find")

        """tabcontrol"""
        # self.tabWidget_2.setTabEnabled(3, False)
        # self.tabWidget_2.setTabEnabled(4, False)
        self.show_rgb_colors()
        
        """__________________________________INFORMATION DATABASE__________________________________________________________________"""

    def create_template_upload_PO(self):
        wb = xw.Book()    
        sheet1 = wb.sheets.add()
        sheet1.name = "Subir PO"
        sheet1.range('A1').value = 'PO'
        sheet1.range('B1').value = 'VENDOR'
        sheet1.range('C1').value = 'AMOUNT'
        sheet1.range('D1').value = 'COMMENT'
        QMessageBox.information(self, "Templete PO", "Se creo el templete correctamente, llenalo para cargarlo posteriormente")

    def load_upload_po(self):
        excel_po, _ = QFileDialog.getOpenFileName(self, "Selecciona la tabla de tu Open Order", "", "Excel files (*.xlsx)")
        if excel_po != "":
            self.excel_po = excel_po
            QMessageBox.information(self, "Archivo de PO", "Fue seleccionado el archivo corretamente")
        else:
            QMessageBox.critical(self, "Seleccion fallida", 'No seleccionaste ningun archivo')

    def validate_run_uploadpo(self):
        now = datetime.now()
        today = now.strftime("%m_%d_%Y")
        if hasattr(self, 'excel_po') and self.excel_po != '':
            excel_data = pd.read_excel(self.excel_po)
            excel_data_columns = excel_data
            excel_data_columns = excel_data_columns.columns.to_list()
            lista = ['PO','VENDOR','AMOUNT','COMMENT']
            if excel_data_columns == lista:
                folder_path = f"C:\\Users\\{cpu}\\Desktop\\Historico_upload_PO\\{today}"
                if os.path.exists(folder_path):
                    pass
                else:
                    self.create_folder_downloads()

                def read_folder_content(folder_path):
                    files = os.listdir(folder_path)
                    pdf_files = [os.path.splitext(file)[0] for file in files if file.endswith('.pdf')]
                    df = pd.DataFrame({'filename': pdf_files})
                    
                    base = pd.merge(excel_data,df,how = 'left',left_on = 'PO',right_on = 'filename')
                    base = base.fillna('')
                    base = base.loc[base['filename']!='']
                    if not base.empty:
                        driver_path = f"C:\\Users\\{cpu}\\Desktop\\driver\\msedgedriver.exe"
                        # driver = webdriver.Edge(executable_path=driver_path)
                        # #driver.get("http://ezanoc1cho:91/PO/PurchaseOrders.aspx")
                        
                        # for index,row in base.iterrows():                   
                        #     po = str(row['PO'])
                        #     vendor =  str(row['VENDOR'])
                        #     amont = float(row['AMOUNT'])
                        #     comment =  str(row['COMMENT'])
                            
                        #     driver.find_element(By.XPATH,"//*[@id='ContentPlaceHolder1_btn_Add']").click()
                        #     driver.find_element(By.XPATH,'//*[@id="txt_m_PO"]').send_keys(po)
                        #     driver.find_element(By.XPATH,'//*[@id="ddl_m_vendor_chosen"]/a/span').click()
                        #     driver.find_element(By.XPATH,"//*[@id='ddl_m_vendor_chosen']/div/div/input").send_keys(vendor)
                        #     driver.find_element(By.XPATH,'//*[@id="txt_m_PO"]').send_keys(po)
                        #     driver.find_element(By.XPATH,'//*[@id="tb_m_amount"]').send_keys(amont)
                        #     driver.find_element(By.XPATH,'//*[@id="tb_m_comments"]').send_keys(comment)
                        
                        
                        # wb = xw.Book()
                        # sheet1 = wb.sheets.add()
                        # sheet1.name = "base"
                        # sheet1.range('A1').value = base 
                    else:
                        QMessageBox.critical(self, "PO en carpeta de descargas", 'No hay PO en la carpeta de descargas')
                    
                                        
                read_folder_content(folder_path)            
                
            else:
                QMessageBox.critical(self, "Seleccion fallida", 'El archivo no correcponde')
        else:
            QMessageBox.critical(self, "Seleccion fallida", 'No seleccionaste ningun archivo')

    def create_folder_downloads(self):
            try:
                now = datetime.now()
                today = now.strftime("%m_%d_%Y")
                cpu = f"{os.environ['USERNAME']}.ONE"

                carpeta_historico = os.path.join(f'C:\\Users\\{cpu}\\Desktop', f'Historico_upload_PO')
                if not os.path.exists(carpeta_historico):
                    os.makedirs(carpeta_historico)
                
                carpeta_historico_day = os.path.join(carpeta_historico, today)
                if not os.path.exists(carpeta_historico_day):
                    os.makedirs(carpeta_historico_day)

            except sqlite3.Error as error:
                QMessageBox.critical(self, "Error en base ", f'No se pudo establecer coneccion {error}')
            
            finally:
                QMessageBox.information(self, "Carpeta Historico", "Se creo la carpeta historico correctamente guarda las PO que quieras subir al sistema")

    def create_a_historic_folder(self):
        
        try:
            now = datetime.now()
            today = now.strftime("%m_%d_%Y")
            cpu = f"{os.environ['USERNAME']}.ONE"

            carpeta_historico = os.path.join(f'C:\\Users\\{cpu}\\Desktop', f'Historico')
            if not os.path.exists(carpeta_historico):
                os.makedirs(carpeta_historico)
            
            carpeta_historico_day = os.path.join(carpeta_historico, today)
            if not os.path.exists(carpeta_historico_day):
                os.makedirs(carpeta_historico_day)
            
            carpeta_need_templates = os.path.join(carpeta_historico_day, 'Templates')
            if not os.path.exists(carpeta_need_templates):
                os.makedirs(carpeta_need_templates)

            def create_excel_file(file_name, content):
                workbook = Workbook()
                sheet = workbook.active

                for row in content:
                    sheet.append(row)

                file_path = os.path.join(carpeta_need_templates, f"{file_name}.xlsx")
                workbook.save(file_path)
                print(f"Excel file '{file_name}.xlsx' created with content.")

            content_1 = [['Buyer','PO Line','PO-No','Ln','Vendor #','Vendor Name','Item-Number','Item-Description',
                                    'Rev','U/M','Ord-Q','Rcvd-Q','Opn-Q','Unit-$','Net-Opn-Val','PO-Dt','Req-Dt','Prom-Dt'
                                    ,'Rev-Pr-Dt','AM','EZA Req Date','LT','Last Note/Status','CONFIRMED RECOVERY','Comments']]


            content_3 = [['Vendor', 'Shortname', 'Copy_1','Send_1','Send_2']]

            create_excel_file("Open Order", content_1)
            # create_excel_file("My Supplier", content_2)
            create_excel_file("Contact Supplier", content_3)
                        
            folder_path = os.path.join(carpeta_historico_day, 'Database')
            if not os.path.exists(folder_path):
                os.makedirs(folder_path)
            
            db_name = f'database_{cpu}.db'
            full_path = os.path.join(folder_path, db_name)
            self.create_database_if_not_exists(full_path)
            
        except sqlite3.Error as error:
            QMessageBox.critical(self, "Error en base ", f'No se pudo establecer coneccion {error}')
            
        finally:
            QMessageBox.information(self, "Carpeta Historico", "Se creo la carpeta historico correctamente asi como los templates necesarios")

    def create_database_if_not_exists(self, database_path):
        try:
            if not os.path.exists(os.path.dirname(database_path)):
                os.makedirs(os.path.dirname(database_path))
            if not os.path.exists(database_path):
                conn = sqlite3.connect(database_path)
                # Additional operations if needed
                self.create_needs_tables()
                conn.close()
                print(f"Database created at {database_path}")
            else:
                print(f"Database already exists at {database_path}")
        except Exception as e:
            print(f"Error occurred while accessing the database: {e}")
            QMessageBox.critical(self, "Error", f"An error occurred: {e}")

    def find_information_selected_table(self):
        now = datetime.now()
        today = now.strftime("%m_%d_%Y")
        my_db = f"C:\\Users\\{cpu}\\Desktop\\Historico\\{today}\\Database\\database_{cpu}.db"
        
        if self.radioButton.isChecked():
            if self.lineEdit_4.text() != '' and self.lineEdit_5.text() != '':
                query = f"""SELECT * FROM open_order WHERE {self.lineEdit_4.text()} = '{self.lineEdit_5.text()}'"""
            else:
                query = "SELECT * FROM open_order"
            try:
                conn = sqlite3.connect(my_db)
                cursor = conn.cursor()
                result = cursor.execute(query)
                self.tableWidget_6.clear()
                self.tableWidget_6.setRowCount(0)
                self.tableWidget_6.setColumnCount(len(cursor.description))
                column_names = [description[0] for description in cursor.description]
                self.tableWidget_6.setHorizontalHeaderLabels(column_names)
                for row_number, row_data in enumerate(result):
                    self.tableWidget_6.insertRow(row_number)
                    for column_number, data in enumerate(row_data):
                        self.tableWidget_6.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
                conn.close()
            except Exception as error:
                QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
            finally:
                pass

        
        elif self.radioButton_2.isChecked():
            if self.lineEdit_4.text() != '' and self.lineEdit_5.text() != '':
                query = f"""SELECT * FROM contact_supplier WHERE {self.lineEdit_4.text()} = '{self.lineEdit_5.text()}'"""
            else:
                query = "SELECT * FROM contact_supplier"
            try:
                conn = sqlite3.connect(my_db)
                cursor = conn.cursor()
                result = cursor.execute(query)
                self.tableWidget_6.clear()
                self.tableWidget_6.setRowCount(0)
                self.tableWidget_6.setColumnCount(len(cursor.description))
                column_names = [description[0] for description in cursor.description]
                self.tableWidget_6.setHorizontalHeaderLabels(column_names)
                for row_number, row_data in enumerate(result):
                    self.tableWidget_6.insertRow(row_number)
                    for column_number, data in enumerate(row_data):
                        self.tableWidget_6.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
                conn.close()
            except Exception as error:
                QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion {error}')
            finally:
                pass

        else:
            QMessageBox.critical(self, "Error en base de datos", f'No seleccionaste ninguna tabla para consultar')

    def mail2(self):
        global mail_value2
        mail_value2 = self.plainTextEdit_2.toPlainText()
        print(mail_value2)
        QMessageBox.information(self, "Éxito", "Fue cargado el cuerpo del correo que sera enviado a los proveedores correctamente")

    def create_needs_tables(self):
        try:
            now = datetime.now()
            today = now.strftime("%m_%d_%Y")
            path_conn_cpu = f"C:\\Users\\{cpu}\\Desktop\\Historico\\{today}\\Database\\database_{cpu}.db"
            conn = sqlite3.connect(path_conn_cpu)
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='open_order'")
            table_exists = cursor.fetchone()
            if table_exists:
                cursor.execute("""DELETE FROM open_order""")
                conn.commit()
            cursor.execute("""CREATE TABLE IF NOT EXISTS open_order(Buyer TEXT,PO_Line TEXT,PO_No TEXT,Ln TEXT,Vendor TEXT,
                    Vendor_Name TEXT,Item_Number TEXT,Item_Description TEXT,Rev TEXT,UM TEXT,
                    Ord_Q TEXT,Rcvd_Q TEXT,Opn_Q TEXT,Unit_Price TEXT,Net_Opn_Val TEXT,PO_Dt TEXT,Req_Dt TEXT,Prom_Dt TEXT,
                    Rev_Pr_Dt TEXT,AM TEXT,EZA_Req_Date TEXT,LT TEXT,Last_Note_Status TEXT,CONFIRMED_RECOVERY TEXT,Comments TEXT)""")
            conn.commit()
            conn.close()
        except Exception as error:
            QMessageBox.critical(self, "Error en la base", f'Error en la base de datos: {error}')
        finally:
            pass
        try:
            now = datetime.now()
            today = now.strftime("%m_%d_%Y")
            path_conn_cpu = f"C:\\Users\\{cpu}\\Desktop\\Historico\\{today}\\Database\\database_{cpu}.db"
            conn = sqlite3.connect(path_conn_cpu)
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='contact_supplier'")
            table_exists = cursor.fetchone()
            if table_exists:
                cursor.execute("""DELETE FROM contact_supplier""")
                conn.commit()
            cursor.execute("""CREATE TABLE IF NOT EXISTS contact_supplier(Vendor TEXT, Shortname TEXT,Copy_1 TEXT,Send_1 TEXT,Send_2	TEXT)""")
            conn.commit()
            conn.close()
        except Exception as error:
            QMessageBox.critical(self, "Error en la base", f'Error en la base de datos: {error}')
        finally:
            pass
        try:
            now = datetime.now()
            today = now.strftime("%m_%d_%Y")
            path_conn_cpu = f"C:\\Users\\{cpu}\\Desktop\\Historico\\{today}\\Database\\database_{cpu}.db"
            conn = sqlite3.connect(path_conn_cpu)
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='colors_format_am'")
            table_exists = cursor.fetchone()
            if table_exists:
                cursor.execute("""DELETE FROM colors_format_am""")
                conn.commit()
            cursor.execute("""CREATE TABLE IF NOT EXISTS colors_format_am(Action_Message TEXT, RGB_Color TEXT)""")
            conn.commit()
            conn.close()
            self.show_rgb_colors()
            QMessageBox.information(self, "Éxito", "Tablas Creadas correctamente")
        except Exception as error:
            QMessageBox.critical(self, "Error en la base", f'Error en la base de datos: {error}')
        finally:
            pass

    def load_open_order_table(self):
        now = datetime.now()
        today = now.strftime("%m_%d_%Y")
        my_db = f"C:\\Users\\{cpu}\\Desktop\\Historico\\{today}\\Database\\database_{cpu}.db"
        my_db_path = f"C:\\Users\\{cpu}\\Desktop\\Historico\\{today}\\Database\\database_{cpu}.db"
        
        if os.path.exists(my_db_path):
            try:
                conn = sqlite3.connect(my_db_path)
                cursor = conn.cursor()
                oo_query ="""SELECT * FROM sqlite_master where type = 'table' AND name = 'open_order'"""
                cursor.execute(oo_query)
                table_exist = cursor.fetchone()
                if table_exist:
                    cursor.execute("""DELETE FROM open_order""")
                    conn.commit()
                cursor.execute("""CREATE TABLE IF NOT EXISTS open_order(
                    Buyer TEXT,
                    PO_Line TEXT,
                    PO_No TEXT,
                    Ln TEXT,
                    Vendor TEXT,
                    Vendor_Name TEXT,
                    Item_Number TEXT,
                    Item_Description TEXT,
                    Rev TEXT,
                    UM TEXT,
                    Ord_Q TEXT,
                    Rcvd_Q TEXT,
                    Opn_Q TEXT,
                    Unit_Price TEXT,
                    Net_Opn_Val TEXT,
                    PO_Dt TEXT,
                    Req_Dt TEXT,
                    Prom_Dt TEXT,
                    Rev_Pr_Dt TEXT,
                    AM TEXT,
                    EZA_Req_Date TEXT,
                    LT TEXT,
                    Last_Note_Status TEXT,
                    CONFIRMED_RECOVERY TEXT,
                    Comments TEXT)""")
                conn.commit()

            except Exception as error:
                QMessageBox.critical(self,"DataBase Error",f"Impossible to reach the DataBase the error is {error}")
            finally:
                pass
        else:
            QMessageBox.warning(self, "Database file not found", f"The database file '{my_db_path}' does not exist.")
            return
        
        load_oo_table, _ = QFileDialog.getOpenFileName(self, "Selecciona la tabla de tu Open Order", "", "Excel files (*.xlsx)")
        
        if load_oo_table != "":
        
            self.load_oo_table = load_oo_table
            oo = pd.read_excel(f'{self.load_oo_table}',dtype=str)
            col_names_oo = oo.columns.to_list()           
            oo_columns_should_be =['Buyer','PO Line','PO-No','Ln','Vendor #','Vendor Name','Item-Number','Item-Description',
                                    'Rev','U/M','Ord-Q','Rcvd-Q','Opn-Q','Unit-$','Net-Opn-Val','PO-Dt','Req-Dt','Prom-Dt'
                                    ,'Rev-Pr-Dt','AM','EZA Req Date','LT','Last Note/Status','CONFIRMED RECOVERY','Comments']

            if oo_columns_should_be == col_names_oo:
                oo = oo.rename(columns={'PO Line':'PO_Line'
                                    ,'PO-No':'PO_No'
                                    ,'Vendor #':'Vendor'
                                    ,'Vendor Name':'Vendor_Name'
                                    ,'Item-Number':'Item_Number'
                                    ,'Item-Description':'Item_Description'
                                    ,'U/M':'UM'
                                    ,'Ord-Q':'Ord_Q'
                                    ,'Rcvd-Q':'Rcvd_Q'
                                    ,'Opn-Q':'Opn_Q'
                                    ,'Unit-$':'Unit_Price'
                                    ,'Net-Opn-Val':'Net_Opn_Val'
                                    ,'PO-Dt':'PO_Dt'
                                    ,'Req-Dt':'Req_Dt'
                                    ,'Prom-Dt':'Prom_Dt'
                                    ,'Rev-Pr-Dt':'Rev_Pr_Dt'
                                    ,'EZA Req Date':'EZA_Req_Date'
                                    ,'Last Note/Status':'Last_Note_Status'
                                    ,'CONFIRMED RECOVERY':'CONFIRMED_RECOVERY'})
                oo = oo.fillna('')
                oo['Last_Note_Status'] = oo['Last_Note_Status'].str.upper()
                # oo['Prom_Dt'] = pd.to_datetime(oo['Prom_Dt'])
                # oo['Prom_Dt'] = oo['Prom_Dt'].dt.strftime('%m/%d/%Y')
                # oo['Rev_Pr_Dt'] = pd.to_datetime(oo['Rev_Pr_Dt'])
                # oo['Rev_Pr_Dt'] = oo['Rev_Pr_Dt'].dt.strftime('%m/%d/%Y')
                oo = oo[['Buyer','PO_Line','PO_No','Ln','Vendor','Vendor_Name','Item_Number','Item_Description',
                        'Rev','UM','Ord_Q','Rcvd_Q','Opn_Q','Unit_Price','Net_Opn_Val','PO_Dt','Req_Dt','Prom_Dt'
                        ,'Rev_Pr_Dt','AM','EZA_Req_Date','LT','Last_Note_Status','CONFIRMED_RECOVERY','Comments']]
                oo = oo.fillna('')
                oo.to_sql('open_order',conn,if_exists='append',index=False)
                QMessageBox.information(self, "Éxito", "Tabla Creada correctamente")
            
            else:
                QMessageBox.critical(self, "Error en el archivo ", 'El archivo no corresponde valida el template') 
        else:
            QMessageBox.critical(self, "Seleccion fallida", 'No seleccionaste ningun archivo')

    def load_contact_supplier(self):
        now = datetime.now()
        today = now.strftime("%m_%d_%Y")
        my_db = f"C:\\Users\\{cpu}\\Desktop\\Historico\\{today}\\Database\\database_{cpu}.db"
        my_db_path = f"C:\\Users\\{cpu}\\Desktop\\Historico\\{today}\\Database\\database_{cpu}.db"
        if os.path.exists(my_db_path):
            try:
                conn = sqlite3.connect(my_db_path)
                cursor = conn.cursor()
                contact_supplier_query ="""SELECT * FROM sqlite_master where type = 'table' AND name = 'contact_supplier'"""
                cursor.execute(contact_supplier_query)
                table_exist = cursor.fetchone()
                if table_exist:
                    cursor.execute("""DELETE FROM contact_supplier""")
                    conn.commit()
                    cursor.execute("""CREATE TABLE IF NOT EXISTS contact_supplier(
                                Vendor TEXT
                                , Shortname TEXT
                                , Copy_1 TEXT
                                , Send_1 TEXT
                                , Send_2 TEXT)""")
            except Exception as error:
                QMessageBox.critical(self,"DataBase Error",f"Impossible to reach the DataBase the error is {error}")
            finally:
                pass
        else:
            QMessageBox.warning(self, "Database file not found", f"The database file '{my_db_path}' does not exist.")
            return
        contact_mail_table, _ = QFileDialog.getOpenFileName(self, "Selecciona la tabla de tu Open Order", "", "Excel files (*.xlsx)")
        
        if contact_mail_table != "":
            self.contact_mail_table = contact_mail_table
            contact_mail_file = pd.read_excel(f'{self.contact_mail_table}',dtype=str)
            contact_mail = contact_mail_file.columns.to_list()
            contact_mail_should_be =['Vendor', 'Shortname', 'Copy_1', 'Send_1', 'Send_2']
            if contact_mail_should_be == contact_mail:

                contact_mail_file['Vendor'] = contact_mail_file['Vendor'].str.upper()
                contact_mail_file['Vendor'] = contact_mail_file['Vendor'].str.strip()
                contact_mail_file['Shortname'] = contact_mail_file['Shortname'].str.upper()
                contact_mail_file['Shortname'] = contact_mail_file['Shortname'].str.strip()
                contact_mail_file = contact_mail_file.fillna('')
                contact_mail_file.to_sql('contact_supplier',conn,if_exists='append',index=False)
                QMessageBox.information(self, "Éxito", "Tabla Creada correctamente")
            
            else:
                QMessageBox.critical(self, "Error en el archivo ", 'El archivo no corresponde') 
        else:
            QMessageBox.critical(self, "Seleccion fallida", 'No seleccionaste ningun archivo')

    def run_open_order_send_supplier_email_good(self):
        print('si estas en el archivo .py correcto')
        self.running_application_warning_enable()
        QApplication.processEvents()
        
        if self.plainTextEdit_2.toPlainText() != '':
            mail_value = self.plainTextEdit_2.toPlainText()
            now = datetime.now()
            today = now.strftime("%m_%d_%Y")
            cpu = os.environ['USERNAME']
            oo, contact = self.get_all_need_information_table()
            carpeta_historico_day = f"C:\\Users\\{cpu}\\Desktop\\Historico\\{today}\\"
            exists = os.path.exists(carpeta_historico_day)
            if exists:
                vendor_code = oo.copy()
                vendor_code = vendor_code[['Vendor']]
                vendor_code = vendor_code.drop_duplicates()

                for vendor in vendor_code['Vendor']:
                    vendor_data = oo[oo['Vendor'] == vendor].reset_index(drop=True) 
                    # Process vendor_data as needed

                    output_filename = os.path.join(carpeta_historico_day, f'OPEN ORDER EZAIR {vendor} {today}.xlsx')
                    vendor_data.to_excel(output_filename, index=False)
                    
                    wb = load_workbook(output_filename)

                    # Access the active sheet
                    sheet = wb.active

                    # Define colors
                    white_text_color = "F0F2F5"
                    black_value_color = "141414"
                    blue_value_color = "0370FC"
                    ro_action_color = "5142F5"
                    ri_action_color = "B30415"

                    # Loop through the rows and apply formatting based on condition
                    for row in range(2, len(vendor_data) + 2):
                        if sheet[f'T{row}'].value == 'RO':
                            for col in range(20, 22):  # Columns T and U are columns 20 and 21
                                    cell = sheet.cell(row=row, column=col)
                                    cell.fill = PatternFill(start_color=ro_action_color, end_color=ro_action_color, fill_type="solid")
                                    cell.font = Font(color=white_text_color)
                        else:
                            if sheet[f'T{row}'].value == 'RI':
                                for col in range(20, 22):  # Columns T and U are columns 20 and 21
                                    cell = sheet.cell(row=row, column=col)
                                    cell.fill = PatternFill(start_color=ri_action_color, end_color=ri_action_color, fill_type="solid")
                                    cell.font = Font(color=white_text_color)
                    # Loop through cells in the header row and apply formatting
                    for cell in sheet[1]:
                        cell.fill = PatternFill(start_color=black_value_color, end_color=black_value_color, fill_type="solid")
                        cell.font = Font(color=white_text_color)

                    # Loop through specific cells and apply formatting
                    for cell in sheet['X1:Y1'][0]:
                        cell.fill = PatternFill(start_color=blue_value_color, end_color=blue_value_color, fill_type="solid")
                        cell.font = Font(color=white_text_color)

                    # Save the workbook
                    wb.save(output_filename)

                    contact = contact.fillna('')
                    for index, row in contact.iterrows():
                        if row['Vendor'] == vendor:
                            shortName = row['Shortname']
                            destinatarios = [row['Send_1'], row['Send_2']]
                            copiados = [row['Copy_1']]
                            output_filename_contact = os.path.join(carpeta_historico_day, f'OPEN ORDER EZAIR {vendor} {today}.xlsx')

                            pythoncom.CoInitialize()
                            outlook = Dispatch('Outlook.Application')
                            mail = outlook.CreateItem(0)
                            mail.To = ";".join(destinatarios)
                            mail.CC = ";".join(copiados)
                            mail.Subject = f'OPEN ORDER EZAIR {shortName}'
                            mail.Body = mail_value  # Use mail_value as the email body
                            attachment = mail.Attachments.Add(output_filename_contact)
                            mail.Send()
                            del outlook
                            print(f'mail to {vendor} was send')
                        else:
                            pass

            self.running_application_warning_disable()
        else:
            QMessageBox.critical(self, "Email Body", 'Please provide the email subject')
            self.running_application_warning_disable()

    def run_open_order_send_supplier_email(self):
        print('si estas en el archivo .py correcto')
        self.running_application_warning_enable()
        QApplication.processEvents()
        
        if self.plainTextEdit_2.toPlainText() != '':
            mail_value = self.plainTextEdit_2.toPlainText()
            now = datetime.now()
            today = now.strftime("%m_%d_%Y")
            cpu = os.environ['USERNAME']
            oo, contact = self.get_all_need_information_table()
            carpeta_historico_day = f"C:\\Users\\{cpu}\\Desktop\\Historico\\{today}\\"
            exists = os.path.exists(carpeta_historico_day)
            if exists:
                vendor_code = oo.copy()
                vendor_code = vendor_code[['Vendor']]
                vendor_code = vendor_code.drop_duplicates()

                for vendor in vendor_code['Vendor']:
                    vendor_data = oo[oo['Vendor'] == vendor].reset_index(drop=True) 
                    # Process vendor_data as needed

                    output_filename = os.path.join(carpeta_historico_day, f'OPEN ORDER EZAIR {vendor} {today}.xlsx')
                    vendor_data.to_excel(output_filename, index=False)
                    
                    wb = load_workbook(output_filename)

                    # Access the active sheet
                    sheet = wb.active
                    column_index = 16  # Assuming column P is the 16th column
                    for row in sheet.iter_rows(min_row=2, min_col=column_index, max_col=column_index):
                        for cell in row:
                            if isinstance(cell.value, datetime):
                                cell.number_format = 'MM/DD/YYYY'
                            elif isinstance(cell.value, str):
                                try:
                                    date_value = datetime.strptime(cell.value, '%Y-%m-%d %H:%M:%S')
                                    cell.value = date_value  # Convert the string to datetime
                                    cell.number_format = 'MM/DD/YYYY'
                                except ValueError:
                                    pass  # Skip cells that cannot be converted to datetime
                                                                
                    column_index = 17  # Assuming column P is the 16th column
                    for row in sheet.iter_rows(min_row=2, min_col=column_index, max_col=column_index):
                        for cell in row:
                            if isinstance(cell.value, datetime):
                                cell.number_format = 'MM/DD/YYYY'
                            elif isinstance(cell.value, str):
                                try:
                                    date_value = datetime.strptime(cell.value, '%Y-%m-%d %H:%M:%S')
                                    cell.value = date_value  # Convert the string to datetime
                                    cell.number_format = 'MM/DD/YYYY'
                                except ValueError:
                                    pass  # Skip cells that cannot be converted to datetime
                    
                    column_index = 18  # Assuming column P is the 16th column
                    for row in sheet.iter_rows(min_row=2, min_col=column_index, max_col=column_index):
                        for cell in row:
                            if isinstance(cell.value, datetime):
                                cell.number_format = 'MM/DD/YYYY'
                            elif isinstance(cell.value, str):
                                try:
                                    date_value = datetime.strptime(cell.value, '%Y-%m-%d %H:%M:%S')
                                    cell.value = date_value  # Convert the string to datetime
                                    cell.number_format = 'MM/DD/YYYY'
                                except ValueError:
                                    pass  # Skip cells that cannot be converted to datetime
                                
                    column_index = 19  # Ass9ming column P is the 16th column
                    for row in sheet.iter_rows(min_row=2, min_col=column_index, max_col=column_index):
                        for cell in row:
                            if isinstance(cell.value, datetime):
                                cell.number_format = 'MM/DD/YYYY'
                            elif isinstance(cell.value, str):
                                try:
                                    date_value = datetime.strptime(cell.value, '%Y-%m-%d %H:%M:%S')
                                    cell.value = date_value  # Convert the string to datetime
                                    cell.number_format = 'MM/DD/YYYY'
                                except ValueError:
                                    pass  # Skip cells that cannot be converted to datetime
                    
                    column_index = 21  # Ass9ming column P is the 16th column
                    for row in sheet.iter_rows(min_row=2, min_col=column_index, max_col=column_index):
                        for cell in row:
                            if isinstance(cell.value, datetime):
                                cell.number_format = 'MM/DD/YYYY'
                            elif isinstance(cell.value, str):
                                try:
                                    date_value = datetime.strptime(cell.value, '%Y-%m-%d %H:%M:%S')
                                    cell.value = date_value  # Convert the string to datetime
                                    cell.number_format = 'MM/DD/YYYY'
                                except ValueError:
                                    pass  # Skip cells that cannot be converted to datetime
                    # Define colors
                    white_text_color = "F0F2F5"
                    black_value_color = "141414"
                    blue_value_color = "0370FC"
                    ro_action_color = "5142F5"
                    ri_action_color = "B30415"

                    # Loop through the rows and apply formatting based on condition
                    for row in range(2, len(vendor_data) + 2):
                        if sheet[f'T{row}'].value == 'RO':
                            for col in range(20, 22):  # Columns T and U are columns 20 and 21
                                    cell = sheet.cell(row=row, column=col)
                                    cell.fill = PatternFill(start_color=ro_action_color, end_color=ro_action_color, fill_type="solid")
                                    cell.font = Font(color=white_text_color)
                        else:
                            if sheet[f'T{row}'].value == 'RI':
                                for col in range(20, 22):  # Columns T and U are columns 20 and 21
                                    cell = sheet.cell(row=row, column=col)
                                    cell.fill = PatternFill(start_color=ri_action_color, end_color=ri_action_color, fill_type="solid")
                                    cell.font = Font(color=white_text_color)
                    # Loop through cells in the header row and apply formatting
                    for cell in sheet[1]:
                        cell.fill = PatternFill(start_color=black_value_color, end_color=black_value_color, fill_type="solid")
                        cell.font = Font(color=white_text_color)

                    # Loop through specific cells and apply formatting
                    for cell in sheet['X1:Y1'][0]:
                        cell.fill = PatternFill(start_color=blue_value_color, end_color=blue_value_color, fill_type="solid")
                        cell.font = Font(color=white_text_color)

                    # Save the workbook
                    wb.save(output_filename)

                    contact = contact.fillna('')
                    for index, row in contact.iterrows():
                        if row['Vendor'] == vendor:
                            shortName = row['Shortname']
                            destinatarios = [row['Send_1'], row['Send_2']]
                            copiados = [row['Copy_1']]
                            output_filename_contact = os.path.join(carpeta_historico_day, f'OPEN ORDER EZAIR {vendor} {today}.xlsx')

                            pythoncom.CoInitialize()
                            outlook = Dispatch('Outlook.Application')
                            mail = outlook.CreateItem(0)
                            mail.To = ";".join(destinatarios)
                            mail.CC = ";".join(copiados)
                            mail.Subject = f'OPEN ORDER EZAIR {shortName}'
                            mail.Body = mail_value  # Use mail_value as the email body
                            attachment = mail.Attachments.Add(output_filename_contact)
                            mail.Send()
                            del outlook
                            print(f'mail to {vendor} was send')
                        else:
                            pass

            self.running_application_warning_disable()
        else:
            QMessageBox.critical(self, "Email Body", 'Please provide the email subject')
            self.running_application_warning_disable()

    """________________________________________SET LAST UPDATE_______________________________________________________________"""

    def get_all_need_information_table(self):
                
        now = datetime.now()
        today = now.strftime("%m_%d_%Y")
        cpu = os.environ['USERNAME']

        my_db = f"C:\\Users\\{cpu}\\Desktop\\Historico\\{today}\\Database\\database_{cpu}.db"
        my_db_path = f"C:\\Users\\{cpu}\\Desktop\\Historico\\{today}\\Database\\database_{cpu}.db"
        try:
            conn = sqlite3.connect(my_db_path)
            query_oo = "SELECT * FROM open_order"
            
            oo = pd.read_sql_query(query_oo, conn)
            oo['Ord_Q'] = oo['Ord_Q'].astype(float)
            oo['Opn_Q'] = oo['Opn_Q'].astype(float)
            oo['Unit_Price'] = oo['Unit_Price'].astype(float)
            oo['LT'] = oo['LT'].astype(int) 
            
            query_contact = "SELECT * FROM contact_supplier"
            contact = pd.read_sql_query(query_contact, conn)
            conn.close()
            return oo,contact
            
        
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion OH {error}')
        finally:
            pass

    def running_application_warning_enable(self):
        running_logo = "J:\\Departments\\Operations\\Apps\\python3\\Icon\\icons8-man-running-96.png"
        pixmap = QPixmap(running_logo)
        self.label_54.setHidden(False)
        self.label_55.setHidden(False)
        pixmap = QPixmap(running_logo)
        scaled_pixmap = pixmap.scaledToWidth(600)
        self.label_54.setPixmap(scaled_pixmap)
        self.label_54.setScaledContents(True)

    def running_application_warning_disable(self):
        self.label_54.setHidden(True)
        self.label_55.setHidden(True)

    def show_rgb_colors(self):
        try:
            now = datetime.now()
            today = now.strftime("%m_%d_%Y")
            path_conn_cpu = f"C:\\Users\\{cpu}\\Desktop\\Historico\\{today}\\Database\\database_{cpu}.db"
            conn = sqlite3.connect(path_conn_cpu)
            cursor = conn.cursor()
            query = "SELECT * FROM colors_format_am"
            result = cursor.execute(query)
            self.tableWidget_8.clear()
            self.tableWidget_8.setRowCount(0)
            self.tableWidget_8.setColumnCount(len(cursor.description))
            column_names = [description[0] for description in cursor.description]
            self.tableWidget_8.setHorizontalHeaderLabels(column_names)
            for row_number, row_data in enumerate(result):
                self.tableWidget_8.insertRow(row_number)
                for column_number, data in enumerate(row_data):
                    self.tableWidget_8.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
            # Ajustar automáticamente el ancho de las columnas
            self.tableWidget_8.resizeColumnsToContents()
            conn.close()
        except Exception as error:
            print('Es necesario crear el folder historico no lo olvides')
        finally:
            pass

    def registrar_color_action_message(self):
        now = datetime.now()
        today = now.strftime("%m_%d_%Y")
        path_conn_cpu = f"C:\\Users\\{cpu}\\Desktop\\Historico\\{today}\\Database\\database_{cpu}.db"
        
        try:
            if self.lineEdit_8.text() != '' and self.lineEdit_9.text() != '':
                conexion = sqlite3.connect(path_conn_cpu)
                cursor = conexion.cursor()
                query = "INSERT INTO colors_format_am (Action_Message,RGB_Color) VALUES (?,?)"
                cursor.execute(query, (str(self.lineEdit_8.text()).upper(),str(self.lineEdit_9.text()).upper()))
                conexion.commit()
                cursor.close()
                conexion.close()
                QMessageBox.information(self, "Cargar Contenedor ", f"Contenedor creado correctamente")
                self.lineEdit_8.clear()
                self.lineEdit_9.clear()
                self.show_rgb_colors()
            else:
                QMessageBox.critical(self, 'Registro fallido', 'Favor de llenar todos los campos')
        except Exception as error:
            QMessageBox.critical(self, 'Registro fallido', f'Error en el registro el error es:{error}')
        finally:
            pass

    def delete_from_action_message_color(self):    
        now = datetime.now()
        today = now.strftime("%m_%d_%Y")
        path_conn_cpu = f"C:\\Users\\{cpu}\\Desktop\\Historico\\{today}\\Database\\database_{cpu}.db"
        
        try:
            if self.lineEdit_8.text() != '':
                conexion = sqlite3.connect(path_conn_cpu)
                cursor = conexion.cursor()
                value_to_delete = str(self.lineEdit_8.text()).upper()
                delete_query = "DELETE FROM colors_format_am WHERE Action_Message = ?"
                cursor.execute(delete_query, (value_to_delete,))
                conexion.commit()
                conexion.close()
                QMessageBox.information(self, "Eliminacion de ", f"Action Message {self.lineEdit_8.text()} eliminado correctamente")
                self.lineEdit_8.clear()
                self.show_rgb_colors()
            else:
                QMessageBox.critical(self, 'Eliminacion fallida', 'Favor de llenar el campo de Action Message')
        except Exception as error:
            QMessageBox.critical(self, 'Registro fallido', f'Error en el registro el error es:{error}')
        finally:
            pass
        
        """-------------------------------clock-------------------------------"""

    def setupUI(self): #CLOCK
        self.label_7 = QLabel(self)
        self.label_7.setGeometry(10, 10, 180, 80)
        font = QFont("Consolas", 32, QFont.Bold)
        self.label_7.setFont(font)

    def updateTime(self): #CLOCK
        current_time = QTime.currentTime().toString("hh:mm:ss")
        self.label_7.setText(current_time)

    def startTimer(self): #CLOCK
        timer = QTimer(self, interval=1000, timeout=self.updateTime)
        timer.start()

class MainApplication3(QMainWindow, Ui_Form): #PRODUCTION SHORTAGES REPORT
    def __init__(self):
        super().__init__()
        self.startTimer()  #CLOCKUi_Form
        self.setupUi(self)
        icon = QIcon(f'{icon_ezair}')
        self.selected_date = None
        self.setWindowIcon(icon)
        self.setWindowTitle("BUYER TOOLS PRD")
        self.show_last_actualization_table()
        self.calendarWidget.clicked.connect(self.date_selected)
        self.pushButton.clicked.connect(self.running_gemba_shortages)
        self.label_45.setHidden(True)
        self.label_44.setHidden(True)

    def date_selected(self):
        selected_date = self.calendarWidget.selectedDate()
        if not selected_date.isNull():
            self.selected_date = selected_date.toString("yyyy-MM-dd")
        else:
            QMessageBox.critical(self, "Error", "Please select a date.")
    
    def get_where_use_information_table(self):
        try:
            conn = sqlite3.connect(path_conn)
            rework_query = "SELECT * FROM where_use_table_plantype"
            where_use = pd.read_sql_query(rework_query, conn)
            where_use = where_use.fillna('')
            where_use['PlanTp']  = where_use['PlanTp'].str.upper().str.strip()
            where_use['Area']  = where_use['Area'].str.upper().str.strip()
            conn.close()
            return where_use
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion vendor {error}')
        finally:
            conn.close()
    
    def get_oh_information_without_qa(self):
        
        try:
            conn = sqlite3.connect(path_conn)
            query_oh = "SELECT * FROM oh_gemba_shortages"
            query_exception_items = "SELECT * FROM ExceptionItems"
            oh = pd.read_sql_query(query_oh, conn)
            exception_items = pd.read_sql_query(query_exception_items, conn)
            oh['OH'] = oh['OH'].astype(float)
            oh.loc[:, 'ItemNo'] = oh['ItemNo'].str.upper()
            exception_items.loc[:, 'ItemNo'] = exception_items['ItemNo'].str.upper()
            oh = pd.merge(oh,exception_items,how='left',on = 'ItemNo')
            oh = oh.fillna('')
            def replace_items(DataFrame):
                for index,row in oh.iterrows():
                    if row['Replace'] !='':
                        DataFrame.at[index,'ItemNo'] = DataFrame.at[index,'Replace']
                    else:
                        pass
            replace_items(oh)

            piv_oh = pd.pivot_table(oh,index = 'ItemNo',values = 'OH',aggfunc = 'sum')
            piv_oh = piv_oh.reset_index()
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion OH {error}')
        finally:
            conn.close()
        return piv_oh
    
    def get_expedite_information_for_gemba_shortages(self):
        try:
            conn = sqlite3.connect(path_conn)
            query_expedite = "SELECT * FROM expedite"
            expedite = pd.read_sql_query(query_expedite, conn)
            expedite.loc[:, 'EntityGroup'] = expedite['EntityGroup'].str.upper()
            expedite.loc[:, 'PlanTp'] = expedite['PlanTp'].str.upper()
            expedite.loc[:, 'MLIKCode'] = expedite['MLIKCode'].str.upper()
            expedite.loc[:, 'AC'] = expedite['AC'].str.upper()
            expedite['llave'] = expedite['EntityGroup'] + expedite['AC']  
            expedite['left'] = expedite['DemandSource'].str[:4]
            expedite['left'] = expedite['left'].str.upper()
            expedite['LT'] = expedite['LT'].astype(float)
            expedite = expedite.loc[(expedite['PlanTp'] !='VMIHDW')]
            expedite = expedite.loc[(expedite['MLIKCode'] =='L')]
            expedite = expedite.loc[(expedite['Ref'] !='Safety Stock')]
            expedite = expedite.loc[(expedite['left'] !='FCST')]
            expedite = expedite.fillna('')
            expedite['MLIKCode'] = expedite['MLIKCode'].str.upper()
            expedite[['Vendor', 'ItemNo']] = expedite[['Vendor', 'ItemNo']].apply(lambda x: x.str.upper().str.strip())
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion Expedite{error}')
        finally:
            conn.close()
        return expedite
    
    def get_in92_information(self):
        try:
            conn = sqlite3.connect(path_conn)
            query_in92 = "SELECT * FROM in92"
            in_9_2 = pd.read_sql_query(query_in92, conn)
            in_9_2 = in_9_2.fillna('')
            in_9_2 = in_9_2[['ItemNo','Vendor','MLIKCode']]
            in_9_2[['ItemNo','Vendor','MLIKCode']] = in_9_2[['ItemNo','Vendor','MLIKCode']].apply(lambda x: x.str.upper().str.strip())
            in_9_2 = in_9_2.fillna('')
            in_9_2 = in_9_2.loc[(in_9_2['MLIKCode']=='L')]
            in_9_2 = in_9_2.loc[(in_9_2['Vendor']!='')]
            in_9_2 = in_9_2.drop_duplicates()
            return in_9_2
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion in92 {error}')
        finally:
            conn.close()
    
    def get_entity_information(self):
        try:
            conn = sqlite3.connect(path_conn)
            query_entity = "SELECT * FROM entity"
            entity = pd.read_sql_query(query_entity, conn)
            entity = entity.fillna('')
            entity.loc[:, 'EntityGroup'] = entity['EntityGroup'].str.upper()
            entity = entity[['EntityGroup','EntityName','Commodity']]
            entity = entity[['EntityGroup','EntityName','Commodity']].apply(lambda x: x.str.upper().str.strip())
            entity = entity.drop_duplicates()
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion entity {error}')
        finally:
            conn.close()
        return entity
    
    def get_vendor_information(self):
        try:
            conn = sqlite3.connect(path_conn)
            query_vendor = "SELECT * FROM Vendor"
            buyer = pd.read_sql_query(query_vendor, conn)
            buyer = buyer.fillna('')
            buyer['Vendor']  = buyer['Vendor'].str.upper().str.strip()
            buyer['Shortname']  = buyer['Shortname'].str.upper().str.strip()
            buyer = buyer[['Vendor','Shortname','Tactical']]
            buyer = buyer[['Vendor','Shortname','Tactical']].apply(lambda x: x.str.upper().str.strip())
            buyer = buyer.drop_duplicates()
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion vendor {error}')
        finally:
            conn.close()
        return buyer
    
    def get_entity_project_information(self):
        try:
            conn = sqlite3.connect(path_conn)
            query_entity_project = "SELECT * FROM entity_project"
            project = pd.read_sql_query(query_entity_project, conn)
            project = project.fillna('')
            project['EntityGroup']  = project['EntityGroup'].str.upper().str.strip()
            project['Project']  = project['Project'].str.upper().str.strip()
            project = project.drop_duplicates()
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion vendor {error}')
        finally:
            conn.close()
        return project
    
    def get_LeanDNA_information(self):
        try:
            conn = sqlite3.connect(path_conn)
            query_actions = "SELECT * FROM LeanDNA"
            leanDNA = pd.read_sql_query(query_actions, conn)
        except Exception as e:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion al acction {e}')
        finally:
            conn.close()
        return leanDNA
    
    def get_oh_qa_information(self):
        try:
            conn = sqlite3.connect(path_conn)
            query_qa_oh = "SELECT * FROM qa_oh"
            qa_oh = pd.read_sql_query(query_qa_oh, conn)
        except Exception as e:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer coneccion al acction {e}')
        finally:
            conn.close()
        return qa_oh

    def show_last_actualization_table(self):
        try:
            conn = sqlite3.connect(path_conn)
            cursor = conn.cursor()
            query = "SELECT * FROM actualizations"
            result = cursor.execute(query)
            
            self.tableWidget.clear()
            self.tableWidget.setRowCount(0)
            self.tableWidget.setColumnCount(len(cursor.description))
            column_names = [description[0] for description in cursor.description]
            self.tableWidget.setHorizontalHeaderLabels(column_names)
            for row_number, row_data in enumerate(result):
                self.tableWidget.insertRow(row_number)
                for column_number, data in enumerate(row_data):
                    self.tableWidget.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
            
            # Ajustar automáticamente el ancho de las columnas
            self.tableWidget.resizeColumnsToContents()
        except Exception as error:
            QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer conexión {error}')
        finally:
            conn.close()
    """________________________________________CROSS INFORMATION REPORTS________________________________________________________"""

    def running_gemba_shortages(self):
        if self.selected_date != None :
            
            self.label_42.setHidden(False)
            self.label_44.setHidden(False)
            now = datetime.now()
            today = now.strftime("%m/%d/%Y %I:%M:%S")
            self.label_44.setText(f'{today}')
            QApplication.processEvents()
            
            
            where_use = self.get_where_use_information_table()
            piv_oh = self.get_oh_information_without_qa() #aqui se modifico para que solo tome material fuera de QA
            expedite = self.get_expedite_information_for_gemba_shortages()
            in_9_2 = self.get_in92_information()
            entity = self.get_entity_information()
            buyer = self.get_vendor_information()
            project = self.get_entity_project_information()
            leanDNA = self.get_LeanDNA_information()
            qa_oh = self.get_oh_qa_information()
            qa_oh['OH'] = qa_oh['OH'].astype(float) 


            leanDNA = leanDNA[['Item_Code','Garden_Grove_C37_Inventory','Bolsa_C91_Inventory','Tijuana_C44_Inventory','Santa_Maria_C61_Inventory','Montreal_CG1_Inventory']]
            new_base = pd.merge(expedite,in_9_2,how = 'left',on = 'ItemNo').fillna('')
            new_base['NEW_VENDOR'] = new_base.apply(lambda row: row['Vendor_y'] if row['Vendor_y'] != '' else row['Vendor_x'], axis=1)
            new_base = new_base.rename(columns={'NEW_VENDOR':'Vendor','MLIKCode_x':'MLIKCode'})
            expedite = new_base
            
            
            expedite['ItemNo'] = expedite['ItemNo'].astype(str)
            expedite['ItemNo'] = expedite['ItemNo'].str.upper()
            
            expedite = expedite[['EntityGroup', 'Project', 'AC', 'ItemNo', 'Description', 'PlanTp','Ref'
                                , 'FillDoc', 'Sort', 'ReqQty', 'DemandSource', 'Unit', 'ReqDate', 'ShipDate'
                                , 'MLIKCode', 'LT', 'STDCost', 'LotSize','llave', 'left', 'Vendor']]

            qa_oh_piv = pd.pivot_table(qa_oh,index = 'ItemNo',values = 'OH',aggfunc='sum')
            qa_oh_piv = qa_oh_piv.reset_index()
            qa_oh_piv = qa_oh_piv.rename(columns={'OH':'QA loc'})

            exp_oh = pd.merge(expedite,piv_oh,how = 'left', on = 'ItemNo')
            exp_oh['OH'] = exp_oh['OH'].fillna(0)
            filtered_df = exp_oh.loc[pd.to_datetime(exp_oh['ReqDate']) <= self.selected_date]
            filtered_df.loc[:, 'EntityGroup'] = filtered_df['EntityGroup'].str.upper()
            filtered_df_entity = pd.merge(filtered_df,entity,how = 'left',on = 'EntityGroup')
            filtered_df_entity['ItemNo'] = filtered_df_entity['ItemNo'].str.upper() 
            filtered_df_entity['EntityName'] = filtered_df_entity['EntityName'].fillna('Agregar Entity') 
            df = filtered_df_entity
            df['ItemNo'] = df['ItemNo'].str.upper() 
            df = df.sort_values(['ItemNo', 'ReqDate'], ascending=[True, True])
            df['OH'] = df['OH'].astype(float)
            df['ReqQty'] = df['ReqQty'].astype(float)
            prev_part_num = None
            prev_balance = 0
            for index, row in df.iterrows():
                curr_part_num = row['ItemNo']
                req_qty = row['ReqQty']
                avail_material = row['OH']
                if prev_part_num == curr_part_num:
                    balance = prev_balance - req_qty
                    df.at[index, 'Balance'] = balance
                    prev_balance = balance
                else:
                    balance = avail_material - req_qty
                    df.at[index, 'Balance'] = balance
                    prev_part_num = curr_part_num
                    prev_balance = balance

            df['Balance'] = df['Balance'].astype(float)
            df['cob'] = df['Balance'].apply(lambda x: 'shortage' if x < 0 else 'coverage by oh')

            df_buyer = pd.merge(df,buyer,how = 'left',on = 'Vendor')
            dupCodes = df_buyer
            dupCodes = dupCodes[['ItemNo','Vendor']]
            dupCodes = dupCodes.drop_duplicates()

            pivdupCodes = pd.pivot_table(dupCodes,index = 'ItemNo',values = 'Vendor',aggfunc = 'count')
            pivdupCodes = pivdupCodes.reset_index()
            pivdupCodes = pivdupCodes.loc[pivdupCodes['Vendor'] != 1]
            df_buyer['Vendor'] = df_buyer['Vendor'].replace('', 'item without VendorCode')
            df_buyer = df_buyer.fillna('')

            def consolidate_name(DataFrame):
                df_buyer['Vendor_name'] = ''
                for index,row in df_buyer.iterrows():
                    if row["Vendor"] == 'item without VendorCode':
                        DataFrame.at[index, "Vendor_name"] = 'item without VendorCode'
                    else:
                        if row["Vendor"] != 'item without VendorCode':
                            DataFrame.at[index, "Vendor_name"]  = row['Shortname']
                        else:
                            if row['Shortname'] == '':
                                DataFrame.at[index, "Vendor_name"] = 'add vendor to bc_File'
            consolidate_name(df_buyer)

            df_buyer = df_buyer.fillna('')
            df_buyer['Vendor_name'] = df_buyer['Vendor_name'].replace('', 'add vendor to bc_File')
            df_buyer['Tactical'] = df_buyer['Tactical'].replace('', 'without owner')
            """Concatenation file"""
            
            try:
                conn = sqlite3.connect(path_conn)
                cursor = conn.cursor()
                query = "SELECT value FROM paths WHERE path = ?"
                selected_path = ('BuyerInformation',)  # Put the value in a tuple
                cursor.execute(query, selected_path)
                result = cursor.fetchone()  # Assuming you expect one row as result
                
                if result is not None:
                    path = result[0]  # Assuming the value you want is in the first column
                else:
                    QMessageBox.critical(self, "Error en Path de comentarios", f'No se pudo encontrar el path de los comentarios{error}')
                
            except Exception as error:
                QMessageBox.critical(self, "Error en base de datos", f'No se pudo establecer conección {error}')
            finally:
                conn.close()
            dfs = []
            # path = path.replace("\\", "\\\\")
            print(path)
            
            
            try:
                for file in os.listdir(path):
                    if file.endswith('.xlsx') or file.endswith('.xls'):
                        df = pd.read_excel(os.path.join(path, file))
                        dfs.append(df)
                df_concat = pd.concat(dfs)
                df_concat1 = df_concat
            except Exception as error:
                QMessageBox.critical(self, "Error", f'No se pudo tomar los cometarios el error es: {error}')
            finally:
                pass
            
            df_concat['ItemNo'] = df_concat['ItemNo'].astype(str)
            df_concat['ItemNo'] = df_concat['ItemNo'].str.upper()
            df_concat['ItemNo'] = df_concat['ItemNo'].str.replace('4500611','04500611')
            
            df_concat['ItemNo'] = df_concat['ItemNo'].str.upper().str.strip()
            df_concat = df_concat[['ItemNo','Shortage date','Root Cause','Comments','ETA']]
            df_concat = df_concat.fillna('')    
            df_buyer_concat = pd.merge(df_buyer,df_concat,how = 'left', on = 'ItemNo')

            gemba = df_buyer_concat

            gemba = gemba.loc[gemba['cob'] != 'coverage by oh']
            gemba['ItemNo'] = gemba['ItemNo'].str.upper() 
            gemba['impacts'] = gemba.groupby('ItemNo')['EntityName'].transform(lambda x: ', '.join(x))
            gemba = gemba[['Tactical','ItemNo','Description','LT','Vendor_name','impacts','cob']].drop_duplicates()
            gemba['impacts'] = gemba['impacts'].str.split(', ').apply(set).str.join(', ') # type: ignore

            df_buyer_concat_project = df_buyer_concat
            df_buyer_concat_project = df_buyer_concat_project.loc[(df_buyer_concat_project['cob'] != 'coverage by oh')]
            df_buyer_concat_project['ItemNo'] = df_buyer_concat_project['ItemNo'].str.upper()

            base2 = pd.merge(df_buyer_concat_project,project,how = 'left',on = 'EntityGroup')
            base2_piv = pd.pivot_table(base2,values = 'ReqQty',aggfunc = 'sum',index = 'ItemNo',columns = 'Project_y')
            base2_piv = base2_piv.reset_index()
            base2_piv = base2_piv.fillna(0)
            base2_piv['total_qty'] = base2_piv['CD-NOTE'] + base2_piv['OEM E1'] + base2_piv['OEM E2'] + base2_piv['SPR']

            gemba['ItemNo'] = gemba['ItemNo'].str.upper().str.upper()
            gemba_proj = pd.merge(gemba,base2_piv,how = 'left',on = 'ItemNo')

            datereqProject = df_buyer_concat
            datereqProject = datereqProject.loc[(datereqProject['cob'] != 'coverage by oh')]
            datereqProject = datereqProject[['EntityGroup','ReqDate','ShipDate','ItemNo','ReqQty']]
            datereqProject1 = pd.merge(datereqProject,project,how = 'left',left_on = 'EntityGroup',right_on = 'EntityGroup')
            datereqProject1 = datereqProject1[['EntityGroup','ReqDate','ShipDate','ItemNo','Project']]
            datereqProject1['ItemNo'] = datereqProject1['ItemNo'].str.upper()
            
            min_date_rows = datereqProject1.groupby(['Project', 'ItemNo'])['ReqDate'].agg(min).reset_index()
            min_date_rows = min_date_rows[['Project','ItemNo','ReqDate']]
            result = pd.merge(min_date_rows, datereqProject1, on=['Project', 'ItemNo', 'ReqDate'], how='inner')
            result['ItemNo'] = result['ItemNo'].str.upper()
            result = result.drop_duplicates()

            resultpiv = pd.pivot_table(result,columns = 'Project',values = 'ReqDate',aggfunc = 'min',index = 'ItemNo')
            resultpiv = resultpiv.reset_index()
            resultpiv = resultpiv.fillna('')

            resultpiv['ItemNo'] = resultpiv['ItemNo'].astype(str)
            resultpiv['ItemNo'] = resultpiv['ItemNo'].str.upper()
            gemba_proj_datesproj= pd.merge(gemba_proj,resultpiv,how = 'left',on = 'ItemNo')

            date_columns = ['CD-NOTE_y', 'OEM E1_y', 'OEM E2_y', 'SPR_y']
            gemba_proj_datesproj[date_columns] = gemba_proj_datesproj[date_columns].apply(pd.to_datetime, errors='coerce')
            gemba_proj_datesproj['min_date'] = gemba_proj_datesproj[date_columns].min(axis=1)

            gemba_proj_datesproj['CD-NOTE_y'] = gemba_proj_datesproj['CD-NOTE_y'].apply(lambda x: x.strftime('%m/%d/%Y') if pd.notnull(x) else '')
            gemba_proj_datesproj['OEM E1_y'] = gemba_proj_datesproj['OEM E1_y'].apply(lambda x: x.strftime('%m/%d/%Y') if pd.notnull(x) else '')
            gemba_proj_datesproj['OEM E2_y'] = gemba_proj_datesproj['OEM E2_y'].apply(lambda x: x.strftime('%m/%d/%Y') if pd.notnull(x) else '')
            gemba_proj_datesproj['SPR_y'] = gemba_proj_datesproj['SPR_y'].apply(lambda x: x.strftime('%m/%d/%Y') if pd.notnull(x) else '')
            gemba_proj_datesproj['min_date'] = gemba_proj_datesproj['min_date'].apply(lambda x: x.strftime('%m/%d/%Y') if pd.notnull(x) else '')
            datestring = datetime.now().strftime('%m/%d/%Y')# type: ignore

            final = pd.merge(gemba_proj_datesproj,df_concat,how='left',on = 'ItemNo')
            final = final.fillna('')
            
            print(final.columns)
            
            final['Shortage date'] = final['Shortage date'].replace('',f'New! {datestring}')

            def format_eta(val):
                if pd.isnull(val) or val == "":
                    return val
                else:
                    try:
                        return pd.to_datetime(val).strftime("%m/%d/%Y")
                    except ValueError:
                        return val
            final["ETA"] = final["ETA"].apply(lambda x: format_eta(x))
            
            final1 = pd.merge(final,leanDNA,how = 'left',left_on = 'ItemNo',right_on = 'Item_Code')
            final1 = final1.rename(columns={'Garden_Grove_C37_Inventory':'Garden_Grove_OH','Bolsa_C91_Inventory':'Bolsa_OH','Tijuana_C44_Inventory':'Tijuana_OH'
                                            ,'Santa_Maria_C61_Inventory':'Santa_Maria_OH','Montreal_CG1_Inventory':'Montreal_OH'})
            final2 = pd.merge(final1,qa_oh_piv,how = 'left',on = 'ItemNo')
            try:
                dfs = []
                path = path.replace("\\", "\\\\")
                for file in os.listdir(path):
                    if file.endswith('.xlsx') or file.endswith('.xls'):
                        df = pd.read_excel(os.path.join(path, file))
                        dfs.append(df)
                
                comments_df = pd.concat(dfs)
                comments_df['ItemNo'] = comments_df['ItemNo'].str.upper().str.strip()
                comments_df = comments_df[['ItemNo','FLAG','Priority 1 Qty']]
            except Exception as error:
                QMessageBox.critical(self, "Error", f'No se pudo tomar los cometarios el error es: {error}')
            finally:
                pass

            final3 = pd.merge(final2,comments_df,how = 'left',on = 'ItemNo')
            
            # final3['Point of use'] = ''
            final3['TOTAL IMPO'] = ''
            final3['Impo Lunes'] = ''
            final3['Impo Martes'] = ''
            final3['Impo Miercoles'] = ''
            final3['Impo Jueves'] = ''
            final3['Impo Viernes'] = ''
            final3['Impo Sabado'] = ''
            
            final3 = final3.rename(columns = {'Status':'Shortage date','impacts':'Commodity'})
            
            final3 = final3[['ItemNo','Description','Vendor_name','Shortage date','Tactical','Commodity','ETA'
                            ,'Priority 1 Qty','total_qty','Root Cause','Comments','FLAG','QA loc'
                            ,'TOTAL IMPO','Impo Lunes','Impo Martes','Impo Miercoles','Impo Jueves','Impo Viernes','Impo Sabado'
                            ,'Garden_Grove_OH','Bolsa_OH','Tijuana_OH','Santa_Maria_OH','Montreal_OH']]

            all_materials_plantype = df_buyer_concat.copy()
            

            all_materials_plantype = all_materials_plantype[['ItemNo','PlanTp']]
            all_materials_plantype['ItemNo'] = all_materials_plantype['ItemNo'].str.upper().str.strip()
            all_materials_plantype['PlanTp'] = all_materials_plantype['PlanTp'].str.upper().str.strip()
            all_materials_plantype = all_materials_plantype.drop_duplicates(subset='ItemNo', keep='first')
            
            base1 = pd.merge(all_materials_plantype,where_use,how ='left', on= 'PlanTp')
            base2 = pd.merge(final3,base1,how ='left', on= 'ItemNo')
            base2 = base2.rename(columns = {'Area':'Point of use'})
            base2 = base2[['ItemNo', 'Description', 'Vendor_name', 'Shortage date', 'Tactical','Commodity','Point of use', 'ETA', 'Priority 1 Qty', 'total_qty', 'Root Cause',
                        'Comments', 'FLAG', 'QA loc', 'TOTAL IMPO', 'Impo Lunes', 'Impo Martes','Impo Miercoles', 'Impo Jueves'
                        , 'Impo Viernes', 'Impo Sabado','Garden_Grove_OH', 'Bolsa_OH', 'Tijuana_OH', 'Santa_Maria_OH','Montreal_OH']]
            
            wb = xw.Book()
            sheet1 = wb.sheets.add()
            sheet1.name = "GembaShortages"
            sheet1.range('A1').value = base2
            column_letter = 'A'
            column_range = sheet1.range(f'{column_letter}:{column_letter}')
            column_range.delete()
            sheet2 = wb.sheets.add()
            sheet2.name = "allmaterials"
            sheet2.range('A1').value = df_buyer_concat
            sheet3 = wb.sheets.add()
            sheet3.name = "checkPN"
            sheet3.range('A1').value = pivdupCodes            
            white_text_color = (240, 242, 245)
            blue_value = (3, 46, 84)
            yellow_value = (245, 241, 10)
            red_value_text = (240, 15, 7)
            green_value = (7, 171, 51)
            orange_value = (230, 130, 0)
            black_value = (20, 20, 20)
            red_value = (207, 23, 2)
            

            for cell in sheet1.range('A1:G1'):
                cell.color = blue_value
                cell.api.Font.Color = xw.utils.rgb_to_int(white_text_color)
            
            for cell in sheet1.range('J1'):
                cell.color = blue_value
                cell.api.Font.Color = xw.utils.rgb_to_int(white_text_color)
            
            for cell in sheet1.range('N1'):
                cell.color = blue_value
                cell.api.Font.Color = xw.utils.rgb_to_int(white_text_color)
            
            for cell in sheet1.range('I1'):
                cell.color = yellow_value
                cell.api.Font.Color = xw.utils.rgb_to_int(red_value_text)
        
            for cell in sheet1.range('H1'):
                cell.color = green_value
                cell.api.Font.Color = xw.utils.rgb_to_int(white_text_color)

            for cell in sheet1.range('K1:L1'):
                cell.color = green_value
                cell.api.Font.Color = xw.utils.rgb_to_int(white_text_color)

            for cell in sheet1.range('P1:U1'):
                cell.color = yellow_value
                cell.api.Font.Color = xw.utils.rgb_to_int(red_value_text)
            
            for cell in sheet1.range('V1:Z1'):
                cell.color = orange_value
                cell.api.Font.Color = xw.utils.rgb_to_int(black_value)
            
            for cell in sheet1.range('M1'):
                cell.color = red_value
                cell.api.Font.Color = xw.utils.rgb_to_int(white_text_color)   
                
            for cell in sheet1.range('O1'):
                cell.color = orange_value
                cell.api.Font.Color = xw.utils.rgb_to_int(red_value_text)  

            datestringsave = datetime.now().strftime('%m_%d')# type: ignore
            writer = pd.ExcelWriter(f'C:\\Users\\{cpu}\\Desktop\\EZA Shortages {datestringsave}.xlsx')            
            base2.to_excel(writer, sheet_name='GembaShortages', index=False)
            df_buyer_concat.to_excel(writer, sheet_name='allmaterials', index=False)
            pivdupCodes.to_excel(writer, sheet_name='checkPN', index=False)
            writer.save()

            self.label_43.setHidden(False)
            self.label_45.setHidden(False)
            now = datetime.now()
            today = now.strftime("%m/%d/%Y %I:%M:%S")
            self.label_45.setText(f'{today}')

            QMessageBox.information(self, "Reporte guardado", "Valida el archivo guardado en tu escritorio")
        else:
            QMessageBox.critical(self, "Error", "Selecciona el rango de covertura del reporte asi como la confirmacion de la actualizacion de la Base")

    """-------------------------------clock-------------------------------"""
    def setupUI(self): #CLOCK
        self.label_7 = QLabel(self)
        self.label_7.setGeometry(10, 10, 180, 80)
        font = QFont("Consolas", 32, QFont.Bold)
        self.label_7.setFont(font)

    def updateTime(self): #CLOCK
        current_time = QTime.currentTime().toString("hh:mm:ss")
        self.label_7.setText(current_time)

    def startTimer(self): #CLOCK
        timer = QTimer(self, interval=1000, timeout=self.updateTime)
        timer.start()

if __name__ == "__main__":
    app = QApplication(sys.argv)
    login_window = LoginWindow()  # Create an instance of the LoginWindow class
    login_window.show()  # Show the login window
    sys.exit(app.exec_())  # Start the application event loop

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainApplication()
    window.show()
    sys.exit(app.exec_())

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainApplication2()
    window.show()
    sys.exit(app.exec_())

if __name__ == "__main__":
    import sys
    app = QApplication(sys.argv)
    main_app3 = MainApplication3()
    main_app3.show()
